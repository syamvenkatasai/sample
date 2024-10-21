#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Thu Jul 6 15:02:18 2022

@author: Gopi Teja B and Venkatesh Madasu

Note: This Folder Monitor is having specific style designed for Medusind Purpose
"""
from datetime import date
import os
import requests
import shutil
import uuid
import re
import subprocess
from reportlab.pdfgen import canvas
from PIL import Image
import os
import numpy as np
import psutil
import xxhash
import json
import ast
import glob
import base64
from PIL import Image
import pandas as pd 
from flask_cors import CORS
from flask import Flask, jsonify, request
from pathlib import Path
from jinja2 import Template
import openpyxl
from openpyxl.styles import NamedStyle
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
# from producer import produce
from db_utils import DB
from time import time as tt
from py_zipkin.zipkin import zipkin_span, ZipkinAttrs, create_http_headers_for_new_span
from py_zipkin.util import generate_random_64bit_string
from pdf2image import convert_from_path
from app import app
from ace_logger import Logging
from datetime import datetime
logging = Logging(name='folder_monitor')

db_config = {
    'host': os.environ['HOST_IP'],
    'user': os.environ['LOCAL_DB_USER'],
    'password': os.environ['LOCAL_DB_PASSWORD'],
    'port': os.environ['LOCAL_DB_PORT']
}


def http_transport(encoded_span):
    body = encoded_span
    requests.post(
        'http://servicebridge:80/zipkin',
        data=body,
        headers={'Content-Type': 'application/x-thrift'},
    )

def measure_memory_usage():
    process = psutil.Process()
    memory_info = process.memory_info()
    return memory_info.rss  # Resident Set Size (RSS) in bytes

def insert_into_audit(case_id, data):
    tenant_id = data.pop('tenant_id')
    db_config['tenant_id'] = tenant_id
    stats_db = DB('stats', **db_config)
    stats_db.insert_dict(data, 'audit')
    return True

def generate_caseid_serial(tenant_id):
    db_config['tenant_id'] = tenant_id
    db = DB('queues', **db_config)

    logging.info('Generating new case ID')
    query = "SELECT `id`,`case_id` FROM `process_queue` ORDER BY `process_queue`.`id` DESC LIMIT 1"

    try:
        existing_case_ids = list(db.execute(query)['case_id'])[0]
        new_case_id = str(int(existing_case_ids) + 1)
    except Exception:
        new_case_id = '100000000'

    logging.info(f'New case ID: {new_case_id}')
    return new_case_id


def generate_caseid(tenant_id, case_type, table_name, db_name):
    db_config['tenant_id'] = tenant_id
    db = DB(f'{db_name}', **db_config)
    logging.info(f"Generating new case ID")
    query = f"SELECT `{case_type}` FROM `{table_name}` where {case_type} like '%%{tenant_id[:3].upper()}%%' ORDER BY `{table_name}`.`{case_type}`"
    existing_case_ids = list(db.execute_(query)[{case_type}])

    new_case_id = tenant_id[:3].upper() + uuid.uuid4().hex.upper()[:7]
    # or condition is added bcz caseid is being created with PO starting
    while new_case_id in existing_case_ids:
        logging.debug(f'`{new_case_id}` already exists. Generating a new one.')
        new_case_id = tenant_id[:3].upper() + uuid.uuid4().hex.upper()[:7]

    logging.info(f'New case ID: {new_case_id}')
    return new_case_id


def file_hash_func(file_path):
    try:
        x = xxhash.xxh64()
        with open(file_path, 'rb') as f:
            file_data = f.read()
        x.update(file_data)

        file_hash = x.hexdigest()
    except Exception:
        file_hash = ''

    return file_hash


def convert_pdf_jpg(file_path, output_path):
    jpg_list = []
    file_full_name = file_path.name
    file_ext = file_path.suffix.lower()
    logging.debug(f"## FM Checkpoint File name: {file_full_name}")
    file_page_object = {}
    try:
        if file_ext == '.pdf':
            with fitz.open(file_path) as new_file:
                # with fitz.open(file_path) as pages:
                #     new_file.insertPDF(pages, to_page=7)
                # new_file.save(file_path)

                for idx, page in enumerate(new_file):
                    page_image_name = f'{file_path.stem}_{str(idx)}.jpg'
                    logging.info(f"page image name {page_image_name}")
                    page_image_path = (f'{output_path}/images/'
                                       + page_image_name)
                    logging.info(f"page image path {page_image_path}")
                    logging.debug(f'Page {str(idx)}: {page_image_path}')
                    mat = fitz.Matrix(300 / 72, 300 / 72)
                    if isinstance(page, fitz.fitz.Page):
                        try:
                            logging.info(f"in try if block")
                            pix = page.getPixmap(matrix=mat)
                            img = Image.frombytes("RGB", (pix.width, pix.height), pix.samples)
                            img.save(page_image_path, format="JPEG")
                        except:
                            logging.info(f"in except if block")
                            pix = page.getPixmap()
                            pix.pil_save(page_image_path)

                    else:
                        try:
                            logging.info(f"in try else block")
                            pix = page.getPixmap(matrix=mat)
                            img = Image.frombytes("RGB", (pix.width, pix.height), pix.samples)
                            img.save(page_image_path, format="JPEG")
                        except:
                            logging.info(f"in except else block")
                            pix = page.getPixmap(matrix=mat)
                            pix.pil_save(page_image_path, 'JPEG')

                    jpg_list.append(page_image_name)
                file_page_object = {file_full_name: jpg_list}
                logging.info(f"file_page_object: {file_page_object}")
    except Exception as e:
        logging.error(f'PDF to image conversion failed with pymupdf: {e}')
        logging.info(f"convertion  pdf to jpg failed {e}")
        # return convert_pdf_jpg_old(file_path, output_path, unique_id)
    return file_page_object


def convert_pdf_jpg_old(file_path, output_path, unique_id):
    jpg_list = []
    file_full_name = file_path.name
    file_ext = file_path.suffix.lower()
    file_page_object = {}
    try:
        if file_ext == '.pdf':

            for idx, page in enumerate(convert_from_path(file_path, 300)):
                page_image_name = f'{file_path.stem}_{str(idx)}.jpg'
                page_image_path = (f'{output_path}/{unique_id}/images/'
                                   + page_image_name)
                logging.debug(f'Page {str(idx)}: {page_image_path}')
                page.save(page_image_path, 'JPEG')
                jpg_list.append(page_image_name)
            file_page_object = {file_full_name: jpg_list}
            logging.info(f"file_page_object: {file_page_object}")
    except Exception as e:
        logging.error(f'PDF to image conversion old method failed: {e}')
    return file_page_object


@app.route('/create_case_id', methods=['POST', 'GET'])
def create_case_id():
    data = request.json
    logging.debug(f"data received {data}")
    tenant_id = data.get('tenant_id', None)
    file = data.get('file', {})
    file_name = file.get('file_name')
    ace = 'ace'
    user = data.get('user', None)
    session_id = data.get('session_id', None)
    source_of_invoice = data.get('move_from', '')
    variables = data.get("variables", {})

    db_config['tenant_id'] = tenant_id

    try:
        memory_before = measure_memory_usage()
        start_time = tt()
    except:
        logging.warning("Failed to start ram and time calc")
        pass
        
    trace_id = generate_random_64bit_string()
        
    attr = ZipkinAttrs(
        trace_id=trace_id,
        span_id=generate_random_64bit_string(),
        parent_span_id=None,
        flags=None,
        is_sampled=False,
        tenant_id=tenant_id
        )

    with zipkin_span(
        service_name='folder_monitor_api',
        span_name='create_case_id',
        transport_handler=http_transport,
        zipkin_attrs=attr,
        port=5010,
        sample_rate=0.5):
        try:
            # Generating the case id
            unique_id = generate_caseid(tenant_id, 'case_id', 'process_queue', 'queues')

            # Creating case id into Database
            queue_db = DB('queues', **db_config)
            extraction_db = DB('extraction', **db_config)
            stats_db = DB('stats', **db_config)

            insert_case_id = ('INSERT INTO `ocr` (`case_id`,`document_id`,`highlight`) '
                              "VALUES (%s,%s,'{}')")
            params = [unique_id,unique_id]
            extraction_db.execute_(insert_case_id, params=params)

            insert_case_id_1 = ('INSERT INTO `process_file` (`case_id`,`document_id`) '
                              "VALUES (%s,%s)")
            params = [unique_id,unique_id]
            queue_db.execute(insert_case_id_1, params=params)

            #AmBank etrade project specific
             # getting child_id from file name
            # child_id_file=''
            # for ele in file_name:
            #     if ele=="_":
            #         break
            #     child_id_file+=ele
            logging.info(f'filename is {file_name}')
            query=f"UPDATE process_file SET file_name = '{file_name}' WHERE case_id='{unique_id}'"
            queue_db.execute(query)

            reponse_data = {"case_id": unique_id, "file_name": file_name}

            # Creating a folder with case id in htdocs
            output_path = Path(f'/app/output/{tenant_id}/assets/pdf/{tenant_id}')
            os.umask(0)
            logging.info(output_path)
            Path(str(output_path / unique_id)).mkdir(parents=True, exist_ok=True)
            os.chmod(str(output_path / unique_id), 0o777)
            logging.debug("folder created in lampp directory")

            shutil.move(output_path / file_name,
                        output_path / unique_id / file_name)
            # api_params = {"variables": {"case_id": {"value": unique_id},
            #                             "file_name": {"value": upload_file_status["file_name"]}}}
            file_path_stored = Path(str(output_path / unique_id / file_name))
            Path(str(output_path / unique_id / 'images')).mkdir(parents=True,
                                                                exist_ok=True)  # creating a image folder
            # splitting pdf into each page
            # try:
            #    image_object = convert_pdf_jpg(
            #        file_path_stored, output_path, unique_id)
            # except:
            #    image_object = convert_pdf_jpg_old(
            #        file_path_stored, output_path, unique_id)

            out_path = f'/app/output/{tenant_id}/assets/pdf/{tenant_id}/{unique_id}'


            logging.debug(f'this is where the case is updated into process queue')
            insert_case_id = 'INSERT INTO `process_queue` ( `case_id`,`document_id`,`file_name`,`file_paths`,`last_updated_by`,`source_of_invoice`,`completed_processes`,`status`) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)'
            params = [unique_id,unique_id, file_name, str(file_path_stored),"system", source_of_invoice,1,"File Uploaded"]
            queue_db.execute(insert_case_id, params=params)

            try:
                file_object=convert_pdf_jpg(file_path_stored,out_path)
                queue_db.execute("update `process_queue` set `imagelist` = %s where `case_id` = %s", params=[json.dumps(file_object), unique_id])
            except Exception as e:
                logging.info(e)

            logging.debug(f'this is where the case is updated into br_comparison_rules')
            insert_case_id_ = 'INSERT INTO `br_comparison_rules` (`case_id`) VALUES (%s)'
            params = [unique_id]
            extraction_db.execute(insert_case_id_, params=params)


            # Audit
#                 query = 'Insert into `audit` (`type`, `last_modified_by`, `table_name`, `changed_data`,`reference_value`,`reference_column`) values (%s,%s,%s,%s,%s,%s)'
#                 params = ['creation', 'folder_monitor', 'process_queue',
#                           json.dumps({"queue": "New Case was generated"}), unique_id,'case_id']
#                 stats_db.execute(query, params=params)

#                 query2= ('INSERT INTO `process_file` ( `case_id`,`document_id`,`file_name`) '
#                                   'VALUES (%s,%s,%s)')
#                 params = [unique_id, unique_id, file_name]
#                 queue_db.execute(query2, params=params)


            final_response_data = {"flag": True, "data": reponse_data}
    
        except Exception:
            logging.exception('Something went wrong while generating case id. Check trace.')
            final_response_data = {'flag': False, 'message': 'System error! Please contact your system administrator.'}
    try:
        memory_after = measure_memory_usage()
        memory_consumed = (memory_after - memory_before) / \
            (1024 * 1024 * 1024)
        end_time = tt()
        memory_consumed = f"{memory_consumed:.10f}"
        logging.info(f"checkpoint memory_after - {memory_after},memory_consumed - {memory_consumed}, end_time - {end_time}")
        time_consumed = str(round(end_time-start_time,3))
    except:
        logging.warning("Failed to calc end of ram and time")
        logging.exception("ram calc went wrong")
        memory_consumed = None
        time_consumed = None
        pass

    # insert audit
    audit_data = {"tenant_id": tenant_id, "user": user, "case_id": unique_id,
                    "api_service": "create_case_id", "service_container": "folder_monitor_api",
                    "changed_data": json.dumps({"queue": "New Case was generated"}),"tables_involved": "process_queue","memory_usage_gb": str(memory_consumed), 
                    "time_consumed_secs": time_consumed, "request_payload": json.dumps(data), 
                    "response_data": reponse_data, "trace_id": trace_id,
                    "session_id": session_id,"status":str(final_response_data['flag'])}
    insert_into_audit(unique_id, audit_data)

    return jsonify(final_response_data)

@app.route('/folder_monitor', methods=['POST', 'GET'])
def folder_monitor():
    """
    Note: If you are using this API the file ingestion should follow below keypoints
    1. folder_structure table need to be filled in io_configuration db
    2. For shared folder path: inside tenant folder of /app/input priority_folder column declared folders should be created
    3. Inside those create as many folders or structures you want the final folder where the file should be picked shoud have `input` name
    """

    data = request.json
    logging.info(f'Data is: {data}')
    # data = data['data']
    tenant_id = data.get('tenant_id', None)
    
    try:
        memory_before = measure_memory_usage()
        start_time = tt()
    except:
        logging.warning("Failed to start ram and time calc")
        pass
        
    
    trace_id = generate_random_64bit_string()

    attr = ZipkinAttrs(
        trace_id=generate_random_64bit_string(),
        span_id=generate_random_64bit_string(),
        parent_span_id=None,
        flags=None,
        is_sampled=False,
        tenant_id=tenant_id if tenant_id is not None else ''
    )

    with zipkin_span(
            service_name='folder_monitor',
            zipkin_attrs=attr,
            span_name='folder_monitor',
            transport_handler=http_transport,
            sample_rate=0.5
    ) as zipkin_context:
        try:
            logging.debug(f'Connecting to tenant {tenant_id}')

            db_config['tenant_id'] = tenant_id
            db = DB('io_configuration', **db_config)

            # input_config = db.get_all(
            #     'input_configuration', condition={'active': 0})
            input_config_qry = f'select * from `input_configuration`'
            input_config=db.execute(input_config_qry)
            output_config = db.get_all('output_configuration')

            # input_config = db.execute_(input_config).set_index('id')
            # #output_config = db.get_all('output_configuration')
            # output_config = f'select * from `output_configuration`'
            # output_config = db.execute_(output_config).set_index('id')

            logging.debug(f'Input Config: {input_config.to_dict()}')
            logging.debug(f'Output Config: {output_config.to_dict()}')

            # Sanity checks
            if (input_config.loc[input_config['type'] == 'Document'].empty
                    or output_config.loc[output_config['type'] == 'Document'].empty):
                message = 'Input/Output not configured in DB.'
                logging.error(message)
                return jsonify({'flag': False, 'message': message})

            for index, row in input_config.iterrows():
                input_path = row['access_1']
                output_path = output_config.ix[row['output']]['access_1']
                workflow = row['workflow']

                logging.debug(f'Input path: {input_path}')
                logging.debug(f'Output path: {output_path}')

                if (input_path is None or not input_path
                        or output_path is None or not output_path):
                    message = 'Input/Output is empty/none in DB.'
                    logging.error(message)
                    return jsonify({'flag': False, 'message': message})

                input_path_str = "/app/input/"+input_path
                output_path = f"/app/output/{tenant_id}/assets/pdf/"+output_path
                input_path = Path(input_path_str)
                output_path = Path(output_path)
                logging.debug(f'Input Absolute path: {input_path}')
                logging.debug(f'Output Absolute path: {output_path}')

                reponse_data = {}

                # Only watch the folder if both are valid directory
                if input_path.is_dir():
                    logging.debug(str(input_path) + '/*')
                    # files = Path(input_path).rglob('*')
                    all_files = get_priority_files(input_path_str)
                    logging.info(f"All files found are: {all_files}")

                    # get first file and send it to pick as priority need to be upload first
                    if len(all_files):
                        files = [all_files[0]]
                    else:
                        files = []

                    file_names = []
                    for file_ in files:
                        logging.debug(f'move from: {file_}')
                        reponse_data['move_from'] = str(file_)
                        filename = file_.name.replace(' ', '_')
                        logging.debug(
                            f'move to: {str(output_path / filename)}')

                        # creating a copy file in error folder and
                        # will be deleted when queue assigned in camunda_api
                        copy_to = Path(file_).parents[1] / 'error/'
                        logging.debug(f'Creating a copy at {copy_to}')
                        shutil.copy(Path(file_), copy_to)

                        shutil.move(Path(file_), Path(output_path) / filename)

                        # Audit
                        # stats_db = DB('stats', **db_config)
                        # query = 'Insert into `audit` (`type`, `last_modified_by`, `changed_data`,`reference_column`,`reference_value`,`table_name`) values (%s,%s,%s,%s,%s,%s)'
                        # params = ['Receiving', 'folder_monitor',
                        #           json.dumps({"queue": "New file received"}), 'filename', filename, 'File Pickup']
                        # stats_db.execute(query, params=params)

                        file_names.append({'file_name': filename})

                    logging.debug(f'Files: {file_names}')
                    reponse_data['files'] = file_names
                    reponse_data['workflow'] = workflow

                    final_response_data = {"flag": True, "data": reponse_data}
                else:
                    message = f'{input_path} not a directory'
                    logging.error(message)
                    final_response_data = {'flag': True, 'message': message}
        except:
            logging.exception(
                'Something went wrong watching folder. Check trace.')
            final_response_data = {'flag': False, 'message': 'System error! Please contact your system administrator.'}
    try:
        memory_after = measure_memory_usage()
        memory_consumed = (memory_after - memory_before) / \
            (1024 * 1024 * 1024)
        end_time = tt()
        memory_consumed = f"{memory_consumed:.10f}"
        logging.info(f"checkpoint memory_after - {memory_after},memory_consumed - {memory_consumed}, end_time - {end_time}")
        time_consumed = str(round(end_time-start_time,3))
    except:
        logging.warning("Failed to calc end of ram and time")
        logging.exception("ram calc went wrong")
        memory_consumed = None
        time_consumed = None
        pass

    # insert audit
    audit_data = {"tenant_id": tenant_id, "user": "", "case_id": "",
                    "api_service": "folder_monitor", "service_container": "folder_monitor_api",
                    "changed_data": json.dumps({"queue": "New file received"}),"tables_involved": "","memory_usage_gb": str(memory_consumed), 
                    "time_consumed_secs": time_consumed, "request_payload": json.dumps(data), 
                    "response_data": reponse_data, "trace_id": trace_id,
                    "session_id": "","status":str(final_response_data['flag'])}
    insert_into_audit("New File Received", audit_data)

    return jsonify(final_response_data)


def get_folder_paths(rootdir, tmp_list=None):

    if tmp_list is None:
        tmp_list = []
    logging.info(f"#### rootdir: {os.scandir(rootdir)}")
    for it in os.scandir(rootdir):
        if it.is_dir():
            tmp_list.append(it.path)
            logging.info(f"###### tmp_list: {tmp_list}, it: {it}")
            get_folder_paths(it, tmp_list)


def filter_priority_input_folders(priority_folder, all_folder_paths):

    filtered_folders = []

    must_folder = priority_folder["priority_folder"].strip()

    end_folders = priority_folder["end_folders"].strip().split(",")

    for folder_path in all_folder_paths:

        if must_folder in folder_path.split('/'):

            end_folder_from_data_list = folder_path.strip().rsplit("/", 1)
            logging.info(
                f"######## end_folder_from_data_list: {end_folder_from_data_list}")

            if len(end_folder_from_data_list) > 1:

                end_folder_from_data = end_folder_from_data_list[1]

            elif len(end_folder_from_data_list) == 1:

                end_folder_from_data = end_folder_from_data_list[0]

            else:
                end_folder_from_data = ""

            if end_folder_from_data != "" and (end_folder_from_data in end_folders):

                filtered_folders.append(folder_path.strip())

    return filtered_folders


def get_files_from_priority_folders(filtered_folders):

    all_files = []

    for folder_path in filtered_folders:

        logging.info(f" ### Checking files in FOLDER:::::: {folder_path}")
        files = list(Path(folder_path).rglob('*.pdf'))
        logging.info(f"######## files: {files}")

        if len(files):

            logging.info(f"### Found files at path: {folder_path}")

            all_files += files

    logging.info(f"##### ALL Files: {all_files}")
    return all_files


def get_priority_files(base_input_path):

    logging.info(f"base input path : {base_input_path}")
    io_configuration_db = DB('io_configuration', **db_config)
    query = "select * from folder_structure"
    priority_folders = io_configuration_db.execute(
        query).to_dict(orient='records')

    priority_folder_list = sorted(
        priority_folders, key=lambda d: d["priority_order"])

    all_folders_from_base_folder = []

    get_folder_paths(base_input_path, all_folders_from_base_folder)

    logging.info(f"#### priority_folder_list: {priority_folder_list}")

    for priority_folder in priority_folder_list:

        filtered_folders = filter_priority_input_folders(
            priority_folder, all_folders_from_base_folder)

        logging.info(f"#### filtered_folders:{filtered_folders}")

        priority_files = get_files_from_priority_folders(filtered_folders)

        logging.info(f"#### priority_files: {priority_files}")

        if len(priority_files):

            return priority_files
    return []


def generate_blob_data(file_path):
    try:
        logging.info("############ Converting the file to blob")
        file_name = file_path.rsplit("/", 1)[1]
        logging.info(f"generating blob for file {file_name}")
        file_blob = ""
        with open(file_path, 'rb') as f:
            logging.info(f"converting to blob for -----------{file_path}")
            file_blob = base64.b64encode(f.read())

        try:
            logging.debug('########### decoding with utf 8 ')
            return_blob = file_blob.decode('utf-8')
            #returning_data['file_name']= chemical_name + '_cover_letter.docx'

            return True, file_name, return_blob
        except Exception:
            message = 'Something went wrong while downloading report.'
            logging.exception(message)
            return False, "", " Error in decoding with utf 8 "

    except Exception as e:
        logging.warning("########### Error in Generating Blob Data")
        logging.exception(e)
        return False, "", " Error in Generating Blob Data"


def remove_all_except_al_num(file_full_name):
    to_return = re.sub('[,.!@#$%^&*()\-=`~\'";:<>/?]',
                       '', file_full_name.lower())
    to_return = to_return.replace(' ', '')
    return to_return


@app.route("/fetch_file_for_image", methods=['POST', 'GET'])
def fetch_file_for_image():
    data = request.json
    logging.info(f"Request Data for fetch file for image: {data}")
    tenant_id = data.get('tenant_id', None)
    case_id = data.get('case_id', None)
    user = data.get('user', None)
    session_id = data.get('session_id', None)
    db_config["tenant_id"] = tenant_id
    queue_db = DB("queues", **db_config)
    exctraction_db = DB("extraction", **db_config)
    try:
        memory_before = measure_memory_usage()
        start_time = tt()
    except:
        logging.warning("Failed to start ram and time calc")
        pass
    try:
        if case_id is None:
            trace_id = generate_random_64bit_string()
        else:
            trace_id = case_id
        
        attr = ZipkinAttrs(
            trace_id=trace_id,
            span_id=generate_random_64bit_string(),
            parent_span_id=None,
            flags=None,
            is_sampled=False,
            tenant_id=tenant_id
        )

        with zipkin_span(
            service_name='folder_monitor_api',
            span_name='fetch_file_for_image',
            transport_handler=http_transport,
            zipkin_attrs=attr,
            port=5010,
            sample_rate=0.5):
            try:
                variables = data.get("variables", {})
                download_type = variables["download_type"]
                query = f"SELECT `document_id` from  `process_queue` where `case_id` = %s and state IS NULL"
                document_id_df = queue_db.execute_(query, params=[case_id])['document_id'].tolist()
                logging.info(f"document ids that we got is {document_id_df}")
            except:
                return{'flag':False,'message':f'{case_id} is missing in the table'}
            
            file_names=[]
            page_info={}
            main_file_names_=[]
            for document_id in document_id_df:
                #this condition will execute in the case of additional files
                page_info_output_list = []
                query = f"select file_name, single_doc_identifiers from process_queue where document_id='{document_id}' and state IS NULL"
                query_data = queue_db.execute_(query)
                file_name_df=query_data['file_name'].to_list()
                file_name_=file_name_df[0]
                file_names=json.loads(file_name_)
                try:
                    page_info_details = list(query_data['single_doc_identifiers'])[0]
                    page_info_details = ast.literal_eval(page_info_details)
                    if len(page_info_details) > 0:
                        for item in page_info_details:
                            for page in range(item["start_page"], item["end_page"] + 1):
                                title = item["file_type"]
                                if title == '':
                                    title = 'Not Detected'
                                page_info_output_list.append(
                                    {"page": page, "title": title})
                    page_info[main_file_names_]= page_info_output_list
                except:
                    pass
            try:
                additional_file_names=[]
                if 'tab_view' in variables:  
                    query1 = f"select `filemanager_file_name` from file_manager where case_id='{case_id}'"
                    try:
                        query1 = f"select `filemanager_file_name` from file_manager where case_id='{case_id}'"
                        query_data = queue_db.execute_(query1)['filemanager_file_name']
                        additional_file_names=json.loads(query_data[0])
                        file_names+=additional_file_names
                    except:
                        query_data = queue_db.execute_(query1)['filemanager_file_name'].to_list()
                        additional_file_names=(query_data[0])
                        file_names.append(additional_file_names)
                    logging.info(f"additional file_names are {additional_file_names}")
                    
                    # for file_ in additional_file_names:
                    #     page_info[file_]= []
            except:
                pass


            file_path_navigates = {}
            return_data = {}
            additional_file_data={}
            logging.info(f"all files are: {file_names}")
            for file_name in file_names:
                logging.info(f"processing for {file_name}")
                file_path = f'/app/output/{tenant_id}/assets/pdf/{tenant_id}/{case_id}/{file_name}'
                logging.info(f"file_path#####:  {file_path}")

                file_path_navigate = f'assets/pdf/{tenant_id}/{case_id}/{file_name}'
                file_path_navigates[file_name] = file_path_navigate

                logging.info(file_path_navigate)
                logging.info(file_path_navigates)

                if download_type == 'blob':
                    blob_status, file_name, file_blob = generate_blob_data(
                        file_path)

                    if not blob_status:
                        jsonify({"flag": False, "message": file_blob})

                    # return_data["file_name"]=file_name
                    try:
                        if additional_file_names:
                            additional_file_data[file_name] = {"blob": file_blob}
                        else:
                            return_data[file_name] = {"blob": file_blob}
                    except:
                        return_data[file_name] = {"blob": file_blob}

                elif download_type == "images":
                    
                    if file_name in additional_file_names:
                        logging.info(f"entered into with {file_name}")
                        query =f"select `filemanager_imagelist` from `file_manager` where `case_id` ='{case_id}' "
                        
                        try:
                            query_data = queue_db.execute_(query)['filemanager_imagelist'].to_list()
                            images_file_names = json.loads(query_data[0])
                            for file in images_file_names:
                                file=json.loads(file)
                                for key,value in file.items():
                                    if key == file_name:
                                        images_list=value
                        except:
                            query_data = queue_db.execute_(query)['filemanager_imagelist']
                            images_file_names = json.loads(query_data[0])
                            images_file_names =json.loads(images_file_names[0])
                            for key,value in images_file_names.items():
                                if key == file_name:
                                    images_list=value
                    else:
                        query =f"select `imagelist` from `process_queue` where `case_id` ='{case_id}'"
                        image_df=queue_db.execute_(query)
                        try:
                            images=image_df.to_dict(orient="records")
                            if images[0]['imagelist']:
                                images_=json.loads(images[0]['imagelist'])
                                try:
                                    images_list=json.loads(images_)[file_name]
                                except:
                                       images_list=images_[file_name]
                        except:
                            # have filepath
                            file_name = Path(file_path).name
                            wkspFldr = f'/app/output/{tenant_id}/assets/pdf/{tenant_id}/{case_id}/images'
                            # wkspFldr = f'/app/output/{ace}/assets/pdf/{tenant_id}/{case_id}'
                            logging.info(f"folder_path is {wkspFldr}")
                            #os.chmod(str(wkspFldr), 0o777)
                            # get the list of images
                            only_file_name = remove_all_except_al_num(file_name)
                            images_list_ = glob.glob(f'{wkspFldr}/*.jpg')
                            logging.info(f"list of images for this file is {images_list_}")
                            images_list = [Path(p).name for p in images_list_]
                            images_list = sorted(images_list, key=lambda x: (len(x), x))
                            # images_list=[]
                        # except Exception as e:
                        #     logging.info(e)
                        #     return_data_ = {"flag": False, "message": "Error in fetching file images"}
                            
                        

                logging.info(f"list of images for the file {file_name} is {images_list}")
                return_data[file_name]=images_list
                    # return it
                #checks for additional files or document set and returns the data seperately
                try:
                    if len(return_data[file_name]) == 0:
                        blob_status, file_name, file_blob = generate_blob_data(
                            file_path)

                        if not blob_status:
                            jsonify({"flag": False, "message": file_blob})

                        # for file_name, file_blob in file_dict.items():
                        #file_blob = f"data:application/pdf;base64,{file_blob}"
                        return_data[file_name] = {"blob": file_blob}
                except Exception as e:
                    logging.info(f"image data is not empty {e}")
                    pass
            logging.info(f"end of fetch file additional file data is {additional_file_data} and main file data is {return_data} ")
            return_data_ = {"flag": True, "data": return_data, "file_path_navigate": file_path_navigates, "pageInfo": page_info, "additonal_files_data" : additional_file_data}

    except Exception as e:
        logging.info(f"something went wrong {e}")
        return_data_ = {"flag": False, "message": "Error in fetching file images"}
        
    try:
        memory_after = measure_memory_usage()
        memory_consumed = (memory_after - memory_before) / \
            (1024 * 1024 * 1024)
        end_time = tt()
        memory_consumed = f"{memory_consumed:.10f}"
        logging.info(f"checkpoint memory_after - {memory_after},memory_consumed - {memory_consumed}, end_time - {end_time}")
        time_consumed = str(round(end_time-start_time,3))
    except:
        logging.info("Failed to calc end of ram and time")
        logging.info("ram calc went wrong")
        memory_consumed = None
        time_consumed = None
        pass
    logging.info(return_data_)

    # Rearranging the "mail_body.pdf" to be the second item in the "data" dictionary
    if 'mail_body.pdf' in return_data_['data']:
        return_data_["last_index"]=['mail_body.pdf']

    logging.info(return_data_)

    logging.info(f"## TD checkpoint memory_consumed: {memory_consumed}, time_consumed: {time_consumed}")
    return return_data_



@app.route('/convert_jpg_pdf', methods=['POST', 'GET'])
def convert_jpg_pdf():
    data = request.json
    case_id = data['case_id']
    tenant_id=data['tenant_id']
    db_config['tenant_id'] = tenant_id
    tenant_id = data.get('tenant_id', None)
    try:
        file_saving_path_case_id=Path(f'/app/output/{tenant_id}/assets/pdf/{tenant_id}/{case_id}')
        extraction_db = DB('queues', **db_config)
        query = f"select `file_name` from `process_queue` where case_id ='{case_id}'"
        query_data = extraction_db.execute_(query)['file_name'].to_list()[0]
        filename=json.loads(query_data)[0]

        if '.jpg' in filename:
            img_path=str(file_saving_path_case_id / filename)
            image = Image.open(img_path)
            pdf = canvas.Canvas(str(file_saving_path_case_id / filename.split('.')[0])+'.pdf', pagesize=image.size)
            pdf.drawInlineImage(img_path, 0, 0, width=image.size[0], height=image.size[1])
            pdf.save()
            logging.debug(f"Successfully made pdf file in the path {str(file_saving_path_case_id / filename.split('.')[0])+'.pdf'}") 
            filename=str(filename.split('.')[0]+'.pdf')
            filename=[filename,"mail_body.pdf"]
            logging.info(F"filename is {filename}")
            query = f"update `process_queue` set `file_name`='{json.dumps(filename)}' where case_id ='{case_id}'"
            query_data = extraction_db.execute_(query)
        elif '.xlsx' in filename or '.xls' in filename:
            convert_docx_to_pdf(str(file_saving_path_case_id / filename), str(file_saving_path_case_id / filename.split('.')[0])+'.pdf')
            filename=str(filename.split('.')[0]+'.pdf')
            filename=[filename,"mail_body.pdf"]
            logging.info(F"filename is {filename}")
            query = f"update `process_queue` set `file_name`='{json.dumps(filename)}' where case_id ='{case_id}'"
            query_data = extraction_db.execute_(query)
        elif '.docx' in filename:
            convert_xlsx_to_pdf(str(file_saving_path_case_id / filename), str(file_saving_path_case_id / filename.split('.')[0])+'.pdf')
            filename=str(filename.split('.')[0]+'.pdf')
            filename=[filename,"mail_body.pdf"]
            logging.info(F"filename is {filename}")
            query = f"update `process_queue` set `file_name`='{json.dumps(filename)}' where case_id ='{case_id}'"
            query_data = extraction_db.execute_(query)

    except Exception as e:
        logging.exception(e)
        logging.info(f"########## error in converting into pdf")
    return  jsonify({"flag":True,"data":{"message":"done saving to pdf"}})


def convert_docx_to_pdf(input_docx, output_pdf):
    try:
        result = subprocess.run(['unoconv', '-f', 'pdf', '-o', output_pdf, input_docx], stdout=subprocess.PIPE, stderr=subprocess.PIPE, check=True)
        output = result.stdout.decode('utf-8')  # Decode bytes to string
        print("Conversion successful!")
        # print(output)  # Print output for debugging
    except subprocess.CalledProcessError as e:
        error_output = e.stderr.decode('utf-8')  # Decode error bytes to string
        print(f"Conversion failed: {e}")
        print(error_output)  # Print error output for debugging


def convert_xlsx_to_pdf(input_xlsx, output_pdf):
    try:
        result = subprocess.run(['unoconv', '-f', 'pdf', '-o', output_pdf, input_xlsx], stdout=subprocess.PIPE, stderr=subprocess.PIPE, check=True)
        output = result.stdout.decode('utf-8')  # Decode bytes to string
        print("Conversion successful!")
        # print(output)  # Print output for debugging
    except subprocess.CalledProcessError as e:
        error_output = e.stderr.decode('utf-8')  # Decode error bytes to string
        print(f"Conversion failed: {e}")
        print(error_output)  # Print error output for debugging


"""
# some issue in save function
def convert_pdf_jpg(file_path, output_path):
    jpg_list = []
    file_full_name = file_path.name
    logging.info(f"file name is {file_full_name}")
    file_ext = file_path.suffix.lower()
    logging.info(f"file type is {file_ext}")
    file_page_object = {}
    file_image_name = remove_all_except_al_num(file_full_name)
    logging.info(f"file image name {file_image_name}")
    try:
        if file_ext == '.pdf':
            with fitz.open(file_path) as new_file:
                # with fitz.open(file_path) as pages:
                #     new_file.insertPDF(pages, to_page=7)
                # new_file.save(file_path)

                for idx, page in enumerate(new_file):
                    page_image_name = f'{file_image_name}_{str(idx)}.jpg'
                    logging.info(f"page image name {page_image_name}")
                    page_image_path = (f'{output_path}/images/'
                                       + page_image_name)
                    logging.info(f'Page {str(idx)}: {page_image_path}')
                    mat = fitz.Matrix(300 / 72, 300 / 72)
                    logging.info(f"pages got is {page},{fitz.fitz.Page}")
                    if isinstance(page, fitz.fitz.Page):
                        logging.info(f"pages got to if")
                        #pix = page.getPixmap(matrix=mat)
                        pix = page.get_pixmap(matrix = mat)
                        pix.pil_save(page_image_path)

                    else:
                        logging.info(f"pages got to else")
                        page.save(page_image_path, 'JPEG')

                    jpg_list.append(page_image_name)
                file_page_object = {file_full_name: jpg_list}
                logging.info(f"file_page_object: {file_page_object}")
    except Exception as e:
        logging.error(f'PDF to image conversion failed with pymupdf: {e}')
        logging.info(e)

    return file_page_object
"""

@app.route('/upload_files', methods=['POST', 'GET'])
def upload_files():
    """
    @author: Kalyani Bendi
    @modifier: Gopi Teja B
    @built at: Am Bank Project Time
    @description: upload of single or multiple files from front end and start the flow 
    """

    data = request.json
    logging.info(f"Request Data is: {data}")
    tenant_id = data.get('tenant_id', None)
    case_id = data.get('case_id', None)
    user = data.get('user', None)
    session_id = data.get('session_id', None)
    request_payload_data = data

    try:
        memory_before = measure_memory_usage()
        start_time = tt()
    except:
        logging.warning("Failed to start ram and time calc")
        pass

    trace_id = generate_random_64bit_string() if case_id is None else case_id
    attr = ZipkinAttrs(
        trace_id=trace_id,
        span_id=generate_random_64bit_string(),
        parent_span_id=None,
        flags=None,
        is_sampled=False,
        tenant_id=tenant_id
    )

    with zipkin_span(
        service_name='folder_monitor_api',
        span_name='upload_files',
        transport_handler=http_transport,
        zipkin_attrs=attr,
        port=5010,
        sample_rate=0.5):

        variables = data.get('variables', {})
        case_creation_details_dict = data.get('fields', {})
        fields_changed = data.get('field_changes',[])
        position = variables.get('position','')

        db_config['tenant_id'] = tenant_id
        queue_db = DB('queues', **db_config)
        extraction_db = DB('extraction', **db_config)
        camunda_port = variables.get('camunda_port',8080)
        camunda_host = variables.get('camunda_host','camundaworkflow')

        deleted_list=[]

        try:
            document_set = data['fields']['casecreation_file_names']
            docset_file_names = list(document_set.keys())
            logging.info(f"==> file names are {docset_file_names}")

            # Generating the case id or fin d the case_id
            case_entry = variables.get('case_entry', None)
            if case_entry == 'New_Case':
                unique_id = generate_caseid(
                    tenant_id, 'case_id', 'process_queue', 'queues')
                logging.info(
                    f"New main case id was generated and case id is {unique_id}")

                # Creating a folder with case id in htdocs
                output_path = Path(
                    f'/app/output/{tenant_id}/assets/pdf/{tenant_id}')
                os.umask(0)
                logging.info(output_path)
                Path(str(output_path / unique_id)).mkdir(parents=True, exist_ok=True)
                os.chmod(str(output_path / unique_id), 0o777)
                logging.debug(
                    f"## FM checkpoint `folder created in lampp directory with {unique_id}`")
                Path(str(output_path / unique_id / 'images')
                    ).mkdir(parents=True, exist_ok=True)  # creating a image folder
            # if case is already existed for the document set
            else:
                unique_id = data.get('case_id', None)

            ocr_dict = {'case_id': unique_id, 'highlight': '{}'}
            br_comparison_rules_dict = {'case_id': unique_id}
            process_queue_dict = {'case_id': unique_id, 'last_updated_by': "System",
                                'completed_processes': 1, 'status': "File Uploaded"}
            process_file_dict = {'case_id': unique_id}

            # if files blob are found create files in specified directory and generate document_id for each file
            if docset_file_names:
                for file_name,blob_data in document_set.items():
                    file_object={}
                    if '.pdf' not in file_name and file_name != 'deletedFile':
                        logging.info(f"File name is something wrong pdf is not there received filename is {file_name}")
                        return {"flag":False, "message": f"File is not in supportive format {file_name}. Please upload pdf file."}
                    
                    elif 'blob' in blob_data:
                        input_path_ = f"/app/output/{tenant_id}/assets/pdf/{tenant_id}/{unique_id}/{file_name}"
                        input_path_ = Path(input_path_)
                        logging.info(f"files are stored in: {input_path_}")

                        doc_blob_data = document_set[file_name]['blob']
                        doc_blob_data = doc_blob_data.split(',', 1)[1]
                        doc_blob_data += '='*(-len(doc_blob_data) % 4)

                        with open(input_path_, 'wb') as f:
                            f.write(base64.b64decode(doc_blob_data))
                            f.close()
                        logging.debug(f'File successfully moved to {input_path_}')
                        file_object=convert_pdf_jpg(Path(input_path_),
                                            f'/app/output/{tenant_id}/assets/pdf/{tenant_id}/{unique_id}')
                        file_object=json.dumps(file_object)
                        logging.info(f"images are created result is {file_object}")

                        logging.debug('File successfully moved')
                        document_unique_id = generate_caseid(
                            tenant_id, 'document_id', 'process_queue', 'queues')

                        logging.info(f"document id is generated for {file_name}")
                        case_creation_details_dict['casecreation_case_status'] = 0
                        case_creation_details_dict['case_id'] = unique_id
                        case_creation_details_dict['document_id'] = document_unique_id
                        case_creation_details_dict['casecreation_file_names'] = file_name
                        case_creation_details_dict['casecreation_updated_by'] = user
                        extraction_db.insert_dict(
                            data=case_creation_details_dict, table='case_creation_details')

                        process_queue_dict['document_id'] = document_unique_id
                        process_queue_dict['file_name'] = file_name
                        process_queue_dict['file_paths'] = input_path_
                        process_queue_dict['status'] = "Files got uploaded"
                        process_queue_dict['imagelist'] = json.dumps(file_object)
                        queue_db.insert_dict(
                            data=process_queue_dict, table='process_queue')

                        process_file_dict['document_id'] = document_unique_id
                        process_file_dict['file_name'] = file_name
                        queue_db.insert_dict(
                            data=process_file_dict, table='process_file')

                        ocr_dict['document_id'] = document_unique_id
                        extraction_db.insert_dict(data=ocr_dict, table='ocr')

                        extraction_db.insert_dict(data=br_comparison_rules_dict, table='br_comparison_rules')

                        logging.info(
                            f"{document_unique_id} is generated for {file_name}")

                    #updating file status as deleted while reuploading
                    elif file_name == 'deletedFile':
                        deleted_list = document_set['deletedFile']
                        logging.info(f"deleted files are: {deleted_list}")
                        #deleted_list=ast.literal_eval(deleted_list)
                        if len(deleted_list) == 1:
                            ccd_query = f"UPDATE case_creation_details SET casecreation_file_status = 'file deleted',casecreation_updated_by='{user}' WHERE casecreation_file_names = '{deleted_list[0]}' AND case_id = '{unique_id}' "
                            extraction_db.execute_(ccd_query)
                            pq_query = f"UPDATE process_queue SET state = 'file deleted' WHERE file_name = '{deleted_list[0]}' AND case_id = '{unique_id}'"
                            queue_db.execute_(pq_query)
                        elif len(deleted_list) > 1:
                            deleted_list_data = tuple(deleted_list)
                            ccd_query = f"UPDATE case_creation_details SET casecreation_file_status = 'file deleted',casecreation_updated_by='{user}' WHERE casecreation_file_names IN {deleted_list_data} AND case_id = '{unique_id}'"
                            extraction_db.execute(ccd_query)

                            pq_query= f"UPDATE process_queue SET state = 'file deleted' WHERE file_name IN {deleted_list_data} AND case_id = '{unique_id}'"
                            queue_db.execute(pq_query)

                            logging.info(f"file status  for case_id {unique_id} got updated in case_creation_details and process_queue table as deleted")
                        else:
                            logging.info(f"## FM no files in deleted_list, came to else condition - {deleted_list}")

                message = "Files Successfully Ingested"
                logging.info(message)

            else:
                message = "No files are received"
                logging.info(message)

            # below logic works if required to hit the flow directly from route call instead of monitoring / scheduler
            if "camunda_host" not in variables or "camunda_port" not in variables or "workflow_name" not in variables:
                logging.debug("camunda_host / camunda_port / workflow_name not configured in variables")
                return {'flag':'False','message':'camunda_host / camunda_port / workflow_name not configured in variables'}

            elif variables is not None and "workflow_name" in variables and position == 'First_Time' and "camunda_host" in variables and "camunda_port" in variables:
                workflow_name = variables.get("workflow_name", None)
                api_params = {"variables": {"case_id": {"value": unique_id},"session_id": {"value": session_id},"tenant_id": {"value": tenant_id},"camunda_host": {"value":camunda_host},"camunda_port": {"value":camunda_port},"user": {"value":user}}}
                request_api = f"http://{camunda_host}:{camunda_port}/rest/engine/default/process-definition/key/{workflow_name}/start"
                headers = {'Content-type': 'application/json; charset=utf-8'}
                logging.info(f"#### FM Hitting the camunda api of {request_api}, api_params = {api_params} ")
                response = requests.post(
                    request_api, json=api_params, headers=headers)
                response_dict = json.loads(response.text)
                logging.info(f"#### Response Dict {response_dict}")


            #updating the changes that are done in reupload poup into the database 
            updated_changes=variables.get("update_changes",None)
            logging.info(f"updated changes {updated_changes} and length of fields changes {len(fields_changed)}")
            if updated_changes == 'True' and len(fields_changed) > 0:
                update = {field: case_creation_details_dict[field] for field in fields_changed  if field != 'casecreation_file_names'}
                where = {
                    'case_id': unique_id
                }
                if update != {}:
                    extraction_db.update('case_creation_details', update=update, where=where)
                    logging.info(f"details got updated in case creation details")
            message="Files uploaded successfully"
            return_data = {'flag': True, 'data':{'document set': docset_file_names,'deleted files' : deleted_list,'message': message,'previous_case_id': unique_id}}

        except Exception:
            case_creation_details_dict['casecreation_case_status'] = 0
            extraction_db.insert_dict(
                data=case_creation_details_dict, table='case_creation_details')
            logging.exception("Something went wrong in upload files")
            return_data = {'flag': False, 'message': 'Something went wrong in uploading files', 'data':{}}

        try:
            memory_after = measure_memory_usage()
            memory_consumed = (memory_after - memory_before) / \
                (1024 * 1024 * 1024)
            end_time = tt()
            memory_consumed = f"{memory_consumed:.10f}"
            logging.info(f"checkpoint memory_after - {memory_after},memory_consumed - {memory_consumed}, end_time - {end_time}")
            time_consumed = str(round(end_time-start_time,3))
        except:
            logging.warning("Failed to calc end of ram and time")
            logging.exception("ram calc went wrong")
            memory_consumed = None
            time_consumed = None
            pass
        request_payload_data.pop('fields')
        logging.info(f"## TD checkpoint memory_consumed: {memory_consumed}, time_consumed: {time_consumed}")
        # insert audit

        audit_data = {"tenant_id": tenant_id, "user": user, "case_id": unique_id, 
                        "api_service": "upload_files", "service_container": "folder_monitor", "changed_data": None,
                        "tables_involved": "","memory_usage_gb": str(memory_consumed), 
                        "time_consumed_secs": time_consumed, "request_payload": json.dumps(request_payload_data), 
                        "response_data": json.dumps(return_data['data']), "trace_id": trace_id, "session_id": session_id,"status":str(return_data['flag'])}
        insert_into_audit(case_id, audit_data)
        return return_data

## While click on the tab name in the left side file  preview it will automatically go to the respected page
@app.route("/tab_navigate", methods=['POST', 'GET'])
def tab_navigate():
    try:
        data = request.json
        logging.info(f"##### tab_navigate: {data}")
        tenant_id = data.get("tenant_id", "")
        case_id = data.get("case_id", "")
        file_name=data.get("file_tab_name", "")
        tab_name = data.get("tab_name", "")
        logging.info(f" tab_name {tab_name}")
        db_config['tenant_id'] = tenant_id
        
        queue_db = DB("queues", **db_config)

        query = f'select `single_doc_identifiers` from `process_queue` where `case_id`="{case_id}" and `file_name`="{file_name}"'
        tab_names = queue_db.execute_(query)['single_doc_identifiers'][0]
        logging.info(f" tab_names is {tab_names}")
        tab_names = json.loads(tab_names)
        logging.info(f" tab_names is {tab_names}")
        logging.info(f" tab_names is {type(tab_names)}")
        
        return_data = {}                 
        for each in tab_names:
            logging.info(f"each: {each}")
            logging.info(f"each['file_type']: {each['file_type']}")
            if tab_name.lower() in each['file_type'].split('_')[0].lower():
                logging.info(f"comes to if")
                return_data = each
                break
        if return_data == {}:
            return jsonify({'flag': False, 'message': f"Could not find this '{tab_name}' file"})
                

        return jsonify({"flag": True, "data": return_data})
    
    except Exception as e:
        logging.info(f"=========> tab_navigate {e}")
        return jsonify({'flag': False, 'message': "Something went wrong"})

@app.route('/supporting_additional_files', methods=['POST', 'GET'])
def supporting_additional_files():
    """
    @author: Kalyani Bendi
    @built at: Am Bank Project Time
    @description: upload of additional files from front end and storing it in database 
    """
    data = request.json
    logging.info(f"Request Data is: {data}")
    tenant_id = data.get('tenant_id', None)
    case_id = data.get('case_id', None)
    user = data.get('user', None)
    session_id = data.get('session_id', None)
    request_payload_data = data
    try:
        memory_before = measure_memory_usage()
        start_time = tt()
    except:
        logging.warning("Failed to start ram and time calc")
        pass
    
    if case_id is None:
        trace_id = generate_random_64bit_string()
    else:
        trace_id = case_id
    
    attr = ZipkinAttrs(
        trace_id=trace_id,
        span_id=generate_random_64bit_string(),
        parent_span_id=None,
        flags=None,
        is_sampled=False,
        tenant_id=tenant_id
    )

    with zipkin_span(
        service_name='folder_monitor_api',
        span_name='supporting_additional_files',
        transport_handler=http_transport,
        zipkin_attrs=attr,
        port=5010,
        sample_rate=0.5):

        variables = data.get('variables', {})
        db_config['tenant_id'] = tenant_id
        queue_db = DB('queues', **db_config)
        try:
            fields = data.get('fields',{})
            comments = None
            additional_file_names = []
            if fields != {}:
                additional_files = fields.get('filemanager_file_name',None)
                additional_file_names = list(additional_files.keys())
                comments = fields.get('filemanager_comments', None)
            logging.info(f"==>additional file names are {additional_file_names}")

            query_check_case = f"select * from file_manager where case_id='{case_id}'"
            result_check_case = queue_db.execute(query_check_case).to_dict(orient='records')
            logging.info(f"checking for case existence {result_check_case}")
            
            if len(result_check_case) == 0:
                file_names_list = []
                imagelist_list = []
                insert_data = { 'case_id' : case_id,
                                'filemanager_stage' : 'additional_files'
                            }
                queue_db.insert_dict(
                        insert_data, 'file_manager')
                logging.info(f"successfully inserted case id and stage into file manager")

            else:
                file_names_list =  ast.literal_eval(result_check_case[0]['filemanager_file_name'])
                imagelist_list = ast.literal_eval(result_check_case[0]['filemanager_imagelist'])
                additional_file_names = list(set(additional_file_names) - set(file_names_list))
                logging.info(f"case already exists additonal files are {additional_file_names}")
            try:
                if additional_file_names:
                    for file_name,blob_data in additional_files.items():
                        if 'blob' in blob_data:
                            input_path_ = f"/app/output/{tenant_id}/assets/pdf/{tenant_id}/{case_id}/{file_name}"
                            input_path_ = Path(input_path_)
                            logging.info(f"files are stored in: {input_path_}")

                            doc_blob_data = additional_files[file_name]['blob']
                            doc_blob_data = doc_blob_data.split(',', 1)[1]
                            doc_blob_data += '='*(-len(doc_blob_data) % 4)

                            with open(input_path_, 'wb') as f:
                                f.write(base64.b64decode(doc_blob_data))
                                f.close()
                            logging.debug(f'File successfully moved to {input_path_}')

                            file_object=convert_pdf_jpg(Path(input_path_),
                                            f'/app/output/{tenant_id}/assets/pdf/{tenant_id}/{case_id}')
                            file_object=json.dumps(file_object)
                            logging.info(f"images are created result is {file_object}")
                            file_names_list.append(file_name)
                            imagelist_list.append(file_object)

                    update_data = { 'filemanager_imagelist' : json.dumps(imagelist_list),
                                    'filemanager_file_name' : json.dumps(file_names_list),
                                    'filemanager_comments' : comments
                                }

                    queue_db.update(table='file_manager', update = update_data,
                                    where={'case_id': case_id})
                    
                    logging.info(f"Additional files successfully stored in file_manager table")
                    
                return_data = {'flag': True, 'data':{'message': 'Successfully uploaded','additional_files': additional_file_names}}
            except:
                logging.info(f"something went wring in accessing additional files")
                return_data = {'flag': False, 'data':{'message': 'Something went wrong in accessing additional files'}}
        
        except Exception:
        
            logging.exception("Something went wrong in uploading additional files")
            return_data = {'flag': False, 'message': 'Something went wrong in uploading additional files'}
        
        try:
            memory_after = measure_memory_usage()
            memory_consumed = (memory_after - memory_before) / \
                (1024 * 1024 * 1024)
            end_time = tt()
            memory_consumed = f"{memory_consumed:.10f}"
            logging.info(f"checkpoint memory_after - {memory_after},memory_consumed - {memory_consumed}, end_time - {end_time}")
            time_consumed = str(round(end_time-start_time,3))
        except:
            logging.warning("Failed to calc end of ram and time")
            logging.exception("ram calc went wrong")
            memory_consumed = None
            time_consumed = None
            pass
        
        logging.info(f"## TD checkpoint memory_consumed: {memory_consumed}, time_consumed: {time_consumed}")
        request_payload_data.pop('fields')
        audit_data = {"tenant_id": tenant_id, "user": user, "case_id": case_id, 
                        "api_service": "upload_files", "service_container": "folder_monitor", "changed_data": None,
                        "tables_involved": "","memory_usage_gb": str(memory_consumed), 
                        "time_consumed_secs": time_consumed, "request_payload": json.dumps(request_payload_data), 
                        "response_data": json.dumps(return_data['data']), "trace_id": trace_id, "session_id": session_id,"status":str(return_data['flag'])}
        insert_into_audit(case_id, audit_data)
        return return_data

def extract_numeric_value(value):
    if isinstance(value, str):
        cleaned_value = re.sub(r'[^\d.]', '', value)  # Remove non-numeric characters except dot
        return pd.to_numeric(cleaned_value, errors='coerce')
    return None

def create_dataframe_from_json(json_data_list):
    columns_data = {}
    nested_keys = ["Debtors 121-150 days", "Debtors 151-180 days", "Debtors > 180 days", "TOTAL_Debitors"]  # Define your nested keys here

    try:
        for idx, json_data in enumerate(json_data_list):
      
            for key, value in json_data.items():
                if value is not None:
                    data_dict = json.loads(value)
                    for nested_key, nested_value in data_dict.items():
                        if 'DEBITORS STATEMENT' in key:
                            column_name = f"{key}_{nested_key}"
                            print(column_name,'-----------')
                        else:
                            column_name = f"{nested_key}"
                        
                        if column_name not in columns_data:
                            columns_data[column_name] = []
                        numeric_value = extract_numeric_value(nested_value)
                        columns_data[column_name].append(numeric_value)
                else:
                    # If the value is None, fill the columns with None or 0, depending on your preference
                    for nested_key in nested_keys:
                        column_name = f"{nested_key}"
                        if column_name not in columns_data:
                            columns_data[column_name] = []
                        columns_data[column_name].append(None)

        # Ensure all lists have the same length by filling with None
        max_length = max(len(data) for data in columns_data.values())
        for key, data in columns_data.items():
            data.extend([None] * (max_length - len(data)))
            
            
        
        df = pd.DataFrame(columns_data)
#         df.fillna(0, inplace=True)  # Fill NaN values with 0 or adjust as needed
        df = df.dropna(how='all')
        # df=df.iloc[0]
        return df

    except Exception as e:
        print(f"Error: {e}")
        return None
    
def convert_to_custom_format(date_str):
    formats = [
        "%d-%m-%Y",
        "%d/%m/%Y",
        "%b %Y",
        "%d %b %Y",
        "%B %Y",
        "%m-%Y",
        "%m/%Y",
        "%d %B, %Y",
        "%B-%Y",
        '%d-%b-%Y %H:%M:%S',
        "%d.%m.%Y"
        
    ]

    for date_format in formats:
        try:
            dt = datetime.strptime(date_str, date_format)
            return dt.strftime("%d-%b-%Y")
        except ValueError as e:
            logging.info(f"{e}###wrong format")
            pass

    return date_str



@app.route('/download_excel', methods=['POST', 'GET'])
def download_excel():
    data=request.json
    db_config['tenant_id'] = 'kmb'
    logging.info(f"Request Data is: {data}")
    queue_db = DB('queues', **db_config)
    queue_id=data.get("queue_id",{})
    extraction_db = DB('extraction', **db_config)
    case_id = data.pop('case_id')
    try:
       
        qry=f"select `segment` from `ocr` where case_id='{case_id}'"
        segment_res = extraction_db.execute_(qry).to_dict(orient="records")[-1]
        segment = segment_res['segment']
        qry=f"SELECT `STOCK STATEMENT_OCR`,`DEBITORS STATEMENT_OCR`,`CREDITORS_OCR`, `PURCHASES_OCR`, `ADVANCES_OCR`, `SALES_OCR`, `ADVANCES DEBTORS_OCR`, `BANK OS_OCR`, `ADVANCES SUPPLIERS_OCR`,`SECURITY DEPOSIT_OCR` FROM `ocr` WHERE  `case_id`='{case_id}';" 
        json_data_list = extraction_db.execute_(qry).to_dict(orient="records")

        qry1=f"SELECT `comments` as Comments, `Due_Date` as 'Stock Due Date',`co_mail_time` as 'Received Date', `cc_limit` as 'CC_LIMIT',`insurance_amount` as 'INSURANCE_AMOUNT' , `Debtors_margin` as 'DEBTORS_MARGIN',`Stock_Margins` as 'STOCK_MARGIN',`deviation_applicable` as 'DEVIATION_APPLICABLE',`Crn` as 'CRN',`customer_name` as 'NAME OF CUSTOMER' ,`request_id` as 'REQUEST_ID' , `Account_Number` as 'Account No',`division` as 'ZMT',`Dp_formula` as 'DP formula',`date` as 'SS_AS_ON_DATE' from ocr where `case_id`='{case_id}';"
        df = extraction_db.execute_(qry1).to_dict(orient="records")[-1]
        received_date_str = df['Received Date']
        stock_due_date=df['Stock Due Date']
        ss_on_date = df['SS_AS_ON_DATE']

        try:
            df['Stock Due Date'] = convert_to_custom_format(stock_due_date)
            df['Received Date'] = convert_to_custom_format(received_date_str)
            df['SS_AS_ON_DATE'] = convert_to_custom_format(ss_on_date)

        except Exception as e:
            print(f'{e}#########exception')
        

        
        df = pd.DataFrame([df])
        values= create_dataframe_from_json(json_data_list)
        values = values.to_dict()
        print(f"#########values are {values} and type is {type(values)}")
        final_result={}
        try:
            for i in values:
                final_result[i]=values[i][0]
            print(f"######final result {final_result}")
        except:
            logging.info(f"exception in appending final_result")
        final_result = pd.DataFrame([final_result])
        if segment == 'Agri':
            df = df.rename(columns={'Account No': 'ACCOUNT_NUMBER'})
            logging.info(f"df##{df}")
            logging.info(f"in if block segement is {segment}")
            # keys = {"CREDITORS_ADVANCES_RECEIVED_FROM_CUSTOMERS": "Adv from Customers", "DEBTORS_ADVANCES_PAID_TO_SUPPLIER": "Adv to Suppliers", "BC_OS_WITH_KMBL_IN_CASE_OF_SUBLIMIT": "BC OS with KMBL - sublimit", "BG_OS_WITH_KMBL_IN_CASE_OF_SUBLIMIT": "BG OS with KMBL - sublimit",  "DEBTORS_0_TO_5_DAYS": "Debtors 0-5 days", "DEBTORS_121_TO_150_DAYS": "Debtors 121-150 days", "DEBTORS_151_TO_180_DAYS": "Debtors 151-180 days", "DEBTORS_6_TO_60_DAYS": "Debtors 6-60 days", "DEBTORS_61_TO_90_DAYS": "Debtors 61-90 days", "DEBTORS_GREATER_THAN_180_DAYS": "Debtors > 180 days", "LC_OS_WITH_KMBL_IN_CASE_OF_SUBLIMIT": "LC OS with KMBL - sublimit", "OUTSTANDING_WITH_OTHER_WC_BANKERS": "OS with other Banks", "PURCHASE_AMOUNT": "Purchases_AMOUNT", "SALES_AMOUNT": "Sales_AMOUNT", "TOTAL_CREDITORS": "Total Creditors", "TOTAL_STOCK": "Total Stock", "WCDL_OS_WITH_KMBL_IN_CASE_OF_SUBLIMIT": "WCDL OS with KMBL- sublimit"}
            keys={'CREDITORS_ADVANCES_RECEIVED_FROM_CUSTOMERS': 'Adv from Customers', 'DEBTORS_ADVANCES_PAID_TO_SUPPLIER': 'Adv to Suppliers', 'BC_OS_WITH_KMBL_IN_CASE_OF_SUBLIMIT': 'BC OS with KMBL - sublimit', 'BG_OS_WITH_KMBL_IN_CASE_OF_SUBLIMIT': 'BG OS with KMBL - sublimit', 'DEBTORS_0_TO_5_DAYS': 'DEBITORS STATEMENT_OCR_Debtors 0-5 days', 'DEBTORS_121_TO_150_DAYS': 'DEBITORS STATEMENT_OCR_Debtors 121-150 days', 'DEBTORS_151_TO_180_DAYS': 'DEBITORS STATEMENT_OCR_Debtors 151-180 days', 'DEBTORS_6_TO_60_DAYS': 'DEBITORS STATEMENT_OCR_Debtors 6-60 days', 'DEBTORS_61_TO_90_DAYS': 'DEBITORS STATEMENT_OCR_Debtors 61-90 days','DEBTORS_91_TO_120_DAYS':'DEBITORS STATEMENT_OCR_Debtors_91_to_120_days','DEBTORS_GREATER_THAN_180_DAYS': 'DEBITORS STATEMENT_OCR_Debtors > 180 days', 'LC_OS_WITH_KMBL_IN_CASE_OF_SUBLIMIT': 'LC OS with KMBL -sublimit', 'OUTSTANDING_WITH_OTHER_WC_BANKERS': 'OS with other Banks', 'PURCHASE_AMOUNT': 'Purchases_Amount', 'SALES_AMOUNT': 'Sales_Amount', 'TOTAL_CREDITORS':'Total Creditors' , 'TOTAL_STOCK': 'Total Stock', 'WCDL_OS_WITH_KMBL_IN_CASE_OF_SUBLIMIT': 'WCDL OS with KMBL- sublimit'}
            merged_data = {key: final_result.get(keys.get(key), np.nan) for key in keys}
            values_dict = {key: value[0] if isinstance(value, pd.Series) else value for key, value in merged_data.items()}
            result_df = pd.DataFrame([values_dict])
          
            res = pd.DataFrame({key: [] for key in ["REQUEST_ID", "CRN", "ACCOUNT_NUMBER","SS_AS_ON_DATE","TOTAL_STOCK", "TOTAL_CREDITORS", "DEBTORS_0_TO_5_DAYS", "DEBTORS_6_TO_60_DAYS", "DEBTORS_61_TO_90_DAYS", "DEBTORS_91_TO_120_DAYS", "DEBTORS_121_TO_150_DAYS", "DEBTORS_151_TO_180_DAYS","DEBTORS_GREATER_THAN_180_DAYS", "DEBTORS_ADVANCES_PAID_TO_SUPPLIER", "CREDITORS_ADVANCES_RECEIVED_FROM_CUSTOMERS", "SALES_AMOUNT", "PURCHASE_AMOUNT", "INSURANCE_AMOUNT", "OUTSTANDING_WITH_OTHER_WC_BANKERS","WCDL_OS_WITH_KMBL_IN_CASE_OF_SUBLIMIT", "LC_OS_WITH_KMBL_IN_CASE_OF_SUBLIMIT", "BC_OS_WITH_KMBL_IN_CASE_OF_SUBLIMIT","BG_OS_WITH_KMBL_IN_CASE_OF_SUBLIMIT", "DEBTORS_MARGIN", "STOCK_MARGIN", "OD_LIMIT", "CC_LIMIT", "DEVIATION_APPLICABLE"  ]})
            result_df = pd.concat([result_df, df], axis=1)   
            # for column, values in result_df.items():
            #     res[column] = values[0]  
            res = res.append(result_df, ignore_index=True)

            dp_query=f"SELECT `dp_formula` from ocr where case_id='{case_id}'"
            dp_res = extraction_db.execute_(dp_query).to_dict(orient="records")[-1]
            logging.info(f"{dp_res}")
            if dp_res['dp_formula'] !=  None:
        
                equation_list = re.findall(r'\b\w+_\w+\b', dp_res['dp_formula'] )
                agri_fields = {'TOTAL_STOCKS': 'Total_stock', 'TOTAL_CREDITORS_AMT': 'Total Creditors', 'DD_0_5': 'DEBITORS STATEMENT_OCR_Debtors 0-5 days', 'DD_6_60': 'DEBITORS STATEMENT_OCR_Debtors 6-60 days', 'DD_61_90': 'DEBITORS STATEMENT_OCR_Debtors 61-90 days', 'DD_91_120': 'DEBITORS STATEMENT_OCR_Debtors_91_to_120_days','DD_121_150': 'DEBITORS STATEMENT_OCR_Debtors 121-150 days', 'DD_151_180': 'DEBITORS STATEMENT_OCR_Debtors 151-180 days', 'DD_GREATER_180': 'DEBITORS STATEMENT_OCR_Debtors > 180 days', 'DD_DEBTORS_ADV_PAID_SUPP': 'Adv to Suppliers', 'CRED_ADV_REC_FROM_CUSTOMERS': 'Adv from Customers', 'INSURANCE_AMT': 'Insurance_amount', 'OS_OTHER_WC_BANKERS': 'Outstanding_with_other_wc_bankers', 'WCDL_WITH_KMBL': 'Wcdl_os_with_kmbl_in_case_of_sublimit', 'LC_WITH_KMBL': 'Lc_os_with_kmbl_in_case_of_sublimit', 'BC_WITH_KMBL': 'Bc_os_with_kmbl_in_case_of_sublimit', 'BG_WITH_KMBL': 'Bg_os_with_kmbl_in_case_of_sublimit'}    
                categories = ["TOTAL_STOCKS", "TOTAL_CREDITORS_AMT", "DD_0_5", "DD_6_60", "DD_61_90", "DD_91_120", "DD_121_150", "DD_151_180", "DD_GREATER_180", "DD_DEBTORS_ADV_PAID_SUPP", "CRED_ADV_REC_FROM_CUSTOMERS", "INSURANCE_AMT", "OS_OTHER_WC_BANKERS", "WCDL_WITH_KMBL", "LC_WITH_KMBL", "BC_WITH_KMBL", "BG_WITH_KMBL"]
                non_dp_formula_fields=[]
                for i in categories:
                    if i not in equation_list:
                        non_dp_formula_fields.append(agri_fields[i])
                # data_dict =  {'Adv from Customers': 'CREDITORS_ADVANCES_RECEIVED_FROM_CUSTOMERS', 'Adv to Suppliers': 'DEBTORS_ADVANCES_PAID_TO_SUPPLIER', 'BC OS with KMBL - sublimit': 'BC_OS_WITH_KMBL_IN_CASE_OF_SUBLIMIT', 'BG OS with KMBL - sublimit': 'BG_OS_WITH_KMBL_IN_CASE_OF_SUBLIMIT', 'Debtors 0-5 days': 'DEBTORS_0_TO_5_DAYS', 'Debtors 121-150 days': 'DEBTORS_121_TO_150_DAYS', 'Debtors 151-180 days': 'DEBTORS_151_TO_180_DAYS', 'Debtors 6-60 days': 'DEBTORS_6_TO_60_DAYS', 'Debtors 61-90 days': 'DEBTORS_61_TO_90_DAYS', 'Debtors_91_to_120_days':'DEBTORS_91_TO_120_DAYS','Debtors > 180 days': 'DEBTORS_GREATER_THAN_180_DAYS', 'LC OS with KMBL -sublimit': 'LC_OS_WITH_KMBL_IN_CASE_OF_SUBLIMIT', 'OS with other Banks': 'OUTSTANDING_WITH_OTHER_WC_BANKERS', 'Purchases_AMOUNT': 'PURCHASE_AMOUNT', 'Sales_AMOUNT': 'SALES_AMOUNT', 'Total Creditors': 'TOTAL_CREDITORS', 'Total Stock': 'TOTAL_STOCK', 'WCDL OS with KMBL- sublimit': 'WCDL_OS_WITH_KMBL_IN_CASE_OF_SUBLIMIT'}
                data_dict =  {'Adv from Customers': 'CREDITORS_ADVANCES_RECEIVED_FROM_CUSTOMERS', 'Adv to Suppliers': 'DEBTORS_ADVANCES_PAID_TO_SUPPLIER', 'Bc_os_with_kmbl_in_case_of_sublimit': 'BC_OS_WITH_KMBL_IN_CASE_OF_SUBLIMIT', 'Bg_os_with_kmbl_in_case_of_sublimit': 'BG_OS_WITH_KMBL_IN_CASE_OF_SUBLIMIT', 'DEBITORS STATEMENT_OCR_Debtors 0-5 days': 'DEBTORS_0_TO_5_DAYS', 'DEBITORS STATEMENT_OCR_Debtors 121-150 days': 'DEBTORS_121_TO_150_DAYS', 'DEBITORS STATEMENT_OCR_Debtors 151-180 days': 'DEBTORS_151_TO_180_DAYS', 'DEBITORS STATEMENT_OCR_Debtors 6-60 days': 'DEBTORS_6_TO_60_DAYS', 'DEBITORS STATEMENT_OCR_Debtors 61-90 days': 'DEBTORS_61_TO_90_DAYS', 'DEBITORS STATEMENT_OCR_Debtors_91_to_120_days':'DEBTORS_91_TO_120_DAYS','DEBITORS STATEMENT_OCR_Debtors > 180 days': 'DEBTORS_GREATER_THAN_180_DAYS', 'Lc_os_with_kmbl_in_case_of_sublimit': 'LC_OS_WITH_KMBL_IN_CASE_OF_SUBLIMIT', 'Outstanding_with_other_wc_bankers': 'OUTSTANDING_WITH_OTHER_WC_BANKERS', 'Purchases_Amount': 'PURCHASE_AMOUNT', 'Sales_Amount': 'SALES_AMOUNT', 'Total_creditors': 'TOTAL_CREDITORS', 'Total_stock': 'TOTAL_STOCK', 'Wcdl_os_with_kmbl_in_case_of_sublimit': 'WCDL_OS_WITH_KMBL_IN_CASE_OF_SUBLIMIT', 'Insurance_amount': 'INSURANCE_AMOUNT'}    # data_dict = {'Total_stock': 'TOTAL_STOCK', 'Total_creditors': 'TOTAL_CREDITORS', 'Debtors_0_to_5_days': 'DEBTORS_0_TO_5_DAYS', 'Debtors_6_to_60_days': 'DEBTORS_6_TO_60_DAYS', 'Debtors_61_to_90_days': 'DEBTORS_61_TO_90_DAYS', 'Debtors_91_to_120_days': 'DEBTORS_91_TO_120_DAYS', 'Debtors_121_to_150_days': 'DEBTORS_121_TO_150_DAYS', 'Debtors_151_to_180_days': 'DEBTORS_151_TO_180_DAYS', 'Debtors_greater_than_180_days': 'DEBTORS_GREATER_THAN_180_DAYS', 'Debtors_advances_paid_to_supplier': 'DEBTORS_ADVANCES_PAID_TO_SUPPLIER', 'Creditors_advances_received_from_customers': 'CREDITORS_ADVANCES_RECEIVED_FROM_CUSTOMERS', 'Insurance_amount': 'INSURANCE_AMOUNT', 'Outstanding_with_other_wc_bankers': 'OUTSTANDING_WITH_OTHER_WC_BANKERS', 'Wcdl_os_with_kmbl_in_case_of_sublimit': 'WCDL_OS_WITH_KMBL_IN_CASE_OF_SUBLIMIT', 'Lc_os_with_kmbl_in_case_of_sublimit': 'LC_OS_WITH_KMBL_IN_CASE_OF_SUBLIMIT', 'Bc_os_with_kmbl_in_case_of_sublimit': 'BC_OS_WITH_KMBL_IN_CASE_OF_SUBLIMIT', 'Bg_os_with_kmbl_in_case_of_sublimit': 'BG_OS_WITH_KMBL_IN_CASE_OF_SUBLIMIT'}

                # data_dict = {'Total_stock': 'TOTAL_STOCK', 'Total_creditors': 'TOTAL_CREDITORS', 'Debtors_0_to_5_days': 'DEBTORS_0_TO_5_DAYS', 'Debtors_6_to_60_days': 'DEBTORS_6_TO_60_DAYS', 'Debtors_61_to_90_days': 'DEBTORS_61_TO_90_DAYS', 'Debtors_91_to_120_days': 'DEBTORS_91_TO_120_DAYS', 'Debtors_121_to_150_days': 'DEBTORS_121_TO_150_DAYS', 'Debtors_151_to_180_days': 'DEBTORS_151_TO_180_DAYS', 'Debtors_greater_than_180_days': 'DEBTORS_GREATER_THAN_180_DAYS', 'Debtors_advances_paid_to_supplier': 'DEBTORS_ADVANCES_PAID_TO_SUPPLIER', 'Creditors_advances_received_from_customers': 'CREDITORS_ADVANCES_RECEIVED_FROM_CUSTOMERS', 'Insurance_amount': 'INSURANCE_AMOUNT', 'Outstanding_with_other_wc_bankers': 'OUTSTANDING_WITH_OTHER_WC_BANKERS', 'Wcdl_os_with_kmbl_in_case_of_sublimit': 'WCDL_OS_WITH_KMBL_IN_CASE_OF_SUBLIMIT', 'Lc_os_with_kmbl_in_case_of_sublimit': 'LC_OS_WITH_KMBL_IN_CASE_OF_SUBLIMIT', 'Bc_os_with_kmbl_in_case_of_sublimit': 'BC_OS_WITH_KMBL_IN_CASE_OF_SUBLIMIT', 'Bg_os_with_kmbl_in_case_of_sublimit': 'BG_OS_WITH_KMBL_IN_CASE_OF_SUBLIMIT'}
                for field in non_dp_formula_fields:
                    
                    res[data_dict[field]] = None
                logging.info(f"result dataframe####  {res}")
            queue_name="agri"
        elif segment == 'Consumer':
           
            logging.info(f"in if block segement is {segment}")
            
            res = pd.DataFrame({key: [] for key in['CRN', 'NAME OF CUSTOMER', 'SS_AS_ON_DATE', 'Raw Material-< 90 DAYS', 'Raw Material-90 - 120 DAYS', 'Raw Material-120 - 180 DAYS', 'Raw Material-> 180 DAYS', 'Raw Material-Unit/Quantity', 'Work In Progress-< 60 DAYS', 'Work In Progress 60- 90 DAYS', 'Work In Progress-90 - 120 DAYS', 'Work In Progress- 120 - 180 DAYS', 'Work In Progress-> 180 DAYS', 'Work In Progress-Unit/Quantity', 'Consumables & Packing Material-< 90 DAYS', 'Consumables & Packing Material-90 - 180 DAYS', 'Consumables & Packing Material-> 180 DAYS', 'Consumables & Packing Material-Unit/Quantity', 'Finished Goods-< 90 DAYS', 'Finished Goods-90 - 120 DAYS', 'Finished Goods-120 - 180 DAYS', 'Work In Progress- 90 - 180 DAYS','Finished Goods-> 180 DAYS', 'Finished Goods-Unit/Quantity', 'Vehicle stock upto 120 days', 'Vehicle stock 120-150 days', 'Spares stock 6-60 DAYS', 'Spares stock < 60 DAYS', 'Total Spares stock', 'Total Stock <30 DAYS', 'Total Stock 30-90 DAYS', 'Total Stock-90 - 120 DAYS', 'Total Stock 120- 180 DAYS', 'Total Stock-> 180 DAYS', 'Total Stock-Unit/Quantity', 'Third party Debtors-< 90 days', 'Third party Debtors-90 - 180 days', 'Third party Debtors-> 180 days', 'Third party Debtors-No. of debtors', 'Spare Debtors less than 60 days', 'Spare Debtors 60-90 days', 'Vehicle debtors', 'Vehicle debtors 6 - 60 days', 'Related party/ group concern Debtors-< 90 days', 'Related party/ group concern Debtors-90 - 180 days', 'Related party/ group concern Debtors-> 180 days', 'Related party/ group concern Debtors-No. of debtors', 'Total debtors Debtors-< 30 days', 'Total debtors Debtors 30- 60 days', 'Total debtors Debtors 60- 90 days', 'Total debtors Debtors-90 - 120 days', 'Total debtors Debtors-120 - 180 days', 'Total debtors Debtors-> 180 days', 'Total debtors Debtors-No. of debtors', 'LC Backed Creditors Amount', 'LC Backed Creditors No of Creditors', 'Other than LC backed Creditors Amount', 'Other than LC backed Creditors No of Creditors', 'spares creditors', 'Total Creditors Amount', 'Total Creditors No of creditors', 'Sales Amount', 'Sales Unit', 'Purchases Amount', 'Purchases Unit', 'Other Channel finance details Bank Name', 'Other Channel details Value', 'Customer Advance', 'Supplier Advance', 'Stock Due Date', 'Outstanding in Other Limits', 'File Name', 'Received Date', 'Account No', 'ZMT', 'DP formula','Comments']})

            keys = {'CRN': 'CRN', 'NAME OF CUSTOMER': 'NAME OF CUSTOMER', 'SS_AS_ON_DATE': 'SS_AS_ON_DATE', 'Raw Material-< 90 DAYS': 'IMPORTED RAW MATERIAL_<90', 'Raw Material-90 - 120 DAYS': 'IMPORTED RAW MATERIAL_91-120', 'Raw Material-120 - 180 DAYS': 'IMPORTED RAW MATERIAL_<180', 'Raw Material-> 180 DAYS': 'IMPORTED RAW MATERIAL_>180', 'Raw Material-Unit/Quantity': 'IMPORTED RAW MATERIAL_Units', 'Work In Progress-< 60 DAYS': 'WORK IN PROGRESS_<60', 'Work In Progress 60- 90 DAYS': 'WORK IN PROGRESS_<90', 'Work In Progress-90 - 120 DAYS': 'WORK IN PROGRESS_<120', 'Work In Progress- 120 - 180 DAYS': 'WORK IN PROGRESS_<180', 'Work In Progress-> 180 DAYS': 'WORK IN PROGRESS_>180','Work In Progress- 90 - 180 DAYS':'WORK IN PROGRESS_<180','Work In Progress-Unit/Quantity': 'WORK IN PROGRESS_Units', 'Consumables & Packing Material-< 90 DAYS': 'CONSUMABLES_<90', 'Consumables & Packing Material-90 - 180 DAYS': 'CONSUMABLES_<180', 'Consumables & Packing Material-> 180 DAYS': 'CONSUMABLES_>180', 'Consumables & Packing Material-Unit/Quantity': 'CONSUMABLES_Units', 'Finished Goods-< 90 DAYS': 'Finished Goods_<120', 'Finished Goods-90 - 120 DAYS': 'Finished Goods_<120', 'Finished Goods-120 - 180 DAYS': 'Finished Goods_<180', 'Finished Goods-> 180 DAYS': 'Finished Goods_>180', 'Finished Goods-Unit/Quantity': 'Finished Goods_Units', 'Total Stock <30 DAYS': 'TOTAL_<30', 'Total Stock 30-90 DAYS': 'TOTAL_<90', 'Total Stock 120- 180 DAYS': 'TOTAL_<180',  'Total Stock-90 - 120 DAYS': 'TOTAL_<120', 'Total Stock-> 180 DAYS': 'TOTAL_>180', 'Total Stock-Unit/Quantity': 'TOTAL_Units','Third party Debtors-< 90 days': 'DEBITORS STATEMENT_OCR_Third Party debtors_<90', 'Third party Debtors-90 - 180 days': 'DEBITORS STATEMENT_OCR_Third Party debtors_<180', 'Third party Debtors-> 180 days': 'DEBITORS STATEMENT_OCR_Third Party debtors_>180', 'Third party Debtors-No. of debtors': 'DEBITORS STATEMENT_OCR_Third Party debtors_No. of debtors', 'Related party/ group concern Debtors-< 90 days': 'DEBITORS STATEMENT_OCR_Related party/group concern Debtors_< 90 days', 'Related party/ group concern Debtors-90 - 180 days': 'DEBITORS STATEMENT_OCR_Related party/group concern Debtors_< 180 days', 'Related party/ group concern Debtors-> 180 days': 'DEBITORS STATEMENT_OCR_Related party/group concern Debtors_> 180 days', 'Related party/ group concern Debtors-No. of debtors': 'DEBITORS STATEMENT_OCR_Related party/group concern Debtors_No. of debtors', 'Total debtors Debtors-< 30 days': 'DEBITORS STATEMENT_OCR_TOTAL_<30', 'Total debtors Debtors 30- 60 days': 'DEBITORS STATEMENT_OCR_TOTAL_<60', 'Total debtors Debtors 60- 90 days': 'DEBITORS STATEMENT_OCR_TOTAL_<90', 'Total debtors Debtors-90 - 120 days': 'DEBITORS STATEMENT_OCR_TOTAL_<120', 'Total debtors Debtors-120 - 180 days': 'DEBITORS STATEMENT_OCR_TOTAL_<180', 'Total debtors Debtors-> 180 days': 'DEBITORS STATEMENT_OCR_TOTAL_>180', 'Total debtors Debtors-No. of debtors': 'TOTAL_No. of debtors','Other than LC backed Creditors Amount': 'OTHER LC BACKED CREDITORS_AMOUNT', 'Total Creditors Amount': 'Total Creditors', 'Total Creditors No of creditors': 'TOTAL_No. of creditors', 'Sales Amount': 'Sales_Amount', 'Sales Unit': 'Sales_Units', 'Purchases Amount': 'Purchases_Amount', 'Purchases Unit': 'Purchases_Units', 'Other Channel finance details Bank Name': 'Other Channel finance Bank Name', 'Customer Advance': 'Customer Advance', 'Supplier Advance': 'Supplier Advance', 'Stock Due Date': 'Stock Due Date', 'Outstanding in Other Limits': 'Outstanding in Other Limits', 'Received Date': 'Received Date', 'Account No': 'ACCOUNT_NUMBER', 'ZMT': 'ZMT', 'DP formula': 'DP formula',"LC Backed Creditors Amount":"LC Backed Creditors_AMOUNT","LC Backed Creditors No of Creditors":"LC Backed Creditors_No. of creditors","Other than LC backed Creditors No of Creditors":"TOTAL_No. of creditors"}

            
            merged_data = {key: final_result.get(keys.get(key), np.nan) for key in keys}
            result_df = pd.DataFrame([merged_data])
            result_df = pd.DataFrame([merged_data])


            result_df = pd.concat([result_df, df], axis=1)
            logging.info(f"#####result dataframe {result_df}")
            for column, values in result_df.items():
                res[column] = values[0]
            others_query=f"SELECT `Sales_ocr`,`purchases_ocr`,`advances_ocr` from ocr where case_id='{case_id}'"
            others_res = extraction_db.execute_(others_query).to_dict(orient="records")[0]
            logging.info(others_res,'##############others_res')
            #sales_keys_set = set(json.loads(others_res['Sales_ocr']).keys())
            #purchases_keys_set = set(json.loads(others_res['purchases_ocr']).keys())
            #advances_keys_set = set(json.loads(others_res['advances_ocr']).keys())
            try:
                sales_keys_set = set(json.loads(others_res['Sales_ocr']).keys())
            except (KeyError, json.JSONDecodeError):
                sales_keys_set = set()

            try:
                purchases_keys_set = set(json.loads(others_res['purchases_ocr']).keys())
            except (KeyError, json.JSONDecodeError):
                purchases_keys_set = set()
            try:
                advances_keys_set = set(json.loads(others_res['advances_ocr']).keys())
            except (KeyError, json.JSONDecodeError):
                advances_keys_set = set()
            # Merging sets
            merged_keys_set = sales_keys_set.union(purchases_keys_set, advances_keys_set)

            # Converting the set to a list
            merged_keys_list = list(merged_keys_set)

            # Printing merged keys
            # print("Merged Keys:", merged_keys_list)
            result_keys_others = [key for key, value in keys.items() if value in merged_keys_list]
            logging.info(result_keys_others,'#####################result keys')

            dp_query=f"SELECT `dp_formula` from ocr where case_id='{case_id}'"
            dp_res = extraction_db.execute_(dp_query).to_dict(orient="records")[-1]
            if dp_res['dp_formula'] != None:
                equation = dp_res['dp_formula'] .replace('-',' ')
                equation =''.join(equation.split('(')[1:])
                equation_list1 = equation.split('+')
                equation_list = []
                for term in equation_list1:
                    terms_with_spaces = term.split()
                    processed_terms = [''.join(char for char in t if char.isalnum() or char == '_') for t in terms_with_spaces]
                    processed_terms = list(filter(None, processed_terms))
                    equation_list.extend(processed_terms)
                logging.info(f"equation_list##{equation_list}")
                consumer_field = {'TS_120_180':'TOTAL_<180','RM_LESS_90': 'IMPORTED RAW MATERIAL_<90', 'RM_120_180':'IMPORTED RAW MATERIAL_<180','RM_90_120': 'IMPORTED RAW MATERIAL_<120', 'RM_90_180': 'IMPORTED RAW MATERIAL_<180', 'RM_GREATER_180': 'IMPORTED RAW MATERIAL_>180', 'WIP_LESS_60': 'WORK IN PROGRESS_<60', 'WIP_60_90': 'WORK IN PROGRESS_<90', 'WIP_90_120': 'WORK IN PROGRESS_<120', 'WIP_LESS_90': 'WORK IN PROGRESS_<90', 'WIP_90_180': 'WORK IN PROGRESS_<180', 'WIP_120_180':'WORK IN PROGRESS_<180','WIP_GREATER_180': 'WORK IN PROGRESS_>180', 'CPM_LESS_90': 'CONSUMABLES_<90', 'CPM_90_180': 'CONSUMABLES_<180', 'CPM_GREATER_180': 'CONSUMABLES_>180', 'FG_LESS_90': 'Finished Goods_<120', 'FG_90_120': 'Finished Goods_<120', 'FG_120_180': 'Finished Goods_<180', 'FG_GREATER_180': 'Finished Goods_>180', 'TS_LESS_30': 'TOTAL_<30', 'TS_30_90': 'TOTAL_<90', 'TS_LESS_90': 'TOTAL_<90', 'TS_GREATER_180': 'TOTAL_>180', 'TPD_LESS_90': 'DEBITORS STATEMENT_OCR_Third Party debtors_<90','TPD_GREATER_90':'DEBITORS STATEMENT_OCR_Third Party debtors_<180', 'TPD_90_180': 'DEBITORS STATEMENT_OCR_Third Party debtors_<180', 'RP_GCD_LESS_90': 'DEBITORS STATEMENT_OCR_Related party/group concern Debtors_< 90 days', 'RP_GCD_90_180': 'DEBITORS STATEMENT_OCR_Related party/group concern Debtors_< 180 days', 'RP_GCD_GREATER_180': 'DEBITORS STATEMENT_OCR_Related party/group concern Debtors_> 180 days', 'TDD_LESS_30': 'DEBITORS STATEMENT_OCR_TOTAL_<30','TDD_30_60': 'DEBITORS STATEMENT_OCR_TOTAL_<60', 'TDD_60_90': 'DEBITORS STATEMENT_OCR_TOTAL_<90', 'TDD_90_120': 'DEBITORS STATEMENT_OCR_TOTAL_<120', 'TDD_GREATER_120': 'DEBITORS STATEMENT_OCR_TOTAL_<180', 'TDD_120_180': 'DEBITORS STATEMENT_OCR_TOTAL_<180', 'TDD_GREATER_180': 'DEBITORS STATEMENT_OCR_TOTAL_>180', 'TOTAL_CREDITORS_AMT': 'Total Creditors', 'CUSTOMER_ADVANCE': 'Customer Advance', 'SUPPLIER_ADVANCE': 'Supplier Advance'}
                categories=['RM_LESS_90', 'RM_90_120', 'RM_120_180', 'RM_GREATER_180', 'WIP_LESS_60', 'WIP_60_90', 'WIP_90_120', 'WIP_120_180', 'WIP_GREATER_180', 'CPM_LESS_90', 'CPM_90_180', 'CPM_GREATER_180', 'FG_LESS_90', 'FG_90_120', 'FG_120_180', 'FG_GREATER_180', 'TS_LESS_30', 'TS_30_90', 'TS_120_180', 'TS_GREATER_180', 'TPD_LESS_90', 'TPD_90_180', 'TPD_GREATER_90', 'RP_GCD_LESS_90', 'RP_GCD_90_180', 'RP_GCD_GREATER_180', 'TDD_LESS_30', 'TDD_30_60', 'TDD_60_90', 'TDD_90_120', 'TDD_120_180', 'TDD_GREATER_180', 'TOTAL_CREDITORS_AMT', 'CUSTOMER_ADVANCE', 'SUPPLIER_ADVANCE']

                non_dp_formula_fields=[]
                for i in categories:
                    if i not in equation_list:
                        non_dp_formula_fields.append(consumer_field[i])
                data_dict= {'IMPORTED RAW MATERIAL_<90': 'Raw Material-< 90 DAYS','IMPORTED RAW MATERIAL_<120': 'Raw Material-90 - 120 DAYS', 'IMPORTED RAW MATERIAL_<180': 'Raw Material-120 - 180 DAYS', 'IMPORTED RAW MATERIAL_>180': 'Raw Material-> 180 DAYS', 'WORK IN PROGRESS_<90': 'Work In Progress 60- 90 DAYS', 'WORK IN PROGRESS_<60': 'Work In Progress-< 60 DAYS', 'WORK IN PROGRESS_<90': 'Work In Progress 60- 90 DAYS', 'WORK IN PROGRESS_<120': 'Work In Progress-90 - 120 DAYS', 'WORK IN PROGRESS_<180': 'Work In Progress- 120 - 180 DAYS', 'WORK IN PROGRESS_>180': 'Work In Progress-> 180 DAYS', 'CONSUMABLES_<90': 'Consumables & Packing Material-< 90 DAYS', 'CONSUMABLES_<180': 'Consumables & Packing Material-90 - 180 DAYS', 'CONSUMABLES_>180': 'Consumables & Packing Material-> 180 DAYS', 'Finished Goods_<120': 'Finished Goods-< 90 DAYS', 'Finished Goods_<120': 'Finished Goods-90 - 120 DAYS', 'Finished Goods_<180': 'Finished Goods-120 - 180 DAYS', 'Finished Goods_>180': 'Finished Goods-> 180 DAYS', 'TOTAL_<30': 'Total Stock <30 DAYS', 'TOTAL_<90': 'Total Stock 30-90 DAYS',  'TOTAL_<180': 'Total Stock 120- 180 DAYS', 'TOTAL_>180': 'Total Stock-> 180 DAYS', 'DEBITORS STATEMENT_OCR_Third Party debtors_<90': 'Third party Debtors-< 90 days', 'DEBITORS STATEMENT_OCR_Third Party debtors_<180': 'Third party Debtors-90 - 180 days', 'DEBITORS STATEMENT_OCR_Related party/group concern Debtors_< 90 days': 'Related party/ group concern Debtors-< 90 days', 'DEBITORS STATEMENT_OCR_Related party/group concern Debtors_< 180 days': 'Related party/ group concern Debtors-90 - 180 days', 'DEBITORS STATEMENT_OCR_Related party/group concern Debtors_> 180 days': 'Related party/ group concern Debtors-> 180 days', 'DEBITORS STATEMENT_OCR_TOTAL_<30': 'Total debtors Debtors-< 30 days', 'DEBITORS STATEMENT_OCR_TOTAL_<60': 'Total debtors Debtors 30- 60 days', 'DEBITORS STATEMENT_OCR_TOTAL_<90': 'Total debtors Debtors 60- 90 days', 'DEBITORS STATEMENT_OCR_TOTAL_<120': 'Total debtors Debtors-90 - 120 days', 'DEBITORS STATEMENT_OCR_TOTAL_<180': 'Total debtors Debtors-120 - 180 days', 'DEBITORS STATEMENT_OCR_TOTAL_>180':'Total debtors Debtors-> 180 days', 'Total Creditors': 'Total Creditors Amount', 'Customer Advance': 'Customer Advance', 'Supplier Advance': 'Supplier Advance'}

                for field in non_dp_formula_fields: 
                                       
                    if 'OTHERS' in equation_list:
                        for M in result_keys_others:
                            if 'Unit' in M:
                                print(M,'###########units')
                                res[M]=None
                                
                            print(M,'others tab fields')
                            if M not in ('Sales Amount','Purchases Amount' ,'Outstanding in Other Limits' ):
                                res[M]=None                                        
                    else:                        
                        for f_other in result_keys_others:
                            print(f_other,'#########f_other')
                            if f_other not in ('Sales Amount','Purchases Amount'):
                                res[f_other]=None
                    res[data_dict[field]] = None

                    res['LC Backed Creditors No of Creditors']=None
                    res['Total Creditors No of creditors']=None
                    # res['Stock Due Date']='######'
                    res['Third party Debtors-No. of debtors']=None
            for i in res:
                if (i not in data_dict.values()) and (i not in df) and (i not in result_keys_others):
                    res[i]=None
            res = res.drop(columns=['CC_LIMIT','INSURANCE_AMOUNT','DEBTORS_MARGIN','STOCK_MARGIN','DEVIATION_APPLICABLE','REQUEST_ID'])
            logging.info(f"###result is {res}")
            logging.info(f"result dataframe####  {res}")
            queue_name="consumer"

        
        if segment == 'RBG':
            qry1=f"SELECT `OUTSTANDING_IN_SPG_CF_INF_TA` as 'OUTSTANDING IN SPG CF INF TA',`Computed_DP` as 'Computed DP',`system_dp` as 'System DP', `Sanction_Limit` as 'Sanction Limit',`Final_DP` as 'Final DP', `Due_Date` as 'Stock Due Date',`co_mail_time` as 'Received Date', `cc_limit` as 'CC_LIMIT',`insurance_amount` as 'INSURANCE_AMOUNT' , `Debtors_margin` as 'DEBTORS_MARGIN',`Stock_Margins` as 'STOCK_MARGIN',`deviation_applicable` as 'DEVIATION APPLICABLE',`Crn` as 'CRN',`customer_name` as 'NAME OF CUSTOMER' ,`request_id` as 'REQUEST ID' , `Account_Number` as 'ACCOUNT NUMBER',`division` as 'ZMT',`Dp_formula` as 'DP Formula',`date` as 'STOCK STATEMENT MONTH AS ON DATE' from ocr where `case_id` = '{case_id}';"
            df = extraction_db.execute_(qry1).to_dict(orient="records")[-1]
            received_date_str = df['Received Date']
            stock_due_date=df['Stock Due Date']
            ss_on_date = df['STOCK STATEMENT MONTH AS ON DATE']

            try:
                
                df['Stock Due Date'] = convert_to_custom_format(stock_due_date)
            
                df['Received Date'] = convert_to_custom_format(received_date_str)
                df['STOCK STATEMENT MONTH AS ON DATE'] = convert_to_custom_format(ss_on_date)

            except Exception as e:
                print(f'{e}#########exception')
        

        
            df = pd.DataFrame([df])
            logging.info(f"in if block segement is {segment}")
            keys={'Last 3 month purchases':'Last 3 Month Purchases','Last 3 month Sales':'Last 3 Month Sales','STOCKS IMPORTED RM 0 5 DAYS': 'Stocks Imported RM 0-5 days', 'STOCKS IMPORTED RM 6 30 DAYS': 'Stocks Imported RM 6-30 days', 'STOCKS IMPORTED RM 31 60 DAYS': 'Stocks Imported RM 31-60 days', 'STOCKS IMPORTED RM 61 90 DAYS': 'Stocks Imported RM 61-90 days', 'STOCKS IMPORTED RM 91 120 DAYS': 'Stocks Imported RM 91-120 days', 'STOCKS IMPORTED RM 121 150 DAYS': 'Stocks Imported 121-150 days', 'STOCKS IMPORTED RM 151 180 DAYS': 'Stocks Imported RM 151-180 days', 'STOCKS IMPORTED RM GREATER THAN180 DAYS': 'Stocks Imported RM >180 days', 'STOCKS WIP 0 5 DAYS': 'Stocks WIP <5 days', 'STOCKS WIP 6 30 DAYS': 'Stocks WIP <30 days', 'STOCKS WIP 31 60 DAYS': 'Stocks WIP <60 days', 'STOCKS WIP 61 90 DAYS': 'Stocks WIP 61-90 days', 'STOCKS WIP 91 120 DAYS': 'Stocks WIP 91-120 days', 'STOCKS WIP 121 150 DAYS': 'Stocks WIP 121-150 days', 'STOCKS WIP 151 180 DAYS': 'Stocks WIP 151-180 days', 'STOCKS WIP GREATER THAN180 DAYS': 'Stocks WIP >180 days', 'STOCKS VEHICLE STOCKS 0 5 DAYS': 'Stocks Vehicle Stocks 0-5 days', 'STOCKS VEHICLE STOCKS 6 30 DAYS': 'Stocks Vehicle Stocks 6-30 days', 'STOCKS VEHICLE STOCKS 31 60 DAYS': 'Stocks Vehicle Stocks 31-60 days', 'STOCKS VEHICLE STOCKS 61 90 DAYS': 'Stocks Vehicle Stocks 61-90 days', 'STOCKS VEHICLE STOCKS 91 120 DAYS': 'Stocks Vehicle Stocks 91-120 days', 'STOCKS VEHICLE STOCKS 121 150 DAYS': 'Stocks Vehicle Stocks 121-150 days', 'STOCKS VEHICLE STOCKS 151 180 DAYS': 'Stocks Vehicle Stocks 151-180 days', 'STOCKS VEHICLE STOCKS GREATER THAN180 DAYS': 'Stocks Vehicle Stocks >180 days', 'STOCKS SPARES STOCKS 0 5 DAYS': 'Stocks Spares Stocks 0-5 days', 'STOCKS SPARES STOCKS 6 30 DAYS': 'Stocks Spares Stocks 6-30 days', 'STOCKS SPARES STOCKS 31 60 DAYS': 'Stocks Spares Stocks 31-60 days', 'STOCKS SPARES STOCKS 61 90 DAYS': 'Stocks Spares Stocks 61-90 days', 'STOCKS SPARES STOCKS 91 120 DAYS': 'Stocks Spares Stocks 91-120 days', 'STOCKS SPARES STOCKS 121 150 DAYS': 'Stocks Spares Stocks 121-150 days', 'STOCKS SPARES STOCKS 151 180 DAYS': 'Stocks Spares Stocks 151 180 days', 'STOCKS SPARES STOCKS GREATER THAN180 DAYS': 'Stocks Spares Stocks >180 days', 'STOCKS FG TRADING 0 5 DAYS': 'Stocks FG Trading 0-5 days', 'STOCKS FG TRADING 6 30 DAYS': 'Stocks FG Trading <30 days', 'STOCKS FG TRADING 31 60 DAYS': 'Stocks FG Trading <60 days', 'STOCKS FG TRADING 61 90 DAYS': 'Stocks FG Trading <90 days', 'STOCKS FG TRADING 91 120 DAYS': 'Stocks FG Trading 91-120 days', 'STOCKS FG TRADING 121 150 DAYS': 'Stocks FG Trading 121-150 days', 'STOCKS FG TRADING 151 180 DAYS': 'Stocks FG Trading 151-180 days', 'STOCKS FG TRADING GREATER THAN180 DAYS': 'Stocks FG Trading >180 days', 'STOCKS CONSUMABLES PACKING MATERIAL': 'Stocks Consumables Packing Material', 'STOCKS STOCK IN TRANSIT': 'Stocks Stock In Transit', 'STOCKS RM LESS THAN 90 DAYS': 'Stocks RM LESS THAN 90 days', 'STOCKS TOTAL SPARES STOCK': 'Stocks Total Spares Stock', 'STOCK LESS THAN 60 DAYS': 'Total Stocks <60 days', 'STOCK LESS THAN 90 DAYS': 'Total Stocks <90 days', 'STOCK LESS THAN 120 DAYS': 'Total Stocks <120 days', 'STOCK LESS THAN 180 DAYS': 'Total Stocks <180 days', 'STOCKS WIP LESS THAN 180 DAYS': 'Stocks WIP LESS THAN 180 days', 'TOTAL STOCKS': 'Total Stocks', 'DEBTORS 0 5 DAYS': 'DEBITORS STATEMENT_OCR_Total Debtors <5 days', 'DEBTORS 6 30 DAYS': 'DEBITORS STATEMENT_OCR_Total Debtors <30 days', 'DEBTORS 31 60 DAYS': 'DEBITORS STATEMENT_OCR_Total Debtors <60 days', 'DEBTORS 61 90 DAYS': 'DEBITORS STATEMENT_OCR_Total Debtors <90 days', 'DEBTORS 91 120 DAYS': 'DEBITORS STATEMENT_OCR_Total Debtors <120 days', 'DEBTORS 121 150 DAYS': 'DEBITORS STATEMENT_OCR_Debtors 121 150 days', 'DEBTORS 151 180 DAYS': 'DEBITORS STATEMENT_OCR_Total Debtors <180 days', 'TOTAL DEBTORS': 'DEBITORS STATEMENT_OCR_Total Debtors', 'DEBTORS DEDUCTION INTERGROUP RECEIVABLES 1': 'DEBITORS STATEMENT_OCR_Debtors Deduction Intergroup Receivables 1', 'DEBTORS DEDUCTION INTERGROUP RECEIVABLES 2': 'DEBITORS STATEMENT_OCR_Debtors Deduction Intergroup Receivables 2', 'DEBTORS LESS THAN 30 DAYS': 'DEBITORS STATEMENT_OCR_DEBITORS STATEMENT_OCR_Total Debtors <30 days', 'DEBTORS LESS THAN 60 DAYS':'DEBITORS STATEMENT_OCR_Total Debtors <60 days', 'DEBTORS LESS THAN 90 DAYS': 'DEBITORS STATEMENT_OCR_Total Debtors <90 days', 'DEBTORS GREATER THAN 90 DAYS': 'DEBITORS STATEMENT_OCR_Total Debtors <120 days', 'DEBTORS GREATER THAN 90 LESS THAN 150 DAYS': 'DEBITORS STATEMENT_OCR_Total Debtors <150 days', 'DEBTORS LESS THAN 120 DAYS': 'DEBITORS STATEMENT_OCR_Total Debtors <120 days', 'DEBTORS GREATER THAN 120 LESS THAN 150 DAYS': 'DEBITORS STATEMENT_OCR_Debtors 121 150 days', 'DEBTORS LESS THAN 180 DAYS': 'DEBITORS STATEMENT_OCR_Total Debtors <180 days', 'TOTAL DEBTORS': 'DEBITORS STATEMENT_OCR_Total Debtors', 'DEBTORS UNBILLED DEBTORS LESS THAN60 DAYS': 'DEBITORS STATEMENT_OCR_Debtors Unbilled Debtors Less Than60 days', 'DEBTORS UNBILLED DEBTORS LESS THAN90DAYS': 'DEBITORS STATEMENT_OCR_Total Debtors_Debtors Unbilled Debtors Less Than90days', 'DEBTORS TOTAL UNBILLED DEBTORS': 'DEBITORS STATEMENT_OCR_Total Debtors_Debtors Total Unbilled Debtors', 'CREDITORS UPTO 90 DAYS': 'Creditors Upto 90 days', 'CREDITORS UPTO 91 150 DAYS': 'Creditors Upto 91 150 days', 'CREDITORS GREATER THAN151 DAYS': 'Creditors Greater Than151 days', 'TOTAL CREDITORS': 'Total Creditors', 'DEBTORS ADVANCE PAID TO SUPPLIER': 'DEBITORS STATEMENT_OCR_Total Debtors_Debtors Advance Paid To Supplier', 'SECURITY DEPOSIT': 'Security Deposit', 'CREDITORS ADVANCES RECEIVED FROM CUSTOMERS': 'Creditors Advances Received From Customers', 'DEBTORS DEDUCTION BILLS DISCOUNTED': 'DEBITORS STATEMENT_OCR_Total Debtors_Debtors Deduction Bills Discounted', 'DEBTORS DEDUCTION BILLS FACTORED ASSIGNED': 'DEBITORS STATEMENT_OCR_Total Debtors_Debtors Deduction Bills Factores Assigned', 'LIMITS OS WITH KMBL BANKS 1': 'Limits OS With KMBL Banks 1', 'LIMITS OS WITH KMBL BANKS 2': 'Limits OS With KMBL Banks 2', 'LIMITS OS WITH KMBL BANKS 3': 'Limits OS With KMBL Banks 3', 'LIMITS OS WITH OTHER BANKS 1': 'Limits OS With Other Banks 1', 'LIMITS OS WITH OTHER BANKS 2': 'Limits OS With Other Banks 2', 'LIMITS OS WITH OTHER BANKS 3': 'Limits OS With Other Banks 3', 'LIMITS OS WITH OTHER BANKS 4': 'Limits OS With Other Banks 4', 'LIMITS OS WITH OTHER BANKS 5': 'Limits OS With Other Banks 5'}
            merged_data = {key: final_result.get(keys.get(key), np.nan) for key in keys}
            values_dict = {key: value[0] if isinstance(value, pd.Series) else value for key, value in merged_data.items()}
            result_df = pd.DataFrame([values_dict])
            print(result_df,'result_df12-------------')
            columns=['REQUEST ID', 'CRN', 'ACCOUNT NUMBER', 'STOCK STATEMENT MONTH AS ON DATE', 'STOCKS IMPORTED RM 0 5 DAYS', 'STOCKS IMPORTED RM 6 30 DAYS', 'STOCKS IMPORTED RM 31 60 DAYS', 'STOCKS IMPORTED RM 61 90 DAYS', 'STOCKS IMPORTED RM 91 120 DAYS', 'STOCKS IMPORTED RM 121 150 DAYS', 'STOCKS IMPORTED RM 151 180 DAYS', 'STOCKS IMPORTED RM GREATER THAN180 DAYS', 'STOCKS WIP 0 5 DAYS', 'STOCKS WIP 6 30 DAYS', 'STOCKS WIP 31 60 DAYS', 'STOCKS WIP 61 90 DAYS', 'STOCKS WIP 91 120 DAYS', 'STOCKS WIP 121 150 DAYS', 'STOCKS WIP 151 180 DAYS', 'STOCKS WIP GREATER THAN180 DAYS', 'STOCKS VEHICLE STOCKS 0 5 DAYS', 'STOCKS VEHICLE STOCKS 6 30 DAYS', 'STOCKS VEHICLE STOCKS 31 60 DAYS', 'STOCKS VEHICLE STOCKS 61 90 DAYS', 'STOCKS VEHICLE STOCKS 91 120 DAYS', 'STOCKS VEHICLE STOCKS 121 150 DAYS', 'STOCKS VEHICLE STOCKS 151 180 DAYS', 'STOCKS VEHICLE STOCKS GREATER THAN180 DAYS', 'STOCKS SPARES STOCKS 0 5 DAYS', 'STOCKS SPARES STOCKS 6 30 DAYS', 'STOCKS SPARES STOCKS 31 60 DAYS', 'STOCKS SPARES STOCKS 61 90 DAYS', 'STOCKS SPARES STOCKS 91 120 DAYS', 'STOCKS SPARES STOCKS 121 150 DAYS', 'STOCKS SPARES STOCKS 151 180 DAYS', 'STOCKS SPARES STOCKS GREATER THAN180 DAYS', 'STOCKS FG TRADING 0 5 DAYS', 'STOCKS FG TRADING 6 30 DAYS', 'STOCKS FG TRADING 31 60 DAYS', 'STOCKS FG TRADING 61 90 DAYS', 'STOCKS FG TRADING 91 120 DAYS', 'STOCKS FG TRADING 121 150 DAYS', 'STOCKS FG TRADING 151 180 DAYS', 'STOCKS FG TRADING GREATER THAN180 DAYS', 'STOCKS CONSUMABLES PACKING MATERIAL', 'STOCKS STOCK IN TRANSIT', 'STOCKS OTHER STOCK 1', 'STOCKS OTHER STOCK 2', 'STOCKS RM LESS THAN 90 DAYS', 'STOCKS RM 120 180 DAYS', 'STOCKS WIP LESS THAN 60 DAYS', 'STOCKS FG LESS THAN 90 DAYS', 'STOCKS FG 120 180 DAYS', 'STOCKS VEHICLE STOCK UPTO 120 DAYS', 'STOCKS SPARES STOCK 6 60 DAYS', 'STOCKS SPARES STOCK LESS THAN 60 DAYS', 'STOCKS TOTAL SPARES STOCK', 'STOCK LESS THAN 60 DAYS', 'STOCK LESS THAN 90 DAYS', 'STOCK LESS THAN 120 DAYS', 'STOCK LESS THAN 180 DAYS', 'STOCKS WIP LESS THAN 180 DAYS', 'TOTAL STOCKS', 'DEBTORS 0 5 DAYS', 'DEBTORS 6 30 DAYS', 'DEBTORS 31 60 DAYS', 'DEBTORS 61 90 DAYS', 'DEBTORS 91 120 DAYS', 'DEBTORS 121 150 DAYS', 'DEBTORS 151 180 DAYS', 'DEBTORS GREATER THAN180 DAYS', 'DEBTORS SPARES DEBTORS 0 5 DAYS', 'DEBTORS SPARES DEBTORS 6 30 DAYS', 'DEBTORS SPARES DEBTORS 31 60 DAYS', 'DEBTORS SPARES DEBTORS 61 90 DAYS', 'DEBTORS SPARES DEBTORS 91 120 DAYS', 'DEBTORS SPARES DEBTORS 121 150 DAYS', 'DEBTORS SPARES DEBTORS 151 180 DAYS', 'DEBTORS SPARES DEBTORS GREATER THAN180 DAYS', 'DEBTORS VEHICLE DEBTORS 0 5 DAYS', 'DEBTORS VEHICLE DEBTORS 6 30 DAYS', 'DEBTORS VEHICLE DEBTORS 31 60 DAYS', 'DEBTORS VEHICLE DEBTORS 61 90 DAYS', 'DEBTORS VEHICLE DEBTORS 91 120 DAYS', 'DEBTORS VEHICLE DEBTORS 121 150 DAYS', 'DEBTORS VEHICLE DEBTORS 151 180 DAYS', 'DEBTORS VEHICLE DEBTORS GREATER THAN180 DAYS', 'DEBTORS SPARES DEBTORS 6 60 DAYS', 'DEBTORS WORKSHOP SERVICE DEBTORS LESS THAN60 DAYS', 'DEBTORS WORKSHOP SERVICE DEBTORS 6 60 DAYS', 'DEBTORS WORKSHOP SERVICE DEBTORS 60 90 DAYS', 'DEBTORS WORKSHOP SERVICE 90 120 DAYS', 'DEBTORS WORKSHOP SERVICE120 150 DAYS', 'DEBTORS VEHICLE DEBTORS 6 60 DAYS', 'DEBTORS VEHICLE DEBTORS LESS THAN 15 DAYS', 'DEBTORS SPARE DEBTORS LESS THAN 15 DAYS', 'DEBTORS VEHICLE DEBTORS LESS THAN 30 DAYS', 'DEBTORS SPARE DEBTORS LESS THAN 30 DAYS', 'DEBTORS DEBTORS EXCHANGE TRACTORS', 'DEBTORS VEHICLE DEBTORS LESS THAN 45 DAYS', 'DEBTORS VEHICLE DEBTORS LESS THAN 60 DAYS', 'DEBTORS SPARES DEBTORS LESS THAN 45 DAYS', 'DEBTORS SPARES DEBTORS LESS THAN 60 DAYS', 'DEBTORS DEDUCTION INTERGROUP RECEIVABLES 1', 'DEBTORS DEDUCTION INTERGROUP RECEIVABLES 2', 'DEBTORS LESS THAN 30 DAYS', 'DEBTORS LESS THAN 60 DAYS', 'DEBTORS LESS THAN 90 DAYS', 'DEBTORS GREATER THAN 90 DAYS', 'DEBTORS GREATER THAN 90 LESS THAN 150 DAYS', 'DEBTORS LESS THAN 120 DAYS', 'DEBTORS GREATER THAN 120 LESS THAN 150 DAYS', 'DEBTORS LESS THAN 180 DAYS', 'DEBTORS EXCHANGE TRACTORS', 'TOTAL DEBTORS', 'DEBTORS UNBILLED DEBTORS LESS THAN60 DAYS', 'DEBTORS UNBILLED DEBTORS LESS THAN90DAYS', 'DEBTORS TOTAL UNBILLED DEBTORS', 'CREDITORS SUNDRY CREDITORS FOR GOODS', 'CREDITORS SUNDRY CREDITORS FOR CAPITAL EXPENSES', 'CREDITORS SUNDRY CREDITORS FOR EXPENSES', 'CREDITORS LC BACKED CREDITORS FOR GOODS', 'CREDITORS LC BACKED CREDITORS FOR CAPITAL EXPENSES', 'CREDITORS LC BACKED CREDITORS FOR EXPENSES', 'CREDITORS OTHER CREDITORS 1', 'CREDITORS OTHER CREDITORS 2', 'CREDITORS SPARES CREDITORS', 'CREDITORS GROUP CONCERN CREDITORS', 'CREDITORS UPTO 90 DAYS', 'CREDITORS UPTO 91 150 DAYS', 'CREDITORS GREATER THAN151 DAYS', 'TOTAL CREDITORS', 'DEBTORS ADVANCE PAID TO SUPPLIER', 'SECURITY DEPOSIT', 'CVD DUTY REFUNDABLE', 'CREDITORS ADVANCES RECEIVED FROM CUSTOMERS', 'DEBTORS ADVANCE TRACTORS', 'DEBTORS DEDUCTION BILLS DISCOUNTED', 'DEBTORS DEDUCTION BILLS FACTORED ASSIGNED', 'LIMITS OS WITH KMBL BANKS 1', 'LIMITS OS WITH KMBL BANKS 2', 'LIMITS OS WITH KMBL BANKS 3', 'LIMITS OS WITH OTHER BANKS 1', 'LIMITS OS WITH OTHER BANKS 2', 'LIMITS OS WITH OTHER BANKS 3', 'LIMITS OS WITH OTHER BANKS 4', 'LIMITS OS WITH OTHER BANKS 5', 'OUTSTANDING IN SPG CF INF TA', 'ADVANCE TRACTORS', 'ADVANCE TRACTORS LESS THAN 90 DAYS', 'ADVANCE TRACTORS GREATER THAN91 DAYS TO 120 DAYS', 'ADVANCE TRACTORS GREATER THAN121 DAYS TO 150 DAYS', 'PURCHASES 1', 'PURCHASES 2', 'PURCHASES 3', 'PURCHASE AMOUNT', 'PURCHASE AMOUNT VEHICLES', 'PURCHASE AMOUNT SPARES', 'PURCHASE AMOUNT WORKSHOP SERVICE', 'TOTAL PURCHASES', 'SALES 1', 'SALES 2', 'SALES 3', 'SALES AMOUNT VEHICLES', 'SALES AMOUNT SPARES', 'SALES AMOUNT WORKSHOP SERVICE', 'TOTAL SALES', 'INVENTORY HOLDING PERIOD', 'COLLECTION PERIOD', 'CREDIT PERIOD', 'NET CYCLE', 'NET CREDIT', 'SALES REALISATION', 'NET CREDIT TO SALES REALISATION', 'OD LIMIT', 'CC LIMIT', 'ADDITIONAL FIELDS NUM1', 'ADDITIONAL FIELDS NUM2', 'ADDITIONAL FIELDS NUM3', 'ADDITIONAL FIELDS NUM4', 'ADDITIONAL FIELDS NUM5', 'ADDITIONAL FIELDS NUM6', 'ADDITIONAL FIELDS NUM7', 'ADDITIONAL FIELDS NUM8', 'ADDITIONAL FIELDS NUM9', 'ADDITIONAL FIELDS NUM10', 'ADDITIONAL FIELDS TXT1', 'ADDITIONAL FIELDS TXT2', 'ADDITIONAL FIELDS TXT3', 'ADDITIONAL FIELDS TXT4', 'ADDITIONAL FIELDS TXT5', 'ADDITIONAL FIELDS TXT6', 'ADDITIONAL FIELDS TXT7', 'ADDITIONAL FIELDS TXT8', 'ADDITIONAL FIELDS TXT9', 'ADDITIONAL FIELDS TXT10', 'DEVIATION APPLICABLE', 'DEBTORS<90 days as per SS', 'Last 3 month Sales', 'Debtors > Sales', 'Last 3 month purchases', 'Stock > Purchases', 'System DP', 'Computed DP', 'Sanction Limit', 'Final DP', 'DP Formula']

            res = pd.DataFrame({key: [] for key in columns})
            result_df = pd.concat([result_df, df], axis=1) 
            print(f"columns list {result_df.columns.tolist()}")
            res = res.append(result_df, ignore_index=True)


            dp_query=f"SELECT `dp_formula` from ocr where case_id='{case_id}'"
            dp_res = extraction_db.execute_(dp_query).to_dict(orient="records")[-1]
            logging.info(f"{dp_res}")
            if dp_res['dp_formula'] !=  None:
        
                equation_list = re.findall(r'\b\w+_\w+\b', dp_res['dp_formula'] )
                rbg_fields={'STOCKS_RM_0_5_DAYS': 'STOCKS IMPORTED RM 0 5 DAYS', 'STOCKS_RM_6_30_DAYS': 'STOCKS IMPORTED RM 6 30 DAYS', 'STOCKS_RM_31_60_DAYS': 'STOCKS IMPORTED RM 31 60 DAYS', 'STOCKS_RM_61_90_DAYS': 'STOCKS IMPORTED RM 61 90 DAYS', 'STOCKS_RM_91_120_DAYS': 'STOCKS IMPORTED RM 91 120 DAYS', 'STOCKS_RM_121_150_DAYS': 'STOCKS IMPORTED RM 121 150 DAYS', 'STOCKS_RM_151_180_DAYS': 'STOCKS IMPORTED RM 151 180 DAYS', 'STOCKS_IMPORTED_RM_GREATER_THAN180_DAYS': 'STOCKS IMPORTED RM GREATER THAN180 DAYS', 'STOCKS_WIP_0_5_DAYS': 'STOCKS WIP 0 5 DAYS', 'STOCKS_WIP_6_30_DAYS': 'STOCKS WIP 6 30 DAYS', 'STOCKS_WIP_31_60_DAYS': 'STOCKS WIP 31 60 DAYS', 'STOCKS_WIP_61_90_DAYS': 'STOCKS WIP 61 90 DAYS', 'STOCKS_WIP_91_120_DAYS': 'STOCKS WIP 91 120 DAYS', 'STOCKS_WIP_121_150_DAYS': 'STOCKS WIP 121 150 DAYS', 'STOCKS_WIP_151_180_DAYS': 'STOCKS WIP 151 180 DAYS', 'STOCKS_WIP_GREATER_THAN180_DAYS': 'STOCKS WIP GREATER THAN180 DAYS', 'STOCKS_VEHICLE_STOCKS_0_5_DAY': 'STOCKS VEHICLE STOCKS 0 5 DAYS', 'STOCKS_VEHICLE_STOCKS_6_30_DAY': 'STOCKS VEHICLE STOCKS 6 30 DAYS', 'STOCKS_VEHICLE_STOCKS_31_60DAY': 'STOCKS VEHICLE STOCKS 31 60 DAYS', 'STOCKS_VEHICLE_STOCKS_61_90DAY': 'STOCKS VEHICLE STOCKS 61 90 DAYS', 'STOCKS_VEHICLE_STOCKS_91_120_DAY': 'STOCKS VEHICLE STOCKS 91 120 DAYS', 'STOCKS_VEHICLE_STOCK121_150DAY': 'STOCKS VEHICLE STOCKS 121 150 DAYS', 'STOCKS_VEHICLE_STOCK151_180DAY': 'STOCKS VEHICLE STOCKS 151 180 DAYS', 'STOCKS_VEHICLE_STOCKS_GREATER_THAN180_DAYS': 'STOCKS VEHICLE STOCKS GREATER THAN180 DAYS', 'STOCKS_SPARES_STOCKS_0_5_DAY': 'STOCKS SPARES STOCKS 0 5 DAYS', 'STOCKS_SPARES_STOCKS_6_30_DAY': 'STOCKS SPARES STOCKS 6 30 DAYS', 'STOCKS_SPARES_STOCKS_31_60_DAY': 'STOCKS SPARES STOCKS 31 60 DAYS', 'STOCKS_SPARES_STOCKS_61_90_DAY': 'STOCKS SPARES STOCKS 61 90 DAYS', 'STOCKS_SPARES_STOCKS_91_120DAY': 'STOCKS SPARES STOCKS 91 120 DAYS', 'STOCKS_SPARES_STOCKS_121_150_DAY': 'STOCKS SPARES STOCKS 121 150 DAYS', 'STOCKS_SPARES_STOCKS151_180DAYs': 'STOCKS SPARES STOCKS 151 180 DAYS', 'STOCKS_SPARES_STOCKS_GREATER_THAN180_DAYS': 'STOCKS SPARES STOCKS GREATER THAN180 DAYS', 'STOCKS_FG_TRADING_0_5_DAYS': 'STOCKS FG TRADING 0 5 DAYS', 'STOCKS_FG_TRADING_6_30_DAYS': 'STOCKS FG TRADING 6 30 DAYS', 'STOCKS_FG_TRADING_31_60_DAYS': 'STOCKS FG TRADING 31 60 DAYS', 'STOCKS_FG_TRADING_61_90_DAYS': 'STOCKS FG TRADING 61 90 DAYS', 'STOCKS_FG_TRADING_91_120_DAYS': 'STOCKS FG TRADING 91 120 DAYS', 'STOCKS_FG_TRADING_121_150_DAYS': 'STOCKS FG TRADING 121 150 DAYS', 'STOCKS_FG_TRADING_151_180_DAYS': 'STOCKS FG TRADING 151 180 DAYS', 'STOCKS_FG_TRADING_GREATER_THAN180_DAYS': 'STOCKS FG TRADING GREATER THAN180 DAYS', 'STK_CONSUMABLES_PKG_MATERIAL': 'STOCKS CONSUMABLES PACKING MATERIAL', 'STOCKS_STOCK_IN_TRANSIT': 'STOCKS STOCK IN TRANSIT', 'STOCKS_RM_LESS_THAN_90_DAYS': 'STOCKS RM LESS THAN 90 DAYS', 'STOCKS_TOTAL_SPARES_STOCK': 'STOCKS TOTAL SPARES STOCK', 'STOCK_LESS_THAN_60_DAYS': 'STOCK LESS THAN 60 DAYS', 'STOCK_LESS_THAN_90_DAYS': 'STOCK LESS THAN 90 DAYS', 'STOCK_LESS_THAN_120_DAYS': 'STOCK LESS THAN 120 DAYS', 'STOCK_LESS_THAN_180_DAYS': 'STOCK LESS THAN 180 DAYS', 'STOCKS_WIP_LESS_THAN_180_DAYS': 'STOCKS WIP LESS THAN 180 DAYS', 'TOTAL_STOCKS': 'TOTAL STOCKS', 'DEBTORS_0_5_DAYS': 'DEBTORS 0 5 DAYS', 'DEBTORS_6_30_DAYS': 'DEBTORS 6 30 DAYS', 'DEBTORS_31_60_DAYS': 'DEBTORS 31 60 DAYS', 'DEBTORS_61_90_DAYS': 'DEBTORS 61 90 DAYS', 'DEBTORS_91_120_DAYS': 'DEBTORS 91 120 DAYS', 'DEBTORS_121_150_DAYS': 'DEBTORS 121 150 DAYS', 'DEBTORS_151_180_DAYS': 'DEBTORS 151 180 DAYS', 'DEBTORS_GREATER_THAN180_DAYS': 'TOTAL DEBTORS', 'DR_DED_INTERGROUP_RECEIVABLES1': 'DEBTORS DEDUCTION INTERGROUP RECEIVABLES 1', 'DR_DED_INTERGROUP_RECEIVABLES2': 'DEBTORS DEDUCTION INTERGROUP RECEIVABLES 2', 'DEBTORS_LESS_THAN_30_DAYS': 'DEBTORS LESS THAN 30 DAYS', 'DEBTORS_LESS_THAN_60_DAYS': 'DEBTORS LESS THAN 60 DAYS', 'DEBTORS_LESS_THAN_90_DAYS': 'DEBTORS LESS THAN 90 DAYS', 'DEBTORS_GRT_THAN_90_DAYS': 'DEBTORS GREATER THAN 90 DAYS', 'DR_GRT_THN_90_LESS_THN_150_DAY': 'DEBTORS GREATER THAN 90 LESS THAN 150 DAYS', 'DEBTORS_LESS_THAN_120_DAYS': 'DEBTORS LESS THAN 120 DAYS', 'DR_GRT_THAN_120LESSTHN_150_DAY': 'DEBTORS GREATER THAN 120 LESS THAN 150 DAYS', 'DEBTORS_LESS_THAN_180_DAYS': 'DEBTORS LESS THAN 180 DAYS', 'TOTAL_DEBTORS': 'TOTAL DEBTORS', 'DR_UNBILLED_DR_LESS_THAN60DAYS': 'DEBTORS UNBILLED DEBTORS LESS THAN60 DAYS', 'DR_UNBILLED_DR_LESS_THAN90DAYS': 'DEBTORS UNBILLED DEBTORS LESS THAN90DAYS', 'DEBTORS_TOTAL_UNBILLED_DEBTORS': 'DEBTORS TOTAL UNBILLED DEBTORS', 'CREDITORS_UPTO_90_DAYS': 'CREDITORS UPTO 90 DAYS', 'CREDITORS_UPTO_91_150_DAYS': 'CREDITORS UPTO 91 150 DAYS', 'CREDITORS_GREATER_THAN151_DAYS': 'CREDITORS GREATER THAN151 DAYS', 'TOTAL_CREDITORS': 'TOTAL CREDITORS', 'DR_ADV_PAID_TO_SUPPLIER': 'DEBTORS ADVANCE PAID TO SUPPLIER', 'SECURITY_DEPOSIT': 'SECURITY DEPOSIT', 'CREDITORS_ADV_REC_FRM_CUST': 'CREDITORS ADVANCES RECEIVED FROM CUSTOMERS', 'DR_DEDUCTION_BILLS_DISC': 'DEBTORS DEDUCTION BILLS DISCOUNTED', 'DR_DEDUCTION_BILLS_FACT_ASSI': 'DEBTORS DEDUCTION BILLS FACTORED ASSIGNED', 'LIMITS_OS_WITH_KMBL_BANKS_1': 'LIMITS OS WITH KMBL BANKS 1', 'LIMITS_OS_WITH_KMBL_BANKS_2': 'LIMITS OS WITH KMBL BANKS 2', 'LIMITS_OS_WITH_KMBL_BANKS_3': 'LIMITS OS WITH KMBL BANKS 3', 'LIMITS_OS_WITH_OTHER_BANKS_1': 'LIMITS OS WITH OTHER BANKS 1', 'LIMITS_OS_WITH_OTHER_BANKS_2': 'LIMITS OS WITH OTHER BANKS 2', 'LIMITS_OS_WITH_OTHER_BANKS_3': 'LIMITS OS WITH OTHER BANKS 3', 'LIMITS_OS_WITH_OTHER_BANKS_4': 'LIMITS OS WITH OTHER BANKS 4', 'LIMITS_OS_WITH_OTHER_BANKS_5': 'LIMITS OS WITH OTHER BANKS 5','STOCKS_FG_LESS_THAN_90_DAYS':'STOCKS FG TRADING 61 90 DAYS'}
                categories=['STOCKS_RM_0_5_DAYS', 'STOCKS_RM_6_30_DAYS', 'STOCKS_RM_31_60_DAYS', 'STOCKS_RM_61_90_DAYS', 'STOCKS_RM_91_120_DAYS', 'STOCKS_RM_121_150_DAYS', 'STOCKS_RM_151_180_DAYS', 'STOCKS_IMPORTED_RM_GREATER_THAN180_DAYS', 'STOCKS_WIP_0_5_DAYS', 'STOCKS_WIP_6_30_DAYS', 'STOCKS_WIP_31_60_DAYS', 'STOCKS_WIP_61_90_DAYS', 'STOCKS_WIP_91_120_DAYS', 'STOCKS_WIP_121_150_DAYS', 'STOCKS_WIP_151_180_DAYS', 'STOCKS_WIP_GREATER_THAN180_DAYS', 'STOCKS_VEHICLE_STOCKS_0_5_DAY', 'STOCKS_VEHICLE_STOCKS_6_30_DAY', 'STOCKS_VEHICLE_STOCKS_31_60DAY', 'STOCKS_VEHICLE_STOCKS_61_90DAY', 'STOCKS_VEHICLE_STOCKS_91_120_DAY', 'STOCKS_VEHICLE_STOCK121_150DAY', 'STOCKS_VEHICLE_STOCK151_180DAY', 'STOCKS_VEHICLE_STOCKS_GREATER_THAN180_DAYS', 'STOCKS_SPARES_STOCKS_0_5_DAY', 'STOCKS_SPARES_STOCKS_6_30_DAY', 'STOCKS_SPARES_STOCKS_31_60_DAY', 'STOCKS_SPARES_STOCKS_61_90_DAY', 'STOCKS_SPARES_STOCKS_91_120DAY', 'STOCKS_SPARES_STOCKS_121_150_DAY', 'STOCKS_SPARES_STOCKS151_180DAYs', 'STOCKS_SPARES_STOCKS_GREATER_THAN180_DAYS', 'STOCKS_FG_TRADING_0_5_DAYS', 'STOCKS_FG_TRADING_6_30_DAYS', 'STOCKS_FG_TRADING_31_60_DAYS', 'STOCKS_FG_TRADING_61_90_DAYS', 'STOCKS_FG_TRADING_91_120_DAYS', 'STOCKS_FG_TRADING_121_150_DAYS', 'STOCKS_FG_TRADING_151_180_DAYS', 'STOCKS_FG_TRADING_GREATER_THAN180_DAYS', 'STK_CONSUMABLES_PKG_MATERIAL', 'STOCKS_STOCK_IN_TRANSIT', 'STOCKS_RM_LESS_THAN_90_DAYS', 'STOCKS_TOTAL_SPARES_STOCK', 'STOCK_LESS_THAN_60_DAYS', 'STOCK_LESS_THAN_90_DAYS', 'STOCK_LESS_THAN_120_DAYS', 'STOCK_LESS_THAN_180_DAYS', 'STOCKS_WIP_LESS_THAN_180_DAYS', 'TOTAL_STOCKS', 'DEBTORS_0_5_DAYS', 'DEBTORS_6_30_DAYS', 'DEBTORS_31_60_DAYS', 'DEBTORS_61_90_DAYS', 'DEBTORS_91_120_DAYS', 'DEBTORS_121_150_DAYS', 'DEBTORS_151_180_DAYS', 'DEBTORS_GREATER_THAN180_DAYS', 'DR_DED_INTERGROUP_RECEIVABLES1', 'DR_DED_INTERGROUP_RECEIVABLES2', 'DEBTORS_LESS_THAN_30_DAYS', 'DEBTORS_LESS_THAN_60_DAYS', 'DEBTORS_LESS_THAN_90_DAYS', 'DEBTORS_GRT_THAN_90_DAYS', 'DR_GRT_THN_90_LESS_THN_150_DAY', 'DEBTORS_LESS_THAN_120_DAYS', 'DR_GRT_THAN_120LESSTHN_150_DAY', 'DEBTORS_LESS_THAN_180_DAYS', 'TOTAL_DEBTORS', 'DR_UNBILLED_DR_LESS_THAN60DAYS', 'DR_UNBILLED_DR_LESS_THAN90DAYS', 'DEBTORS_TOTAL_UNBILLED_DEBTORS', 'CREDITORS_UPTO_90_DAYS', 'CREDITORS_UPTO_91_150_DAYS', 'CREDITORS_GREATER_THAN151_DAYS', 'TOTAL_CREDITORS', 'DR_ADV_PAID_TO_SUPPLIER', 'SECURITY_DEPOSIT', 'CREDITORS_ADV_REC_FRM_CUST', 'DR_DEDUCTION_BILLS_DISC', 'DR_DEDUCTION_BILLS_FACT_ASSI', 'LIMITS_OS_WITH_KMBL_BANKS_1', 'LIMITS_OS_WITH_KMBL_BANKS_2', 'LIMITS_OS_WITH_KMBL_BANKS_3', 'LIMITS_OS_WITH_OTHER_BANKS_1', 'LIMITS_OS_WITH_OTHER_BANKS_2', 'LIMITS_OS_WITH_OTHER_BANKS_3', 'LIMITS_OS_WITH_OTHER_BANKS_4', 'LIMITS_OS_WITH_OTHER_BANKS_5','STOCKS_FG_LESS_THAN_90_DAYS']
                non_dp_formula_fields=[]
                for i in categories:
                    if i not in equation_list:                     
                        non_dp_formula_fields.append(rbg_fields[i])
                for field in non_dp_formula_fields:
                    res[field] = None
                res = res.drop(columns=['Stock Due Date','Received Date','CC_LIMIT','INSURANCE_AMOUNT','DEBTORS_MARGIN','STOCK_MARGIN','NAME OF CUSTOMER','ZMT'])

            logging.info(f"result dataframe####  {res}")
            queue_name="rbg"

        status_qry=f'UPDATE `process_queue` SET `status`="Report Generated" WHERE `case_id`="{case_id}"'
        queue_db.execute(status_qry)
        temp_dir = Path('./app/temp_dir').absolute()
        temp_dir.mkdir(parents=True, exist_ok=True)
        temp_file = 'temp.xlsx'

        os.system(f'chattr -i {temp_dir}')
        os.system(f'chmod -R 777 {temp_dir}')

        res.to_excel(temp_dir / temp_file, index=False)
        workbook = openpyxl.load_workbook(temp_dir / temp_file)
        sheet = workbook.active
        if queue_name == 'rbg':
           
            sheet['GR2'] = "=IF(DK3>GQ3, GQ3, DK3)"
            sheet['GT2'] = "=IF(BK3>GS3, GS3, BK3)"
            workbook.save(temp_dir / temp_file)
            

        # Create a new workbook
        workbook_modified = openpyxl.Workbook()
        worksheet_modified = workbook_modified.active

        # Copy values and styles from the original worksheet to the new worksheet
        for row_num, row in enumerate(sheet.iter_rows(min_row=1), start=1):
            for col_num, cell in enumerate(row, start=1):
                new_cell = worksheet_modified.cell(row=row_num, column=col_num, value=cell.value)
                if isinstance(cell.value, (int, float)):
                    new_cell.number_format = '0'
        workbook_modified.save(temp_dir / temp_file)
                

        
        with open(temp_dir / temp_file, 'rb') as f:

            export_data_blob = base64.b64encode(f.read())

        filename = f'{case_id}_{queue_name}_istock_output.xlsx'
        response = {
            'flag': True,
            'data': {
                    'blob': export_data_blob.decode('utf-8'),
                    'file_name': filename
                }
        }

        audit_data = {"tenant_id": "kmb", "user": "", "case_id": case_id, 
                        "api_service": "download_excel", "service_container": "folder_monitor", "changed_data": None,
                        "tables_involved": "","memory_usage_gb": None, 
                        "time_consumed_secs": None, "request_payload": None, 
                        "response_data": None, "trace_id": case_id, "session_id": None,"status":str(response['flag'])}
        insert_into_audit(case_id, audit_data)

        return response
    except Exception:
        logging.exception(f'Something went wrong exporting data')
        return {'flag': False, 'message': 'Unable to export data.'}
    


@app.route('/download_template_excel', methods=['POST', 'GET'])
def download_template_excel():
    data=request.json
    db_config['tenant_id'] = 'kmb'
    logging.info(f"Request Data is: {data}")
    queue_db = DB('queues', **db_config)
    queue_id=data.get("queue_id",{})
    extraction_db = DB('extraction', **db_config)
    case_id = data.pop('case_id')
    try:
    
        qry=f"select `wbg_template_name` from `ocr` where case_id='{case_id}'"
        segment_res = extraction_db.execute_(qry).to_dict(orient="records")[-1]
        template_name= segment_res['wbg_template_name']
        qry=f"SELECT `STOCK STATEMENT_OCR`,`DEBITORS STATEMENT_OCR`,`CREDITORS_OCR`, `PURCHASES_OCR`, `ADVANCES_OCR`, `SALES_OCR`, `ADVANCES DEBTORS_OCR`, `BANK OS_OCR`, `ADVANCES SUPPLIERS_OCR`,`SECURITY DEPOSIT_OCR` FROM `ocr` WHERE  `case_id`='{case_id}';" 
        json_data_list = extraction_db.execute_(qry).to_dict(orient="records")

        qry1=f"SELECT `Due_Date` as 'Due_Date', `Last_3_month_Sales` as 'Last_three_month_Sales',`Last_3_month_purchases` as 'Last_three_month_Purchases',`wbg_Channel_Finance` as 'channel_finance',`wbg_db_backed_facility` as 'db_backed_facility',`wbg_margin_dp_with_kmbl` as 'margin_dp_with_kmbl',`wbg_total_ass_funding` as 'total_ass_funding',`wbg_net_eli_lo_ass_cov_bank` as 'net_eli_lo_ass_cov_bank',`wbg_average_asset_cover_tl` as 'average_asset_cover_tl',`wbg_asset_of_the_above` as 'asset_of_the_above',`wbg_balance_asset_for_dp_back` as 'balance_asset_for_dp_back',`wbg_asset_for_dp_backed_other_bank` as 'asset_for_dp_backed_other_bank',`wbg_balance_sec_to_kmbl` as 'balance_sec_to_kmbl',`wbg_margin` as 'margin',`wbg_dp_available_for_kmbl` as 'dp_available_for_kmbl',`wbg_immoveable_property` as 'immoveable_property',`wbg_capital_mark_related_exposure` as 'capital_mark_related_exposure',`wbg_dealer_finance` as 'dealer_finance',`wbg_others` as 'others',`wbg_total_exclusion` as 'total_exclusion',`wbg_loan_assets_bond_deb` as 'loan_assets_bond_deb',`wbg_net_add_assets_bond_deben` as 'net_add_assets_bond_deben',`wbg_net_stock_dp_eligible` as 'net_stock_dp_eligible',`wbg_dp_on_stock` as 'dp_on_stock',`wbg_chargeable_debtors` as 'chargeable_debtors',`wbg_export_rec_beyond_days` as 'export_rec_beyond_days',`wbg_domestic_rec_beyond_days` as 'domestic_rec_beyond_days',`wbg_net_export_receviables` as 'net_export_receviables',`wbg_net_domestic_receviables` as 'net_domestic_receviables',`wbg_Kotak_PO_Funding` as 'kotak_po_funding',`wbg_net_dp` as 'net_dp',`wbg_debtors_applicable_dp` as 'debtors_applicable_dp',`wbg_eligible_debtors` as 'eligible_debtors',`wbg_less_margin` as 'less_margin',`wbg_stock_applicable_dp` as 'stock_applicable_dp',`wbg_slow_obsolete_stock` as 'slow_obsolete_stock',`wbg_eligible_stock` as 'eligible_stock',`wbg_gt_stock_90` as 'stock_gt_90',`wbg_final_dp_adhoc` as 'final_dp_adhoc',`wbg_adhoc` as 'adhoc',`wbg_final_dp` as 'final_dp',`wbg_dp_limit_outside_consortium` as 'limit_consortium',`wbg_dp_available_for_outside_consortium` as 'dp_available_for_outside_consor',`wbg_total_dp` as 'Total_Dp',`wbg_dp_for_outside_consortium` as 'outside_consortium',`wbg_final_dp_within_consortium` as 'final_dp_consortium',`wbg_dp_limit_consortium` as 'dp_limit_consortium',`wbg_advised_lead_bank` as 'advised_as_lead_bank',`wbg_net_dp_per_our_share` as 'net_dp_our_share',`wbg_our_share` as 'our_share',`wbg_total_dp` as 'Total_Dp',`wbg_less_margin_25` as 'less_margin_25',`wbg_net_debtors` as 'Net_debtors',`wbg_considered_for_dp` as 'considered_dp',`wbg_net_stock_dp` as 'Net_stock',`wbg_less_margin_25` as 'less_margin_25',`wbg_EPC_PCFC_stock` as 'epc_pcfc_stock',`wbg_PBD_Utilised_from_other_Bank` as 'pbd_utilised',`wbg_buyers_credit_for_wc` as 'buyers_credit_wc',`wbg_total_stock` as 'Total_stock',`cc_limit` as 'CC_LIMIT',`insurance_amount` as 'INSURANCE_AMOUNT' , `Debtors_margin` as 'DEBTORS_MARGIN',`Stock_Margins` as 'STOCK_MARGIN',`deviation_applicable` as 'DEVIATION_APPLICABLE',`Crn` as 'CRN',`customer_name` as 'NAME_OF_CUSTOMER' ,`request_id` as 'REQUEST_ID' , `Account_Number` as 'ACCOUNT_NUMBER',`division` as 'ZMT',`Dp_formula` as 'dp_formula' from ocr where `case_id`='{case_id}';"
        df = extraction_db.execute_(qry1).to_dict(orient="records")[-1]
        values= create_dataframe_from_json(json_data_list)
        values = values.to_dict()
        final_result={}
        for i in values:
            final_result[i]=values[i][0]
        print(final_result)
        df_data = {**df, **final_result}        
        df_data['Date']=date.today()
        # Define the Jinja template
        logging.info(f"df_data--{df_data}")
        data_underscore = {key.replace(' ', '_'): value for key, value in df_data.items()}
        def get_variation_value(variation_list):
            return next((data_underscore.get(variation, ' ') for variation in variation_list), ' ')

        store_spares_value_variations = ["Stores_&_Spares", "Stocks_Consumables_Packing_Material_Total", "Stores_and_Consumables"]
        work_in_progress_variations=["Stocks_WIP_Total","stocks_in_process"]
        finished_goods_variations=["Stocks_FG_Trading_Total","Finished_Goods"]
        Sales_variations=["Last_3_Month_Sales","Sales"]
        Raw_Materials_variations=["Total_Raw_Material","Total_Raw_Material"]
        Stock_in_Transit_variations = ["Stock_in_Transit","Stock_in_tansit"]
        Group_Co_debtors_variations = ["Related_party/group_concern_Debtors_Total","Book_Debts_on_Group"]
        Total_debtors_variations_variations = ["DEBITORS_STATEMENT_OCR_Total_Debtors","DEBITORS_STATEMENT_OCR_Gross_Book_Debts"]
        Stores_Spares_variations =["Stocks_Consumables_Packing_Material_Total","Stores_and_Consumables"]
        Total_Stock_variations =["Total_Stocks","Total_Stock"]
        # Example usage for each variation
        store_spares_value = get_variation_value(store_spares_value_variations)
        work_in_progress_value = get_variation_value(work_in_progress_variations)
        finished_goods_value = get_variation_value(finished_goods_variations)
        sales_value = get_variation_value(Sales_variations)
        raw_materials_value = get_variation_value(Raw_Materials_variations)
        stock_in_transit_value = get_variation_value(Stock_in_Transit_variations)
        group_co_debtors_value = get_variation_value(Group_Co_debtors_variations)
        total_debtors_value = get_variation_value(Total_debtors_variations_variations)
        stores_spares_value = get_variation_value(Stores_Spares_variations)
        total_stock_value = get_variation_value(Total_Stock_variations)
        df_data = {
            'Last_three_month_Sales': data_underscore.get('Last_three_month_Sales', ' '),
            'Last_three_month_Purchases': data_underscore.get('Last_3_Month_Purchases', ' '),
            'channel_finance': data_underscore.get('channel_finance', ' '),
            'db_backed_facility': data_underscore.get('db_backed_facility', ' '),
            'margin_dp_with_kmbl': data_underscore.get('margin_dp_with_kmbl', ' '),
            'total_ass_funding': data_underscore.get('total_ass_funding', ' '),
            'net_eli_lo_ass_cov_bank': data_underscore.get('net_eli_lo_ass_cov_bank', ' '),
            'average_asset_cover_tl': data_underscore.get('average_asset_cover_tl', ' '),
            'asset_of_the_above': data_underscore.get('asset_of_the_above', ' '),
            'balance_asset_for_dp_back': data_underscore.get('balance_asset_for_dp_back', ' '),
            'asset_for_dp_backed_other_bank': data_underscore.get('asset_for_dp_backed_other_bank', ' '),
            'balance_sec_to_kmbl': data_underscore.get('balance_sec_to_kmbl', ' '),
            'dp_available_for_kmbl': data_underscore.get('dp_available_for_kmbl', ' '),
            'immoveable_property': data_underscore.get('immoveable_property', ' '),
            'capital_mark_related_exposure': data_underscore.get('capital_mark_related_exposure', ' '),
            'dealer_finance': data_underscore.get('dealer_finance', ' '),
            'others': data_underscore.get('others', ' '),
            'total_exclusion': data_underscore.get('total_exclusion', ' '),
            'loan_assets_bond_deb': data_underscore.get('loan_assets_bond_deb', ' '),
            'net_add_assets_bond_deben': data_underscore.get('net_add_assets_bond_deben', ' '),
            'net_stock_dp_eligible': data_underscore.get('net_stock_dp_eligible', ' '),
            'dp_formula': data_underscore.get('dp_formula', ' '),
            'dp_on_stock': data_underscore.get('dp_on_stock', ' '),
            'chargeable_debtors': data_underscore.get('chargeable_debtors', ' '),
            'export_rec_beyond_days': data_underscore.get('export_rec_beyond_days', ' '),
            'domestic_rec_beyond_days': data_underscore.get('domestic_rec_beyond_days', ' '),
            'net_export_receviables': data_underscore.get('net_export_receviables', ' '),
            'net_domestic_receviables': data_underscore.get('net_domestic_receviables', ' '),
            'stock_gt_90': data_underscore.get('stock_gt_90', ' '),
            'pbd_utilised': data_underscore.get('pbd_utilised', ' '),
            'epc_pcfc_stock': data_underscore.get('epc_pcfc_stock', ' '),
            'Net_stock': data_underscore.get('Net_stock', ' '),
            'less_margin_25': data_underscore.get('less_margin_25', ' '),
            'considered_dp': data_underscore.get('considered_dp', ' '),
            'Net_debtors': data_underscore.get('Net_debtors', ' '),
            'Total_Dp': data_underscore.get('Total_Dp', ' '),
            'net_dp_our_share': data_underscore.get('net_dp_our_share', ' '),
            'advised_as_lead_bank': data_underscore.get('advised_as_lead_bank', ' '),
            'dp_limit_consortium': data_underscore.get('dp_limit_consortium', ' '),
            'final_dp_consortium': data_underscore.get('final_dp_consortium', ' '),
            'outside_consortium': data_underscore.get('outside_consortium', ' '),
            'dp_available_for_outside_consor': data_underscore.get('dp_available_for_outside_consor', ' '),
            'limit_consortium': data_underscore.get('limit_consortium', ' '),
            'dp_limit_consortium': data_underscore.get('dp_limit_consortium', ' '),
            'final_dp': data_underscore.get('final_dp', ' '),
            'adhoc': data_underscore.get('adhoc', ' '),
            'final_dp_adhoc': data_underscore.get('final_dp_adhoc', ' '),
            'buyers_credit_wc': data_underscore.get('buyers_credit_wc', ' '),
            'Total_stock': total_stock_value,
            'eligible_stock': data_underscore.get('eligible_stock', ' '),
            'slow_obsolete_stock': data_underscore.get('slow_obsolete_stock', ' '),
            'stock_applicable_dp': data_underscore.get('stock_applicable_dp', ' '),
            'less_margin': data_underscore.get('less_margin', ' '),
            'eligible_debtors': data_underscore.get('eligible_debtors', ' '),
            'debtors_applicable_dp': data_underscore.get('debtors_applicable_dp', ' '),
            'net_dp': data_underscore.get('net_dp', ' '),
            'kotak_po_funding': data_underscore.get('kotak_po_funding', ' '),
            'DEBTORS_MARGIN': data_underscore.get('DEBTORS_MARGIN', ' '),
            'STOCK_MARGIN': data_underscore.get('STOCK_MARGIN', ' '),
            'NAME_OF_CUSTOMER': data_underscore.get('NAME_OF_CUSTOMER', ' '),
            'Crn': data_underscore.get('CRN', ' '),
            'ACCOUNT_NUMBER': data_underscore.get('ACCOUNT_NUMBER', ' '),
            'Due_Date': data_underscore.get('Due_Date', ' '),
            'Raw_Material': raw_materials_value,
            'Work_in_Process': data_underscore.get('work_in_progress_value', ' '),
            'Finished_Goods': finished_goods_value,
            'TOTAL_Debitors': total_debtors_value ,
            'TOTAL_Creditors': data_underscore.get('Total_Creditors', ' '),
            'Stores_Spares':stores_spares_value,
            'Other_stock': data_underscore.get('Other stock', ' '),
            'Stock_in_Transit': stock_in_transit_value,
            'TOTAL_CREDITORS': data_underscore.get('Total_Creditors', ' '),
            'LC_BACKED_CREDITORS_FOR_GOODS': data_underscore.get('LC_Backed_Creditors_For_Goods', ' '),
            'Less_Creditors': data_underscore.get('Less_Creditors', ' '),
            'LC_Creditors': data_underscore.get('LC Creditors(not included in creditors above)', ' '),
            'Sales': sales_value,
            'Purchases': data_underscore.get('Last_3_Month_Purchases', ' '),
            'Advance_paid_to_suppliers':data_underscore.get('Advance_paid_to_suppliers', ' '),
            'Advance_from_Customer':data_underscore.get('Advances_received_from_customer', ' '),
            'group_co_debtors' : group_co_debtors_value
        }

        print(f'{df_data}#######before df_data')
        for key, value in df_data.items():
            if value is None:
                df_data[key] = 0

        print(f'{df_data}########after df_data')

        #template1
        if '111111' in template_name:


            template_data1 = """



    <table style="width:100%">
    <tr><th></th><td></td><td></td></tr>
        <tr><th></th><td></td><td></td></tr>
        <tr style="height:600px";"font-size: 20px;"><td></td><td></td><td style="font-size: 20px;">DP STOCK STATEMENT</td><td></td><td></td></tr>
        <tr style="height:600px"><td></td><td>Name of Borrower</td><td>{{NAME_OF_CUSTOMER}}</td><td>{{Date}}</td></tr>
        <tr style="height:600px"><td></td><td>Margin</td><td>Stock</td><td>{{STOCK_MARGIN}}</td></tr >
        <tr style="height:600px"><td></td><td></td><td>Debtors</td><td>{{DEBTORS_MARGIN}}</td></tr >
        <tr style="height:600px"><td></td><td></td><td></td><td></td> </tr >
        <tr style="height:600px"><td></td><td></td><td>Details for the month </td><td></td></tr >
        <tr style="height:600px";"color: red;"><td></td><td>Raw Material</td><td>{{Raw_Material}}</td></tr>
        <tr style="height:600px"><td></td><td style="color: blue;">Work in Process</td><td>{{Work_in_Process}}</td></tr>
        <tr style="height:600px"><td></td><td>Finished Goods</td><td>{{Finished_Goods}}</td></tr>
        <tr style="height:600px"><td></td><td>Stores & Spares</td><td>{{Stores_Spares}}</td></tr>
        <tr style="height:600px"><td></td><td>Other stock</td><td>{{Other_stock}}</td></tr>
        <tr style="height:600px"><td></td><td>Stock in Transit</td><td>{{Stock_in_Transit}}</td></tr>
        <tr style="height:600px"><td></td><td>Total Stock</td><td></td></tr>
        <tr style="height:600px"><td></td><td>Creditors(net of advances)</td><td>{{TOTAL_CREDITORS}}</td></tr>
        <tr style="height:600px"><td></td><td>Buyers Credit for working capital</td><td>{{buyers_credit_wc}}</td></tr>
        <tr style="height:600px"><td></td><td>LC Creditors(not included in creditors above)</td><td>{{LC_BACKED_CREDITORS_FOR_GOODS}}</td></tr>
        <tr style="height:600px"><td></td><td>PBD Utilised from other Bank</td><td>{{pbd_utilised}}</td></tr>
        <tr style="height:600px"><td></td><td>Add Advance paid to suppliers</td><td>{{Advance_paid_to_suppliers}}</td></tr>
        <tr style="height:600px"><td></td><td>Less Channel Finance </td><td>{{channel_finance}}</td></tr>
        <tr style="height:600px"><td></td><td>EPC/PCFC stock</td><td>{{epc_pcfc_stock}}</td></tr>
        <tr style="height:600px"><td></td><td>Net Stock</td><td>{{Net_stock}}</td></tr>
        <tr style="height:600px"><td></td><td>Less: Margin</td><td>{{less_margin}}</td></tr>
        <tr style="height:600px"><td></td><td>DP on stock</td><td>{{dp_on_stock}}</td></tr>
        <tr style="height:600px"><td></td><td>Total debtors</td><td>{{TOTAL_Debitors}}</td></tr>
        <tr style="height:600px"><td></td><td>Non Group co. drs.   > 90 days as per term sheet  days</td><td>{{wb_non_groupco_drs_days}}</td></tr>
        <tr style="height:600px"><td></td><td>Bill discounted debtors</td><td>{{wb_bill_discounted}}</td></tr>
        <tr style="height:600px"><td></td><td>Group concern debtors</td><td>{{group_co_debtors}}</td></tr>
        <tr style="height:600px"><td></td><td>Advance from Customer</td><td>{{Advance_from_Customer}}</td></tr>
        <tr style="height:600px"><td></td><td>Net Debtors</td><td>{{Net_debtors}}</td></tr>
        <tr style="height:600px"><td></td><td>Less: Margin</td><td>{{less_margin_25}}</td></tr>
        <tr style="height:600px"><td></td><td>DP on debtors</td><td>{{debtors_applicable_dp}}</td></tr>
        <tr style="height:600px"><td></td><td>Total DP</td><td>{{Total_Dp}}</td></tr>
        <tr style="height:600px"><td></td><td>Our Share in consortium</td><td>{{wb_our_share_in_consortium}}</td></tr>
        <tr style="height:600px"><td></td><td>DP as per our share in consortium</td><td>{{net_dp_our_share}}</td></tr>
        <tr style="height:600px"><td></td><td>DP as advised by lead bank</td><td>{{advised_as_lead_bank}}</td></tr>
        <tr style="height:600px"><td></td><td>DP limit</td><td>{{wb_dp_limit}}</td></tr>
        <tr style="height:600px"><td></td><td>DP to be set</td><td>{{final_dp}}</td></tr>
        <tr style="height:600px"><td></td><td>Adhoc</td><td>{{adhoc}}</td></tr>
        <tr style="height:600px"><td></td><td>Final DP(With adhoc)</td><td>{{final_dp_adhoc}}</td></tr>
        <tr style="height:600px"><td></td><td>Other WC bankers outstanding</td><td>{{wb_other_wc_bankers_os}}</td></tr>
        <tr style="height:600px"><td></td><td>Kotak PO Funding</td><td>{{kotak_po_funding}}</td></tr>
        <tr style="height:600px"><td></td><td>Sales</td><td>{{Sales}}</td></tr>
        <tr style="height:600px"><td></td><td>Purchases</td><td>{{Purchases}}</td></tr>

    </table>

    """




            print(df_data)
            # Sample data for the template
            data = df_data
            template = Template(template_data1)
            rendered_template = template.render(data)
            # # Create a Jinja template object
            # template = Template(template_data1)


            # # Render the template with data
            # rendered_template = template.render(data)
            # print(rendered_template)
            # Parse the HTML table into a DataFrame
            dfs = pd.read_html(rendered_template)

            # Get the DataFrame from the list of DataFrames
            df = dfs[0]
            df.columns = [" ", " ", " ","","",]

            # print(df)


            output_path = Path(
                        f'/app/output/kmb/assets/pdf/kmb/{case_id}')
            os.umask(0)
            logging.info(f"---{output_path}---")
            df.to_excel(f"{output_path}/output.xlsx", index=False )
            
            
            logging.info(f"---saved---")
            workbook = openpyxl.load_workbook(f"{output_path}/output.xlsx",data_only=True)
            # Select the desired sheet
            sheet = workbook.active
            sheet.title = "Main_Format1"# You can also specify the sheet name like workbook['Sheet1']
            for row in range(10, 47):
                cell_address = 'C' + str(row)
                cell = sheet[cell_address]
                if cell.value is not None and isinstance(cell.value, str):
                    try:
                        cell.value = float(cell.value)  # Convert text to integer
                    except ValueError:
                        # Handle non-convertible values here if needed
                        pass


            sheet.column_dimensions['B'].width = 60
            sheet.column_dimensions['c'].width = 60
            sheet.column_dimensions['d'].width = 40
            sheet['C16'] = '=SUM(C10:C15)'
            sheet['C24'] = '=C16-C17-C18-C19-D25+C21-C22-C23'
            sheet['C25']='=IF(C24<0,0,C24*D6)'
            sheet['C26'] ='=IF(C24<0,0,C24-C25)'
            sheet['C32']='=IF(C24<0,C27-C28-C29-C30-C31+C24,C27-SUM(C28,C29,C30,C31))'
            sheet['C33']='=C32*D7'
            sheet['C34']='=C32-C33'
            sheet['C35']='=IF(C24<0,C34,C26+C34)'
            sheet['C37']='=C35* C36'
            sheet['C40']='=IF(38= 0, MIN(C37, C39), MIN(C38 ,C39))'
            sheet['C42']='=IF(C40+C41>C39,C39,C40+C41)'
            # sheet['C24'] = '=C16-C17-C18-C19-C20+C21-C22-C23'
            # sheet['C25']='=IF(C24 <0, 0, C24*D6)'
            # sheet['C26'] ='=IF(C24<0, 0,C24-C25)'
            # sheet['C32']='=IF(C24<0, C27 - C28 - C29- C30 -C31 + C24, C24 - SUM(C28, C29, C30, C31))'
            # sheet['C33']='=C32*D7'
            # sheet['C34']='=C32-C33'
            # sheet['C35']='=IF(C24<0,C34,C26+C34)'
            # sheet['C37']='=C35* C36'
            # sheet['C40']='=IF(38= 0, MIN(C37, C39), MIN(C38 ,C39))'
            # sheet['C42']='=IF(C40+ C41> C39, C39, C40 + C41)'

            
            img_path = Path(f'/app/output/kmb/assets/pdf/kotak.png')
            img = openpyxl.drawing.image.Image(img_path)
            img.anchor = "B3"  
            for row in sheet.iter_rows(min_row=9, max_row=46, min_col=3, max_col=3):
                for cell in row:
                    cell.fill = PatternFill(start_color="53b553", end_color="53b553", fill_type="solid")
                    

            for row in sheet.iter_rows(min_row=9, max_row=20, min_col=3, max_col=3):
                for cell in row:
                    cell.fill = PatternFill(start_color="53b553", end_color="53b553", fill_type="solid")
                    
                    # Green color
            for row in sheet.iter_rows(min_row=4, max_row=4, min_col=3, max_col=3):
                for cell in row:
                    cell.font = Font(size=16, bold=True, color="090a0a")
            for row in sheet.iter_rows(min_row=16, max_row=16, min_col=2, max_col=2):
                for cell in row:
                    cell.font = Font(size=10, bold=True, color="090a0a")        
                    
            for row in sheet.iter_rows(min_row=9, max_row=9, min_col=2, max_col=4):
                
                for cell in row:
                    cell.fill = PatternFill(start_color="626a80", end_color="626a80", fill_type="solid")
                    cell.font = Font(size=14, bold=True, color="FFFFFF")
            for row in sheet.iter_rows(min_row=6, max_row=7, min_col=2, max_col=3):
                for cell in row:
                    cell.fill = PatternFill(start_color="626a80", end_color="626a80", fill_type="solid")   
                    cell.font = Font(size=14, bold=True, color="FFFFFF")
            for row in sheet.iter_rows(min_row=5, max_row=5, min_col=2, max_col=2):
                for cell in row:
                    cell.fill = PatternFill(start_color="626a80", end_color="626a80", fill_type="solid")
                    cell.font = Font(size=14, bold=True, color="FFFFFF")

            # Save the changes to the workbook

            workbook.save(f"{output_path}/output.xlsx")
            try:
                workbook = openpyxl.load_workbook(f"{output_path}/output.xlsx")
                sheet = workbook.active
                workbook_modified = openpyxl.Workbook()
                worksheet_modified = workbook_modified.active

                # Copy values and styles from the original worksheet to the new worksheet
                for row_num, row in enumerate(sheet.iter_rows(min_row=1), start=1):
                    for col_num, cell in enumerate(row, start=1):
                        new_cell = worksheet_modified.cell(row=row_num, column=col_num, value=cell.value)
                        if isinstance(cell.value, (int, float)):
                            new_cell.number_format = '0'
                workbook_modified.save(f"{output_path}/output.xlsx")
            except Exception as e:
                logging.info(f'######{e} exception')

            logging.info(f"completed sheet 1-----")

            Heads= [
                "For the Month", "Due Date", "CRN", "Account Number", "Total Stock", "Stock Applicable for DP", 
                "Stock < 90", "Stock < 120", "Total Creditors", "Creditors Applicable for DP", 
                "Total Debtors", "Debtors Applicable for DP", "Debtors < 90", "Debtors < 120", 
                "Total Sales", "Total Purchases", "Other Bank Limit", "System DP", "System Limit", 
                "CC Limit", "Computed DP", "Final DP", "Stock <30 Days", "Stock Between 31 and 60 Days", 
                "Stock Between 61 and 90 Days", "Stock Between 91 and 120 Days", "Stock Between 121 and 180 Days", 
                "Stock > 180 Days", "Debtors <30 Days", "Debtors Between 31 and 60 Days", 
                "Debtors Between 61 and 90 Days", "Debtors Between 91 and 120 Days", 
                "Debtors Between 121 and 180 Days", "Debtors > 180 Days", "Sales Spares Workshop", 
                "Sales Amount Vehicle", "Purchase Spares Workshop", "Debtors > 90 Days Spares Workshop", 
                "Debtors > 90 Days Vehicle", "Debtors Spares Between 91 to 120", 
                "Debtors Vehicle Between 91 to 120", "Total Debtors Spares", "Total Spares Stock", 
                "Total Spares Stock <= 120", "Total Spares Stock <= 90", "Total Vehicle Stock", 
                "Total Vehicle Stock <= 120", "Total Vehicle Stock <= 90", 
                "Debtors Advance Paid to suppliers", "Purchase Vehicles", "Total Debtors Vehicles", 
                "Total Stock <=120", "Total Stock <=90", "Advance Received from customers", 
                "Bill discounted debtors", "Channel Finance", "Debtors 0 to 60 Days", 
                "Debtors 0 to 5 Days", "Debtors 6 to 30 Days", "Debtors 6 to 60 Days", 
                "Debtors < 150 days", "Debtors < 45 days", "Debtors Specific", 
                "Debtors of Vehicle < 30 days", "Debtors of Vehicle < 60 days", 
                "Debtors of spares < 60 days", "Group Co debtors", "Insurance Amount", 
                "Limits Os With KMBL Banks 1", "Outstanding with other WC Bankers", 
                "Previous month Total Creditors", "Previous month Total Debtors", 
                "Previous month Total Stock", "Purchase 1 Month prior", "Purchase 2 Month prior", 
                "Purchase 3 Month prior", "Purchase 4 Month prior", "Purchase 5 Month prior", 
                "Purchase Spares 1 Month Prior", "Purchase Spares 2 Month Prior", 
                "Purchase Spares 3 Month Prior", "Purchase Spares 4 Month Prior", 
                "Purchase Spares 5 Month Prior", "Purchase Vehicles 1 Month Prior", 
                "Purchase Vehicles 2 Month Prior", "Purchase Vehicles 3 Month Prior", 
                "Purchase Vehicles 4 Month Prior", "Purchase Vehicles 5 Month Prior", 
                "RM Stock < 60 days", "RM Stock < 90 days", "Sales 1 Month Prior", 
                "Sales 2 Month Prior", "Sales 3 Month Prior", "Sales 4 Month Prior", 
                "Sales 5 Month Prior", "Sales Spares 1 Month Prior", "Sales Spares 2 Month Prior", 
                "Sales Spares 3 Month Prior", "Sales Spares 4 Month Prior", 
                "Sales Spares 5 Month Prior", "Sales Vehicles 1 Month Prior", 
                "Sales Vehicles 2 Month Prior", "Sales Vehicles 3 Month Prior", 
                "Sales Vehicles 4 Month Prior", "Sales Vehicles 5 Month Prior", "Security Deposit", 
                "Stock Between 0 and 60 Days", "Stock Specific", "Stock in Transit < 90 Days", 
                "Stock of Vehicle < 60 days", "Stock of spares < 60 days", "Total Purchase Spares", 
                "Total Purchase Vehicles", "Total Sales Spares", "Total Sales Vehicles", 
                "Unbilled Debtors"
            ]
            # Number of columns
            num_columns = len(Heads)

            # Create 'Values' list with empty strings
            values_list = ["-"] * num_columns
            print(values_list)

            # Update the 'Values' key in the data2 dictionary


            # Print the updated data2 dictionary
            data2={'Heads':Heads,'Values':values_list}
            print(data2)



            # data2={'Heads':Heads,'Values':Values}
            df2 = pd.DataFrame(data2)
            # df2.to_excel("Downloads/output.xlsx", index=False )
            print(df2)
            new_sheet = workbook.create_sheet(title="Key Metrics")

            new_sheet.column_dimensions['A'].width = 50
            new_sheet.column_dimensions['B'].width = 50

            for row in dataframe_to_rows(df2, index=False, header=True):
                new_sheet.append(row)
            for row in new_sheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=2):
                for cell in row:
                    cell.fill = PatternFill(start_color="626a80", end_color="626a80", fill_type="solid")

                    
            if df_data['Due_Date'] is None:
                df_data['Due_Date']=0
            if df_data['Crn'] is None:
                df_data['Crn']=0
            if df_data['ACCOUNT_NUMBER'] is None:
                df_data['ACCOUNT_NUMBER']=0

            new_sheet['B3']=df_data['Due_Date']
            new_sheet['B4']=df_data['Crn']
            new_sheet['B5']=df_data['ACCOUNT_NUMBER']
            new_sheet['B6']='=Main_Format1!C16'
            new_sheet['B7']='=Main_Format1!C26'
            new_sheet['B10']='=SUM(Main_Format1!C17,Main_Format1!C18,Main_Format1!A8)'
            new_sheet['B14']='=Main_Format1!C26-Main_Format1!C28'
            new_sheet['B17']='=Main_Format1!C27'
            new_sheet['B13']='=Main_Format1!C34'
            new_sheet['B16']='=Main_Format1!C45'
            new_sheet['B17']='=Main_Format1!C46'
            new_sheet['B18']='=Main_Format1!C43'
            new_sheet['B21']='=Main_Format1!C39'
            new_sheet['B22']='=Main_Format1!C42'
            new_sheet['B23']='=Main_Format1!C16'
            new_sheet['B50']='=Main_Format1!C21'
            new_sheet['B55']='=Main_Format1!C31'
            new_sheet['B56']='=Main_Format1!C29'
            new_sheet['B57']='=Main_Format1!C22'
            new_sheet['B68']='=Main_Format1!C30'
            new_sheet['B68']='=Main_Format1!C30'
            new_sheet['B68']='=Main_Format1!C30'
            new_sheet['B70']='=Main_Format1!C44'
            new_sheet['B71']='=Main_Format1!C43' 
            report_name="Kotak_WBG_111111_Output_Mapping_V1.0"
        
        #template2
        if 'WBG_100000_MASTER' in template_name:
            template_data1 = """

<table style="width:100%">
            <tr><th></th><td></td><td></td></tr>
                <tr><th></th><td></td><td></td></tr>
                <tr style="height:600px";"font-size: 20px;"><td></td><td></td><td style="font-size: 20px;">DP STOCK STATEMENT</td><td></td><td></td></tr>
                <tr style="height:600px"><td></td><td>Name of Borrower</td><td>{{NAME_OF_CUSTOMER}}</td><td>{{Date}}</td></tr>
                <tr style="height:600px"><td></td><td>Margin</td><td>Stock</td><td>{{STOCK_MARGIN}}</td></tr >
                <tr style="height:600px"><td></td><td></td><td>Debtors</td><td>{{DEBTORS_MARGIN}}</td></tr >
                <tr style="height:600px"><td></td><td></td><td></td><td></td> </tr >
                <tr style="height:600px"><td></td><td></td><td>Details for the month </td><td></td></tr >
                <tr style="height:600px";"color: red;"><td></td><td>Raw Material</td><td>{{Raw_Material}}</td></tr>
                <tr style="height:600px"><td></td><td style="color: blue;">Work in Process</td><td>{{Work_in_Process}}</td></tr>
                <tr style="height:600px"><td></td><td>Finished Goods</td><td>{{Finished_Goods}}</td></tr>
                <tr style="height:600px"><td></td><td>Stores & Spares</td><td>{{Stores_Spares}}</td></tr>
                <tr style="height:600px"><td></td><td>Other stock</td><td>{{Other_stock}}</td></tr>
                <tr style="height:600px"><td></td><td>Stock in Transit</td><td>{{Stock_in_Transit}}</td></tr>
                <tr style="height:600px"><td></td><td>Total Stock</td><td>{{Total_stock}}</td></tr>
                <tr style="height:600px"><td></td><td>Less: Creditors</td><td>{{Less_Creditors}}</td></tr>
                <tr style="height:600px"><td></td><td>Buyers Credit for working capital</td><td>{{buyers_credit_wc}}</td></tr>
                <tr style="height:600px"><td></td><td>LC Creditors(not included in creditors above)</td><td>{{LC_Creditors}}</td></tr>
                <tr style="height:600px"><td></td><td>PBD Utilised from other Bank</td><td>{{pbd_utilised}}</td></tr>
                <tr style="height:600px"><td></td><td>EPC/PCFC stock</td><td>{{epc_pcfc_stock}}</td></tr>
                <tr style="height:600px"><td></td><td>Channel Finance</td><td>{{channel_finance}}</td></tr>
                <tr style="height:600px"><td></td><td>Advance paid to suppliers</td><td>{{Advance_paid_to_suppliers}}</td></tr>
                <tr style="height:600px"><td></td><td>Stock applicable for DP</td><td>{{debtors_applicable_dp}}</td></tr>
                <tr style="height:600px"><td></td><td>Less Margin</td><td>{{less_margin}}</td></tr>
                <tr style="height:600px"><td></td><td>Net stock for DP</td><td>{{Net_stock}}</td></tr>
                <tr style="height:600px"><td></td><td> </td><td></td></tr>
                <tr style="height:600px"><td></td><td>Total debtors</td><td>{{TOTAL_Debitors}}</td></tr>
                <tr style="height:600px"><td></td><td>Less: Non Group co. drs. > 90 days as per term sheet days</td><td>{{wb_non_groupco_drs_days}}</td></tr>
                <tr style="height:600px"><td></td><td>Less : Bill discounted debtors</td><td>{{billed_discounted_debtors}}</td></tr>
                <tr style="height:600px"><td></td><td>Less : Group concern debtors</td><td>{{group_co_debtors}}</td></tr>
                <tr style="height:600px"><td></td><td>Less: Advance from Customer</td><td>{{Advance_from_Customer}}</td></tr>
                <tr style="height:600px"><td></td><td>Net Debtors</td><td>{{Net_debtors}}</td></tr>
                <tr style="height:600px"><td></td><td>Less Margin 25%</td><td>{{less_margin_25}}</td></tr>
                <tr style="height:600px"><td></td><td>Debtors applicable for DP</td><td>{{debtors_applicable_dp}}</td></tr>
                <tr style="height:600px"><td></td><td>Total DP</td><td>{{Total_Dp}}</td></tr>
                <tr style="height:600px"><td></td><td>Less: Other WC bankers outstanding</td><td>{{wb_dp_limit}}</td></tr>
                <tr style="height:600px"><td></td><td>Less: Kotak PO Funding</td><td>{{kotak_po_funding}}</td></tr>
                <tr style="height:600px"><td></td><td>Net DP</td><td>{{net_dp}}</td></tr>
                <tr style="height:600px"><td></td><td>DP Limit</td><td>{{dp_limit_consortium}}</td></tr>
                <tr style="height:600px"><td></td><td>Final DP</td><td>{{final_dp}}</td></tr>
                <tr style="height:600px"><td></td><td>Adhoc</td><td>{{adhoc}}</td></tr>
                <tr style="height:600px"><td></td><td>Final DP (with Adhoc)</td><td>{{final_dp_adhoc}}</td></tr>
                <tr style="height:600px"><td></td><td>Sales</td><td>{{Sales}}</td></tr>
                <tr style="height:600px"><td></td><td>Purchases</td><td>{{Purchases}}</td></tr>
                
                
            </table>

"""
            print(df_data)
            # Sample data for the template
            data = df_data
            template = Template(template_data1)
            rendered_template = template.render(data)
            # # Create a Jinja template object
            # template = Template(template_data1)


            # # Render the template with data
            # rendered_template = template.render(data)
            # print(rendered_template)
            # Parse the HTML table into a DataFrame
            dfs = pd.read_html(rendered_template)

            # Get the DataFrame from the list of DataFrames
            df = dfs[0]
            df.columns = [" ", " ", " ","","",]

            # print(df)



            output_path = Path(
                        f'/app/output/kmb/assets/pdf/kmb/{case_id}')
            os.umask(0)
            logging.info(f"---{output_path}---")
            df.to_excel(f"{output_path}/output.xlsx", index=False )
            
            
            logging.info(f"---saved---")
            workbook = openpyxl.load_workbook(f"{output_path}/output.xlsx",data_only=True)

            # Select the desired sheet
            sheet = workbook.active
            sheet.title = "Main_Format1"# You can also specify the sheet name like workbook['Sheet1']
            for row in range(10, 47):
                cell_address = 'C' + str(row)
                cell = sheet[cell_address]
                if cell.value is not None and isinstance(cell.value, str):
                    try:
                        print(type(cell.value),'------------------------------------------')
                        cell.value = float(cell.value) # Convert text to integer
                    except ValueError as e:
                        print(e,'------------------------------{e}')
                        # Handle non-convertible values here if needed
                        pass



            sheet.column_dimensions['B'].width = 60
            sheet.column_dimensions['c'].width = 60
            sheet.column_dimensions['d'].width = 40
            sheet['C16'] = '=SUM(C10:C15)'
            sheet['C24'] = '=C16-C17-C18-C19-C20-C21-C22+C23'
            sheet['C25'] = '=IF(C24<0,0,C24*D6)'
            sheet['C26'] = '=IF(C24<0,0,C24-C25)'
            sheet['C33'] = '=IF(C24<0,C28-C29-C30-C31-C32+C24,C28-C29-C30-C31-C32)'
            sheet['C34'] = '=C33*D7'
            sheet['C35'] = '=C33-C34'
            sheet['C36'] = '=IF(C24<0,C35,C26+C35)'
            sheet['C39'] = '=C36-C37-C38'
            sheet['C41'] = '=MIN(C40,C39)'
            sheet['C43'] = '=IF(C41+C42>C40,C40,C41+C42)'


            img_path = Path(f'/app/output/kmb/assets/pdf/kmb/kotak.PNG')
            img = openpyxl.drawing.image.Image(img_path)
            img.anchor = "B3"  

            for row in sheet.iter_rows(min_row=9, max_row=45, min_col=3, max_col=3):
                for cell in row:
                    cell.fill = PatternFill(start_color="53b553", end_color="53b553", fill_type="solid")
                    
                    # Green color
            for row in sheet.iter_rows(min_row=4, max_row=4, min_col=3, max_col=3):
                for cell in row:
                    cell.font = Font(size=16, bold=True, color="090a0a")
            for row in sheet.iter_rows(min_row=16, max_row=16, min_col=2, max_col=2):
                for cell in row:
                    cell.font = Font(size=10, bold=True, color="090a0a")        
                    
            for row in sheet.iter_rows(min_row=9, max_row=9, min_col=2, max_col=4):
                
                for cell in row:
                    cell.fill = PatternFill(start_color="626a80", end_color="626a80", fill_type="solid")
                    cell.font = Font(size=14, bold=True, color="FFFFFF")
            for row in sheet.iter_rows(min_row=6, max_row=7, min_col=2, max_col=3):
                for cell in row:
                    cell.fill = PatternFill(start_color="626a80", end_color="626a80", fill_type="solid")   
                    cell.font = Font(size=14, bold=True, color="FFFFFF")
            for row in sheet.iter_rows(min_row=5, max_row=5, min_col=2, max_col=2):
                for cell in row:
                    cell.fill = PatternFill(start_color="626a80", end_color="626a80", fill_type="solid")
                    cell.font = Font(size=14, bold=True, color="FFFFFF")

            # Save the changes to the workbook



            workbook.save(f"{output_path}/output.xlsx")
            logging.info(f"completed sheet 1-----")


            Heads= [
                "For the Month", "Due Date", "CRN", "Account Number", "Total Stock", "Stock Applicable for DP", 
                "Stock < 90", "Stock < 120", "Total Creditors", "Creditors Applicable for DP", 
                "Total Debtors", "Debtors Applicable for DP", "Debtors < 90", "Debtors < 120", 
                "Total Sales", "Total Purchases", "Other Bank Limit", "System DP", "System Limit", 
                "CC Limit", "Computed DP", "Final DP", "Stock <30 Days", "Stock Between 31 and 60 Days", 
                "Stock Between 61 and 90 Days", "Stock Between 91 and 120 Days", "Stock Between 121 and 180 Days", 
                "Stock > 180 Days", "Debtors <30 Days", "Debtors Between 31 and 60 Days", 
                "Debtors Between 61 and 90 Days", "Debtors Between 91 and 120 Days", 
                "Debtors Between 121 and 180 Days", "Debtors > 180 Days", "Sales Spares Workshop", 
                "Sales Amount Vehicle", "Purchase Spares Workshop", "Debtors > 90 Days Spares Workshop", 
                "Debtors > 90 Days Vehicle", "Debtors Spares Between 91 to 120", 
                "Debtors Vehicle Between 91 to 120", "Total Debtors Spares", "Total Spares Stock", 
                "Total Spares Stock <= 120", "Total Spares Stock <= 90", "Total Vehicle Stock", 
                "Total Vehicle Stock <= 120", "Total Vehicle Stock <= 90", 
                "Debtors Advance Paid to suppliers", "Purchase Vehicles", "Total Debtors Vehicles", 
                "Total Stock <=120", "Total Stock <=90", "Advance Received from customers", 
                "Bill discounted debtors", "Channel Finance", "Debtors 0 to 60 Days", 
                "Debtors 0 to 5 Days", "Debtors 6 to 30 Days", "Debtors 6 to 60 Days", 
                "Debtors < 150 days", "Debtors < 45 days", "Debtors Specific", 
                "Debtors of Vehicle < 30 days", "Debtors of Vehicle < 60 days", 
                "Debtors of spares < 60 days", "Group Co debtors", "Insurance Amount", 
                "Limits Os With KMBL Banks 1", "Outstanding with other WC Bankers", 
                "Previous month Total Creditors", "Previous month Total Debtors", 
                "Previous month Total Stock", "Purchase 1 Month prior", "Purchase 2 Month prior", 
                "Purchase 3 Month prior", "Purchase 4 Month prior", "Purchase 5 Month prior", 
                "Purchase Spares 1 Month Prior", "Purchase Spares 2 Month Prior", 
                "Purchase Spares 3 Month Prior", "Purchase Spares 4 Month Prior", 
                "Purchase Spares 5 Month Prior", "Purchase Vehicles 1 Month Prior", 
                "Purchase Vehicles 2 Month Prior", "Purchase Vehicles 3 Month Prior", 
                "Purchase Vehicles 4 Month Prior", "Purchase Vehicles 5 Month Prior", 
                "RM Stock < 60 days", "RM Stock < 90 days", "Sales 1 Month Prior", 
                "Sales 2 Month Prior", "Sales 3 Month Prior", "Sales 4 Month Prior", 
                "Sales 5 Month Prior", "Sales Spares 1 Month Prior", "Sales Spares 2 Month Prior", 
                "Sales Spares 3 Month Prior", "Sales Spares 4 Month Prior", 
                "Sales Spares 5 Month Prior", "Sales Vehicles 1 Month Prior", 
                "Sales Vehicles 2 Month Prior", "Sales Vehicles 3 Month Prior", 
                "Sales Vehicles 4 Month Prior", "Sales Vehicles 5 Month Prior", "Security Deposit", 
                "Stock Between 0 and 60 Days", "Stock Specific", "Stock in Transit < 90 Days", 
                "Stock of Vehicle < 60 days", "Stock of spares < 60 days", "Total Purchase Spares", 
                "Total Purchase Vehicles", "Total Sales Spares", "Total Sales Vehicles", 
                "Unbilled Debtors"
            ]
            # Number of columns
            num_columns = len(Heads)

            # Create 'Values' list with empty strings
            values_list = ["-"] * num_columns
            print(values_list)

            # Update the 'Values' key in the data2 dictionary


            # Print the updated data2 dictionary
            data2={'Heads':Heads,'Values':values_list}
            print(data2)



            # data2={'Heads':Heads,'Values':Values}
            df2 = pd.DataFrame(data2)
            # df2.to_excel("Downloads/output.xlsx", index=False )
            print(df2)
            new_sheet = workbook.create_sheet(title="Key Metrics")

            new_sheet.column_dimensions['A'].width = 50
            new_sheet.column_dimensions['B'].width = 50

            for row in dataframe_to_rows(df2, index=False, header=True):
                new_sheet.append(row)
            for row in new_sheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=2):
                for cell in row:
                    cell.fill = PatternFill(start_color="626a80", end_color="626a80", fill_type="solid")

                    

            new_sheet['B3']=df_data['Due_Date']
            new_sheet['B4']=df_data['Crn']
            new_sheet['B5']=df_data['ACCOUNT_NUMBER']
            new_sheet['B6']='=Main_Format1!C16'
            new_sheet['B7']='=Main_Format1!C26'
            new_sheet['B10']='=SUM(Main_Format1!C17,Main_Format1!C18,Main_Format1!C19)'
            new_sheet['B17']='=Main_Format1!C27'
            new_sheet['B12']='=Main_Format1!C28'
            new_sheet['B13']='=Main_Format1!C35'
            new_sheet['B8']='=Main_Format1!C28-Main_Format1!C29'
            new_sheet['B9']='=Main_Format1!C28-Main_Format1!C29'
            new_sheet['B16']='=Main_Format1!C44'
            new_sheet['B17']='=Main_Format1!C45'
            new_sheet['B18']='=Main_Format1!C37'
            new_sheet['B21']='=Main_Format1!C40'
            new_sheet['B22']='=Main_Format1!C39'
            new_sheet['B23']='=Main_Format1!C43'
            new_sheet['B50']='=Main_Format1!C23'
            new_sheet['B55']='=Main_Format1!C32'
            new_sheet['B56']='=Main_Format1!C30'
            new_sheet['B57']='=Main_Format1!C22'
            new_sheet['B68']='=Main_Format1!C31'
            new_sheet['B70']='=Main_Format1!C38'
            new_sheet['B71']='=Main_Format1!C37' 

            report_name="Kotak_WBG_100000_Output_Mapping_V1.0"

        #template3
        if '22222' in template_name:
            template_data1 = """



    <table style="width:100%">
    <tr><th></th><td></td><td></td></tr>
        <tr><th></th><td></td><td></td></tr>
        <tr style="height:600px";"font-size: 20px;"><td></td><td></td><td style="font-size: 20px;">DP STOCK STATEMENT</td><td></td><td></td></tr>
        <tr style="height:600px"><td></td><td>Name of Borrower</td><td>{{NAME_OF_CUSTOMER}}</td><td>{{Date}}</td></tr>
<tr style="height:600px"><td></td><td>Margin</td><td>Stock</td><td>{{STOCK_MARGIN}}</td></tr >
<tr style="height:600px"><td></td><td></td><td>Debtors</td><td>{{DEBTORS_MARGIN}}</td></tr >
<tr style="height:600px"><td></td><td></td><td></td><td></td> </tr >
<tr style="height:600px"><td></td><td></td><td>Details for the month </td><td></td></tr >
<tr style="height:600px";"color: red;"><td></td><td>Raw Material</td><td>{{Raw_Material}}</td></tr>
<tr style="height:600px"><td></td><td style="color: blue;">Work in Process</td><td>{{Work_in_Process}}</td></tr>
<tr style="height:600px"><td></td><td>Finished Goods</td><td>{{Finished_Goods}}</td></tr>
<tr style="height:600px"><td></td><td>Stores & Spares</td><td>{{Stores_Spares}}</td></tr>
<tr style="height:600px"><td></td><td>Other stock</td><td>{{Other_stock}}</td></tr>
<tr style="height:600px"><td></td><td>Stock in Transit</td><td>{{Stock_in_Transit}}</td></tr>
<tr style="height:600px"><td></td><td>Total Stock</td><td>{{Total_stock}}</td></tr>
<tr style="height:600px"><td></td><td>Less: Creditors</td><td>{{Less_Creditors}}</td></tr>
<tr style="height:600px"><td></td><td>Buyers Credit for working capital</td><td>{{buyers_credit_wc}}</td></tr>
<tr style="height:600px"><td></td><td>LC Creditors(not included in creditors above)</td><td>{{LC_Creditors}}</td></tr>
<tr style="height:600px"><td></td><td>PBD Utilised from other Bank</td><td>{{pbd_utilised}}</td></tr>
<tr style="height:600px"><td></td><td>EPC/PCFC stock</td><td>{{epc_pcfc_stock}}</td></tr>
<tr style="height:600px"><td></td><td>Net stock</td><td>{{Net_stock}}</td></tr>
<tr style="height:600px"><td></td><td>Less Margin 25%</td><td>{{less_margin_25}}</td></tr>
<tr style="height:600px"><td>A</td><td>Considered for DP</td><td>{{considered_dp}}</td></tr>
<tr style="height:600px"><td></td><td></td><td></td></tr>
<tr style="height:600px"><td></td><td>Total debtors</td><td>{{TOTAL_Debitors}}</td></tr>
<tr style="height:600px"><td></td><td>Non Group co. drs.>120 days as per term sheet  days</td><td>{{wb_non_groupco_drs_days}}</td></tr>
<tr style="height:600px"><td></td><td> </td><td></td></tr>
<tr style="height:600px"><td></td><td>Less : Bill discounted debtors</td><td>{{Bill_discunted}}</td></tr>
<tr style="height:600px"><td></td><td>Less : Group concern debtors</td><td>{{group_co_debtors}}</td></tr>
<tr style="height:600px"><td></td><td>Less: Advance from Customer</td><td>{{Advance_from_Customer}}</td></tr>
<tr style="height:600px"><td></td><td>Net Debtors</td><td>{{Net_debtors}}</td></tr>
<tr style="height:600px"><td></td><td>Less Margin 25%</td><td>{{less_margin_25}}</td></tr>
<tr style="height:600px"><td></td><td>Considered for DP</td><td>{{considered_dp}}</td></tr>
<tr style="height:600px"><td></td><td>Total DP</td><td>{{Total_Dp}}</td></tr>
<tr style="height:600px"><td></td><td>Our Share</td><td>{{our_share}}</td></tr>
<tr style="height:600px"><td></td><td>Net DP as per our share (if lead bank letter not received)</td><td>{{net_dp_our_share}}</td></tr>
<tr style="height:600px"><td></td><td>advised_as_lead_bank</td><td>{{advised_as_lead_bank}}</td></tr>
<tr style="height:600px"><td></td><td>DP Limit(Consortium)</td><td>{{dp_limit_consortium}}</td></tr>
<tr style="height:600px"><td></td><td>Final DP within consortium</td><td>{{final_dp_consortium}}</td></tr>
<tr style="height:600px"><td></td><td>DP for outside consortium</td><td>{{outside_consortium}}</td></tr>
<tr style="height:600px"><td></td><td>Total DP</td><td>{{Total_Dp}}</td></tr>
<tr style="height:600px"><td></td><td>Less: Bank o/s</td><td>{{less_bank_os}}</td></tr>
<tr style="height:600px"><td></td><td>DP available for outside consortium</td><td>{{dp_available_for_outside_consor}}</td></tr>
<tr style="height:600px"><td></td><td>DP Limit( outside Consortium)</td><td>{{limit_consortium}}</td></tr>
<tr style="height:600px"><td></td><td>DP for outside consortium</td><td>{{outside_consortium}}</td></tr>
<tr style="height:600px"><td></td><td>Final DP</td><td>{{final_dp}}</td></tr>
<tr style="height:600px"><td></td><td>Adhoc</td><td>{{adhoc}}</td></tr>
<tr style="height:600px"><td></td><td>Final DP (with Adhoc)</td><td>{{final_dp_adhoc}}</td></tr>
<tr style="height:600px"><td></td><td>Sales</td><td>{{Sales}}</td></tr>
<tr style="height:600px"><td></td><td>Purchases</td><td>{{Purchases}}</td></tr>
</table>

    """

            print(df_data)
            # Sample data for the template
            data = df_data
            template = Template(template_data1)
            rendered_template = template.render(data)
            # # Create a Jinja template object
            # template = Template(template_data1)


            # # Render the template with data
            # rendered_template = template.render(data)
            # print(rendered_template)
            # Parse the HTML table into a DataFrame
            dfs = pd.read_html(rendered_template)

            # Get the DataFrame from the list of DataFrames
            df = dfs[0]
            df.columns = [" ", " ", " ","","",]

            # print(df)



            output_path = Path(
                        f'/app/output/kmb/assets/pdf/kmb/{case_id}')
            os.umask(0)
            logging.info(f"---{output_path}---")
            df.to_excel(f"{output_path}/output.xlsx", index=False )
            
            
            logging.info(f"---saved---")
            workbook = openpyxl.load_workbook(f"{output_path}/output.xlsx",data_only=True)

            # Select the desired sheet
            sheet = workbook.active
            sheet.title = "Main_Format1"# You can also specify the sheet name like workbook['Sheet1']
            for row in range(10, 47):
                cell_address = 'C' + str(row)
                cell = sheet[cell_address]
                if cell.value is not None and isinstance(cell.value, str):
                    try:
                        print(type(cell.value),'------------------------------------------')
                        cell.value = float(cell.value) # Convert text to integer
                    except ValueError as e:
                        print(e,'------------------------------{e}')
                        # Handle non-convertible values here if needed
                        pass



            sheet.column_dimensions['B'].width = 60
            sheet.column_dimensions['c'].width = 60
            sheet.column_dimensions['d'].width = 40
            sheet['C16'] = '=SUM(C10:C15)'
            sheet['C22'] = '=C16-C17-C18-C19-C20-C21'
            sheet['C23'] = '=IF(C22<=0,0,C22*D6)'
            sheet['C24'] = '=IF(C22<0,0,C22-C23)'
            sheet['C32'] = '=IF(C22<0,C26-C27-C29-C30-C31+C22,C26-C27-C29-C30-C31)'
            sheet['C33']='=C32*D7'
            sheet['C34'] ='=C32-C33'
            sheet['C35']='=IF(C22<0,C34,C34+C34)'
            sheet['C40']='=IF(C38<0,MIN(C37,C39),MIN(C38,C39))'
            sheet['C44']='=C42-C43'
            sheet['C46']='=MIN(C44,C45)'
            sheet['C37']='=C35* C36'
            sheet['C47']='=C40+C41'
            sheet['C49']='=C47+C48'
        

            img_path = Path(f'/app/output/kmb/assets/pdf/kmb/kotak.PNG')
            img = openpyxl.drawing.image.Image(img_path)
            img.anchor = "B3"  

            for row in sheet.iter_rows(min_row=9, max_row=51, min_col=3, max_col=3):
                for cell in row:
                    cell.fill = PatternFill(start_color="53b553", end_color="53b553", fill_type="solid")
                    
                    # Green color
            for row in sheet.iter_rows(min_row=4, max_row=4, min_col=3, max_col=3):
                for cell in row:
                    cell.font = Font(size=16, bold=True, color="090a0a")
            for row in sheet.iter_rows(min_row=16, max_row=16, min_col=2, max_col=2):
                for cell in row:
                    cell.font = Font(size=10, bold=True, color="090a0a")        
                    
            for row in sheet.iter_rows(min_row=9, max_row=9, min_col=2, max_col=4):
                
                for cell in row:
                    cell.fill = PatternFill(start_color="626a80", end_color="626a80", fill_type="solid")
                    cell.font = Font(size=14, bold=True, color="FFFFFF")
            for row in sheet.iter_rows(min_row=6, max_row=7, min_col=2, max_col=3):
                for cell in row:
                    cell.fill = PatternFill(start_color="626a80", end_color="626a80", fill_type="solid")   
                    cell.font = Font(size=14, bold=True, color="FFFFFF")
            for row in sheet.iter_rows(min_row=5, max_row=5, min_col=2, max_col=2):
                for cell in row:
                    cell.fill = PatternFill(start_color="626a80", end_color="626a80", fill_type="solid")
                    cell.font = Font(size=14, bold=True, color="FFFFFF")

            # Save the changes to the workbook



            workbook.save(f"{output_path}/output.xlsx")
            logging.info(f"completed sheet 1-----")


            Heads= [
                "For the Month", "Due Date", "CRN", "Account Number", "Total Stock", "Stock Applicable for DP", 
                "Stock < 90", "Stock < 120", "Total Creditors", "Creditors Applicable for DP", 
                "Total Debtors", "Debtors Applicable for DP", "Debtors < 90", "Debtors < 120", 
                "Total Sales", "Total Purchases", "Other Bank Limit", "System DP", "System Limit", 
                "CC Limit", "Computed DP", "Final DP", "Stock <30 Days", "Stock Between 31 and 60 Days", 
                "Stock Between 61 and 90 Days", "Stock Between 91 and 120 Days", "Stock Between 121 and 180 Days", 
                "Stock > 180 Days", "Debtors <30 Days", "Debtors Between 31 and 60 Days", 
                "Debtors Between 61 and 90 Days", "Debtors Between 91 and 120 Days", 
                "Debtors Between 121 and 180 Days", "Debtors > 180 Days", "Sales Spares Workshop", 
                "Sales Amount Vehicle", "Purchase Spares Workshop", "Debtors > 90 Days Spares Workshop", 
                "Debtors > 90 Days Vehicle", "Debtors Spares Between 91 to 120", 
                "Debtors Vehicle Between 91 to 120", "Total Debtors Spares", "Total Spares Stock", 
                "Total Spares Stock <= 120", "Total Spares Stock <= 90", "Total Vehicle Stock", 
                "Total Vehicle Stock <= 120", "Total Vehicle Stock <= 90", 
                "Debtors Advance Paid to suppliers", "Purchase Vehicles", "Total Debtors Vehicles", 
                "Total Stock <=120", "Total Stock <=90", "Advance Received from customers", 
                "Bill discounted debtors", "Channel Finance", "Debtors 0 to 60 Days", 
                "Debtors 0 to 5 Days", "Debtors 6 to 30 Days", "Debtors 6 to 60 Days", 
                "Debtors < 150 days", "Debtors < 45 days", "Debtors Specific", 
                "Debtors of Vehicle < 30 days", "Debtors of Vehicle < 60 days", 
                "Debtors of spares < 60 days", "Group Co debtors", "Insurance Amount", 
                "Limits Os With KMBL Banks 1", "Outstanding with other WC Bankers", 
                "Previous month Total Creditors", "Previous month Total Debtors", 
                "Previous month Total Stock", "Purchase 1 Month prior", "Purchase 2 Month prior", 
                "Purchase 3 Month prior", "Purchase 4 Month prior", "Purchase 5 Month prior", 
                "Purchase Spares 1 Month Prior", "Purchase Spares 2 Month Prior", 
                "Purchase Spares 3 Month Prior", "Purchase Spares 4 Month Prior", 
                "Purchase Spares 5 Month Prior", "Purchase Vehicles 1 Month Prior", 
                "Purchase Vehicles 2 Month Prior", "Purchase Vehicles 3 Month Prior", 
                "Purchase Vehicles 4 Month Prior", "Purchase Vehicles 5 Month Prior", 
                "RM Stock < 60 days", "RM Stock < 90 days", "Sales 1 Month Prior", 
                "Sales 2 Month Prior", "Sales 3 Month Prior", "Sales 4 Month Prior", 
                "Sales 5 Month Prior", "Sales Spares 1 Month Prior", "Sales Spares 2 Month Prior", 
                "Sales Spares 3 Month Prior", "Sales Spares 4 Month Prior", 
                "Sales Spares 5 Month Prior", "Sales Vehicles 1 Month Prior", 
                "Sales Vehicles 2 Month Prior", "Sales Vehicles 3 Month Prior", 
                "Sales Vehicles 4 Month Prior", "Sales Vehicles 5 Month Prior", "Security Deposit", 
                "Stock Between 0 and 60 Days", "Stock Specific", "Stock in Transit < 90 Days", 
                "Stock of Vehicle < 60 days", "Stock of spares < 60 days", "Total Purchase Spares", 
                "Total Purchase Vehicles", "Total Sales Spares", "Total Sales Vehicles", 
                "Unbilled Debtors"
            ]
            # Number of columns
            num_columns = len(Heads)

            # Create 'Values' list with empty strings
            values_list = ["-"] * num_columns
            print(values_list)

            # Update the 'Values' key in the data2 dictionary


            # Print the updated data2 dictionary
            data2={'Heads':Heads,'Values':values_list}
            print(data2)



            # data2={'Heads':Heads,'Values':Values}
            df2 = pd.DataFrame(data2)
            # df2.to_excel("Downloads/output.xlsx", index=False )
            print(df2)
            new_sheet = workbook.create_sheet(title="Key Metrics")

            new_sheet.column_dimensions['A'].width = 50
            new_sheet.column_dimensions['B'].width = 50





            for row in dataframe_to_rows(df2, index=False, header=True):
                new_sheet.append(row)
            for row in new_sheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=2):
                for cell in row:
                    cell.fill = PatternFill(start_color="626a80", end_color="626a80", fill_type="solid")

                    

            new_sheet['B3']=df_data['Due_Date']
            new_sheet['B4']=df_data['Crn']
            new_sheet['B5']=df_data['ACCOUNT_NUMBER']
            new_sheet['B6']='=Main_Format1!C16'
            new_sheet['B7']='=Main_Format1!C26'
            new_sheet['B10']='=SUM(Main_Format1!C17,Main_Format1!C18,Main_Format1!C19)'
            new_sheet['B17']='=Main_Format1!C27'
            new_sheet['B12']='=Main_Format1!C28'
            new_sheet['B13']='=Main_Format1!C35'
            new_sheet['B8']='=Main_Format1!C28-Main_Format1!C29'
            new_sheet['B9']='=Main_Format1!C28-Main_Format1!C29'
            new_sheet['B16']='=Main_Format1!C44'
            new_sheet['B17']='=Main_Format1!C45'
            new_sheet['B18']='=Main_Format1!C37'
            new_sheet['B21']='=Main_Format1!C40'
            new_sheet['B22']='=Main_Format1!C39'
            new_sheet['B23']='=Main_Format1!C43'
            new_sheet['B50']='=Main_Format1!C23'
            new_sheet['B55']='=Main_Format1!C32'
            new_sheet['B56']='=Main_Format1!C30'
            new_sheet['B57']='=Main_Format1!C22'
            new_sheet['B68']='=Main_Format1!C31'
            new_sheet['B70']='=Main_Format1!C38'
            new_sheet['B71']='=Main_Format1!C37' 

            report_name="Kotak_WBG_222222_CONSORTIUM_Output_Mapping_V1.0"

        #template4
        if 'WBG_100000_JASANI' in template_name:
            template_data1 = """


            <table style="width:100%">
            <tr><th></th><td></td><td></td></tr>
                <tr><th></th><td></td><td></td></tr>
                <tr style="height:600px";"font-size: 20px;"><td></td><td></td><td style="font-size: 20px;">DP STOCK STATEMENT</td><td></td><td></td></tr>
                <tr style="height:600px"><td></td><td>Name of Borrower</td><td>{{NAME_OF_CUSTOMER}}</td><td></td></tr>
                <tr style="height:600px"><td></td><td>Margin</td><td>Stock</td><td>{{STOCK_MARGIN}}</td></tr >
                <tr style="height:600px"><td></td><td></td><td>Debtors</td><td>{{DEBTORS_MARGIN}}</td></tr >
                <tr style="height:600px"><td></td><td></td><td></td><td></td> </tr >
                <tr style="height:600px"><td></td><td>Details for the month</td><td>EPC-Old & New Limits</td><td>CC/WCDL- Limits secured by MF</td><td></td></tr >
                <tr style="height:600px";"color: red;"><td></td><td>Raw Material</td><td>{{Raw_Material}}</td></tr>
                <tr style="height:600px"><td></td><td style="color: blue;">Work in Process</td><td>{{Work_in_Process}}</td></tr>
                <tr style="height:600px"><td></td><td>Finished Goods</td><td>{{Finished_Goods}}</td></tr>
                <tr style="height:600px"><td></td><td>Stores & Spares</td><td>{{Stores_Spares}}</td></tr>
                <tr style="height:600px"><td></td><td>Other stock</td><td>{{Other_stock}}</td></tr>
                <tr style="height:600px"><td></td><td>Stock in Transit</td><td>{{Stock_in_Transit}}</td></tr>
                <tr style="height:600px"><td></td><td>Total Stock</td><td>{{Total_stock}}</td></tr>
                <tr style="height:600px"><td></td><td>Less: Creditors</td><td>{{Less_Creditors}}</td></tr>
                <tr style="height:600px"><td></td><td>Buyers Credit for working capital</td><td>{{buyers_credit_wc}}</td></tr>
                <tr style="height:600px"><td></td><td>LC Creditors(not included in creditors above)</td><td>{{LC_Creditors}}</td></tr>
                <tr style="height:600px"><td></td><td>PBD Utilised from other Bank</td><td>{{pbd_utilised}}</td></tr>
                <tr style="height:600px"><td></td><td>EPC/PCFC stock</td><td>{{epc_pcfc_stock}}</td></tr>
                <tr style="height:600px"><td></td><td>Stock applicable for DP</td><td>{{stock_applicable_dp}}</td></tr>
                <tr style="height:600px"><td></td><td>Less Margin</td><td>{{less_margin}}</td></tr>
                <tr style="height:600px"><td>A</td><td>Net stock for DP</td><td>{{Net_stock}}</td></tr>
                <tr style="height:600px"><td></td><td></td><td></td></tr>
                <tr style="height:600px"><td></td><td>Total debtors</td><td>{{TOTAL_Debitors}}</td></tr>   
                <tr style="height:600px"><td></td><td>Less: Non Group co. drs.> 120 days as per term sheet  days</td><td>{{wb_non_groupco_drs_days}}</td></tr>             
                <tr style="height:600px"><td></td><td>Less : Bill discounted debtors</td><td>{{bill_discounted_debtors}}</td></tr>
                <tr style="height:600px"><td></td><td>Less : Group concern debtors</td><td>{{group_co_debtors}}</td></tr>
                <tr style="height:600px"><td></td><td>Less: Advance from Customer</td><td>{{Advance_from_Customer}}</td></tr>       
                <tr style="height:600px"><td>B</td><td>Net Debtors</td><td>{{Net_debtors}}</td></tr>       
                <tr style="height:600px"><td></td><td>Less Margin</td><td>{{less_margin}}</td></tr>      
                <tr style="height:600px"><td></td><td>Debtors applicable for DP</td><td>{{debtors_applicable_dp}}</td></tr>
                <tr style="height:600px"><td>(A+B)</td><td>Total DP</td><td>{{Total_Dp}}</td></tr>  
                <tr style="height:600px"><td></td><td>Less: Other WC bankers outstanding </td><td>{{less_other_wc_bankers}}</td></tr>   
                <tr style="height:600px"><td></td><td>Less: Kotak PO Funding</td><td>{{kotak_po_funding}}</td></tr>
                <tr style="height:600px"><td></td><td>Net DP</td><td>{{net_dp}}</td></tr>           
                <tr style="height:600px"><td></td><td>DP Limit</td><td>{{dp_limit_consortium}}</td></tr>     
                <tr style="height:600px"><td></td><td>Final DP</td><td>{{final_dp}}</td></tr>
                <tr style="height:600px"><td></td><td>Adhoc</td><td>{{adhoc}}</td></tr>
                <tr style="height:600px"><td></td><td>Final DP (with Adhoc)</td><td>{{final_dp_adhoc}}</td></tr>
                <tr style="height:600px"><td></td><td>Sales</td><td>{{Sales}}</td></tr>
                <tr style="height:600px"><td></td><td>Purchases</td><td>{{Purchases}}</td></tr>
                
                
            </table>

"""

            print(df_data)
            # Sample data for the template
            data = df_data
            template = Template(template_data1)
            rendered_template = template.render(data)
            # # Create a Jinja template object
            # template = Template(template_data1)


            # # Render the template with data
            # rendered_template = template.render(data)
            # print(rendered_template)
            # Parse the HTML table into a DataFrame
            dfs = pd.read_html(rendered_template)

            # Get the DataFrame from the list of DataFrames
            df = dfs[0]
            df.columns = [" ", " ", " ","","",]

            # print(df)
            output_path = Path(
                        f'/app/output/kmb/assets/pdf/kmb/{case_id}')
            os.umask(0)
            logging.info(f"---{output_path}---")
            df.to_excel(f"{output_path}/output.xlsx", index=False )
            
            
            logging.info(f"---saved---")
            workbook = openpyxl.load_workbook(f"{output_path}/output.xlsx",data_only=True)

            # Select the desired sheet
            sheet = workbook.active
            sheet.title = "Main_Format1"# You can also specify the sheet name like workbook['Sheet1']
            for row in range(10, 47):
                cell_address = 'C' + str(row)
                cell = sheet[cell_address]
                if cell.value is not None and isinstance(cell.value, str):
                    try:
                        print(type(cell.value),'------------------------------------------')
                        cell.value = float(cell.value) # Convert text to integer
                    except ValueError as e:
                        print(e,'------------------------------{e}')
                        # Handle non-convertible values here if needed
                        pass



            sheet.column_dimensions['B'].width = 60
            sheet.column_dimensions['c'].width = 60
            sheet.column_dimensions['d'].width = 40
            sheet['C16'] = '=SUM(C10:C15)'
            sheet['C22'] = '=C16-C17-C18-C19-C20-C21'
            sheet['C23'] = '=IF(C22<0,0,C22*D6)'
            sheet['C24'] = '=IF(C22<0,0,C22-C23)'
            sheet['C31'] = '=IF(C22<0,C26-C27-C28-C29-C30+C22,C26-C27-C28-C29-C30)'
            sheet['C32'] = '=C31*D7'
            sheet['C33'] = '=C31-C32'
            sheet['C34'] = '=IF(C22<0,C33,C24+C33)'
            sheet['C37'] = '=C34-C35-C36'
            sheet['C39'] = '=MIN(C38,C37)'
            sheet['C41'] = '=IF(C39+C40>C38,C38,C39+C40)'
   

            img_path = Path(f'/app/output/kmb/assets/pdf/kmb/kotak.PNG')
            img = openpyxl.drawing.image.Image(img_path)
            img.anchor = "B3"  

            for row in sheet.iter_rows(min_row=9, max_row=43, min_col=3, max_col=3):
                for cell in row:
                    cell.fill = PatternFill(start_color="53b553", end_color="53b553", fill_type="solid")

                    # Green color
            for row in sheet.iter_rows(min_row=4, max_row=4, min_col=3, max_col=3):
                for cell in row:
                    cell.font = Font(size=16, bold=True, color="090a0a")
            for row in sheet.iter_rows(min_row=16, max_row=16, min_col=2, max_col=2):
                for cell in row:
                    cell.font = Font(size=10, bold=True, color="090a0a")        

            for row in sheet.iter_rows(min_row=9, max_row=9, min_col=2, max_col=4):

                for cell in row:
                    cell.fill = PatternFill(start_color="626a80", end_color="626a80", fill_type="solid")
                    cell.font = Font(size=14, bold=True, color="FFFFFF")
            for row in sheet.iter_rows(min_row=6, max_row=7, min_col=2, max_col=3):
                for cell in row:
                    cell.fill = PatternFill(start_color="626a80", end_color="626a80", fill_type="solid")   
                    cell.font = Font(size=14, bold=True, color="FFFFFF")
            for row in sheet.iter_rows(min_row=5, max_row=5, min_col=2, max_col=2):
                for cell in row:
                    cell.fill = PatternFill(start_color="626a80", end_color="626a80", fill_type="solid")
                    cell.font = Font(size=14, bold=True, color="FFFFFF")

            # Save the changes to the workbook



            workbook.save(f"{output_path}/output.xlsx")
            logging.info(f"completed sheet 1-----")


            Heads= [
                "For the Month", "Due Date", "CRN", "Account Number", "Total Stock", "Stock Applicable for DP", 
                "Stock < 90", "Stock < 120", "Total Creditors", "Creditors Applicable for DP", 
                "Total Debtors", "Debtors Applicable for DP", "Debtors < 90", "Debtors < 120", 
                "Total Sales", "Total Purchases", "Other Bank Limit", "System DP", "System Limit", 
                "CC Limit", "Computed DP", "Final DP", "Stock <30 Days", "Stock Between 31 and 60 Days", 
                "Stock Between 61 and 90 Days", "Stock Between 91 and 120 Days", "Stock Between 121 and 180 Days", 
                "Stock > 180 Days", "Debtors <30 Days", "Debtors Between 31 and 60 Days", 
                "Debtors Between 61 and 90 Days", "Debtors Between 91 and 120 Days", 
                "Debtors Between 121 and 180 Days", "Debtors > 180 Days", "Sales Spares Workshop", 
                "Sales Amount Vehicle", "Purchase Spares Workshop", "Debtors > 90 Days Spares Workshop", 
                "Debtors > 90 Days Vehicle", "Debtors Spares Between 91 to 120", 
                "Debtors Vehicle Between 91 to 120", "Total Debtors Spares", "Total Spares Stock", 
                "Total Spares Stock <= 120", "Total Spares Stock <= 90", "Total Vehicle Stock", 
                "Total Vehicle Stock <= 120", "Total Vehicle Stock <= 90", 
                "Debtors Advance Paid to suppliers", "Purchase Vehicles", "Total Debtors Vehicles", 
                "Total Stock <=120", "Total Stock <=90", "Advance Received from customers", 
                "Bill discounted debtors", "Channel Finance", "Debtors 0 to 60 Days", 
                "Debtors 0 to 5 Days", "Debtors 6 to 30 Days", "Debtors 6 to 60 Days", 
                "Debtors < 150 days", "Debtors < 45 days", "Debtors Specific", 
                "Debtors of Vehicle < 30 days", "Debtors of Vehicle < 60 days", 
                "Debtors of spares < 60 days", "Group Co debtors", "Insurance Amount", 
                "Limits Os With KMBL Banks 1", "Outstanding with other WC Bankers", 
                "Previous month Total Creditors", "Previous month Total Debtors", 
                "Previous month Total Stock", "Purchase 1 Month prior", "Purchase 2 Month prior", 
                "Purchase 3 Month prior", "Purchase 4 Month prior", "Purchase 5 Month prior", 
                "Purchase Spares 1 Month Prior", "Purchase Spares 2 Month Prior", 
                "Purchase Spares 3 Month Prior", "Purchase Spares 4 Month Prior", 
                "Purchase Spares 5 Month Prior", "Purchase Vehicles 1 Month Prior", 
                "Purchase Vehicles 2 Month Prior", "Purchase Vehicles 3 Month Prior", 
                "Purchase Vehicles 4 Month Prior", "Purchase Vehicles 5 Month Prior", 
                "RM Stock < 60 days", "RM Stock < 90 days", "Sales 1 Month Prior", 
                "Sales 2 Month Prior", "Sales 3 Month Prior", "Sales 4 Month Prior", 
                "Sales 5 Month Prior", "Sales Spares 1 Month Prior", "Sales Spares 2 Month Prior", 
                "Sales Spares 3 Month Prior", "Sales Spares 4 Month Prior", 
                "Sales Spares 5 Month Prior", "Sales Vehicles 1 Month Prior", 
                "Sales Vehicles 2 Month Prior", "Sales Vehicles 3 Month Prior", 
                "Sales Vehicles 4 Month Prior", "Sales Vehicles 5 Month Prior", "Security Deposit", 
                "Stock Between 0 and 60 Days", "Stock Specific", "Stock in Transit < 90 Days", 
                "Stock of Vehicle < 60 days", "Stock of spares < 60 days", "Total Purchase Spares", 
                "Total Purchase Vehicles", "Total Sales Spares", "Total Sales Vehicles", 
                "Unbilled Debtors"
            ]
            # Number of columns
            num_columns = len(Heads)

            # Create 'Values' list with empty strings
            values_list = ["-"] * num_columns
            print(values_list)

            # Update the 'Values' key in the data2 dictionary


            # Print the updated data2 dictionary
            data2={'Heads':Heads,'Values':values_list}
            print(data2)



            # data2={'Heads':Heads,'Values':Values}
            df2 = pd.DataFrame(data2)
            # df2.to_excel("Downloads/output.xlsx", index=False )
            print(df2)
            new_sheet = workbook.create_sheet(title="Key Metrics")

            new_sheet.column_dimensions['A'].width = 50
            new_sheet.column_dimensions['B'].width = 50
            for row in dataframe_to_rows(df2, index=False, header=True):
                new_sheet.append(row)
            for row in new_sheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=2):
                for cell in row:
                    cell.fill = PatternFill(start_color="626a80", end_color="626a80", fill_type="solid")

                    

            new_sheet['B3']=df_data['Due_Date']
            new_sheet['B4']=df_data['Crn']
            new_sheet['B5']=df_data['ACCOUNT_NUMBER']
            new_sheet['B6']='=Main_Format1!C16'
            new_sheet['B7']='=Main_Format1!C26'
            new_sheet['B10']='=SUM(Main_Format1!C17,Main_Format1!C18,Main_Format1!C19)'
            new_sheet['B17']='=Main_Format1!C27'
            new_sheet['B12']='=Main_Format1!C28'
            new_sheet['B13']='=Main_Format1!C35'
            new_sheet['B8']='=Main_Format1!C28-Main_Format1!C29'
            new_sheet['B9']='=Main_Format1!C28-Main_Format1!C29'
            new_sheet['B16']='=Main_Format1!C44'
            new_sheet['B17']='=Main_Format1!C45'
            new_sheet['B18']='=Main_Format1!C37'
            new_sheet['B21']='=Main_Format1!C40'
            new_sheet['B22']='=Main_Format1!C39'
            new_sheet['B23']='=Main_Format1!C43'
            new_sheet['B50']='=Main_Format1!C23'
            new_sheet['B55']='=Main_Format1!C32'
            new_sheet['B56']='=Main_Format1!C30'
            new_sheet['B57']='=Main_Format1!C22'
            new_sheet['B68']='=Main_Format1!C31'
            new_sheet['B70']='=Main_Format1!C38'
            new_sheet['B71']='=Main_Format1!C37' 

            report_name="Kotak_WBG_100000_JASANI_Output_Mapping_V1.0"

        #template5
        if 'WBG_444444_PRIVI' in template_name:
            template_data1 ="""<table style="width:100%">
    <tr><th></th><td></td><td></td></tr>
        <tr><th></th><td></td><td></td></tr>
        <tr style="height:600px";"font-size: 20px;"><td></td><td></td><td style="font-size: 20px;">DP STOCK STATEMENT</td><td>Division 1</td><td></td></tr>
        <tr style="height:600px"><td></td><td>Name of Borrower</td><td>{{NAME_OF_CUSTOMER}}</td><td></td></tr>
        <tr style="height:600px"><td></td><td></td><td></td><td>As per SANCTION</td></tr >
        <tr style="height:600px"><td></td><td>Details for the month </td><td>Statement value</td><td>Margin</td><td>Net value</td></tr>
        <tr style="height:600px"><td></td><td>Raw Material</td><td>{{Raw_Material}}</td><td>{{STOCK_MARGIN}}</td></tr>
        <tr style="height:600px"><td></td><td>Work in Process</td><td>{{Work_in_Process}}</td><td>{{STOCK_MARGIN}}</td></tr>
        <tr style="height:600px"><td></td><td>Finished Goods</td><td>{{Finished_Goods}}</td><td>{{STOCK_MARGIN}}</td></tr>
        <tr style="height:600px"><td></td><td>Stores & Spares</td><td>{{Stores_Spares}}</td><td>{{STOCK_MARGIN}}</td></tr>
        <tr style="height:600px"><td></td><td>Stock in Transit</td><td>{{Stock_in_Transit}}</td><td>{{STOCK_MARGIN}}</td></tr>
        <tr style="height:600px"><td></td><td>Other stock</td><td>{{Other_stock}}</td><td>{{STOCK_MARGIN}}</td></tr>   
        <tr style="height:600px"><td></td><td>Total Stock</td><td>{{Total_stock}}</td></tr>
        <tr style="height:600px"><td></td><td>Less: Total Creditors(BC/LC) </td><td>{{Creditors}}</td></tr>
        <tr style="height:600px"><td></td><td>PBD Utilised from other Bank</td><td>{{pbd_utilised}}</td></tr>
        <tr style="height:600px"><td></td><td>Advance paid to suppliers </td><td>{{Advance_paid_to_suppliers}}</td></tr>
        <tr style="height:600px"><td></td><td>Channel Finance </td><td>{{channel_finance}}</td></tr>  
        <tr style="height:600px"><td></td><td>EPC/PCFC stock</td><td>{{epc_pcfc_stock}}</td></tr>
        <tr style="height:600px"><td></td><td>Net stock</td><td>{{Net_stock}}</td></tr>
        <tr style="height:600px"><td>A</td><td>Drawing power I</td><td>{{drawing_power_1}}</td></tr>
        <tr style="height:600px"><td></td><td></td><td></td></tr>     
        <tr style="height:600px"><td></td><td>Total debtors</td><td>{{TOTAL_Debitors}}</td></tr>         
        <tr style="height:600px"><td></td><td>Domestic receivables</td><td>{{wb_non_groupco_drs_days}}</td></tr>             
        <tr style="height:600px"><td></td><td>Domestic receivables beyond cover period</td><td>{{domestic_rec_beyond_days}}</td></tr>
        <tr style="height:600px"><td></td><td>Group Co debtors</td><td>{{group_co_debtors}}</td></tr>
        <tr style="height:600px"><td></td><td>Bill Discounted</td><td>{{bill_discounted}}</td></tr>
        <tr style="height:600px"><td></td><td>Less: Deductibles(Advances)</td><td>{{deductibles}}</td></tr>
        <tr style="height:600px"><td>B</td><td>Net Domestic Receivables</td><td>{{net_domestic_receviables}}</td><td>{{STOCK_MARGIN}}</td></tr>
        <tr style="height:600px"><td></td><td></td><td></td></tr>  
        <tr style="height:600px"><td></td><td>Export receivables</td><td>{{export_receviables}}</td><td>{{STOCK_MARGIN}}</td></tr>
        <tr style="height:600px"><td></td><td>Export receivables beyond cover period</td><td>{{export_receviables_beyond}}</td></tr>
        <tr style="height:600px"><td></td><td>less: Deductibles</td><td>{{less_deductibles}}</td></tr>
        <tr style="height:600px"><td></td><td>Net Export Receivables</td><td>{{net_export_receviables}}</td></tr>
        <tr style="height:600px"><td></td><td></td><td></td></tr> 
        
        <tr style="height:600px"><td></td><td>Domestic receivables upto … days</td><td>{{domestic_receviableees_upto}}</td></tr>
        <tr style="height:600px"><td></td><td>Domestic receivables beyond … days</td><td>{{domestic_rec_beyond_days}}</td></tr>
        <tr style="height:600px"><td></td><td>Export receivables upto … days</td><td>{{export_receive_beyond_days}}</td></tr>
        <tr style="height:600px"><td></td><td>Export receivables beyond … days</td><td>{{export_rec_beyond_days}}</td></tr>
        <tr style="height:600px"><td></td><td></td><td></td></tr>    
        <tr style="height:600px"><td>B</td><td>chargeaeble_debtors</td><td>{{chargeaeble_debtors}}</td></tr>
        <tr style="height:600px"><td>B</td><td>Drawing power II</td><td>{{Drawing_power}}</td></tr>     
        <tr style="height:600px"><td>(A+B)</td><td>Total Drawing power (I +II)</td><td>{{Total_drawing}}</td></tr> 
        <tr style="height:600px"><td></td><td>Less: Other WC bankers outstanding </td><td>{{less_other_wc_bankers}}</td></tr>   
        <tr style="height:600px"><td></td><td>Less: Kotak PO Funding</td><td>{{kotak_po_funding}}</td></tr>
        <tr style="height:600px"><td></td><td>Net DP</td><td>{{net_dp}}</td></tr>           
        <tr style="height:600px"><td></td><td>DP Limit</td><td>{{dp_limit_consortium}}</td></tr>     
        <tr style="height:600px"><td></td><td>Final DP</td><td>{{final_dp}}</td></tr>
        <tr style="height:600px"><td></td><td>Adhoc</td><td>{{adhoc}}</td></tr>
        <tr style="height:600px"><td></td><td>Final DP (with Adhoc)</td><td>{{final_dp_adhoc}}</td></tr>
        <tr style="height:600px"><td></td><td>Sales</td><td>{{Sales}}</td></tr>
        <tr style="height:600px"><td></td><td>Purchases</td><td>{{Purchases}}</td></tr>
        
    </table>"""
            print(df_data)
            # Sample data for the template
            data = df_data
            template = Template(template_data1)
            rendered_template = template.render(data)
            # # Create a Jinja template object
            # template = Template(template_data1)


            # # Render the template with data
            # rendered_template = template.render(data)
            # print(rendered_template)
            # Parse the HTML table into a DataFrame
            dfs = pd.read_html(rendered_template)

            # Get the DataFrame from the list of DataFrames
            df = dfs[0]
            df.columns = [" ", " ", " ","","",]

            # print(df)
            output_path = Path(
                        f'/app/output/kmb/assets/pdf/kmb/{case_id}')
            os.umask(0)
            logging.info(f"---{output_path}---")
            df.to_excel(f"{output_path}/output.xlsx", index=False )
            
            
            logging.info(f"---saved---")
            workbook = openpyxl.load_workbook(f"{output_path}/output.xlsx",data_only=True)

            # Select the desired sheet
            sheet = workbook.active
            sheet.title = "Main_Format1"# You can also specify the sheet name like workbook['Sheet1']
            for row in range(10, 47):
                cell_address = 'C' + str(row)
                cell = sheet[cell_address]
                if cell.value is not None and isinstance(cell.value, str):
                    try:
                        print(type(cell.value),'------------------------------------------')
                        cell.value = float(cell.value) # Convert text to integer
                    except ValueError as e:
                        print(e,'------------------------------{e}')
                        # Handle non-convertible values here if needed
                        pass


            sheet.column_dimensions['B'].width = 60
            sheet.column_dimensions['c'].width = 60
            sheet.column_dimensions['d'].width = 40
            sheet.column_dimensions['E'].width = 60
            sheet['E14'] = '=SUM(E8:E13)'
            sheet['E29'] = '=E24-E25-E26-E27-E28'
            sheet['E34']='=E31-E32-E33'

            img_path = Path(f'/app/output/kmb/assets/pdf/kmb/kotak.PNG')
            img = openpyxl.drawing.image.Image(img_path)
            img.anchor = "B3"  

            for row in sheet.iter_rows(min_row=8, max_row=52, min_col=5, max_col=5):
                for cell in row:
                    cell.fill = PatternFill(start_color="53b553", end_color="53b553", fill_type="solid")

                    # Green color
            for row in sheet.iter_rows(min_row=4, max_row=4, min_col=3, max_col=3):
                for cell in row:
                    cell.font = Font(size=16, bold=True, color="090a0a")
            for row in sheet.iter_rows(min_row=16, max_row=16, min_col=2, max_col=2):
                for cell in row:
                    cell.font = Font(size=10, bold=True, color="090a0a")        

            for row in sheet.iter_rows(min_row=6, max_row=6, min_col=2, max_col=3):
                for cell in row:
                    cell.fill = PatternFill(start_color="626a80", end_color="626a80", fill_type="solid")   
                    cell.font = Font(size=14, bold=True, color="FFFFFF")
            for row in sheet.iter_rows(min_row=7, max_row=7, min_col=1, max_col=5):
                for cell in row:
                    cell.fill = PatternFill(start_color="626a80", end_color="626a80", fill_type="solid")   
                    cell.font = Font(size=14, bold=True, color="FFFFFF")
            for row in sheet.iter_rows(min_row=5, max_row=5, min_col=2, max_col=2):
                for cell in row:
                    cell.fill = PatternFill(start_color="626a80", end_color="626a80", fill_type="solid")
                    cell.font = Font(size=14, bold=True, color="FFFFFF")
            for row in sheet.iter_rows(min_row=21, max_row=21, min_col=1, max_col=5):
                for cell in row:
                    cell.fill = PatternFill(start_color="626a80", end_color="626a80", fill_type="solid")
                    cell.font = Font(size=14, bold=True, color="FFFFFF")
        
            for row in sheet.iter_rows(min_row=42, max_row=43, min_col=1, max_col=5):
                for cell in row:
                    cell.fill = PatternFill(start_color="626a80", end_color="626a80", fill_type="solid")
                    cell.font = Font(size=14, bold=True, color="FFFFFF")
            for row in sheet.iter_rows(min_row=46, max_row=50, min_col=1, max_col=5):
                for cell in row:
                    cell.fill = PatternFill(start_color="626a80", end_color="626a80", fill_type="solid")
                    cell.font = Font(size=14, bold=True, color="FFFFFF")

        
            # Save the changes to the workbook



            workbook.save(f"{output_path}/output.xlsx")
            logging.info(f"completed sheet 1-----")


            Heads= [
                "For the Month", "Due Date", "CRN", "Account Number", "Total Stock", "Stock Applicable for DP", 
                "Stock < 90", "Stock < 120", "Total Creditors", "Creditors Applicable for DP", 
                "Total Debtors", "Debtors Applicable for DP", "Debtors < 90", "Debtors < 120", 
                "Total Sales", "Total Purchases", "Other Bank Limit", "System DP", "System Limit", 
                "CC Limit", "Computed DP", "Final DP", "Stock <30 Days", "Stock Between 31 and 60 Days", 
                "Stock Between 61 and 90 Days", "Stock Between 91 and 120 Days", "Stock Between 121 and 180 Days", 
                "Stock > 180 Days", "Debtors <30 Days", "Debtors Between 31 and 60 Days", 
                "Debtors Between 61 and 90 Days", "Debtors Between 91 and 120 Days", 
                "Debtors Between 121 and 180 Days", "Debtors > 180 Days", "Sales Spares Workshop", 
                "Sales Amount Vehicle", "Purchase Spares Workshop", "Debtors > 90 Days Spares Workshop", 
                "Debtors > 90 Days Vehicle", "Debtors Spares Between 91 to 120", 
                "Debtors Vehicle Between 91 to 120", "Total Debtors Spares", "Total Spares Stock", 
                "Total Spares Stock <= 120", "Total Spares Stock <= 90", "Total Vehicle Stock", 
                "Total Vehicle Stock <= 120", "Total Vehicle Stock <= 90", 
                "Debtors Advance Paid to suppliers", "Purchase Vehicles", "Total Debtors Vehicles", 
                "Total Stock <=120", "Total Stock <=90", "Advance Received from customers", 
                "Bill discounted debtors", "Channel Finance", "Debtors 0 to 60 Days", 
                "Debtors 0 to 5 Days", "Debtors 6 to 30 Days", "Debtors 6 to 60 Days", 
                "Debtors < 150 days", "Debtors < 45 days", "Debtors Specific", 
                "Debtors of Vehicle < 30 days", "Debtors of Vehicle < 60 days", 
                "Debtors of spares < 60 days", "Group Co debtors", "Insurance Amount", 
                "Limits Os With KMBL Banks 1", "Outstanding with other WC Bankers", 
                "Previous month Total Creditors", "Previous month Total Debtors", 
                "Previous month Total Stock", "Purchase 1 Month prior", "Purchase 2 Month prior", 
                "Purchase 3 Month prior", "Purchase 4 Month prior", "Purchase 5 Month prior", 
                "Purchase Spares 1 Month Prior", "Purchase Spares 2 Month Prior", 
                "Purchase Spares 3 Month Prior", "Purchase Spares 4 Month Prior", 
                "Purchase Spares 5 Month Prior", "Purchase Vehicles 1 Month Prior", 
                "Purchase Vehicles 2 Month Prior", "Purchase Vehicles 3 Month Prior", 
                "Purchase Vehicles 4 Month Prior", "Purchase Vehicles 5 Month Prior", 
                "RM Stock < 60 days", "RM Stock < 90 days", "Sales 1 Month Prior", 
                "Sales 2 Month Prior", "Sales 3 Month Prior", "Sales 4 Month Prior", 
                "Sales 5 Month Prior", "Sales Spares 1 Month Prior", "Sales Spares 2 Month Prior", 
                "Sales Spares 3 Month Prior", "Sales Spares 4 Month Prior", 
                "Sales Spares 5 Month Prior", "Sales Vehicles 1 Month Prior", 
                "Sales Vehicles 2 Month Prior", "Sales Vehicles 3 Month Prior", 
                "Sales Vehicles 4 Month Prior", "Sales Vehicles 5 Month Prior", "Security Deposit", 
                "Stock Between 0 and 60 Days", "Stock Specific", "Stock in Transit < 90 Days", 
                "Stock of Vehicle < 60 days", "Stock of spares < 60 days", "Total Purchase Spares", 
                "Total Purchase Vehicles", "Total Sales Spares", "Total Sales Vehicles", 
                "Unbilled Debtors"
            ]
            # Number of columns
            num_columns = len(Heads)

            # Create 'Values' list with empty strings
            values_list = ["-"] * num_columns
            print(values_list)

            # Update the 'Values' key in the data2 dictionary


            # Print the updated data2 dictionary
            data2={'Heads':Heads,'Values':values_list}
            print(data2)



            # data2={'Heads':Heads,'Values':Values}
            df2 = pd.DataFrame(data2)
            # df2.to_excel("Downloads/output.xlsx", index=False )
            print(df2)
            new_sheet = workbook.create_sheet(title="Key Metrics")

            new_sheet.column_dimensions['A'].width = 50
            new_sheet.column_dimensions['B'].width = 50

            for row in dataframe_to_rows(df2, index=False, header=True):
                new_sheet.append(row)
            for row in new_sheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=2):
                for cell in row:
                    cell.fill = PatternFill(start_color="626a80", end_color="626a80", fill_type="solid")

                    

            new_sheet['B3']=df_data['Due_Date']
            new_sheet['B4']=df_data['Crn']
            new_sheet['B5']=df_data['ACCOUNT_NUMBER']
            new_sheet['B6']='=Main_Format1!C16'
            new_sheet['B7']='=Main_Format1!C26'
            new_sheet['B10']='=SUM(Main_Format1!C17,Main_Format1!C18,Main_Format1!C19)'
            new_sheet['B17']='=Main_Format1!C27'
            new_sheet['B12']='=Main_Format1!C28'
            new_sheet['B13']='=Main_Format1!C35'
            new_sheet['B8']='=Main_Format1!C28-Main_Format1!C29'
            new_sheet['B9']='=Main_Format1!C28-Main_Format1!C29'
            new_sheet['B16']='=Main_Format1!C44'
            new_sheet['B17']='=Main_Format1!C45'
            new_sheet['B18']='=Main_Format1!C37'
            new_sheet['B21']='=Main_Format1!C40'
            new_sheet['B22']='=Main_Format1!C39'
            new_sheet['B23']='=Main_Format1!C43'
            new_sheet['B50']='=Main_Format1!C23'
            new_sheet['B55']='=Main_Format1!C32'
            new_sheet['B56']='=Main_Format1!C30'
            new_sheet['B57']='=Main_Format1!C22'
            new_sheet['B68']='=Main_Format1!C31'
            new_sheet['B70']='=Main_Format1!C38'
            new_sheet['B71']='=Main_Format1!C37' 

            report_name="Kotak_WBG_444444_PRIVI_Output_Mapping_V1.0"

        #template6
        if 'WBG_444444_MULTIPLE_MARGIN' in template_name:
            template_data1 ="""<table style="width:100%">
    <tr><th></th><td></td><td></td></tr>
        <tr><th></th><td></td><td></td></tr>
        <tr style="height:600px";"font-size: 20px;"><td></td><td></td><td style="font-size: 20px;">DP STOCK STATEMENT</td><td>Division 1</td><td></td></tr>
        <tr style="height:600px"><td></td><td>Name of Borrower</td><td>{{NAME_OF_CUSTOMER}}</td><td></td></tr>
        <tr style="height:600px"><td></td><td></td><td></td><td>As per SANCTION</td></tr >
        <tr style="height:600px"><td></td><td>Details for the month </td><td>Statement value</td><td>Margin</td><td>Net value</td></tr>
        <tr style="height:600px"><td></td><td>Raw Material</td><td>{{Raw_Material}}</td><td>{{STOCK_MARGIN}}</td></tr>
        <tr style="height:600px"><td></td><td>Work in Process</td><td>{{Work_in_Process}}</td><td>{{STOCK_MARGIN}}</td></tr>
        <tr style="height:600px"><td></td><td>Finished Goods</td><td>{{Finished_Goods}}</td><td>{{STOCK_MARGIN}}</td></tr>
        <tr style="height:600px"><td></td><td>Stores & Spares</td><td>{{Stores_Spares}}</td><td>{{STOCK_MARGIN}}</td></tr>
        <tr style="height:600px"><td></td><td>Stock in Transit</td><td>{{Stock_in_Transit}}</td><td>{{STOCK_MARGIN}}</td></tr>
        <tr style="height:600px"><td></td><td>Other stock</td><td>{{Other_stock}}</td><td>{{STOCK_MARGIN}}</td></tr>   
        <tr style="height:600px"><td></td><td>Total Stock</td><td>{{Total_stock}}</td></tr>
        <tr style="height:600px"><td></td><td>Less: Total Creditors(BC/LC) </td><td>{{Creditors}}</td></tr>
        <tr style="height:600px"><td></td><td>PBD Utilised from other Bank</td><td>{{pbd_utilised}}</td></tr>
        <tr style="height:600px"><td></td><td>Advance paid to suppliers </td><td>{{Advance_paid_to_suppliers}}</td></tr>
        <tr style="height:600px"><td></td><td>Channel Finance </td><td>{{channel_finance}}</td></tr>  
        <tr style="height:600px"><td></td><td>EPC/PCFC stock</td><td>{{epc_pcfc_stock}}</td></tr>
        <tr style="height:600px"><td></td><td>Net stock</td><td>{{Net_stock}}</td></tr>
        <tr style="height:600px"><td>A</td><td>Drawing power I</td><td>{{drawing_power_1}}</td></tr>
        <tr style="height:600px"><td></td><td></td><td></td></tr>     
        <tr style="height:600px"><td></td><td>Total debtors</td><td>{{TOTAL_Debitors}}</td></tr>         
        <tr style="height:600px"><td></td><td>Domestic receivables</td><td>{{wb_non_groupco_drs_days}}</td></tr>             
        <tr style="height:600px"><td></td><td>Domestic receivables beyond cover period</td><td>{{domestic_rec_beyond_days}}</td></tr>
        <tr style="height:600px"><td></td><td>Group Co debtors</td><td>{{group_co_debtors}}</td></tr>
        <tr style="height:600px"><td></td><td>Bill Discounted</td><td>{{bill_discounted}}</td></tr>
        <tr style="height:600px"><td></td><td>Less: Deductibles(Advances)</td><td>{{deductibles}}</td></tr>
        <tr style="height:600px"><td>B</td><td>Net Domestic Receivables</td><td>{{net_domestic_receviables}}</td><td>{{STOCK_MARGIN}}</td></tr>
        <tr style="height:600px"><td></td><td></td><td></td></tr>  
        <tr style="height:600px"><td></td><td>Export receivables</td><td>{{export_receviables}}</td><td>{{STOCK_MARGIN}}</td></tr>
        <tr style="height:600px"><td></td><td>Export receivables beyond cover period</td><td>{{export_receviables_beyond}}</td></tr>
        <tr style="height:600px"><td></td><td>less: Deductibles</td><td>{{less_deductibles}}</td></tr>
        <tr style="height:600px"><td></td><td>Net Export Receivables</td><td>{{net_export_receviables}}</td></tr>
        <tr style="height:600px"><td></td><td></td><td></td></tr> 
        
        <tr style="height:600px"><td></td><td>Domestic receivables upto … days</td><td>{{domestic_receviableees_upto}}</td></tr>
        <tr style="height:600px"><td></td><td>Domestic receivables beyond … days</td><td>{{domestic_rec_beyond_days}}</td></tr>
        <tr style="height:600px"><td></td><td>Export receivables upto … days</td><td>{{export_receive_beyond_days}}</td></tr>
        <tr style="height:600px"><td></td><td>Export receivables beyond … days</td><td>{{export_rec_beyond_days}}</td></tr>
        <tr style="height:600px"><td></td><td></td><td></td></tr>    
        <tr style="height:600px"><td>B</td><td>chargeaeble_debtors</td><td>{{chargeaeble_debtors}}</td></tr>
        <tr style="height:600px"><td>B</td><td>Drawing power II</td><td>{{Drawing_power}}</td></tr>     
        <tr style="height:600px"><td>(A+B)</td><td>Total Drawing power (I +II)</td><td>{{Total_drawing}}</td></tr> 
        <tr style="height:600px"><td></td><td>Less: Other WC bankers outstanding </td><td>{{less_other_wc_bankers}}</td></tr>   
        <tr style="height:600px"><td></td><td>Less: Kotak PO Funding</td><td>{{kotak_po_funding}}</td></tr>
        <tr style="height:600px"><td></td><td>Net DP</td><td>{{net_dp}}</td></tr>           
        <tr style="height:600px"><td></td><td>DP Limit</td><td>{{dp_limit_consortium}}</td></tr>     
        <tr style="height:600px"><td></td><td>Final DP</td><td>{{final_dp}}</td></tr>
        <tr style="height:600px"><td></td><td>Adhoc</td><td>{{adhoc}}</td></tr>
        <tr style="height:600px"><td></td><td>Final DP (with Adhoc)</td><td>{{final_dp_adhoc}}</td></tr>
        <tr style="height:600px"><td></td><td>Sales</td><td>{{Sales}}</td></tr>
        <tr style="height:600px"><td></td><td>Purchases</td><td>{{Purchases}}</td></tr>
        
    </table>"""
            print(df_data)
            # Sample data for the template
            data = df_data
            template = Template(template_data1)
            rendered_template = template.render(data)
            # # Create a Jinja template object
            # template = Template(template_data1)


            # # Render the template with data
            # rendered_template = template.render(data)
            # print(rendered_template)
            # Parse the HTML table into a DataFrame
            dfs = pd.read_html(rendered_template)

            # Get the DataFrame from the list of DataFrames
            df = dfs[0]
            df.columns = [" ", " ", " ","","",]

            # print(df)
            output_path = Path(
                        f'/app/output/kmb/assets/pdf/kmb/{case_id}')
            os.umask(0)
            logging.info(f"---{output_path}---")
            df.to_excel(f"{output_path}/output.xlsx", index=False )
            
            
            logging.info(f"---saved---")
            workbook = openpyxl.load_workbook(f"{output_path}/output.xlsx",data_only=True)

            # Select the desired sheet
            sheet = workbook.active
            sheet.title = "Main_Format1"# You can also specify the sheet name like workbook['Sheet1']
            for row in range(10, 47):
                cell_address = 'C' + str(row)
                cell = sheet[cell_address]
                if cell.value is not None and isinstance(cell.value, str):
                    try:
                        print(type(cell.value),'------------------------------------------')
                        cell.value = float(cell.value) # Convert text to integer
                    except ValueError as e:
                        print(e,'------------------------------{e}')
                        # Handle non-convertible values here if needed
                        pass


            sheet.column_dimensions['B'].width = 60
            sheet.column_dimensions['c'].width = 40
            sheet.column_dimensions['d'].width = 40
            sheet.column_dimensions['e'].width = 40
            sheet['C16'] = '=SUM(C10:C15)'
            sheet['C24'] = '=C16-C17-C18-C19-C20+C21-C22-C23'
            sheet['C25']='=IF(C24 <0, 0, C24*D6)'
            sheet['C26'] ='=IF(C24<0, 0,C24-C25)'
            sheet['C32']='=IF(C24<0, C27 - C28 - C29- C30 -C31 + C24, C24 - SUM(C28, C29, C30, C31))'
            sheet['C33']='=C32*D7'
            sheet['C34']='=C32-C33'
            sheet['C35']='=IF(C24<0,C34,C26+C34)'
            sheet['C37']='=C35* C36'
            sheet['C40']='=IF(38= 0, MIN(C37, C39), MIN(C38 ,C39))'
            sheet['C42']='=IF(C40+ C41> C39, C39, C40 + C41)'
            sheet['E8']='=C8*D8'
            sheet['E9']='=C9*D9'
            sheet['E9']='=C10*D10'
            sheet['E9']='=C11*D11'
            sheet['E9']='=C12*D12'
            sheet['E9']='=C13*D13'
            sheet['E34']='=C34*D34'
            sheet['E14']='=SUM(E8:E13)'
            sheet['E20']='=E14-E15'
            sheet['E21']='=E20'
            sheet['E41']='=E24+E25+E31+E32+E29+E34'
            sheet['E42']='=E41'
            sheet['E46']='=E42-E44-E45'
            sheet['E48']='=MIN(E46,E47)'
            sheet['E50']='=IF(E48+E49>E47,E47,E48+E49)'
    
    
            img_path = Path(f'/app/output/kmb/assets/pdf/kmb/kotak.PNG')
            img = openpyxl.drawing.image.Image(img_path)
            img.anchor = "B3"  

            for row in sheet.iter_rows(min_row=8, max_row=52, min_col=5, max_col=5):
                for cell in row:
                    cell.fill = PatternFill(start_color="53b553", end_color="53b553", fill_type="solid")

                    # Green color
            for row in sheet.iter_rows(min_row=4, max_row=4, min_col=3, max_col=3):
                for cell in row:
                    cell.font = Font(size=16, bold=True, color="090a0a")
            for row in sheet.iter_rows(min_row=16, max_row=16, min_col=2, max_col=2):
                for cell in row:
                    cell.font = Font(size=10, bold=True, color="090a0a")        

            for row in sheet.iter_rows(min_row=6, max_row=6, min_col=2, max_col=3):
                for cell in row:
                    cell.fill = PatternFill(start_color="626a80", end_color="626a80", fill_type="solid")   
                    cell.font = Font(size=14, bold=True, color="FFFFFF")
            for row in sheet.iter_rows(min_row=7, max_row=7, min_col=1, max_col=5):
                for cell in row:
                    cell.fill = PatternFill(start_color="626a80", end_color="626a80", fill_type="solid")   
                    cell.font = Font(size=14, bold=True, color="FFFFFF")
            for row in sheet.iter_rows(min_row=5, max_row=5, min_col=2, max_col=2):
                for cell in row:
                    cell.fill = PatternFill(start_color="626a80", end_color="626a80", fill_type="solid")
                    cell.font = Font(size=14, bold=True, color="FFFFFF")
            for row in sheet.iter_rows(min_row=21, max_row=21, min_col=1, max_col=5):
                for cell in row:
                    cell.fill = PatternFill(start_color="626a80", end_color="626a80", fill_type="solid")
                    cell.font = Font(size=14, bold=True, color="FFFFFF")
        
            for row in sheet.iter_rows(min_row=42, max_row=43, min_col=1, max_col=5):
                for cell in row:
                    cell.fill = PatternFill(start_color="626a80", end_color="626a80", fill_type="solid")
                    cell.font = Font(size=14, bold=True, color="FFFFFF")
            for row in sheet.iter_rows(min_row=46, max_row=50, min_col=1, max_col=5):
                for cell in row:
                    cell.fill = PatternFill(start_color="626a80", end_color="626a80", fill_type="solid")
                    cell.font = Font(size=14, bold=True, color="FFFFFF")


            # Save the changes to the workbook
            workbook.save(f"{output_path}/output.xlsx")
            logging.info(f"completed sheet 1-----")


            Heads= [
                "For the Month", "Due Date", "CRN", "Account Number", "Total Stock", "Stock Applicable for DP", 
                "Stock < 90", "Stock < 120", "Total Creditors", "Creditors Applicable for DP", 
                "Total Debtors", "Debtors Applicable for DP", "Debtors < 90", "Debtors < 120", 
                "Total Sales", "Total Purchases", "Other Bank Limit", "System DP", "System Limit", 
                "CC Limit", "Computed DP", "Final DP", "Stock <30 Days", "Stock Between 31 and 60 Days", 
                "Stock Between 61 and 90 Days", "Stock Between 91 and 120 Days", "Stock Between 121 and 180 Days", 
                "Stock > 180 Days", "Debtors <30 Days", "Debtors Between 31 and 60 Days", 
                "Debtors Between 61 and 90 Days", "Debtors Between 91 and 120 Days", 
                "Debtors Between 121 and 180 Days", "Debtors > 180 Days", "Sales Spares Workshop", 
                "Sales Amount Vehicle", "Purchase Spares Workshop", "Debtors > 90 Days Spares Workshop", 
                "Debtors > 90 Days Vehicle", "Debtors Spares Between 91 to 120", 
                "Debtors Vehicle Between 91 to 120", "Total Debtors Spares", "Total Spares Stock", 
                "Total Spares Stock <= 120", "Total Spares Stock <= 90", "Total Vehicle Stock", 
                "Total Vehicle Stock <= 120", "Total Vehicle Stock <= 90", 
                "Debtors Advance Paid to suppliers", "Purchase Vehicles", "Total Debtors Vehicles", 
                "Total Stock <=120", "Total Stock <=90", "Advance Received from customers", 
                "Bill discounted debtors", "Channel Finance", "Debtors 0 to 60 Days", 
                "Debtors 0 to 5 Days", "Debtors 6 to 30 Days", "Debtors 6 to 60 Days", 
                "Debtors < 150 days", "Debtors < 45 days", "Debtors Specific", 
                "Debtors of Vehicle < 30 days", "Debtors of Vehicle < 60 days", 
                "Debtors of spares < 60 days", "Group Co debtors", "Insurance Amount", 
                "Limits Os With KMBL Banks 1", "Outstanding with other WC Bankers", 
                "Previous month Total Creditors", "Previous month Total Debtors", 
                "Previous month Total Stock", "Purchase 1 Month prior", "Purchase 2 Month prior", 
                "Purchase 3 Month prior", "Purchase 4 Month prior", "Purchase 5 Month prior", 
                "Purchase Spares 1 Month Prior", "Purchase Spares 2 Month Prior", 
                "Purchase Spares 3 Month Prior", "Purchase Spares 4 Month Prior", 
                "Purchase Spares 5 Month Prior", "Purchase Vehicles 1 Month Prior", 
                "Purchase Vehicles 2 Month Prior", "Purchase Vehicles 3 Month Prior", 
                "Purchase Vehicles 4 Month Prior", "Purchase Vehicles 5 Month Prior", 
                "RM Stock < 60 days", "RM Stock < 90 days", "Sales 1 Month Prior", 
                "Sales 2 Month Prior", "Sales 3 Month Prior", "Sales 4 Month Prior", 
                "Sales 5 Month Prior", "Sales Spares 1 Month Prior", "Sales Spares 2 Month Prior", 
                "Sales Spares 3 Month Prior", "Sales Spares 4 Month Prior", 
                "Sales Spares 5 Month Prior", "Sales Vehicles 1 Month Prior", 
                "Sales Vehicles 2 Month Prior", "Sales Vehicles 3 Month Prior", 
                "Sales Vehicles 4 Month Prior", "Sales Vehicles 5 Month Prior", "Security Deposit", 
                "Stock Between 0 and 60 Days", "Stock Specific", "Stock in Transit < 90 Days", 
                "Stock of Vehicle < 60 days", "Stock of spares < 60 days", "Total Purchase Spares", 
                "Total Purchase Vehicles", "Total Sales Spares", "Total Sales Vehicles", 
                "Unbilled Debtors"
            ]
            # Number of columns
            num_columns = len(Heads)

            # Create 'Values' list with empty strings
            values_list = ["-"] * num_columns
            print(values_list)

            # Update the 'Values' key in the data2 dictionary


            # Print the updated data2 dictionary
            data2={'Heads':Heads,'Values':values_list}
            print(data2)



            # data2={'Heads':Heads,'Values':Values}
            df2 = pd.DataFrame(data2)
            # df2.to_excel("Downloads/output.xlsx", index=False )
            print(df2)
            new_sheet = workbook.create_sheet(title="Key Metrics")

            new_sheet.column_dimensions['A'].width = 50
            new_sheet.column_dimensions['B'].width = 50

            for row in dataframe_to_rows(df2, index=False, header=True):
                new_sheet.append(row)
            for row in new_sheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=2):
                for cell in row:
                    cell.fill = PatternFill(start_color="626a80", end_color="626a80", fill_type="solid")

                    

            new_sheet['B3']=df_data['Due_Date']
            new_sheet['B4']=df_data['Crn']
            new_sheet['B5']=df_data['ACCOUNT_NUMBER']
            new_sheet['B6']='=Main_Format1!C14'
            new_sheet['B7']='=Main_Format1!C21'
            new_sheet['B10']='=Main_Format1!C15'
            new_sheet['B17']='=Main_Format1!C27'
            new_sheet['B12']='=Main_Format1!C23'
            new_sheet['B13']='=Main_Format1!C35'
            new_sheet['B14']='=Main_Format1!C29-Main_Format1!C37+Main_Format1!A31-Main_Format1!C39'
            new_sheet['B15']='=Main_Format1!C24-Main_Format1!C37+Main_Format1!C31-Main_Format1!C39'
            new_sheet['B16']='=Main_Format1!C51'
            new_sheet['B17']='=Main_Format1!C52'
            new_sheet['B18']='=Main_Format1!C44'
            new_sheet['B21']='=Main_Format1!C47'
            new_sheet['B22']='=Main_Format1!C46'
            new_sheet['B23']='=Main_Format1!C48'
            new_sheet['B50']='=Main_Format1!C17'
            new_sheet['B55']='=Main_Format1!C28'
            new_sheet['B56']='=Main_Format1!C27'
            new_sheet['B57']='=Main_Format1!C18'
            new_sheet['B68']='=Main_Format1!C26'
            new_sheet['B70']='=Main_Format1!C45'
            new_sheet['B71']='=Main_Format1!C44' 
            

            report_name="Kotak_WBG_444444_Multiple_Margin_Output_Mapping_V1.0"


        #template7
        if 'WBG_333333_MASTER' in template_name:
            template_data1 = """



            

        <table style="width:100%">
            <tr><th></th><td></td><td></td></tr>
                <tr><th></th><td></td><td></td></tr>
                <tr style="height:600px";"font-size: 20px;"><td></td><td></td><td style="font-size: 20px;">DP STOCK STATEMENT</td><td></td><td></td></tr>
                <tr style="height:600px"><td></td><td>Name of Borrower</td><td>{{NAME_OF_CUSTOMER}}</td><td></td></tr>
                <tr style="height:600px"><td></td><td>Margin</td><td>Stock</td><td>{{STOCK_MARGIN}}</td></tr >
                <tr style="height:600px"><td></td><td></td><td>Debtors</td><td>{{DEBTORS_MARGIN}}</td></tr >
                <tr style="height:600px"><td></td><td></td><td></td><td></td> </tr >
                <tr style="height:600px"><td></td><td></td><td>Details for the month </td><td>EPC-Old & New Limits</td></tr >
                <tr style="height:600px";"color: red;"><td></td><td>Raw Material</td><td>{{Raw_Material}}</td></tr>
                <tr style="height:600px"><td></td><td style="color: blue;">Work in Process</td><td>{{Work_in_Process}}</td></tr>
                <tr style="height:600px"><td></td><td>Finished Goods</td><td>{{Finished_Goods}}</td></tr>
                <tr style="height:600px"><td></td><td>Stores & Spares</td><td>{{Stores_Spares}}</td></tr>
                <tr style="height:600px"><td></td><td>Other stock</td><td>{{Other_stock}}</td></tr>
                <tr style="height:600px"><td></td><td>Stock in Transit</td><td>{{Stock_in_Transit}}</td></tr>
                <tr style="height:600px"><td></td><td>Total Stock</td><td>{{Total_stock}}</td></tr>
                <tr style="height:600px"><td></td><td>Less: Creditors</td><td>{{Less_Creditors}}</td></tr>
                <tr style="height:600px"><td></td><td>Less: inc in creditors by 14.22%</td><td>{{less_inc}}</td></tr>
                <tr style="height:600px"><td></td><td>LC Creditors(not included in creditors above)</td><td>{{LC_Creditors}}</td></tr>
                <tr style="height:600px"><td></td><td>Advance paid to suppliers</td><td>{{Advance_paid_to_suppliers}}</td></tr>
                <tr style="height:600px"><td></td><td>PBD Utilised from other Bank</td><td>{{pbd_utilised}}</td></tr>
                <tr style="height:600px"><td></td><td>EPC/PCFC stock</td><td>{{epc_pcfc_stock}}</td></tr>
                <tr style="height:600px"><td>A</td><td>Net stock</td><td>{{Net_stock}}</td></tr>
                <tr style="height:600px"><td></td><td>Less Margin 25%</td><td>{{less_margin_25}}</td></tr>
                <tr style="height:600px"><td></td><td>Considered for DP</td><td>{{considered_dp}}</td></tr>
                <tr style="height:600px"><td></td><td></td><td></td></tr>
                <tr style="height:600px"><td></td><td>Total debtors</td><td>{{TOTAL_Debitors}}</td></tr>   
                <tr style="height:600px"><td></td><td>Less: Non Group co. drs.> 90/120/180 days as per term sheet  days</td><td>{{wb_non_groupco_drs_days}}</td></tr>             
                <tr style="height:600px"><td></td><td>Less : Bill discounted debtors</td><td>{{bill_discounted_debtors}}</td></tr>
                <tr style="height:600px"><td></td><td>Less : Group concern debtors</td><td>{{group_co_debtors}}</td></tr>
                <tr style="height:600px"><td></td><td>Less: Advance from Customer</td><td>{{Advance_from_Customer}}</td></tr>       
                <tr style="height:600px"><td>B</td><td>Net Debtors</td><td>{{Net_debtors}}</td></tr>       
                <tr style="height:600px"><td></td><td>Less Margin 25%</td><td>{{less_margin_25}}</td></tr>      
                <tr style="height:600px"><td></td><td>Debtors applicable for DP</td><td>{{debtors_applicable_dp}}</td></tr>
                <tr style="height:600px"><td></td><td>Considered for DP</td><td>{{considered_dp}}</td></tr>
                <tr style="height:600px"><td>(A+B)</td><td>Total DP</td><td>{{Total_Dp}}</td></tr>  
                <tr style="height:600px"><td></td><td>Less: Other WC bankers outstanding </td><td>{{less_other_wc_bankers}}</td></tr>   
                <tr style="height:600px"><td></td><td>Less: Kotak PO Funding</td><td>{{kotak_po_funding}}</td></tr>
                <tr style="height:600px"><td></td><td>Net DP</td><td>{{net_dp}}</td></tr>           
                <tr style="height:600px"><td></td><td>DP Limit</td><td>{{dp_limit_consortium}}</td></tr>     
                <tr style="height:600px"><td></td><td>Final DP</td><td>{{final_dp}}</td></tr>
                <tr style="height:600px"><td></td><td>Adhoc</td><td>{{adhoc}}</td></tr>
                <tr style="height:600px"><td></td><td>Final DP (with Adhoc)</td><td>{{final_dp_adhoc}}</td></tr>
                <tr style="height:600px"><td></td><td>Sales</td><td>{{Sales}}</td></tr>
                <tr style="height:600px"><td></td><td>Purchases</td><td>{{Purchases}}</td></tr>
            </table>"""
           
            print(df_data)
            # Sample data for the template
            data = df_data
            template = Template(template_data1)
            rendered_template = template.render(data)
            # # Create a Jinja template object
            # template = Template(template_data1)


            # # Render the template with data
            # rendered_template = template.render(data)
            # print(rendered_template)
            # Parse the HTML table into a DataFrame
            dfs = pd.read_html(rendered_template)

            # Get the DataFrame from the list of DataFrames
            df = dfs[0]
            df.columns = [" ", " ", " ","","",]

            # print(df)
            output_path = Path(
                        f'/app/output/kmb/assets/pdf/kmb/{case_id}')
            os.umask(0)
            logging.info(f"---{output_path}---")
            df.to_excel(f"{output_path}/output.xlsx", index=False )
            
            
            logging.info(f"---saved---")
            workbook = openpyxl.load_workbook(f"{output_path}/output.xlsx",data_only=True)

            # Select the desired sheet
            sheet = workbook.active
            sheet.title = "Main_Format1"# You can also specify the sheet name like workbook['Sheet1']
            for row in range(10, 47):
                cell_address = 'C' + str(row)
                cell = sheet[cell_address]
                if cell.value is not None and isinstance(cell.value, str):
                    try:
                        print(type(cell.value),'------------------------------------------')
                        cell.value = float(cell.value) # Convert text to integer
                    except ValueError as e:
                        print(e,'------------------------------{e}')
                        # Handle non-convertible values here if needed
                        pass



            sheet.column_dimensions['B'].width = 60
            sheet.column_dimensions['c'].width = 60
            sheet.column_dimensions['d'].width = 40
            sheet['C16'] = '=SUM(C10:C15)'
            sheet['C24'] = '=C16-C17-C18-C19-C20+C21-C22-C23'
            sheet['C25']='=IF(C24 <0, 0, C24*D6)'
            sheet['C26'] ='=IF(C24<0, 0,C24-C25)'
            sheet['C32']='=IF(C24<0, C27 - C28 - C29- C30 -C31 + C24, C24 - SUM(C28, C29, C30, C31))'
            sheet['C33']='=C32*D7'
            sheet['C34']='=C32-C33'
            sheet['C35']='=IF(C24<0,C34,C26+C34)'
            sheet['C37']='=C35* C36'
            sheet['C40']='=IF(38= 0, MIN(C37, C39), MIN(C38 ,C39))'
            sheet['C42']='=IF(C40+ C41> C39, C39, C40 + C41)'


            img_path = Path(f'/app/output/kmb/assets/pdf/kmb/kotak.PNG')
            img = openpyxl.drawing.image.Image(img_path)
            img.anchor = "B3"  

            for row in sheet.iter_rows(min_row=9, max_row=45, min_col=3, max_col=3):
                for cell in row:
                    cell.fill = PatternFill(start_color="53b553", end_color="53b553", fill_type="solid")
                    
                    # Green color
            for row in sheet.iter_rows(min_row=4, max_row=4, min_col=3, max_col=3):
                for cell in row:
                    cell.font = Font(size=16, bold=True, color="090a0a")
            for row in sheet.iter_rows(min_row=16, max_row=16, min_col=2, max_col=2):
                for cell in row:
                    cell.font = Font(size=10, bold=True, color="090a0a")        
                    
            for row in sheet.iter_rows(min_row=9, max_row=9, min_col=2, max_col=4):
                
                for cell in row:
                    cell.fill = PatternFill(start_color="626a80", end_color="626a80", fill_type="solid")
                    cell.font = Font(size=14, bold=True, color="FFFFFF")
            for row in sheet.iter_rows(min_row=6, max_row=7, min_col=2, max_col=3):
                for cell in row:
                    cell.fill = PatternFill(start_color="626a80", end_color="626a80", fill_type="solid")   
                    cell.font = Font(size=14, bold=True, color="FFFFFF")
            for row in sheet.iter_rows(min_row=5, max_row=5, min_col=2, max_col=2):
                for cell in row:
                    cell.fill = PatternFill(start_color="626a80", end_color="626a80", fill_type="solid")
                    cell.font = Font(size=14, bold=True, color="FFFFFF")

            # Save the changes to the workbook



            workbook.save(f"{output_path}/output.xlsx")
            logging.info(f"completed sheet 1-----")


            Heads= [
                "For the Month", "Due Date", "CRN", "Account Number", "Total Stock", "Stock Applicable for DP", 
                "Stock < 90", "Stock < 120", "Total Creditors", "Creditors Applicable for DP", 
                "Total Debtors", "Debtors Applicable for DP", "Debtors < 90", "Debtors < 120", 
                "Total Sales", "Total Purchases", "Other Bank Limit", "System DP", "System Limit", 
                "CC Limit", "Computed DP", "Final DP", "Stock <30 Days", "Stock Between 31 and 60 Days", 
                "Stock Between 61 and 90 Days", "Stock Between 91 and 120 Days", "Stock Between 121 and 180 Days", 
                "Stock > 180 Days", "Debtors <30 Days", "Debtors Between 31 and 60 Days", 
                "Debtors Between 61 and 90 Days", "Debtors Between 91 and 120 Days", 
                "Debtors Between 121 and 180 Days", "Debtors > 180 Days", "Sales Spares Workshop", 
                "Sales Amount Vehicle", "Purchase Spares Workshop", "Debtors > 90 Days Spares Workshop", 
                "Debtors > 90 Days Vehicle", "Debtors Spares Between 91 to 120", 
                "Debtors Vehicle Between 91 to 120", "Total Debtors Spares", "Total Spares Stock", 
                "Total Spares Stock <= 120", "Total Spares Stock <= 90", "Total Vehicle Stock", 
                "Total Vehicle Stock <= 120", "Total Vehicle Stock <= 90", 
                "Debtors Advance Paid to suppliers", "Purchase Vehicles", "Total Debtors Vehicles", 
                "Total Stock <=120", "Total Stock <=90", "Advance Received from customers", 
                "Bill discounted debtors", "Channel Finance", "Debtors 0 to 60 Days", 
                "Debtors 0 to 5 Days", "Debtors 6 to 30 Days", "Debtors 6 to 60 Days", 
                "Debtors < 150 days", "Debtors < 45 days", "Debtors Specific", 
                "Debtors of Vehicle < 30 days", "Debtors of Vehicle < 60 days", 
                "Debtors of spares < 60 days", "Group Co debtors", "Insurance Amount", 
                "Limits Os With KMBL Banks 1", "Outstanding with other WC Bankers", 
                "Previous month Total Creditors", "Previous month Total Debtors", 
                "Previous month Total Stock", "Purchase 1 Month prior", "Purchase 2 Month prior", 
                "Purchase 3 Month prior", "Purchase 4 Month prior", "Purchase 5 Month prior", 
                "Purchase Spares 1 Month Prior", "Purchase Spares 2 Month Prior", 
                "Purchase Spares 3 Month Prior", "Purchase Spares 4 Month Prior", 
                "Purchase Spares 5 Month Prior", "Purchase Vehicles 1 Month Prior", 
                "Purchase Vehicles 2 Month Prior", "Purchase Vehicles 3 Month Prior", 
                "Purchase Vehicles 4 Month Prior", "Purchase Vehicles 5 Month Prior", 
                "RM Stock < 60 days", "RM Stock < 90 days", "Sales 1 Month Prior", 
                "Sales 2 Month Prior", "Sales 3 Month Prior", "Sales 4 Month Prior", 
                "Sales 5 Month Prior", "Sales Spares 1 Month Prior", "Sales Spares 2 Month Prior", 
                "Sales Spares 3 Month Prior", "Sales Spares 4 Month Prior", 
                "Sales Spares 5 Month Prior", "Sales Vehicles 1 Month Prior", 
                "Sales Vehicles 2 Month Prior", "Sales Vehicles 3 Month Prior", 
                "Sales Vehicles 4 Month Prior", "Sales Vehicles 5 Month Prior", "Security Deposit", 
                "Stock Between 0 and 60 Days", "Stock Specific", "Stock in Transit < 90 Days", 
                "Stock of Vehicle < 60 days", "Stock of spares < 60 days", "Total Purchase Spares", 
                "Total Purchase Vehicles", "Total Sales Spares", "Total Sales Vehicles", 
                "Unbilled Debtors"
            ]
            # Number of columns
            num_columns = len(Heads)

            # Create 'Values' list with empty strings
            values_list = ["-"] * num_columns
            print(values_list)

            # Update the 'Values' key in the data2 dictionary


            # Print the updated data2 dictionary
            data2={'Heads':Heads,'Values':values_list}
            print(data2)



            # data2={'Heads':Heads,'Values':Values}
            df2 = pd.DataFrame(data2)
            # df2.to_excel("Downloads/output.xlsx", index=False )
            print(df2)
            new_sheet = workbook.create_sheet(title="Key Metrics")

            new_sheet.column_dimensions['A'].width = 50
            new_sheet.column_dimensions['B'].width = 50

            for row in dataframe_to_rows(df2, index=False, header=True):
                new_sheet.append(row)
            for row in new_sheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=2):
                for cell in row:
                    cell.fill = PatternFill(start_color="626a80", end_color="626a80", fill_type="solid")

                    

            new_sheet['B3']=df_data['Due_Date']
            new_sheet['B4']=df_data['Crn']
            new_sheet['B5']=df_data['ACCOUNT_NUMBER']
            new_sheet['B6']='=Main_Format1!C16'
            new_sheet['B7']='=Main_Format1!C26'
            new_sheet['B10']='=SUM(Main_Format1!C17,Main_Format1!C18,Main_Format1!C19)'
            new_sheet['B17']='=Main_Format1!C27'
            new_sheet['B12']='=Main_Format1!C28'
            new_sheet['B13']='=Main_Format1!C35'
            new_sheet['B8']='=Main_Format1!C28-Main_Format1!C29'
            new_sheet['B9']='=Main_Format1!C28-Main_Format1!C29'
            new_sheet['B16']='=Main_Format1!C44'
            new_sheet['B17']='=Main_Format1!C45'
            new_sheet['B18']='=Main_Format1!C37'
            new_sheet['B21']='=Main_Format1!C40'
            new_sheet['B22']='=Main_Format1!C39'
            new_sheet['B23']='=Main_Format1!C43'
            new_sheet['B50']='=Main_Format1!C23'
            new_sheet['B55']='=Main_Format1!C32'
            new_sheet['B56']='=Main_Format1!C30'
            new_sheet['B57']='=Main_Format1!C22'
            new_sheet['B68']='=Main_Format1!C31'
            new_sheet['B70']='=Main_Format1!C38'
            new_sheet['B71']='=Main_Format1!C37' 

            report_name="Kotak_WBG_333333_Output_Mapping_V1.0"

        #template8
        if 'CMB-SME_100000_MASTER' in  template_name:
            template_data1 = """

            <table style="width:100%">
            <tr><th></th><td></td><td></td></tr>
                <tr><th></th><td></td><td></td></tr>
                <tr style="height:600px";"font-size: 20px;"><td></td><td></td><td style="font-size: 20px;">DP STOCK STATEMENT</td><td></td><td></td></tr>
                <tr style="height:600px"><td></td><td>Name of Borrower</td><td>{{NAME_OF_CUSTOMER}}</td><td>{{Date}}</td></tr>
                <tr style="height:600px"><td></td><td>Margin</td><td>Stock</td><td>{{STOCK_MARGIN}}</td></tr >
                <tr style="height:600px"><td></td><td></td><td>Debtors</td><td>{{DEBTORS_MARGIN}}</td></tr >
                <tr style="height:600px"><td></td><td></td><td>Dp Formula</td><td>{{dp_formula}}</td> </tr >
                <tr style="height:600px"><td></td><td></td><td>Details for the month </td><td></td></tr >
                <tr style="height:600px";"color: red;"><td></td><td>Raw Material</td><td>{{Raw_Material}}</td></tr>
                <tr style="height:600px"><td></td><td style="color: blue;">Work in Process</td><td>{{Work_in_Process}}</td></tr>
                <tr style="height:600px"><td></td><td>Finished Goods</td><td>{{Finished_Goods}}</td></tr>
                <tr style="height:600px"><td></td><td>Stores & Spares</td><td>{{Stores_Spares}}</td></tr>
                <tr style="height:600px"><td></td><td>Other stock</td><td>{{Other_stock}}</td></tr>
                <tr style="height:600px"><td></td><td>Stock in Transit</td><td>{{Stock_in_Transit}}</td></tr>
                <tr style="height:600px"><td></td><td>Total Stock</td><td>{{Total_stock}}</td></tr>        
                <tr style="height:600px"><td></td><td>Less: Stock > 90/120/150/180 days as per term sheet  days</td><td>{{stock_gt_90}}</td></tr>
                <tr style="height:600px"><td></td><td>Eligible Stock</td><td>{{eligible_stock}}</td></tr>
                <tr style="height:600px"><td></td><td>Less: Creditors</td><td>{{Less_Creditors}}</td></tr>
                <tr style="height:600px"><td></td><td>Add: Advance paid to suppliers</td><td>{{Advance_paid_to_suppliers}}</td></tr>
                <tr style="height:600px"><td></td><td>Buyers Credit for working capital</td><td>{{buyers_credit_wc}}</td></tr>
                <tr style="height:600px"><td></td><td>LC Creditors(not included in creditors above)</td><td>{{LC_Creditors}}</td></tr>
                <tr style="height:600px"><td></td><td>PBD Utilised from other Bank</td><td>{{pbd_utilised}}</td></tr>
                <tr style="height:600px"><td></td><td>EPC/PCFC stock</td><td>{{epc_pcfc_stock}}</td></tr>
                <tr style="height:600px"><td></td><td>Less: Slow/ Obsolete stock</td><td>{{slow_obsolete_stock}}</td></tr>
                <tr style="height:600px"><td></td><td>Stock applicable for DP</td><td>{{stock_applicable_dp}}</td></tr>
                <tr style="height:600px"><td></td><td>Less Margin</td><td>{{less_margin}}</td></tr>
                <tr style="height:600px"><td>A</td><td>Net stock for DP</td><td>{{Net_stock}}</td></tr>
                <tr style="height:600px"><td></td><td></td><td></td></tr>
                <tr style="height:600px"><td></td><td>Total debtors</td><td>{{TOTAL_Debitors}}</td></tr>      
                <tr style="height:600px"><td></td><td>Less: Non Group co. drs.> 90/120/150/180 days as per term sheet  days</td><td>{{wb_non_groupco_drs_days}}</td></tr>
                <tr style="height:600px"><td></td><td>Eligible Debtors</td><td>{{eligible_debtors}}</td></tr>             
                <tr style="height:600px"><td></td><td>Less : Bill discounted debtors</td><td>{{bill_discounted_debtors}}</td></tr>
                <tr style="height:600px"><td></td><td>Less : Group concern debtors</td><td>{{group_co_debtors}}</td></tr>
                <tr style="height:600px"><td></td><td>Less: Advance from Customer</td><td>{{Advance_from_Customer}}</td></tr>       
                <tr style="height:600px"><td>B</td><td>Net Debtors</td><td>{{Net_debtors}}</td></tr>
                <tr style="height:600px"><td></td><td>Less Margin</td><td>{{less_margin}}</td></tr>      
                <tr style="height:600px"><td></td><td>Debtors applicable for DP</td><td>{{debtors_applicable_dp}}</td></tr>
                <tr style="height:600px"><td>(A+B)</td><td>Total DP</td><td>{{Total_Dp}}</td></tr>   
                <tr style="height:600px"><td></td><td>Less: Other WC bankers outstanding </td><td>{{less_other_wc_bankers}}</td></tr>  
                <tr style="height:600px"><td></td><td>Less: Limit o/s with KKMBL</td><td>{{limit_os_with_kmbl}}</td></tr>
                <tr style="height:600px"><td></td><td>Net DP</td><td>{{net_dp}}</td></tr>           
                <tr style="height:600px"><td></td><td>DP Limit</td><td>{{dp_limit_consortium}}</td></tr>
                <tr style="height:600px"><td></td><td>Final DP</td><td>{{final_dp}}</td></tr>
                <tr style="height:600px"><td></td><td>Adhoc</td><td>{{adhoc}}</td></tr>
                <tr style="height:600px"><td></td><td>Final DP (with Adhoc)</td><td>{{final_dp_adhoc}}</td></tr>
                <tr style="height:600px"><td></td><td>Sales</td><td>{{Sales}}</td></tr>
                <tr style="height:600px"><td></td><td>Purchases</td><td>{{Purchases}}</td></tr>
                <tr style="height:600px"><td></td><td>Last 3 Months Sales</td><td>{{Last_three_month_Sales}}</td></tr>
                <tr style="height:600px"><td></td><td>Last 3  Months Purchases</td><td>{{Last_three_month_Purchases}}</td></tr>
                
                
            </table>

"""
            print(df_data)
            # Sample data for the template
            data = df_data
            template = Template(template_data1)
            rendered_template = template.render(data)
            # # Create a Jinja template object
            # template = Template(template_data1)


            # # Render the template with data
            # rendered_template = template.render(data)
            # print(rendered_template)
            # Parse the HTML table into a DataFrame
            dfs = pd.read_html(rendered_template)

            # Get the DataFrame from the list of DataFrames
            df = dfs[0]
            df.columns = [" ", " ", " ","","",]

            # print(df)



            output_path = Path(
                        f'/app/output/kmb/assets/pdf/kmb/{case_id}')
            os.umask(0)
            logging.info(f"---{output_path}---")
            df.to_excel(f"{output_path}/output.xlsx", index=False )
            
            
            logging.info(f"---saved---")
            workbook = openpyxl.load_workbook(f"{output_path}/output.xlsx",data_only=True)

            # Select the desired sheet
            sheet = workbook.active
            sheet.title = "Main_Format1"# You can also specify the sheet name like workbook['Sheet1']
            for row in range(10, 47):
                cell_address = 'C' + str(row)
                cell = sheet[cell_address]
                if cell.value is not None and isinstance(cell.value, str):
                    try:
                        print(type(cell.value),'------------------------------------------')
                        cell.value = float(cell.value) # Convert text to integer
                    except ValueError as e:
                        print(e,'------------------------------{e}')
                        # Handle non-convertible values here if needed
                        pass



            sheet.column_dimensions['B'].width = 60
            sheet.column_dimensions['c'].width = 60
            sheet.column_dimensions['d'].width = 40
            sheet['C16'] = '=SUM(C10:C15)'
            sheet['C18'] = '=C16-C17'
            sheet['C26'] = '=C18-C19-C21-C22-C23-C24+C20-C25'
            sheet['C27'] = '=IF(C26<0,0,C26*D6)'
            sheet['C28'] = '=IF(C26<0,0,C26-C27)'
            sheet['C32'] = '=C30-C31'
            sheet['C36'] = '=IF(C26<0,C30-C31-C33-C34-C35+C26,C30-C31-C33-C34-C35)'
            sheet['C37'] = '=IF(C36<0,0,C36*D7)'
            sheet['C38'] = '=C36-C37'
            sheet['C39'] = '=IF(C26<0,C38,C28+C38)'
            sheet['C42'] = '=C39-C40-C41'
            sheet['C44'] = '=MIN(C43,C42)'
            sheet['C46'] = '=IF(C44+C45>C43,C43,C44+C45)'


            img_path = Path(f'/app/output/kmb/assets/pdf/kmb/kotak.PNG')
            img = openpyxl.drawing.image.Image(img_path)
            img.anchor = "B3"  

            for row in sheet.iter_rows(min_row=9, max_row=49, min_col=3, max_col=3):
                for cell in row:
                    cell.fill = PatternFill(start_color="53b553", end_color="53b553", fill_type="solid")
                    
                    # Green color
            for row in sheet.iter_rows(min_row=4, max_row=4, min_col=3, max_col=3):
                for cell in row:
                    cell.font = Font(size=16, bold=True, color="090a0a")
            for row in sheet.iter_rows(min_row=16, max_row=16, min_col=2, max_col=2):
                for cell in row:
                    cell.font = Font(size=10, bold=True, color="090a0a")        
                    
            for row in sheet.iter_rows(min_row=9, max_row=9, min_col=2, max_col=4):
                
                for cell in row:
                    cell.fill = PatternFill(start_color="626a80", end_color="626a80", fill_type="solid")
                    cell.font = Font(size=14, bold=True, color="FFFFFF")
            for row in sheet.iter_rows(min_row=6, max_row=7, min_col=2, max_col=3):
                for cell in row:
                    cell.fill = PatternFill(start_color="626a80", end_color="626a80", fill_type="solid")   
                    cell.font = Font(size=14, bold=True, color="FFFFFF")
            for row in sheet.iter_rows(min_row=5, max_row=5, min_col=2, max_col=2):
                for cell in row:
                    cell.fill = PatternFill(start_color="626a80", end_color="626a80", fill_type="solid")
                    cell.font = Font(size=14, bold=True, color="FFFFFF")

            # Save the changes to the workbook



            workbook.save(f"{output_path}/output.xlsx")
            logging.info(f"completed sheet 1-----")


            Heads= [
                "For the Month", "Due Date", "CRN", "Account Number", "Total Stock", "Stock Applicable for DP", 
                "Stock < 90", "Stock < 120", "Total Creditors", "Creditors Applicable for DP", 
                "Total Debtors", "Debtors Applicable for DP", "Debtors < 90", "Debtors < 120", 
                "Total Sales", "Total Purchases", "Other Bank Limit", "System DP", "System Limit", 
                "CC Limit", "Computed DP", "Final DP", "Stock <30 Days", "Stock Between 31 and 60 Days", 
                "Stock Between 61 and 90 Days", "Stock Between 91 and 120 Days", "Stock Between 121 and 180 Days", 
                "Stock > 180 Days", "Debtors <30 Days", "Debtors Between 31 and 60 Days", 
                "Debtors Between 61 and 90 Days", "Debtors Between 91 and 120 Days", 
                "Debtors Between 121 and 180 Days", "Debtors > 180 Days", "Sales Spares Workshop", 
                "Sales Amount Vehicle", "Purchase Spares Workshop", "Debtors > 90 Days Spares Workshop", 
                "Debtors > 90 Days Vehicle", "Debtors Spares Between 91 to 120", 
                "Debtors Vehicle Between 91 to 120", "Total Debtors Spares", "Total Spares Stock", 
                "Total Spares Stock <= 120", "Total Spares Stock <= 90", "Total Vehicle Stock", 
                "Total Vehicle Stock <= 120", "Total Vehicle Stock <= 90", 
                "Debtors Advance Paid to suppliers", "Purchase Vehicles", "Total Debtors Vehicles", 
                "Total Stock <=120", "Total Stock <=90", "Advance Received from customers", 
                "Bill discounted debtors", "Channel Finance", "Debtors 0 to 60 Days", 
                "Debtors 0 to 5 Days", "Debtors 6 to 30 Days", "Debtors 6 to 60 Days", 
                "Debtors < 150 days", "Debtors < 45 days", "Debtors Specific", 
                "Debtors of Vehicle < 30 days", "Debtors of Vehicle < 60 days", 
                "Debtors of spares < 60 days", "Group Co debtors", "Insurance Amount", 
                "Limits Os With KMBL Banks 1", "Outstanding with other WC Bankers", 
                "Previous month Total Creditors", "Previous month Total Debtors", 
                "Previous month Total Stock", "Purchase 1 Month prior", "Purchase 2 Month prior", 
                "Purchase 3 Month prior", "Purchase 4 Month prior", "Purchase 5 Month prior", 
                "Purchase Spares 1 Month Prior", "Purchase Spares 2 Month Prior", 
                "Purchase Spares 3 Month Prior", "Purchase Spares 4 Month Prior", 
                "Purchase Spares 5 Month Prior", "Purchase Vehicles 1 Month Prior", 
                "Purchase Vehicles 2 Month Prior", "Purchase Vehicles 3 Month Prior", 
                "Purchase Vehicles 4 Month Prior", "Purchase Vehicles 5 Month Prior", 
                "RM Stock < 60 days", "RM Stock < 90 days", "Sales 1 Month Prior", 
                "Sales 2 Month Prior", "Sales 3 Month Prior", "Sales 4 Month Prior", 
                "Sales 5 Month Prior", "Sales Spares 1 Month Prior", "Sales Spares 2 Month Prior", 
                "Sales Spares 3 Month Prior", "Sales Spares 4 Month Prior", 
                "Sales Spares 5 Month Prior", "Sales Vehicles 1 Month Prior", 
                "Sales Vehicles 2 Month Prior", "Sales Vehicles 3 Month Prior", 
                "Sales Vehicles 4 Month Prior", "Sales Vehicles 5 Month Prior", "Security Deposit", 
                "Stock Between 0 and 60 Days", "Stock Specific", "Stock in Transit < 90 Days", 
                "Stock of Vehicle < 60 days", "Stock of spares < 60 days", "Total Purchase Spares", 
                "Total Purchase Vehicles", "Total Sales Spares", "Total Sales Vehicles", 
                "Unbilled Debtors"
            ]
            # Number of columns
            num_columns = len(Heads)

            # Create 'Values' list with empty strings
            values_list = ["-"] * num_columns
            print(values_list)

            # Update the 'Values' key in the data2 dictionary


            # Print the updated data2 dictionary
            data2={'Heads':Heads,'Values':values_list}
            print(data2)



            # data2={'Heads':Heads,'Values':Values}
            df2 = pd.DataFrame(data2)
            # df2.to_excel("Downloads/output.xlsx", index=False )
            print(df2)
            new_sheet = workbook.create_sheet(title="Key Metrics")

            new_sheet.column_dimensions['A'].width = 50
            new_sheet.column_dimensions['B'].width = 50

            for row in dataframe_to_rows(df2, index=False, header=True):
                new_sheet.append(row)
            for row in new_sheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=2):
                for cell in row:
                    cell.fill = PatternFill(start_color="626a80", end_color="626a80", fill_type="solid")

                    

            new_sheet['B3']=df_data['Due_Date']
            new_sheet['B4']=df_data['Crn']
            new_sheet['B5']=df_data['ACCOUNT_NUMBER']
            new_sheet['B6']='=Main_Format1!C16'
            new_sheet['B7']='=Main_Format1!C26'
            new_sheet['B10']='=SUM(Main_Format1!C17,Main_Format1!C18,Main_Format1!C19)'
            new_sheet['B17']='=Main_Format1!C27'
            new_sheet['B12']='=Main_Format1!C28'
            new_sheet['B13']='=Main_Format1!C35'
            new_sheet['B8']='=Main_Format1!C28-Main_Format1!C29'
            new_sheet['B9']='=Main_Format1!C28-Main_Format1!C29'
            new_sheet['B16']='=Main_Format1!C44'
            new_sheet['B17']='=Main_Format1!C45'
            new_sheet['B18']='=Main_Format1!C37'
            new_sheet['B21']='=Main_Format1!C40'
            new_sheet['B22']='=Main_Format1!C39'
            new_sheet['B23']='=Main_Format1!C43'
            new_sheet['B50']='=Main_Format1!C23'
            new_sheet['B55']='=Main_Format1!C32'
            new_sheet['B56']='=Main_Format1!C30'
            new_sheet['B57']='=Main_Format1!C22'
            new_sheet['B68']='=Main_Format1!C31'
            new_sheet['B70']='=Main_Format1!C38'
            new_sheet['B71']='=Main_Format1!C37' 

            report_name="Kotak_CMB_SME_100000_Output_Mapping_V1.0"

        #template9
        if 'CMB_SME_444444_MULTIPLE_MARGIN' in template_name:
            template_data1 ="""<table style="width:100%">
    <tr><th></th><td></td><td></td></tr>
        <tr><th></th><td></td><td></td></tr>
        <tr style="height:600px";"font-size: 20px;"><td></td><td></td><td style="font-size: 20px;">DP STOCK STATEMENT</td><td>Division 1</td><td></td></tr>
        <tr style="height:600px"><td></td><td>Name of Borrower</td><td>{{NAME_OF_CUSTOMER}}</td><td></td></tr>
        <tr style="height:600px"><td></td><td></td><td></td><td>As per SANCTION</td></tr >
        <tr style="height:600px"><td></td><td>Details for the month </td><td>Statement value</td><td>Margin</td><td>Net value</td></tr>
        <tr style="height:600px"><td></td><td>Raw Material</td><td>{{Raw_Material}}</td><td>{{STOCK_MARGIN}}</td></tr>
        <tr style="height:600px"><td></td><td>Work in Process</td><td>{{Work_in_Process}}</td><td>{{STOCK_MARGIN}}</td></tr>
        <tr style="height:600px"><td></td><td>Finished Goods</td><td>{{Finished_Goods}}</td><td>{{STOCK_MARGIN}}</td></tr>
        <tr style="height:600px"><td></td><td>Stores & Spares</td><td>{{Stores_Spares}}</td><td>{{STOCK_MARGIN}}</td></tr>
        <tr style="height:600px"><td></td><td>Stock in Transit</td><td>{{Stock_in_Transit}}</td><td>{{STOCK_MARGIN}}</td></tr>
        <tr style="height:600px"><td></td><td>Other stock</td><td>{{Other_stock}}</td><td>{{STOCK_MARGIN}}</td></tr>   
        <tr style="height:600px"><td></td><td>Total Stock</td><td>{{Total_stock}}</td></tr>
        <tr style="height:600px"><td></td><td>Less: Total Creditors(BC/LC) </td><td>{{Creditors}}</td></tr>
        <tr style="height:600px"><td></td><td>PBD Utilised from other Bank</td><td>{{pbd_utilised}}</td></tr>
        <tr style="height:600px"><td></td><td>Advance paid to suppliers </td><td>{{Advance_paid_to_suppliers}}</td></tr>
        <tr style="height:600px"><td></td><td>Channel Finance </td><td>{{channel_finance}}</td></tr>  
        <tr style="height:600px"><td></td><td>EPC/PCFC stock</td><td>{{epc_pcfc_stock}}</td></tr>
        <tr style="height:600px"><td></td><td>Net stock</td><td>{{Net_stock}}</td></tr>
        <tr style="height:600px"><td>A</td><td>Drawing power I</td><td>{{drawing_power_1}}</td></tr>
        <tr style="height:600px"><td></td><td></td><td></td></tr>     
        <tr style="height:600px"><td></td><td>Total debtors</td><td>{{TOTAL_Debitors}}</td></tr>         
        <tr style="height:600px"><td></td><td>Domestic receivables</td><td>{{wb_non_groupco_drs_days}}</td></tr>             
        <tr style="height:600px"><td></td><td>Domestic receivables beyond cover period</td><td>{{domestic_rec_beyond_days}}</td></tr>
        <tr style="height:600px"><td></td><td>Group Co debtors</td><td>{{group_co_debtors}}</td></tr>
        <tr style="height:600px"><td></td><td>Bill Discounted</td><td>{{bill_discounted}}</td></tr>
        <tr style="height:600px"><td></td><td>Less: Deductibles(Advances)</td><td>{{deductibles}}</td></tr>
        <tr style="height:600px"><td>B</td><td>Net Domestic Receivables</td><td>{{net_domestic_receviables}}</td><td>{{STOCK_MARGIN}}</td></tr>
        <tr style="height:600px"><td></td><td></td><td></td></tr>  
        <tr style="height:600px"><td></td><td>Export receivables</td><td>{{export_receviables}}</td><td>{{STOCK_MARGIN}}</td></tr>
        <tr style="height:600px"><td></td><td>Export receivables beyond cover period</td><td>{{export_receviables_beyond}}</td></tr>
        <tr style="height:600px"><td></td><td>less: Deductibles</td><td>{{less_deductibles}}</td></tr>
        <tr style="height:600px"><td></td><td>Net Export Receivables</td><td>{{net_export_receviables}}</td></tr>
        <tr style="height:600px"><td></td><td></td><td></td></tr> 
        
        <tr style="height:600px"><td></td><td>Domestic receivables upto … days</td><td>{{domestic_receviableees_upto}}</td></tr>
        <tr style="height:600px"><td></td><td>Domestic receivables beyond … days</td><td>{{domestic_rec_beyond_days}}</td></tr>
        <tr style="height:600px"><td></td><td>Export receivables upto … days</td><td>{{export_receive_beyond_days}}</td></tr>
        <tr style="height:600px"><td></td><td>Export receivables beyond … days</td><td>{{export_rec_beyond_days}}</td></tr>
        <tr style="height:600px"><td></td><td></td><td></td></tr>    
        <tr style="height:600px"><td>B</td><td>chargeaeble_debtors</td><td>{{chargeaeble_debtors}}</td></tr>
        <tr style="height:600px"><td>B</td><td>Drawing power II</td><td>{{Drawing_power}}</td></tr>     
        <tr style="height:600px"><td>(A+B)</td><td>Total Drawing power (I +II)</td><td>{{Total_drawing}}</td></tr> 
        <tr style="height:600px"><td></td><td>Less: Other WC bankers outstanding </td><td>{{less_other_wc_bankers}}</td></tr>   
        <tr style="height:600px"><td></td><td>Less: Kotak PO Funding</td><td>{{kotak_po_funding}}</td></tr>
        <tr style="height:600px"><td></td><td>Net DP</td><td>{{net_dp}}</td></tr>           
        <tr style="height:600px"><td></td><td>DP Limit</td><td>{{dp_limit_consortium}}</td></tr>     
        <tr style="height:600px"><td></td><td>Final DP</td><td>{{final_dp}}</td></tr>
        <tr style="height:600px"><td></td><td>Adhoc</td><td>{{adhoc}}</td></tr>
        <tr style="height:600px"><td></td><td>Final DP (with Adhoc)</td><td>{{final_dp_adhoc}}</td></tr>
        <tr style="height:600px"><td></td><td>Sales</td><td>{{Sales}}</td></tr>
        <tr style="height:600px"><td></td><td>Purchases</td><td>{{Purchases}}</td></tr>
        
    </table>"""
     
            print(df_data)
            data = df_data
            template = Template(template_data1)
            rendered_template = template.render(data)
      
            dfs = pd.read_html(rendered_template)

            # Get the DataFrame from the list of DataFrames
            df = dfs[0]
            df.columns = [" ", " ", " ","","",]

            # print(df)



            output_path = Path(
                        f'/app/output/kmb/assets/pdf/kmb/{case_id}')
            os.umask(0)
            logging.info(f"---{output_path}---")
            df.to_excel(f"{output_path}/output.xlsx", index=False )
            
            
            logging.info(f"---saved---")
            workbook = openpyxl.load_workbook(f"{output_path}/output.xlsx",data_only=True)

            # Select the desired sheet
            sheet = workbook.active
            sheet.title = "Main_Format1"# You can also specify the sheet name like workbook['Sheet1']
            for row in range(10, 47):
                cell_address = 'C' + str(row)
                cell = sheet[cell_address]
                if cell.value is not None and isinstance(cell.value, str):
                    try:
                        print(type(cell.value),'------------------------------------------')
                        cell.value = float(cell.value) # Convert text to integer
                    except ValueError as e:
                        print(e,'------------------------------{e}')
                        # Handle non-convertible values here if needed
                        pass


            sheet.column_dimensions['B'].width = 60
            sheet.column_dimensions['c'].width = 40
            sheet.column_dimensions['d'].width = 40
            sheet.column_dimensions['e'].width = 40
            sheet['C16'] = '=SUM(C10:C15)'
            sheet['C24'] = '=C16-C17-C18-C19-C20+C21-C22-C23'
            sheet['C25']='=IF(C24 <0, 0, C24*D6)'
            sheet['C26'] ='=IF(C24<0, 0,C24-C25)'
            sheet['C32']='=IF(C24<0, C27 - C28 - C29- C30 -C31 + C24, C24 - SUM(C28, C29, C30, C31))'
            sheet['C33']='=C32*D7'
            sheet['C34']='=C32-C33'
            sheet['C35']='=IF(C24<0,C34,C26+C34)'
            sheet['C37']='=C35* C36'
            sheet['C40']='=IF(38= 0, MIN(C37, C39), MIN(C38 ,C39))'
            sheet['C42']='=IF(C40+ C41> C39, C39, C40 + C41)'
            sheet['E8']='=C8*D8'
            sheet['E9']='=C9*D9'
            sheet['E9']='=C10*D10'
            sheet['E9']='=C11*D11'
            sheet['E9']='=C12*D12'
            sheet['E9']='=C13*D13'
            sheet['E34']='=C34*D34'
            sheet['E14']='=SUM(E8:E13)'
            sheet['E20']='=E14-E15'
            sheet['E21']='=E20'
            sheet['E41']='=E24+E25+E31+E32+E29+E34'
            sheet['E42']='=E41'
            sheet['E46']='=E42-E44-E45'
            sheet['E48']='=MIN(E46,E47)'
            sheet['E50']='=IF(E48+E49>E47,E47,E48+E49)'
            

            img_path = Path(f'/app/output/kmb/assets/pdf/kmb/kotak.PNG')
            img = openpyxl.drawing.image.Image(img_path)
            img.anchor = "B3"  

            for row in sheet.iter_rows(min_row=8, max_row=52, min_col=5, max_col=5):
                for cell in row:
                    cell.fill = PatternFill(start_color="53b553", end_color="53b553", fill_type="solid")

                    # Green color
            for row in sheet.iter_rows(min_row=4, max_row=4, min_col=3, max_col=3):
                for cell in row:
                    cell.font = Font(size=16, bold=True, color="090a0a")
            for row in sheet.iter_rows(min_row=16, max_row=16, min_col=2, max_col=2):
                for cell in row:
                    cell.font = Font(size=10, bold=True, color="090a0a")        

            for row in sheet.iter_rows(min_row=6, max_row=6, min_col=2, max_col=3):
                for cell in row:
                    cell.fill = PatternFill(start_color="626a80", end_color="626a80", fill_type="solid")   
                    cell.font = Font(size=14, bold=True, color="FFFFFF")
            for row in sheet.iter_rows(min_row=7, max_row=7, min_col=1, max_col=5):
                for cell in row:
                    cell.fill = PatternFill(start_color="626a80", end_color="626a80", fill_type="solid")   
                    cell.font = Font(size=14, bold=True, color="FFFFFF")
            for row in sheet.iter_rows(min_row=5, max_row=5, min_col=2, max_col=2):
                for cell in row:
                    cell.fill = PatternFill(start_color="626a80", end_color="626a80", fill_type="solid")
                    cell.font = Font(size=14, bold=True, color="FFFFFF")
            for row in sheet.iter_rows(min_row=21, max_row=21, min_col=1, max_col=5):
                for cell in row:
                    cell.fill = PatternFill(start_color="626a80", end_color="626a80", fill_type="solid")
                    cell.font = Font(size=14, bold=True, color="FFFFFF")
        
            for row in sheet.iter_rows(min_row=42, max_row=43, min_col=1, max_col=5):
                for cell in row:
                    cell.fill = PatternFill(start_color="626a80", end_color="626a80", fill_type="solid")
                    cell.font = Font(size=14, bold=True, color="FFFFFF")
            for row in sheet.iter_rows(min_row=46, max_row=50, min_col=1, max_col=5):
                for cell in row:
                    cell.fill = PatternFill(start_color="626a80", end_color="626a80", fill_type="solid")
                    cell.font = Font(size=14, bold=True, color="FFFFFF")

            # Save the changes to the workbook



            workbook.save(f"{output_path}/output.xlsx")
            logging.info(f"completed sheet 1-----")


            Heads= [
                "For the Month", "Due Date", "CRN", "Account Number", "Total Stock", "Stock Applicable for DP", 
                "Stock < 90", "Stock < 120", "Total Creditors", "Creditors Applicable for DP", 
                "Total Debtors", "Debtors Applicable for DP", "Debtors < 90", "Debtors < 120", 
                "Total Sales", "Total Purchases", "Other Bank Limit", "System DP", "System Limit", 
                "CC Limit", "Computed DP", "Final DP", "Stock <30 Days", "Stock Between 31 and 60 Days", 
                "Stock Between 61 and 90 Days", "Stock Between 91 and 120 Days", "Stock Between 121 and 180 Days", 
                "Stock > 180 Days", "Debtors <30 Days", "Debtors Between 31 and 60 Days", 
                "Debtors Between 61 and 90 Days", "Debtors Between 91 and 120 Days", 
                "Debtors Between 121 and 180 Days", "Debtors > 180 Days", "Sales Spares Workshop", 
                "Sales Amount Vehicle", "Purchase Spares Workshop", "Debtors > 90 Days Spares Workshop", 
                "Debtors > 90 Days Vehicle", "Debtors Spares Between 91 to 120", 
                "Debtors Vehicle Between 91 to 120", "Total Debtors Spares", "Total Spares Stock", 
                "Total Spares Stock <= 120", "Total Spares Stock <= 90", "Total Vehicle Stock", 
                "Total Vehicle Stock <= 120", "Total Vehicle Stock <= 90", 
                "Debtors Advance Paid to suppliers", "Purchase Vehicles", "Total Debtors Vehicles", 
                "Total Stock <=120", "Total Stock <=90", "Advance Received from customers", 
                "Bill discounted debtors", "Channel Finance", "Debtors 0 to 60 Days", 
                "Debtors 0 to 5 Days", "Debtors 6 to 30 Days", "Debtors 6 to 60 Days", 
                "Debtors < 150 days", "Debtors < 45 days", "Debtors Specific", 
                "Debtors of Vehicle < 30 days", "Debtors of Vehicle < 60 days", 
                "Debtors of spares < 60 days", "Group Co debtors", "Insurance Amount", 
                "Limits Os With KMBL Banks 1", "Outstanding with other WC Bankers", 
                "Previous month Total Creditors", "Previous month Total Debtors", 
                "Previous month Total Stock", "Purchase 1 Month prior", "Purchase 2 Month prior", 
                "Purchase 3 Month prior", "Purchase 4 Month prior", "Purchase 5 Month prior", 
                "Purchase Spares 1 Month Prior", "Purchase Spares 2 Month Prior", 
                "Purchase Spares 3 Month Prior", "Purchase Spares 4 Month Prior", 
                "Purchase Spares 5 Month Prior", "Purchase Vehicles 1 Month Prior", 
                "Purchase Vehicles 2 Month Prior", "Purchase Vehicles 3 Month Prior", 
                "Purchase Vehicles 4 Month Prior", "Purchase Vehicles 5 Month Prior", 
                "RM Stock < 60 days", "RM Stock < 90 days", "Sales 1 Month Prior", 
                "Sales 2 Month Prior", "Sales 3 Month Prior", "Sales 4 Month Prior", 
                "Sales 5 Month Prior", "Sales Spares 1 Month Prior", "Sales Spares 2 Month Prior", 
                "Sales Spares 3 Month Prior", "Sales Spares 4 Month Prior", 
                "Sales Spares 5 Month Prior", "Sales Vehicles 1 Month Prior", 
                "Sales Vehicles 2 Month Prior", "Sales Vehicles 3 Month Prior", 
                "Sales Vehicles 4 Month Prior", "Sales Vehicles 5 Month Prior", "Security Deposit", 
                "Stock Between 0 and 60 Days", "Stock Specific", "Stock in Transit < 90 Days", 
                "Stock of Vehicle < 60 days", "Stock of spares < 60 days", "Total Purchase Spares", 
                "Total Purchase Vehicles", "Total Sales Spares", "Total Sales Vehicles", 
                "Unbilled Debtors"
            ]
            # Number of columns
            num_columns = len(Heads)

            # Create 'Values' list with empty strings
            values_list = ["-"] * num_columns
            print(values_list)

            # Update the 'Values' key in the data2 dictionary


            # Print the updated data2 dictionary
            data2={'Heads':Heads,'Values':values_list}
            print(data2)



            # data2={'Heads':Heads,'Values':Values}
            df2 = pd.DataFrame(data2)
            # df2.to_excel("Downloads/output.xlsx", index=False )
            print(df2)
            new_sheet = workbook.create_sheet(title="Key Metrics")

            new_sheet.column_dimensions['A'].width = 50
            new_sheet.column_dimensions['B'].width = 50


            for row in dataframe_to_rows(df2, index=False, header=True):
                new_sheet.append(row)
            for row in new_sheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=2):
                for cell in row:
                    cell.fill = PatternFill(start_color="626a80", end_color="626a80", fill_type="solid")

                    

            new_sheet['B3']=df_data['Due_Date']
            new_sheet['B4']=df_data['Crn']
            new_sheet['B5']=df_data['ACCOUNT_NUMBER']
            new_sheet['B6']='=Main_Format1!C16'
            new_sheet['B7']='=Main_Format1!C26'
            new_sheet['B10']='=SUM(Main_Format1!C17,Main_Format1!C18,Main_Format1!C19)'
            new_sheet['B17']='=Main_Format1!C27'
            new_sheet['B12']='=Main_Format1!C28'
            new_sheet['B13']='=Main_Format1!C35'
            new_sheet['B8']='=Main_Format1!C28-Main_Format1!C29'
            new_sheet['B9']='=Main_Format1!C28-Main_Format1!C29'
            new_sheet['B16']='=Main_Format1!C44'
            new_sheet['B17']='=Main_Format1!C45'
            new_sheet['B18']='=Main_Format1!C37'
            new_sheet['B21']='=Main_Format1!C40'
            new_sheet['B22']='=Main_Format1!C39'
            new_sheet['B23']='=Main_Format1!C43'
            new_sheet['B50']='=Main_Format1!C23'
            new_sheet['B55']='=Main_Format1!C32'
            new_sheet['B56']='=Main_Format1!C30'
            new_sheet['B57']='=Main_Format1!C22'
            new_sheet['B68']='=Main_Format1!C31'
            new_sheet['B70']='=Main_Format1!C38'
            new_sheet['B71']='=Main_Format1!C37' 

            report_name="Kotak_CMB_SME_444444_Output_Mapping_V1.0"

        #template10
        if 'WBG_555555_BFIG' in template_name:
            template_data1 = """

<table style="width:100%">
  <tr style="height:600px";"font-size: 20px;"><td></td><td><b>NAME OF THE COMPANY :</b></td><td>{{NAME_OF_CUSTOMER}}</td><td style="font-size: 20px;"></td><td></td></tr>
                <tr style="height:600px"><td></td><td>Date</td><td>{{Due_Date}}</td><td></td></tr>  
                <tr style="height:600px"><td></td><td>Facilities: DP Backed facility</td><td>{{db_backed_facility}}</td><td></td></tr> 
                <tr style="height:600px"><td></td><td>Margin requirement for DP with KMBL</td><td>{{margin_dp_with_kmbl}}</td><td></td></tr>    
                <tr style="height:600px"><td></td><td>Asset cover requirement for KMBL (no. of times)</td><td>{{asset_cover_requirement}}</td><td></td></tr>
                <tr style="height:600px"><td></td><td></td><td></td></tr>
                <tr style="height:600px"><td></td><td>Total Available loan assets:</td><td>{{total_loan_asset}}</td><td></td></tr> 
                <tr style="height:600px"><td></td><td>Add: other assets if any</td><td>{{add_asset}}</td><td></td></tr>
                <tr style="height:600px"><td></td><td><b>Total available  assets for funding :</b></td><td>{{total_ass_funding}}</td><td></td></tr>
                <tr style="height:600px"><td></td><td>Loan assets not eligible for bank finance:</td><td>{{Dude_Date}}</td><td></td></tr>  
                <tr style="height:600px"><td></td><td>Less: Security Reqd. For other non-bank lenders</td><td>{{less_security}}</td><td></td></tr> 
                <tr style="height:600px"><td></td><td></td><td></td><td></td></tr>
                <tr style="height:600px"><td></td><td></td><td></td><td></td></tr>
                <tr style="height:600px"><td></td><td><b>Net eligible loan assets considered for asset cover (a) to Banks </b></td><td>{{net_eli_lo_ass_cov_bank}}</td><td></td></tr>  
                <tr style="height:600px"><td></td><td><b>(b) Limits/Borrowings from other Banks/NBFCs/Fis:</b></td><td>{{limits_from_banks}}</td><td></td></tr>  
                <tr style="height:600px"><td></td><td>Outstanding of  secured facilities which are not DP backed (TL  Banks/FIs )</td><td>{{outstand_secured}}</td><td></td></tr>  
                <tr style="height:600px"><td></td><td>Average Asset cover for TL  ( no. times)</td><td>{{average_asset_cover_tl}}</td><td></td></tr>  
                <tr style="height:600px"><td></td><td>Asset required for the above </td><td>{{asset_of_the_above}}</td><td></td></tr>  
                <tr style="height:600px"><td></td><td>Balance Asset available for DP backed facilities</td><td>{{balance_asset_for_dp_back}}</td><td></td></tr> 
                <tr style="height:600px"><td></td><td><b>Asset  required for DP backed facilities of other banks</b></td><td>{{asset_for_dp_backed_other_bank}}</td><td></td></tr> 
                <tr style="height:600px"><td></td><td><b>Balance security available to KMBL</b></td><td>{{balance_sec_to_kmbl}}</td><td></td></tr> 
                <tr style="height:600px"><td></td><td><b>Margin</b></td><td>{{margin}}</td><td></td></tr> 
                <tr style="height:600px"><td></td><td><b>DP available for KMBL</b></td><td>{{dp_available_for_kmbl}}</td><td></td></tr> 
                <tr style="height:600px"><td></td><td><b>DP to be set</b></td><td>{{dp_to_be_set}}</td><td></td></tr> 
                <tr style="height:600px"><td></td><td></td><td></td><td></td></tr>
                <tr style="height:600px"><td></td><td><b>Loan assets not eligible for Bank Financing</b></td><td>{{loan_assets_not_eligible}}</td><td></td></tr> 
                <tr style="height:600px"><td></td><td>Immoveable Property</td><td>{{immoveable_property}}</td><td></td></tr> 
                <tr style="height:600px"><td></td><td>Capital Market related exposure</td><td>{{capital_mark_related_exposure}}</td><td></td></tr> 
                <tr style="height:600px"><td></td><td>Dealer Finance</td><td>{{dealer_finance}}</td><td></td></tr> 
                <tr style="height:600px"><td></td><td>Others</td><td>{{others}}</td><td></td></tr>
                <tr style="height:600px"><td></td><td><b>Total Exclusion</b></td><td>{{total_exclusion}}</td><td></td></tr>
                <tr style="height:600px"><td></td><td>Loan assets  Required  to Bond/Debentures</td><td>{{loan_assets_bond_deb}}</td><td></td></tr>
                <tr style="height:600px"><td></td><td><b>Net additional assets required for Bond / Debebtures</b></td><td>{{net_add_assets_bond_deben}}</td><td></td></tr>
                
            </table>
            
            
            """
            print(df_data)
            # Sample data for the template
            data = df_data
            template = Template(template_data1)
            rendered_template = template.render(data)
            # # Create a Jinja template object
            # template = Template(template_data1)


            # # Render the template with data
            # rendered_template = template.render(data)
            # print(rendered_template)
            # Parse the HTML table into a DataFrame
            dfs = pd.read_html(rendered_template)

            # Get the DataFrame from the list of DataFrames
            df = dfs[0]
            df.columns = [" ", " ", " ","","",]

            # print(df)



            output_path = Path(
                        f'/app/output/kmb/assets/pdf/kmb/{case_id}')
            os.umask(0)
            logging.info(f"---{output_path}---")
            df.to_excel(f"{output_path}/output.xlsx", index=False )
            
            
            logging.info(f"---saved---")
            workbook = openpyxl.load_workbook(f"{output_path}/output.xlsx",data_only=True)

            # Select the desired sheet
            sheet = workbook.active
            sheet.title = "Main_Format1"# You can also specify the sheet name like workbook['Sheet1']
            for row in range(10, 47):
                cell_address = 'C' + str(row)
                cell = sheet[cell_address]
                if cell.value is not None and isinstance(cell.value, str):
                    try:
                        print(type(cell.value),'------------------------------------------')
                        cell.value = float(cell.value) # Convert text to integer
                    except ValueError as e:
                        print(e,'------------------------------{e}')
                        # Handle non-convertible values here if needed
                        pass



           
            sheet.column_dimensions['B'].width = 60
            sheet.column_dimensions['c'].width = 60
            sheet.column_dimensions['d'].width = 40
            sheet['C10'] = '=C8+C9'
            sheet['C15'] = '=C10-C11-C12'
            sheet['C18'] = '=C15/C17'
            sheet['C20'] = '=C15-C19'
            sheet['C22'] = '=C20-C21'
            sheet['C23'] = '=C22*C5'
            sheet['C24'] = '=C22-C23'
            sheet['C25'] = '=MIN(C24,C4)'
            sheet['C32'] = '=SUM(C28,C29,C30,C31)'
            sheet['C34'] = '=C33-C32'
            img_path = Path(f'/app/output/kmb/assets/pdf/kmb/kotak.PNG')
            img = openpyxl.drawing.image.Image(img_path)
            img.anchor = "B3"  

            # Save the changes to the workbook



            workbook.save(f"{output_path}/output.xlsx")
            logging.info(f"completed sheet 1-----")


            Heads= [
                "For the Month", "Due Date", "CRN", "Account Number", "Total Stock", "Stock Applicable for DP", 
                "Stock < 90", "Stock < 120", "Total Creditors", "Creditors Applicable for DP", 
                "Total Debtors", "Debtors Applicable for DP", "Debtors < 90", "Debtors < 120", 
                "Total Sales", "Total Purchases", "Other Bank Limit", "System DP", "System Limit", 
                "CC Limit", "Computed DP", "Final DP", "Stock <30 Days", "Stock Between 31 and 60 Days", 
                "Stock Between 61 and 90 Days", "Stock Between 91 and 120 Days", "Stock Between 121 and 180 Days", 
                "Stock > 180 Days", "Debtors <30 Days", "Debtors Between 31 and 60 Days", 
                "Debtors Between 61 and 90 Days", "Debtors Between 91 and 120 Days", 
                "Debtors Between 121 and 180 Days", "Debtors > 180 Days", "Sales Spares Workshop", 
                "Sales Amount Vehicle", "Purchase Spares Workshop", "Debtors > 90 Days Spares Workshop", 
                "Debtors > 90 Days Vehicle", "Debtors Spares Between 91 to 120", 
                "Debtors Vehicle Between 91 to 120", "Total Debtors Spares", "Total Spares Stock", 
                "Total Spares Stock <= 120", "Total Spares Stock <= 90", "Total Vehicle Stock", 
                "Total Vehicle Stock <= 120", "Total Vehicle Stock <= 90", 
                "Debtors Advance Paid to suppliers", "Purchase Vehicles", "Total Debtors Vehicles", 
                "Total Stock <=120", "Total Stock <=90", "Advance Received from customers", 
                "Bill discounted debtors", "Channel Finance", "Debtors 0 to 60 Days", 
                "Debtors 0 to 5 Days", "Debtors 6 to 30 Days", "Debtors 6 to 60 Days", 
                "Debtors < 150 days", "Debtors < 45 days", "Debtors Specific", 
                "Debtors of Vehicle < 30 days", "Debtors of Vehicle < 60 days", 
                "Debtors of spares < 60 days", "Group Co debtors", "Insurance Amount", 
                "Limits Os With KMBL Banks 1", "Outstanding with other WC Bankers", 
                "Previous month Total Creditors", "Previous month Total Debtors", 
                "Previous month Total Stock", "Purchase 1 Month prior", "Purchase 2 Month prior", 
                "Purchase 3 Month prior", "Purchase 4 Month prior", "Purchase 5 Month prior", 
                "Purchase Spares 1 Month Prior", "Purchase Spares 2 Month Prior", 
                "Purchase Spares 3 Month Prior", "Purchase Spares 4 Month Prior", 
                "Purchase Spares 5 Month Prior", "Purchase Vehicles 1 Month Prior", 
                "Purchase Vehicles 2 Month Prior", "Purchase Vehicles 3 Month Prior", 
                "Purchase Vehicles 4 Month Prior", "Purchase Vehicles 5 Month Prior", 
                "RM Stock < 60 days", "RM Stock < 90 days", "Sales 1 Month Prior", 
                "Sales 2 Month Prior", "Sales 3 Month Prior", "Sales 4 Month Prior", 
                "Sales 5 Month Prior", "Sales Spares 1 Month Prior", "Sales Spares 2 Month Prior", 
                "Sales Spares 3 Month Prior", "Sales Spares 4 Month Prior", 
                "Sales Spares 5 Month Prior", "Sales Vehicles 1 Month Prior", 
                "Sales Vehicles 2 Month Prior", "Sales Vehicles 3 Month Prior", 
                "Sales Vehicles 4 Month Prior", "Sales Vehicles 5 Month Prior", "Security Deposit", 
                "Stock Between 0 and 60 Days", "Stock Specific", "Stock in Transit < 90 Days", 
                "Stock of Vehicle < 60 days", "Stock of spares < 60 days", "Total Purchase Spares", 
                "Total Purchase Vehicles", "Total Sales Spares", "Total Sales Vehicles", 
                "Unbilled Debtors"
            ]
            # Number of columns
            num_columns = len(Heads)

            # Create 'Values' list with empty strings
            values_list = ["-"] * num_columns
            print(values_list)

            # Update the 'Values' key in the data2 dictionary


            # Print the updated data2 dictionary
            data2={'Heads':Heads,'Values':values_list}
            print(data2)



            # data2={'Heads':Heads,'Values':Values}
            df2 = pd.DataFrame(data2)
            # df2.to_excel("Downloads/output.xlsx", index=False )
            print(df2)
            new_sheet = workbook.create_sheet(title="Key Metrics")

            new_sheet.column_dimensions['A'].width = 50
            new_sheet.column_dimensions['B'].width = 50





            for row in dataframe_to_rows(df2, index=False, header=True):
                new_sheet.append(row)
            for row in new_sheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=2):
                for cell in row:
                    cell.fill = PatternFill(start_color="626a80", end_color="626a80", fill_type="solid")

                    

            new_sheet['B3']=df_data['Due_Date']
            new_sheet['B4']=df_data['Crn']
            new_sheet['B5']=df_data['ACCOUNT_NUMBER']
            new_sheet['B6']='=Main_Format1!C16'
            new_sheet['B7']='=Main_Format1!C26'
            new_sheet['B10']='=SUM(Main_Format1!C17,Main_Format1!C18,Main_Format1!C19)'
            new_sheet['B17']='=Main_Format1!C27'
            new_sheet['B12']='=Main_Format1!C28'
            new_sheet['B13']='=Main_Format1!C35'
            new_sheet['B8']='=Main_Format1!C28-Main_Format1!C29'
            new_sheet['B9']='=Main_Format1!C28-Main_Format1!C29'
            new_sheet['B16']='=Main_Format1!C44'
            new_sheet['B17']='=Main_Format1!C45'
            new_sheet['B18']='=Main_Format1!C37'
            new_sheet['B21']='=Main_Format1!C40'
            new_sheet['B22']='=Main_Format1!C39'
            new_sheet['B23']='=Main_Format1!C43'
            new_sheet['B50']='=Main_Format1!C23'
            new_sheet['B55']='=Main_Format1!C32'
            new_sheet['B56']='=Main_Format1!C30'
            new_sheet['B57']='=Main_Format1!C22'
            new_sheet['B68']='=Main_Format1!C31'
            new_sheet['B70']='=Main_Format1!C38'
            new_sheet['B71']='=Main_Format1!C37' 

            report_name="Kotak_WBG_555555_BFIG_Output_Mapping_V1.0"



        if '444445 output mapping' in template_name:
            template_data1 = """



            <table style="width:100%">
            <tr><th></th><td></td><td></td></tr>
                <tr><th></th><td></td><td></td></tr>
                <tr style="height:600px";"font-size: 20px;"><td></td><td></td><td style="font-size: 20px;">DP STOCK STATEMENT</td><td>Division 1</td><td></td></tr>
                <tr style="height:600px"><td></td><td>Name of Borrower</td><td>{{NAME_OF_CUSTOMER}}</td><td></td></tr>
                <tr style="height:600px"><td></td><td></td><td></td><td>As per SANCTION</td></tr >
                <tr style="height:600px"><td></td><td></td><td>Details for the month </td><td>Statement value</td></tr >
                <tr style="height:600px";"color: red;"><td></td><td>Raw Material</td><td>{{Raw_Material}}</td></tr>
                <tr style="height:600px"><td></td><td style="color: blue;">Work in Process</td><td>{{stocks_in_process_Closing_Stock}}</td></tr>
                <tr style="height:600px"><td></td><td>Finished Goods</td><td>{{Finished_Goods_Closing_Stock}}</td></tr>
                <tr style="height:600px"><td></td><td>Stores & Spares</td><td>{{Stores_Spares}}</td></tr>
                <tr style="height:600px"><td></td><td>Stock in Transit</td><td>{{Stock_in_Transit}}</td></tr>
                <tr style="height:600px"><td></td><td>Other stock</td><td>{{Other_stock}}</td></tr>   
                <tr style="height:600px"><td></td><td>Total Stock</td><td>{{Total_stock}}</td></tr>
                <tr style="height:600px"><td></td><td>Less: Total Creditors(BC/LC) </td><td>{{Creditors}}</td></tr>
                <tr style="height:600px"><td></td><td>PBD Utilised from other Bank</td><td>{{pbd_utilised}}</td></tr>
                <tr style="height:600px"><td></td><td>Advance paid to suppliers </td><td>{{buyers_credittt_wc}}</td></tr>
                <tr style="height:600px"><td></td><td>Channel Finance </td><td>{{channel_finance}}</td></tr>  
                <tr style="height:600px"><td></td><td>EPC/PCFC stock</td><td>{{epc_pcfc_stock}}</td></tr>
                <tr style="height:600px"><td></td><td>Net stock</td><td>{{Net_stock}}</td></tr>
                <tr style="height:600px"><td>A</td><td>Drawing power I</td><td>{{wbgg_less_margin_25}}</td></tr>
                <tr style="height:600px"><td></td><td></td><td></td></tr>     
                <tr style="height:600px"><td></td><td>Total debtors</td><td>{{TOTAL_Debitors}}</td></tr>         
                <tr style="height:600px"><td></td><td>Domestic receivables</td><td>{{wb_non_groupco_drs_days}}</td></tr>             
                <tr style="height:600px"><td></td><td>Domestic receivables beyond cover period</td><td>{{domestic_rec_beyond_days}}</td></tr>
                <tr style="height:600px"><td></td><td>Group Co debtors</td><td>{{wb_net_debtors}}</td></tr>
                <tr style="height:600px"><td></td><td>Bill Discounted</td><td>{{wb_net_debtors}}</td></tr>
                <tr style="height:600px"><td></td><td>Less: Deductibles(Advances)</td><td>{{wb_net_debtors}}</td></tr>
                <tr style="height:600px"><td>B</td><td>Net Domestic Receivables</td><td>{{net_domestic_receviables}}</td></tr>
                <tr style="height:600px"><td></td><td></td><td></td></tr>  
                <tr style="height:600px"><td></td><td>Export receivables</td><td>{{net_domestic_receviableees}}</td></tr>
                <tr style="height:600px"><td></td><td>Export receivables beyond cover period</td><td>{{net_domestic_receviableees}}</td></tr>
                <tr style="height:600px"><td></td><td>less: Deductibles</td><td>{{net_domestic_receviableees}}</td></tr>
                <tr style="height:600px"><td></td><td>Net Export Receivables</td><td>{{net_export_receviables}}</td></tr>
                <tr style="height:600px"><td></td><td></td><td></td></tr> 
                
                <tr style="height:600px"><td></td><td>Domestic receivables upto … days</td><td>{{net_domestic_receviableees}}</td></tr>
                <tr style="height:600px"><td></td><td>Domestic receivables beyond … days</td><td>{{domestic_rec_beyond_days}}</td></tr>
                <tr style="height:600px"><td></td><td>Export receivables upto … days</td><td>{{net_domestic_receviableees}}</td></tr>
                <tr style="height:600px"><td></td><td>Export receivables beyond … days</td><td>{{export_rec_beyond_days}}</td></tr>
                <tr style="height:600px"><td></td><td></td><td></td></tr>    
                <tr style="height:600px"><td>B</td><td>Drawing power II</td><td>{{chargeaeble_debtors}}</td></tr>     
                <tr style="height:600px"><td>(A+B)</td><td>Total Drawing power (I +II)</td><td>{{Todtal}}</td></tr> 
                <tr style="height:600px"><td></td><td>Less: Other WC bankers outstanding </td><td>{{our_shareee}}</td></tr>   
                <tr style="height:600px"><td></td><td>Less: Kotak PO Funding</td><td>{{kotak_po_funding}}</td></tr>
                <tr style="height:600px"><td></td><td>Net DP</td><td>{{net_dp}}</td></tr>           
                <tr style="height:600px"><td></td><td>DP Limit</td><td>{{dppp_limit_consortium}}</td></tr>     
                <tr style="height:600px"><td></td><td>Final DP</td><td>{{final_dp}}</td></tr>
                <tr style="height:600px"><td></td><td>Adhoc</td><td>{{adhoc}}</td></tr>
                <tr style="height:600px"><td></td><td>Final DP (with Adhoc)</td><td>{{final_dp_adhoc}}</td></tr>
                <tr style="height:600px"><td></td><td>Sales</td><td>{{Sales}}</td></tr>
                <tr style="height:600px"><td></td><td>Purchases</td><td>{{Purchases}}</td></tr>
                
                
            </table>
            
            """




    

           





            print(df_data)
            # Sample data for the template
            data = df_data
            template = Template(template_data1)
            rendered_template = template.render(data)
            # # Create a Jinja template object
            # template = Template(template_data1)


            # # Render the template with data
            # rendered_template = template.render(data)
            # print(rendered_template)
            # Parse the HTML table into a DataFrame
            dfs = pd.read_html(rendered_template)

            # Get the DataFrame from the list of DataFrames
            df = dfs[0]
            df.columns = [" ", " ", " ","","",]

            # print(df)



            output_path = Path(
                        f'/app/output/kmb/assets/pdf/kmb/{case_id}')
            os.umask(0)
            logging.info(f"---{output_path}---")
            df.to_excel(f"{output_path}/output.xlsx", index=False )
            
            
            logging.info(f"---saved---")
            workbook = openpyxl.load_workbook(f"{output_path}/output.xlsx",data_only=True)

            # Select the desired sheet
            sheet = workbook.active
            sheet.title = "Main_Format1"# You can also specify the sheet name like workbook['Sheet1']
            for row in range(10, 47):
                cell_address = 'C' + str(row)
                cell = sheet[cell_address]
                if cell.value is not None and isinstance(cell.value, str):
                    try:
                        print(type(cell.value),'------------------------------------------')
                        cell.value = float(cell.value) # Convert text to integer
                    except ValueError as e:
                        print(e,'------------------------------{e}')
                        # Handle non-convertible values here if needed
                        pass



            sheet.column_dimensions['B'].width = 60
            sheet.column_dimensions['c'].width = 60
            sheet.column_dimensions['d'].width = 40
            sheet['C16'] = '=SUM(C10:C15)'
            sheet['C24'] = '=C16-C17-C18-C19-C20+C21-C22-C23'
            sheet['C25']='=IF(C24 <0, 0, C24*D6)'
            sheet['C26'] ='=IF(C24<0, 0,C24-C25)'
            sheet['C32']='=IF(C24<0, C27 - C28 - C29- C30 -C31 + C24, C24 - SUM(C28, C29, C30, C31))'
            sheet['C33']='=C32*D7'
            sheet['C34']='=C32-C33'
            sheet['C35']='=IF(C24<0,C34,C26+C34)'
            sheet['C37']='=C35* C36'
            sheet['C40']='=IF(38= 0, MIN(C37, C39), MIN(C38 ,C39))'
            sheet['C42']='=IF(C40+ C41> C39, C39, C40 + C41)'


            img_path = Path(f'/app/output/kmb/assets/pdf/kmb/kotak.PNG')
            img = openpyxl.drawing.image.Image(img_path)
            img.anchor = "B3"  

            for row in sheet.iter_rows(min_row=9, max_row=45, min_col=3, max_col=3):
                for cell in row:
                    cell.fill = PatternFill(start_color="53b553", end_color="53b553", fill_type="solid")
                    
                    # Green color
            for row in sheet.iter_rows(min_row=4, max_row=4, min_col=3, max_col=3):
                for cell in row:
                    cell.font = Font(size=16, bold=True, color="090a0a")
            for row in sheet.iter_rows(min_row=16, max_row=16, min_col=2, max_col=2):
                for cell in row:
                    cell.font = Font(size=10, bold=True, color="090a0a")        
                    
            for row in sheet.iter_rows(min_row=9, max_row=9, min_col=2, max_col=4):
                
                for cell in row:
                    cell.fill = PatternFill(start_color="626a80", end_color="626a80", fill_type="solid")
                    cell.font = Font(size=14, bold=True, color="FFFFFF")
            for row in sheet.iter_rows(min_row=6, max_row=7, min_col=2, max_col=3):
                for cell in row:
                    cell.fill = PatternFill(start_color="626a80", end_color="626a80", fill_type="solid")   
                    cell.font = Font(size=14, bold=True, color="FFFFFF")
            for row in sheet.iter_rows(min_row=5, max_row=5, min_col=2, max_col=2):
                for cell in row:
                    cell.fill = PatternFill(start_color="626a80", end_color="626a80", fill_type="solid")
                    cell.font = Font(size=14, bold=True, color="FFFFFF")

            # Save the changes to the workbook



            workbook.save(f"{output_path}/output.xlsx")
            logging.info(f"completed sheet 1-----")


            Heads= [
                "For the Month", "Due Date", "CRN", "Account Number", "Total Stock", "Stock Applicable for DP", 
                "Stock < 90", "Stock < 120", "Total Creditors", "Creditors Applicable for DP", 
                "Total Debtors", "Debtors Applicable for DP", "Debtors < 90", "Debtors < 120", 
                "Total Sales", "Total Purchases", "Other Bank Limit", "System DP", "System Limit", 
                "CC Limit", "Computed DP", "Final DP", "Stock <30 Days", "Stock Between 31 and 60 Days", 
                "Stock Between 61 and 90 Days", "Stock Between 91 and 120 Days", "Stock Between 121 and 180 Days", 
                "Stock > 180 Days", "Debtors <30 Days", "Debtors Between 31 and 60 Days", 
                "Debtors Between 61 and 90 Days", "Debtors Between 91 and 120 Days", 
                "Debtors Between 121 and 180 Days", "Debtors > 180 Days", "Sales Spares Workshop", 
                "Sales Amount Vehicle", "Purchase Spares Workshop", "Debtors > 90 Days Spares Workshop", 
                "Debtors > 90 Days Vehicle", "Debtors Spares Between 91 to 120", 
                "Debtors Vehicle Between 91 to 120", "Total Debtors Spares", "Total Spares Stock", 
                "Total Spares Stock <= 120", "Total Spares Stock <= 90", "Total Vehicle Stock", 
                "Total Vehicle Stock <= 120", "Total Vehicle Stock <= 90", 
                "Debtors Advance Paid to suppliers", "Purchase Vehicles", "Total Debtors Vehicles", 
                "Total Stock <=120", "Total Stock <=90", "Advance Received from customers", 
                "Bill discounted debtors", "Channel Finance", "Debtors 0 to 60 Days", 
                "Debtors 0 to 5 Days", "Debtors 6 to 30 Days", "Debtors 6 to 60 Days", 
                "Debtors < 150 days", "Debtors < 45 days", "Debtors Specific", 
                "Debtors of Vehicle < 30 days", "Debtors of Vehicle < 60 days", 
                "Debtors of spares < 60 days", "Group Co debtors", "Insurance Amount", 
                "Limits Os With KMBL Banks 1", "Outstanding with other WC Bankers", 
                "Previous month Total Creditors", "Previous month Total Debtors", 
                "Previous month Total Stock", "Purchase 1 Month prior", "Purchase 2 Month prior", 
                "Purchase 3 Month prior", "Purchase 4 Month prior", "Purchase 5 Month prior", 
                "Purchase Spares 1 Month Prior", "Purchase Spares 2 Month Prior", 
                "Purchase Spares 3 Month Prior", "Purchase Spares 4 Month Prior", 
                "Purchase Spares 5 Month Prior", "Purchase Vehicles 1 Month Prior", 
                "Purchase Vehicles 2 Month Prior", "Purchase Vehicles 3 Month Prior", 
                "Purchase Vehicles 4 Month Prior", "Purchase Vehicles 5 Month Prior", 
                "RM Stock < 60 days", "RM Stock < 90 days", "Sales 1 Month Prior", 
                "Sales 2 Month Prior", "Sales 3 Month Prior", "Sales 4 Month Prior", 
                "Sales 5 Month Prior", "Sales Spares 1 Month Prior", "Sales Spares 2 Month Prior", 
                "Sales Spares 3 Month Prior", "Sales Spares 4 Month Prior", 
                "Sales Spares 5 Month Prior", "Sales Vehicles 1 Month Prior", 
                "Sales Vehicles 2 Month Prior", "Sales Vehicles 3 Month Prior", 
                "Sales Vehicles 4 Month Prior", "Sales Vehicles 5 Month Prior", "Security Deposit", 
                "Stock Between 0 and 60 Days", "Stock Specific", "Stock in Transit < 90 Days", 
                "Stock of Vehicle < 60 days", "Stock of spares < 60 days", "Total Purchase Spares", 
                "Total Purchase Vehicles", "Total Sales Spares", "Total Sales Vehicles", 
                "Unbilled Debtors"
            ]
            # Number of columns
            num_columns = len(Heads)

            # Create 'Values' list with empty strings
            values_list = ["-"] * num_columns
            print(values_list)

            # Update the 'Values' key in the data2 dictionary


            # Print the updated data2 dictionary
            data2={'Heads':Heads,'Values':values_list}
            print(data2)



            # data2={'Heads':Heads,'Values':Values}
            df2 = pd.DataFrame(data2)
            # df2.to_excel("Downloads/output.xlsx", index=False )
            print(df2)
            new_sheet = workbook.create_sheet(title="Key Metrics")

            new_sheet.column_dimensions['A'].width = 50
            new_sheet.column_dimensions['B'].width = 50





            for row in dataframe_to_rows(df2, index=False, header=True):
                new_sheet.append(row)
            for row in new_sheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=2):
                for cell in row:
                    cell.fill = PatternFill(start_color="626a80", end_color="626a80", fill_type="solid")

                    

            new_sheet['B2']=df_data['Due_Date']
            new_sheet['B3']=df_data['Crn']
            new_sheet['B4']=df_data['ACCOUNT_NUMBER']
            new_sheet['B6']='=Main_Format1!C16'
            new_sheet['B7']='=Main_Format1!C26'
            new_sheet['B10']='=SUM(Main_Format1!C17,Main_Format1!C18,Main_Format1!C19)'
            new_sheet['B17']='=Main_Format1!C27'
            new_sheet['B12']='=Main_Format1!C28'
            new_sheet['B13']='=Main_Format1!C35'
            new_sheet['B8']='=Main_Format1!C28-Main_Format1!C29'
            new_sheet['B9']='=Main_Format1!C28-Main_Format1!C29'
            new_sheet['B16']='=Main_Format1!C44'
            new_sheet['B17']='=Main_Format1!C45'
            new_sheet['B18']='=Main_Format1!C37'
            new_sheet['B21']='=Main_Format1!C40'
            new_sheet['B22']='=Main_Format1!C39'
            new_sheet['B23']='=Main_Format1!C43'
            new_sheet['B50']='=Main_Format1!C23'
            new_sheet['B55']='=Main_Format1!C32'
            new_sheet['B56']='=Main_Format1!C30'
            new_sheet['B57']='=Main_Format1!C22'
            new_sheet['B68']='=Main_Format1!C31'
            new_sheet['B70']='=Main_Format1!C38'
            new_sheet['B71']='=Main_Format1!C37' 

            report_name="Kotak_WBG_10_Output_Mapping_V1.0"


    

        
           



            
  
        workbook.save(f"{output_path}/output.xlsx")
        print("completed sheet2")
       
        with open(f"{output_path}/output.xlsx", 'rb') as f:
            export_data_blob = base64.b64encode(f.read())
        status_qry1=f'UPDATE `process_queue` SET `status`="Report Generated" WHERE `case_id`="{case_id}"'
        queue_db.execute(status_qry1)
        filename = f'{case_id}_{report_name}.xlsx'
        response = {
            'flag': True,
            'data': {
                    'blob': export_data_blob.decode('utf-8'),
                    'file_name': filename
                }
        }
        audit_data = {"tenant_id": "kmb", "user": "", "case_id": case_id, 
                        "api_service": "download_excel", "service_container": "folder_monitor", "changed_data": None,
                        "tables_involved": "","memory_usage_gb": None, 
                        "time_consumed_secs": None, "request_payload": None, 
                        "response_data": None, "trace_id": case_id, "session_id": None,"status":str(response['flag'])}
        insert_into_audit(case_id, audit_data)




        return response
    except Exception:
        logging.exception(f'Something went wrong exporting data')
        return {'flag': False, 'message': 'Unable to export data.'}    
    

@app.route('/dp_formula_calculation', methods=['POST', 'GET'])
def dp_formula_calculation():
    data=request.json
    db_config['tenant_id'] = 'kmb'
    logging.info(f"Request Data is: {data}")
    queue_db = DB('queues', **db_config)
    queue_id=data.get("queue_id",{})
    extraction_db = DB('extraction', **db_config)
    case_id = data.pop('case_id') 
    try:
        qry=f"select `segment` from `ocr` where case_id='{case_id}'"
        segment_res = extraction_db.execute_(qry).to_dict(orient="records")[-1]
        segment = segment_res['segment']
        qry=f"SELECT `STOCK STATEMENT_OCR`,`DEBITORS STATEMENT_OCR`,`CREDITORS_OCR`, `PURCHASES_OCR`, `ADVANCES_OCR`, `SALES_OCR`, `ADVANCES DEBTORS_OCR`, `BANK OS_OCR`, `ADVANCES SUPPLIERS_OCR`,`SECURITY DEPOSIT_OCR` FROM `ocr` WHERE  `case_id`='{case_id}';" 
        json_data_list = extraction_db.execute_(qry).to_dict(orient="records")

        qry1=f"SELECT `case_id` as 'CASE_ID' ,`comments` as Comments, `Due_Date` as 'Stock Due Date',`co_mail_time` as 'Received Date', `cc_limit` as 'CC_LIMIT',`insurance_amount` as 'INSURANCE_AMOUNT' , `Debtors_margin` as 'DEBTORS_MARGIN',`Stock_Margins` as 'STOCK_MARGIN',`deviation_applicable` as 'DEVIATION_APPLICABLE',`Crn` as 'CRN',`customer_name` as 'NAME OF CUSTOMER' ,`request_id` as 'REQUEST_ID' , `Account_Number` as 'Account No',`division` as 'ZMT',`Dp_formula` as 'DP formula',`date` as 'SS_AS_ON_DATE' from ocr where `case_id`='{case_id}';"
        df = extraction_db.execute_(qry1).to_dict(orient="records")[-1]
        received_date_str = df['Received Date']
        stock_due_date=df['Stock Due Date']
        ss_on_date = df['SS_AS_ON_DATE']

        try:
            df['Stock Due Date'] = convert_to_custom_format(stock_due_date)
            df['Received Date'] = convert_to_custom_format(received_date_str)
            df['SS_AS_ON_DATE'] = convert_to_custom_format(ss_on_date)
        except Exception as e:
            print(f'{e}#########exception')

        
        df = pd.DataFrame([df])
        values= create_dataframe_from_json(json_data_list)
        values = values.to_dict()
        print(f"#########values are {values} and type is {type(values)}")
        final_result={}
        try:
            for i in values:
                final_result[i]=values[i][0]
            print(f"######final result {final_result}")
        except:
            logging.info(f"exception in appending final_result")
        final_result = pd.DataFrame([final_result])
        if segment == 'Agri':
            df = df.rename(columns={'Account No': 'ACCOUNT_NUMBER'})
            logging.info(f"df##{df}")
            logging.info(f"in if block segement is {segment}")
            # keys = {"CREDITORS_ADVANCES_RECEIVED_FROM_CUSTOMERS": "Adv from Customers", "DEBTORS_ADVANCES_PAID_TO_SUPPLIER": "Adv to Suppliers", "BC_OS_WITH_KMBL_IN_CASE_OF_SUBLIMIT": "BC OS with KMBL - sublimit", "BG_OS_WITH_KMBL_IN_CASE_OF_SUBLIMIT": "BG OS with KMBL - sublimit",  "DEBTORS_0_TO_5_DAYS": "Debtors 0-5 days", "DEBTORS_121_TO_150_DAYS": "Debtors 121-150 days", "DEBTORS_151_TO_180_DAYS": "Debtors 151-180 days", "DEBTORS_6_TO_60_DAYS": "Debtors 6-60 days", "DEBTORS_61_TO_90_DAYS": "Debtors 61-90 days", "DEBTORS_GREATER_THAN_180_DAYS": "Debtors > 180 days", "LC_OS_WITH_KMBL_IN_CASE_OF_SUBLIMIT": "LC OS with KMBL - sublimit", "OUTSTANDING_WITH_OTHER_WC_BANKERS": "OS with other Banks", "PURCHASE_AMOUNT": "Purchases_AMOUNT", "SALES_AMOUNT": "Sales_AMOUNT", "TOTAL_CREDITORS": "Total Creditors", "TOTAL_STOCK": "Total Stock", "WCDL_OS_WITH_KMBL_IN_CASE_OF_SUBLIMIT": "WCDL OS with KMBL- sublimit"}
            keys={'CREDITORS_ADVANCES_RECEIVED_FROM_CUSTOMERS': 'Adv from Customers', 'DEBTORS_ADVANCES_PAID_TO_SUPPLIER': 'Adv to Suppliers', 'BC_OS_WITH_KMBL_IN_CASE_OF_SUBLIMIT': 'BC OS with KMBL - sublimit', 'BG_OS_WITH_KMBL_IN_CASE_OF_SUBLIMIT': 'BG OS with KMBL - sublimit', 'DEBTORS_0_TO_5_DAYS': 'DEBITORS STATEMENT_OCR_Debtors 0-5 days', 'DEBTORS_121_TO_150_DAYS': 'DEBITORS STATEMENT_OCR_Debtors 121-150 days', 'DEBTORS_151_TO_180_DAYS': 'DEBITORS STATEMENT_OCR_Debtors 151-180 days', 'DEBTORS_6_TO_60_DAYS': 'DEBITORS STATEMENT_OCR_Debtors 6-60 days', 'DEBTORS_61_TO_90_DAYS': 'DEBITORS STATEMENT_OCR_Debtors 61-90 days','DEBTORS_91_TO_120_DAYS':'DEBITORS STATEMENT_OCR_Debtors_91_to_120_days','DEBTORS_GREATER_THAN_180_DAYS': 'DEBITORS STATEMENT_OCR_Debtors > 180 days', 'LC_OS_WITH_KMBL_IN_CASE_OF_SUBLIMIT': 'LC OS with KMBL -sublimit', 'OUTSTANDING_WITH_OTHER_WC_BANKERS': 'OS with other Banks', 'PURCHASE_AMOUNT': 'Purchases_Amount', 'SALES_AMOUNT': 'Sales_Amount', 'TOTAL_CREDITORS':'Total Creditors' , 'TOTAL_STOCK': 'Total Stock', 'WCDL_OS_WITH_KMBL_IN_CASE_OF_SUBLIMIT': 'WCDL OS with KMBL- sublimit'}
            merged_data = {key: final_result.get(keys.get(key), np.nan) for key in keys}
            values_dict = {key: value[0] if isinstance(value, pd.Series) else value for key, value in merged_data.items()}
            result_df = pd.DataFrame([values_dict])
          
            res = pd.DataFrame({key: [] for key in ["CASE_ID","REQUEST_ID", "CRN", "ACCOUNT_NUMBER","SS_AS_ON_DATE","TOTAL_STOCK", "TOTAL_CREDITORS", "DEBTORS_0_TO_5_DAYS", "DEBTORS_6_TO_60_DAYS", "DEBTORS_61_TO_90_DAYS", "DEBTORS_91_TO_120_DAYS", "DEBTORS_121_TO_150_DAYS", "DEBTORS_151_TO_180_DAYS","DEBTORS_GREATER_THAN_180_DAYS", "DEBTORS_ADVANCES_PAID_TO_SUPPLIER", "CREDITORS_ADVANCES_RECEIVED_FROM_CUSTOMERS", "SALES_AMOUNT", "PURCHASE_AMOUNT", "INSURANCE_AMOUNT", "OUTSTANDING_WITH_OTHER_WC_BANKERS","WCDL_OS_WITH_KMBL_IN_CASE_OF_SUBLIMIT", "LC_OS_WITH_KMBL_IN_CASE_OF_SUBLIMIT", "BC_OS_WITH_KMBL_IN_CASE_OF_SUBLIMIT","BG_OS_WITH_KMBL_IN_CASE_OF_SUBLIMIT", "DEBTORS_MARGIN", "STOCK_MARGIN", "OD_LIMIT", "CC_LIMIT", "DEVIATION_APPLICABLE"  ]})
            result_df = pd.concat([result_df, df], axis=1)   
            # for column, values in result_df.items():
            #     res[column] = values[0]  
            res = res.append(result_df, ignore_index=True)


            dp_query=f"SELECT `dp_formula` from ocr where case_id='{case_id}'"
            dp_res = extraction_db.execute_(dp_query).to_dict(orient="records")[-1]
            logging.info(f"{dp_res}")
            if dp_res['dp_formula'] !=  None:
        
                equation_list = re.findall(r'\b\w+_\w+\b', dp_res['dp_formula'] )
                agri_fields = {'TOTAL_STOCKS': 'Total_stock', 'TOTAL_CREDITORS_AMT': 'Total Creditors', 'DD_0_5': 'DEBITORS STATEMENT_OCR_Debtors 0-5 days', 'DD_6_60': 'DEBITORS STATEMENT_OCR_Debtors 6-60 days', 'DD_61_90': 'DEBITORS STATEMENT_OCR_Debtors 61-90 days', 'DD_91_120': 'DEBITORS STATEMENT_OCR_Debtors_91_to_120_days','DD_121_150': 'DEBITORS STATEMENT_OCR_Debtors 121-150 days', 'DD_151_180': 'DEBITORS STATEMENT_OCR_Debtors 151-180 days', 'DD_GREATER_180': 'DEBITORS STATEMENT_OCR_Debtors > 180 days', 'DD_DEBTORS_ADV_PAID_SUPP': 'Adv to Suppliers', 'CRED_ADV_REC_FROM_CUSTOMERS': 'Adv from Customers', 'INSURANCE_AMT': 'Insurance_amount', 'OS_OTHER_WC_BANKERS': 'Outstanding_with_other_wc_bankers', 'WCDL_WITH_KMBL': 'Wcdl_os_with_kmbl_in_case_of_sublimit', 'LC_WITH_KMBL': 'Lc_os_with_kmbl_in_case_of_sublimit', 'BC_WITH_KMBL': 'Bc_os_with_kmbl_in_case_of_sublimit', 'BG_WITH_KMBL': 'Bg_os_with_kmbl_in_case_of_sublimit'}    
                categories = ["TOTAL_STOCKS", "TOTAL_CREDITORS_AMT", "DD_0_5", "DD_6_60", "DD_61_90", "DD_91_120", "DD_121_150", "DD_151_180", "DD_GREATER_180", "DD_DEBTORS_ADV_PAID_SUPP", "CRED_ADV_REC_FROM_CUSTOMERS", "INSURANCE_AMT", "OS_OTHER_WC_BANKERS", "WCDL_WITH_KMBL", "LC_WITH_KMBL", "BC_WITH_KMBL", "BG_WITH_KMBL"]
                non_dp_formula_fields=[]
                for i in categories:
                    if i not in equation_list:
                        non_dp_formula_fields.append(agri_fields[i])
                # data_dict =  {'Adv from Customers': 'CREDITORS_ADVANCES_RECEIVED_FROM_CUSTOMERS', 'Adv to Suppliers': 'DEBTORS_ADVANCES_PAID_TO_SUPPLIER', 'BC OS with KMBL - sublimit': 'BC_OS_WITH_KMBL_IN_CASE_OF_SUBLIMIT', 'BG OS with KMBL - sublimit': 'BG_OS_WITH_KMBL_IN_CASE_OF_SUBLIMIT', 'Debtors 0-5 days': 'DEBTORS_0_TO_5_DAYS', 'Debtors 121-150 days': 'DEBTORS_121_TO_150_DAYS', 'Debtors 151-180 days': 'DEBTORS_151_TO_180_DAYS', 'Debtors 6-60 days': 'DEBTORS_6_TO_60_DAYS', 'Debtors 61-90 days': 'DEBTORS_61_TO_90_DAYS', 'Debtors_91_to_120_days':'DEBTORS_91_TO_120_DAYS','Debtors > 180 days': 'DEBTORS_GREATER_THAN_180_DAYS', 'LC OS with KMBL -sublimit': 'LC_OS_WITH_KMBL_IN_CASE_OF_SUBLIMIT', 'OS with other Banks': 'OUTSTANDING_WITH_OTHER_WC_BANKERS', 'Purchases_AMOUNT': 'PURCHASE_AMOUNT', 'Sales_AMOUNT': 'SALES_AMOUNT', 'Total Creditors': 'TOTAL_CREDITORS', 'Total Stock': 'TOTAL_STOCK', 'WCDL OS with KMBL- sublimit': 'WCDL_OS_WITH_KMBL_IN_CASE_OF_SUBLIMIT'}
                data_dict =  {'Adv from Customers': 'CREDITORS_ADVANCES_RECEIVED_FROM_CUSTOMERS', 'Adv to Suppliers': 'DEBTORS_ADVANCES_PAID_TO_SUPPLIER', 'Bc_os_with_kmbl_in_case_of_sublimit': 'BC_OS_WITH_KMBL_IN_CASE_OF_SUBLIMIT', 'Bg_os_with_kmbl_in_case_of_sublimit': 'BG_OS_WITH_KMBL_IN_CASE_OF_SUBLIMIT', 'DEBITORS STATEMENT_OCR_Debtors 0-5 days': 'DEBTORS_0_TO_5_DAYS', 'DEBITORS STATEMENT_OCR_Debtors 121-150 days': 'DEBTORS_121_TO_150_DAYS', 'DEBITORS STATEMENT_OCR_Debtors 151-180 days': 'DEBTORS_151_TO_180_DAYS', 'DEBITORS STATEMENT_OCR_Debtors 6-60 days': 'DEBTORS_6_TO_60_DAYS', 'DEBITORS STATEMENT_OCR_Debtors 61-90 days': 'DEBTORS_61_TO_90_DAYS', 'DEBITORS STATEMENT_OCR_Debtors_91_to_120_days':'DEBTORS_91_TO_120_DAYS','DEBITORS STATEMENT_OCR_Debtors > 180 days': 'DEBTORS_GREATER_THAN_180_DAYS', 'Lc_os_with_kmbl_in_case_of_sublimit': 'LC_OS_WITH_KMBL_IN_CASE_OF_SUBLIMIT', 'Outstanding_with_other_wc_bankers': 'OUTSTANDING_WITH_OTHER_WC_BANKERS', 'Purchases_Amount': 'PURCHASE_AMOUNT', 'Sales_Amount': 'SALES_AMOUNT', 'Total_creditors': 'TOTAL_CREDITORS', 'Total_stock': 'TOTAL_STOCK', 'Wcdl_os_with_kmbl_in_case_of_sublimit': 'WCDL_OS_WITH_KMBL_IN_CASE_OF_SUBLIMIT', 'Insurance_amount': 'INSURANCE_AMOUNT'}    # data_dict = {'Total_stock': 'TOTAL_STOCK', 'Total_creditors': 'TOTAL_CREDITORS', 'Debtors_0_to_5_days': 'DEBTORS_0_TO_5_DAYS', 'Debtors_6_to_60_days': 'DEBTORS_6_TO_60_DAYS', 'Debtors_61_to_90_days': 'DEBTORS_61_TO_90_DAYS', 'Debtors_91_to_120_days': 'DEBTORS_91_TO_120_DAYS', 'Debtors_121_to_150_days': 'DEBTORS_121_TO_150_DAYS', 'Debtors_151_to_180_days': 'DEBTORS_151_TO_180_DAYS', 'Debtors_greater_than_180_days': 'DEBTORS_GREATER_THAN_180_DAYS', 'Debtors_advances_paid_to_supplier': 'DEBTORS_ADVANCES_PAID_TO_SUPPLIER', 'Creditors_advances_received_from_customers': 'CREDITORS_ADVANCES_RECEIVED_FROM_CUSTOMERS', 'Insurance_amount': 'INSURANCE_AMOUNT', 'Outstanding_with_other_wc_bankers': 'OUTSTANDING_WITH_OTHER_WC_BANKERS', 'Wcdl_os_with_kmbl_in_case_of_sublimit': 'WCDL_OS_WITH_KMBL_IN_CASE_OF_SUBLIMIT', 'Lc_os_with_kmbl_in_case_of_sublimit': 'LC_OS_WITH_KMBL_IN_CASE_OF_SUBLIMIT', 'Bc_os_with_kmbl_in_case_of_sublimit': 'BC_OS_WITH_KMBL_IN_CASE_OF_SUBLIMIT', 'Bg_os_with_kmbl_in_case_of_sublimit': 'BG_OS_WITH_KMBL_IN_CASE_OF_SUBLIMIT'}

                # data_dict = {'Total_stock': 'TOTAL_STOCK', 'Total_creditors': 'TOTAL_CREDITORS', 'Debtors_0_to_5_days': 'DEBTORS_0_TO_5_DAYS', 'Debtors_6_to_60_days': 'DEBTORS_6_TO_60_DAYS', 'Debtors_61_to_90_days': 'DEBTORS_61_TO_90_DAYS', 'Debtors_91_to_120_days': 'DEBTORS_91_TO_120_DAYS', 'Debtors_121_to_150_days': 'DEBTORS_121_TO_150_DAYS', 'Debtors_151_to_180_days': 'DEBTORS_151_TO_180_DAYS', 'Debtors_greater_than_180_days': 'DEBTORS_GREATER_THAN_180_DAYS', 'Debtors_advances_paid_to_supplier': 'DEBTORS_ADVANCES_PAID_TO_SUPPLIER', 'Creditors_advances_received_from_customers': 'CREDITORS_ADVANCES_RECEIVED_FROM_CUSTOMERS', 'Insurance_amount': 'INSURANCE_AMOUNT', 'Outstanding_with_other_wc_bankers': 'OUTSTANDING_WITH_OTHER_WC_BANKERS', 'Wcdl_os_with_kmbl_in_case_of_sublimit': 'WCDL_OS_WITH_KMBL_IN_CASE_OF_SUBLIMIT', 'Lc_os_with_kmbl_in_case_of_sublimit': 'LC_OS_WITH_KMBL_IN_CASE_OF_SUBLIMIT', 'Bc_os_with_kmbl_in_case_of_sublimit': 'BC_OS_WITH_KMBL_IN_CASE_OF_SUBLIMIT', 'Bg_os_with_kmbl_in_case_of_sublimit': 'BG_OS_WITH_KMBL_IN_CASE_OF_SUBLIMIT'}
                for field in non_dp_formula_fields:
                    
                    res[data_dict[field]] = None
                logging.info(f"result dataframe####  {res}")
            res=res.to_dict(orient='list')
            logging.info(f"result dataframe####  {res}")
            query=f"update `ocr` set `dp_formula_json` = '{json.dumps(res)}' where `case_id` ='{case_id}'"
            extraction_db.execute_(query)
            logging.info(f"result dataframe####  {res}")
         
            queue_name="agri"
        elif segment == 'Consumer':
            logging.info(f"in if block segement is {segment}")
            
            res = pd.DataFrame({key: [] for key in['CASE_ID','CRN', 'NAME OF CUSTOMER', 'SS_AS_ON_DATE', 'Raw Material-< 90 DAYS', 'Raw Material-90 - 120 DAYS', 'Raw Material-120 - 180 DAYS', 'Raw Material-> 180 DAYS', 'Raw Material-Unit/Quantity', 'Work In Progress-< 60 DAYS', 'Work In Progress 60- 90 DAYS', 'Work In Progress-90 - 120 DAYS', 'Work In Progress- 120 - 180 DAYS', 'Work In Progress-> 180 DAYS', 'Work In Progress-Unit/Quantity', 'Consumables & Packing Material-< 90 DAYS', 'Consumables & Packing Material-90 - 180 DAYS', 'Consumables & Packing Material-> 180 DAYS', 'Consumables & Packing Material-Unit/Quantity', 'Finished Goods-< 90 DAYS', 'Finished Goods-90 - 120 DAYS', 'Finished Goods-120 - 180 DAYS', 'Work In Progress- 90 - 180 DAYS','Finished Goods-> 180 DAYS', 'Finished Goods-Unit/Quantity', 'Vehicle stock upto 120 days', 'Vehicle stock 120-150 days', 'Spares stock 6-60 DAYS', 'Spares stock < 60 DAYS', 'Total Spares stock', 'Total Stock <30 DAYS', 'Total Stock 30-90 DAYS', 'Total Stock-90 - 120 DAYS', 'Total Stock 120- 180 DAYS', 'Total Stock-> 180 DAYS', 'Total Stock-Unit/Quantity', 'Third party Debtors-< 90 days', 'Third party Debtors-90 - 180 days', 'Third party Debtors-> 180 days', 'Third party Debtors-No. of debtors', 'Spare Debtors less than 60 days', 'Spare Debtors 60-90 days', 'Vehicle debtors', 'Vehicle debtors 6 - 60 days', 'Related party/ group concern Debtors-< 90 days', 'Related party/ group concern Debtors-90 - 180 days', 'Related party/ group concern Debtors-> 180 days', 'Related party/ group concern Debtors-No. of debtors', 'Total debtors Debtors-< 30 days', 'Total debtors Debtors 30- 60 days', 'Total debtors Debtors 60- 90 days', 'Total debtors Debtors-90 - 120 days', 'Total debtors Debtors-120 - 180 days', 'Total debtors Debtors-> 180 days', 'Total debtors Debtors-No. of debtors', 'LC Backed Creditors Amount', 'LC Backed Creditors No of Creditors', 'Other than LC backed Creditors Amount', 'Other than LC backed Creditors No of Creditors', 'spares creditors', 'Total Creditors Amount', 'Total Creditors No of creditors', 'Sales Amount', 'Sales Unit', 'Purchases Amount', 'Purchases Unit', 'Other Channel finance details Bank Name', 'Other Channel details Value', 'Customer Advance', 'Supplier Advance', 'Stock Due Date', 'Outstanding in Other Limits', 'File Name', 'Received Date', 'Account No', 'ZMT', 'DP formula','Comments']})

            keys = {'CASE_ID':'CASE_ID','CRN': 'CRN', 'NAME OF CUSTOMER': 'NAME OF CUSTOMER', 'SS_AS_ON_DATE': 'SS_AS_ON_DATE', 'Raw Material-< 90 DAYS': 'IMPORTED RAW MATERIAL_<90', 'Raw Material-90 - 120 DAYS': 'IMPORTED RAW MATERIAL_<120', 'Raw Material-120 - 180 DAYS': 'IMPORTED RAW MATERIAL_<180', 'Raw Material-> 180 DAYS': 'IMPORTED RAW MATERIAL_>180', 'Raw Material-Unit/Quantity': 'IMPORTED RAW MATERIAL_Units', 'Work In Progress-< 60 DAYS': 'WORK IN PROGRESS_<60', 'Work In Progress 60- 90 DAYS': 'WORK IN PROGRESS_<90', 'Work In Progress-90 - 120 DAYS': 'WORK IN PROGRESS_<120', 'Work In Progress- 120 - 180 DAYS': 'WORK IN PROGRESS_<180', 'Work In Progress-> 180 DAYS': 'WORK IN PROGRESS_>180','Work In Progress- 90 - 180 DAYS':'WORK IN PROGRESS_<180','Work In Progress-Unit/Quantity': 'WORK IN PROGRESS_Units', 'Consumables & Packing Material-< 90 DAYS': 'CONSUMABLES_<90', 'Consumables & Packing Material-90 - 180 DAYS': 'CONSUMABLES_<180', 'Consumables & Packing Material-> 180 DAYS': 'CONSUMABLES_>180', 'Consumables & Packing Material-Unit/Quantity': 'CONSUMABLES_Units', 'Finished Goods-< 90 DAYS': 'Finished Goods_<120', 'Finished Goods-90 - 120 DAYS': 'Finished Goods_<120', 'Finished Goods-120 - 180 DAYS': 'Finished Goods_<180', 'Finished Goods-> 180 DAYS': 'Finished Goods_>180', 'Finished Goods-Unit/Quantity': 'Finished Goods_Units', 'Total Stock <30 DAYS': 'TOTAL_<30', 'Total Stock 30-90 DAYS': 'TOTAL_<90', 'Total Stock 120- 180 DAYS': 'TOTAL_<180',  'Total Stock-90 - 120 DAYS': 'TOTAL_<120', 'Total Stock-> 180 DAYS': 'TOTAL_>180', 'Total Stock-Unit/Quantity': 'TOTAL_Units','Third party Debtors-< 90 days': 'DEBITORS STATEMENT_OCR_Third Party debtors_<90', 'Third party Debtors-90 - 180 days': 'DEBITORS STATEMENT_OCR_Third Party debtors_<180', 'Third party Debtors-> 180 days': 'DEBITORS STATEMENT_OCR_Third Party debtors_>180', 'Third party Debtors-No. of debtors': 'DEBITORS STATEMENT_OCR_Third Party debtors_No. of debtors', 'Related party/ group concern Debtors-< 90 days': 'DEBITORS STATEMENT_OCR_Related party/group concern Debtors_< 90 days', 'Related party/ group concern Debtors-90 - 180 days': 'DEBITORS STATEMENT_OCR_Related party/group concern Debtors_< 180 days', 'Related party/ group concern Debtors-> 180 days': 'DEBITORS STATEMENT_OCR_Related party/group concern Debtors_> 180 days', 'Related party/ group concern Debtors-No. of debtors': 'DEBITORS STATEMENT_OCR_Related party/group concern Debtors_No. of debtors', 'Total debtors Debtors-< 30 days': 'DEBITORS STATEMENT_OCR_TOTAL_<30', 'Total debtors Debtors 30- 60 days': 'DEBITORS STATEMENT_OCR_TOTAL_<60', 'Total debtors Debtors 60- 90 days': 'DEBITORS STATEMENT_OCR_TOTAL_<90', 'Total debtors Debtors-90 - 120 days': 'DEBITORS STATEMENT_OCR_TOTAL_<120', 'Total debtors Debtors-120 - 180 days': 'DEBITORS STATEMENT_OCR_TOTAL_<180', 'Total debtors Debtors-> 180 days': 'DEBITORS STATEMENT_OCR_TOTAL_>180', 'Total debtors Debtors-No. of debtors': 'TOTAL_No. of debtors','Other than LC backed Creditors Amount': 'OTHER LC BACKED CREDITORS_AMOUNT', 'Total Creditors Amount': 'Total Creditors', 'Total Creditors No of creditors': 'TOTAL_No. of creditors', 'Sales Amount': 'Sales_Amount', 'Sales Unit': 'Sales_Units', 'Purchases Amount': 'Purchases_Amount', 'Purchases Unit': 'Purchases_Units', 'Other Channel finance details Bank Name': 'Other Channel finance Bank Name', 'Customer Advance': 'Customer Advance', 'Supplier Advance': 'Supplier Advance', 'Stock Due Date': 'Stock Due Date', 'Outstanding in Other Limits': 'Outstanding in Other Limits', 'Received Date': 'Received Date', 'Account No': 'ACCOUNT_NUMBER', 'ZMT': 'ZMT', 'DP formula': 'DP formula',"LC Backed Creditors Amount":"LC Backed Creditors_AMOUNT","LC Backed Creditors No of Creditors":"LC Backed Creditors_No. of creditors","Other than LC backed Creditors No of Creditors":"TOTAL_No. of creditors"}

            
            merged_data = {key: final_result.get(keys.get(key), np.nan) for key in keys}
            result_df = pd.DataFrame([merged_data])
            result_df = pd.DataFrame([merged_data])


            result_df = pd.concat([result_df, df], axis=1)
            logging.info(f"#####result dataframe {result_df}")
            for column, values in result_df.items():
                res[column] = values[0]
            others_query=f"SELECT `Sales_ocr`,`purchases_ocr`,`advances_ocr` from ocr where case_id='{case_id}'"
            others_res = extraction_db.execute_(others_query).to_dict(orient="records")[0]
            logging.info(others_res,'##############others_res')
            #sales_keys_set = set(json.loads(others_res['Sales_ocr']).keys())
            #purchases_keys_set = set(json.loads(others_res['purchases_ocr']).keys())
            #advances_keys_set = set(json.loads(others_res['advances_ocr']).keys())
            try:
                sales_keys_set = set(json.loads(others_res['Sales_ocr']).keys())
            except (KeyError, json.JSONDecodeError):
                sales_keys_set = set()

            try:
                purchases_keys_set = set(json.loads(others_res['purchases_ocr']).keys())
            except (KeyError, json.JSONDecodeError):
                purchases_keys_set = set()
            try:
                advances_keys_set = set(json.loads(others_res['advances_ocr']).keys())
            except (KeyError, json.JSONDecodeError):
                advances_keys_set = set()
            # Merging sets
            merged_keys_set = sales_keys_set.union(purchases_keys_set, advances_keys_set)

            # Converting the set to a list
            merged_keys_list = list(merged_keys_set)

            # Printing merged keys
            # print("Merged Keys:", merged_keys_list)
            result_keys_others = [key for key, value in keys.items() if value in merged_keys_list]
            logging.info(result_keys_others,'#####################result keys')

            dp_query=f"SELECT `dp_formula` from ocr where case_id='{case_id}'"
            dp_res = extraction_db.execute_(dp_query).to_dict(orient="records")[-1]
            if dp_res['dp_formula'] != None:
                equation = dp_res['dp_formula'] .replace('-',' ')
                equation =''.join(equation.split('(')[1:])
                equation_list1 = equation.split('+')
                equation_list = []
                for term in equation_list1:
                    terms_with_spaces = term.split()
                    processed_terms = [''.join(char for char in t if char.isalnum() or char == '_') for t in terms_with_spaces]
                    processed_terms = list(filter(None, processed_terms))
                    equation_list.extend(processed_terms)
                logging.info(f"equation_list##{equation_list}")
                consumer_field = {'TS_120_180':'TOTAL_<180','RM_LESS_90': 'IMPORTED RAW MATERIAL_<90', 'RM_120_180':'IMPORTED RAW MATERIAL_<180','RM_90_120': 'IMPORTED RAW MATERIAL_<120', 'RM_90_180': 'IMPORTED RAW MATERIAL_<180', 'RM_GREATER_180': 'IMPORTED RAW MATERIAL_>180', 'WIP_LESS_60': 'WORK IN PROGRESS_<60', 'WIP_60_90': 'WORK IN PROGRESS_<90', 'WIP_90_120': 'WORK IN PROGRESS_<120', 'WIP_LESS_90': 'WORK IN PROGRESS_<90', 'WIP_90_180': 'WORK IN PROGRESS_<180', 'WIP_120_180':'WORK IN PROGRESS_<180','WIP_GREATER_180': 'WORK IN PROGRESS_>180', 'CPM_LESS_90': 'CONSUMABLES_<90', 'CPM_90_180': 'CONSUMABLES_<180', 'CPM_GREATER_180': 'CONSUMABLES_>180', 'FG_LESS_90': 'Finished Goods_<120', 'FG_90_120': 'Finished Goods_<120', 'FG_120_180': 'Finished Goods_<180', 'FG_GREATER_180': 'Finished Goods_>180', 'TS_LESS_30': 'TOTAL_<30', 'TS_30_90': 'TOTAL_<90', 'TS_LESS_90': 'TOTAL_<90', 'TS_GREATER_180': 'TOTAL_>180', 'TPD_LESS_90': 'DEBITORS STATEMENT_OCR_Third Party debtors_<90','TPD_GREATER_90':'DEBITORS STATEMENT_OCR_Third Party debtors_<180', 'TPD_90_180': 'DEBITORS STATEMENT_OCR_Third Party debtors_<180', 'RP_GCD_LESS_90': 'DEBITORS STATEMENT_OCR_Related party/group concern Debtors_< 90 days', 'RP_GCD_90_180': 'DEBITORS STATEMENT_OCR_Related party/group concern Debtors_< 180 days', 'RP_GCD_GREATER_180': 'DEBITORS STATEMENT_OCR_Related party/group concern Debtors_> 180 days', 'TDD_LESS_30': 'DEBITORS STATEMENT_OCR_TOTAL_<30','TDD_30_60': 'DEBITORS STATEMENT_OCR_TOTAL_<60', 'TDD_60_90': 'DEBITORS STATEMENT_OCR_TOTAL_<90', 'TDD_90_120': 'DEBITORS STATEMENT_OCR_TOTAL_<120', 'TDD_GREATER_120': 'DEBITORS STATEMENT_OCR_TOTAL_<180', 'TDD_120_180': 'DEBITORS STATEMENT_OCR_TOTAL_<180', 'TDD_GREATER_180': 'DEBITORS STATEMENT_OCR_TOTAL_>180', 'TOTAL_CREDITORS_AMT': 'Total Creditors', 'CUSTOMER_ADVANCE': 'Customer Advance', 'SUPPLIER_ADVANCE': 'Supplier Advance'}
                categories=['RM_LESS_90', 'RM_90_120', 'RM_120_180', 'RM_GREATER_180', 'WIP_LESS_60', 'WIP_60_90', 'WIP_90_120', 'WIP_120_180', 'WIP_GREATER_180', 'CPM_LESS_90', 'CPM_90_180', 'CPM_GREATER_180', 'FG_LESS_90', 'FG_90_120', 'FG_120_180', 'FG_GREATER_180', 'TS_LESS_30', 'TS_30_90', 'TS_120_180', 'TS_GREATER_180', 'TPD_LESS_90', 'TPD_90_180', 'TPD_GREATER_90', 'RP_GCD_LESS_90', 'RP_GCD_90_180', 'RP_GCD_GREATER_180', 'TDD_LESS_30', 'TDD_30_60', 'TDD_60_90', 'TDD_90_120', 'TDD_120_180', 'TDD_GREATER_180', 'TOTAL_CREDITORS_AMT', 'CUSTOMER_ADVANCE', 'SUPPLIER_ADVANCE']

                non_dp_formula_fields=[]
                for i in categories:
                    if i not in equation_list:
                        non_dp_formula_fields.append(consumer_field[i])
                data_dict= {'IMPORTED RAW MATERIAL_<90': 'Raw Material-< 90 DAYS','IMPORTED RAW MATERIAL_<120': 'Raw Material-90 - 120 DAYS', 'IMPORTED RAW MATERIAL_<180': 'Raw Material-120 - 180 DAYS', 'IMPORTED RAW MATERIAL_>180': 'Raw Material-> 180 DAYS', 'WORK IN PROGRESS_<90': 'Work In Progress 60- 90 DAYS', 'WORK IN PROGRESS_<60': 'Work In Progress-< 60 DAYS', 'WORK IN PROGRESS_<90': 'Work In Progress 60- 90 DAYS', 'WORK IN PROGRESS_<120': 'Work In Progress-90 - 120 DAYS', 'WORK IN PROGRESS_<180': 'Work In Progress- 120 - 180 DAYS', 'WORK IN PROGRESS_>180': 'Work In Progress-> 180 DAYS', 'CONSUMABLES_<90': 'Consumables & Packing Material-< 90 DAYS', 'CONSUMABLES_<180': 'Consumables & Packing Material-90 - 180 DAYS', 'CONSUMABLES_>180': 'Consumables & Packing Material-> 180 DAYS', 'Finished Goods_<120': 'Finished Goods-< 90 DAYS', 'Finished Goods_<120': 'Finished Goods-90 - 120 DAYS', 'Finished Goods_<180': 'Finished Goods-120 - 180 DAYS', 'Finished Goods_>180': 'Finished Goods-> 180 DAYS', 'TOTAL_<30': 'Total Stock <30 DAYS', 'TOTAL_<90': 'Total Stock 30-90 DAYS',  'TOTAL_<180': 'Total Stock 120- 180 DAYS', 'TOTAL_>180': 'Total Stock-> 180 DAYS', 'DEBITORS STATEMENT_OCR_Third Party debtors_<90': 'Third party Debtors-< 90 days', 'DEBITORS STATEMENT_OCR_Third Party debtors_<180': 'Third party Debtors-90 - 180 days', 'DEBITORS STATEMENT_OCR_Related party/group concern Debtors_< 90 days': 'Related party/ group concern Debtors-< 90 days', 'DEBITORS STATEMENT_OCR_Related party/group concern Debtors_< 180 days': 'Related party/ group concern Debtors-90 - 180 days', 'DEBITORS STATEMENT_OCR_Related party/group concern Debtors_> 180 days': 'Related party/ group concern Debtors-> 180 days', 'DEBITORS STATEMENT_OCR_TOTAL_<30': 'Total debtors Debtors-< 30 days', 'DEBITORS STATEMENT_OCR_TOTAL_<60': 'Total debtors Debtors 30- 60 days', 'DEBITORS STATEMENT_OCR_TOTAL_<90': 'Total debtors Debtors 60- 90 days', 'DEBITORS STATEMENT_OCR_TOTAL_<120': 'Total debtors Debtors-90 - 120 days', 'DEBITORS STATEMENT_OCR_TOTAL_<180': 'Total debtors Debtors-120 - 180 days', 'DEBITORS STATEMENT_OCR_TOTAL_>180':'Total debtors Debtors-> 180 days', 'Total Creditors': 'Total Creditors Amount', 'Customer Advance': 'Customer Advance', 'Supplier Advance': 'Supplier Advance'}

                for field in non_dp_formula_fields: 
                                       
                    if 'OTHERS' in equation_list:
                        for M in result_keys_others:
                            if 'Unit' in M:
                                print(M,'###########units')
                                res[M]=None
                                
                            print(M,'others tab fields')
                            if M not in ('Sales Amount','Purchases Amount' ,'Outstanding in Other Limits' ):
                                res[M]=None                                        
                    else:                        
                        for f_other in result_keys_others:
                            print(f_other,'#########f_other')
                            if f_other not in ('Sales Amount','Purchases Amount'):
                                res[f_other]=None
                    res[data_dict[field]] = None
                    res['LC Backed Creditors No of Creditors']=None
                    res['Total Creditors No of creditors']=None
                    # res['Stock Due Date']='######'
                    res['Third party Debtors-No. of debtors']=None
            for i in res:
                if (i not in data_dict.values()) and (i not in df) and (i not in result_keys_others):
                    res[i]=None
            res = res.drop(columns=['CC_LIMIT','INSURANCE_AMOUNT','DEBTORS_MARGIN','STOCK_MARGIN','DEVIATION_APPLICABLE','REQUEST_ID'])
            logging.info(f"###result is {res}")
           
            
            res=res.to_dict(orient='list')
            logging.info(f"result dataframe####  {res}")
            query=f"update `ocr` set `dp_formula_json` = '{json.dumps(res)}' where `case_id` ='{case_id}'"
            extraction_db.execute_(query)
            logging.info(f"result dataframe####  {res}")
            queue_name="consumer"

        
        if segment == 'RBG':
            qry1=f"SELECT `OUTSTANDING_IN_SPG_CF_INF_TA` as 'OUTSTANDING IN SPG CF INF TA',`Computed_DP` as 'Computed DP',`system_dp` as 'System DP', `Sanction_Limit` as 'Sanction Limit',`Final_DP` as 'Final DP', `Due_Date` as 'Stock Due Date',`co_mail_time` as 'Received Date', `cc_limit` as 'CC_LIMIT',`insurance_amount` as 'INSURANCE_AMOUNT' , `Debtors_margin` as 'DEBTORS_MARGIN',`Stock_Margins` as 'STOCK_MARGIN',`deviation_applicable` as 'DEVIATION APPLICABLE',`Crn` as 'CRN',`customer_name` as 'NAME OF CUSTOMER' ,`request_id` as 'REQUEST ID' , `Account_Number` as 'ACCOUNT NUMBER',`division` as 'ZMT',`Dp_formula` as 'DP Formula',`date` as 'STOCK STATEMENT MONTH AS ON DATE' from ocr where `case_id` = '{case_id}';"
            df = extraction_db.execute_(qry1).to_dict(orient="records")[-1]
            received_date_str = df['Received Date']
            stock_due_date=df['Stock Due Date']
            ss_on_date = df['STOCK STATEMENT MONTH AS ON DATE']

            try:
                
                df['Stock Due Date'] = convert_to_custom_format(stock_due_date)
            
                df['Received Date'] = convert_to_custom_format(received_date_str)
                df['STOCK STATEMENT MONTH AS ON DATE'] = convert_to_custom_format(ss_on_date)

            except Exception as e:
                print(f'{e}#########exception')
        

        
            df = pd.DataFrame([df])
            logging.info(f"in if block segement is {segment}")
            keys={'Last 3 month purchases':'Last 3 Month Purchases','Last 3 month Sales':'Last 3 Month Sales','STOCKS IMPORTED RM 0 5 DAYS': 'Stocks Imported RM 0-5 days', 'STOCKS IMPORTED RM 6 30 DAYS': 'Stocks Imported RM 6-30 days', 'STOCKS IMPORTED RM 31 60 DAYS': 'Stocks Imported RM 31-60 days', 'STOCKS IMPORTED RM 61 90 DAYS': 'Stocks Imported RM 61-90 days', 'STOCKS IMPORTED RM 91 120 DAYS': 'Stocks Imported RM 91-120 days', 'STOCKS IMPORTED RM 121 150 DAYS': 'Stocks Imported 121-150 days', 'STOCKS IMPORTED RM 151 180 DAYS': 'Stocks Imported RM 151-180 days', 'STOCKS IMPORTED RM GREATER THAN180 DAYS': 'Stocks Imported RM >180 days', 'STOCKS WIP 0 5 DAYS': 'Stocks WIP <5 days', 'STOCKS WIP 6 30 DAYS': 'Stocks WIP <30 days', 'STOCKS WIP 31 60 DAYS': 'Stocks WIP <60 days', 'STOCKS WIP 61 90 DAYS': 'Stocks WIP 61-90 days', 'STOCKS WIP 91 120 DAYS': 'Stocks WIP 91-120 days', 'STOCKS WIP 121 150 DAYS': 'Stocks WIP 121-150 days', 'STOCKS WIP 151 180 DAYS': 'Stocks WIP 151-180 days', 'STOCKS WIP GREATER THAN180 DAYS': 'Stocks WIP >180 days', 'STOCKS VEHICLE STOCKS 0 5 DAYS': 'Stocks Vehicle Stocks 0-5 days', 'STOCKS VEHICLE STOCKS 6 30 DAYS': 'Stocks Vehicle Stocks 6-30 days', 'STOCKS VEHICLE STOCKS 31 60 DAYS': 'Stocks Vehicle Stocks 31-60 days', 'STOCKS VEHICLE STOCKS 61 90 DAYS': 'Stocks Vehicle Stocks 61-90 days', 'STOCKS VEHICLE STOCKS 91 120 DAYS': 'Stocks Vehicle Stocks 91-120 days', 'STOCKS VEHICLE STOCKS 121 150 DAYS': 'Stocks Vehicle Stocks 121-150 days', 'STOCKS VEHICLE STOCKS 151 180 DAYS': 'Stocks Vehicle Stocks 151-180 days', 'STOCKS VEHICLE STOCKS GREATER THAN180 DAYS': 'Stocks Vehicle Stocks >180 days', 'STOCKS SPARES STOCKS 0 5 DAYS': 'Stocks Spares Stocks 0-5 days', 'STOCKS SPARES STOCKS 6 30 DAYS': 'Stocks Spares Stocks 6-30 days', 'STOCKS SPARES STOCKS 31 60 DAYS': 'Stocks Spares Stocks 31-60 days', 'STOCKS SPARES STOCKS 61 90 DAYS': 'Stocks Spares Stocks 61-90 days', 'STOCKS SPARES STOCKS 91 120 DAYS': 'Stocks Spares Stocks 91-120 days', 'STOCKS SPARES STOCKS 121 150 DAYS': 'Stocks Spares Stocks 121-150 days', 'STOCKS SPARES STOCKS 151 180 DAYS': 'Stocks Spares Stocks 151 180 days', 'STOCKS SPARES STOCKS GREATER THAN180 DAYS': 'Stocks Spares Stocks >180 days', 'STOCKS FG TRADING 0 5 DAYS': 'Stocks FG Trading 0-5 days', 'STOCKS FG TRADING 6 30 DAYS': 'Stocks FG Trading <30 days', 'STOCKS FG TRADING 31 60 DAYS': 'Stocks FG Trading <60 days', 'STOCKS FG TRADING 61 90 DAYS': 'Stocks FG Trading <90 days', 'STOCKS FG TRADING 91 120 DAYS': 'Stocks FG Trading 91-120 days', 'STOCKS FG TRADING 121 150 DAYS': 'Stocks FG Trading 121-150 days', 'STOCKS FG TRADING 151 180 DAYS': 'Stocks FG Trading 151-180 days', 'STOCKS FG TRADING GREATER THAN180 DAYS': 'Stocks FG Trading >180 days', 'STOCKS CONSUMABLES PACKING MATERIAL': 'Stocks Consumables Packing Material', 'STOCKS STOCK IN TRANSIT': 'Stocks Stock In Transit', 'STOCKS RM LESS THAN 90 DAYS': 'Stocks RM LESS THAN 90 days', 'STOCKS TOTAL SPARES STOCK': 'Stocks Total Spares Stock', 'STOCK LESS THAN 60 DAYS': 'Total Stocks <60 days', 'STOCK LESS THAN 90 DAYS': 'Total Stocks <90 days', 'STOCK LESS THAN 120 DAYS': 'Total Stocks <120 days', 'STOCK LESS THAN 180 DAYS': 'Total Stocks <180 days', 'STOCKS WIP LESS THAN 180 DAYS': 'Stocks WIP LESS THAN 180 days', 'TOTAL STOCKS': 'Total Stocks', 'DEBTORS 0 5 DAYS': 'DEBITORS STATEMENT_OCR_Total Debtors <5 days', 'DEBTORS 6 30 DAYS': 'DEBITORS STATEMENT_OCR_Total Debtors <30 days', 'DEBTORS 31 60 DAYS': 'DEBITORS STATEMENT_OCR_Total Debtors <60 days', 'DEBTORS 61 90 DAYS': 'DEBITORS STATEMENT_OCR_Total Debtors <90 days', 'DEBTORS 91 120 DAYS': 'DEBITORS STATEMENT_OCR_Total Debtors <120 days', 'DEBTORS 121 150 DAYS': 'DEBITORS STATEMENT_OCR_Debtors 121 150 days', 'DEBTORS 151 180 DAYS': 'DEBITORS STATEMENT_OCR_Total Debtors <180 days', 'TOTAL DEBTORS': 'DEBITORS STATEMENT_OCR_Total Debtors', 'DEBTORS DEDUCTION INTERGROUP RECEIVABLES 1': 'DEBITORS STATEMENT_OCR_Debtors Deduction Intergroup Receivables 1', 'DEBTORS DEDUCTION INTERGROUP RECEIVABLES 2': 'DEBITORS STATEMENT_OCR_Debtors Deduction Intergroup Receivables 2', 'DEBTORS LESS THAN 30 DAYS': 'DEBITORS STATEMENT_OCR_DEBITORS STATEMENT_OCR_Total Debtors <30 days', 'DEBTORS LESS THAN 60 DAYS':'DEBITORS STATEMENT_OCR_Total Debtors <60 days', 'DEBTORS LESS THAN 90 DAYS': 'DEBITORS STATEMENT_OCR_Total Debtors <90 days', 'DEBTORS GREATER THAN 90 DAYS': 'DEBITORS STATEMENT_OCR_Total Debtors <120 days', 'DEBTORS GREATER THAN 90 LESS THAN 150 DAYS': 'DEBITORS STATEMENT_OCR_Total Debtors <150 days', 'DEBTORS LESS THAN 120 DAYS': 'DEBITORS STATEMENT_OCR_Total Debtors <120 days', 'DEBTORS GREATER THAN 120 LESS THAN 150 DAYS': 'DEBITORS STATEMENT_OCR_Debtors 121 150 days', 'DEBTORS LESS THAN 180 DAYS': 'DEBITORS STATEMENT_OCR_Total Debtors <180 days', 'TOTAL DEBTORS': 'DEBITORS STATEMENT_OCR_Total Debtors', 'DEBTORS UNBILLED DEBTORS LESS THAN60 DAYS': 'DEBITORS STATEMENT_OCR_Debtors Unbilled Debtors Less Than60 days', 'DEBTORS UNBILLED DEBTORS LESS THAN90DAYS': 'DEBITORS STATEMENT_OCR_Total Debtors_Debtors Unbilled Debtors Less Than90days', 'DEBTORS TOTAL UNBILLED DEBTORS': 'DEBITORS STATEMENT_OCR_Total Debtors_Debtors Total Unbilled Debtors', 'CREDITORS UPTO 90 DAYS': 'Creditors Upto 90 days', 'CREDITORS UPTO 91 150 DAYS': 'Creditors Upto 91 150 days', 'CREDITORS GREATER THAN151 DAYS': 'Creditors Greater Than151 days', 'TOTAL CREDITORS': 'Total Creditors', 'DEBTORS ADVANCE PAID TO SUPPLIER': 'DEBITORS STATEMENT_OCR_Total Debtors_Debtors Advance Paid To Supplier', 'SECURITY DEPOSIT': 'Security Deposit', 'CREDITORS ADVANCES RECEIVED FROM CUSTOMERS': 'Creditors Advances Received From Customers', 'DEBTORS DEDUCTION BILLS DISCOUNTED': 'DEBITORS STATEMENT_OCR_Total Debtors_Debtors Deduction Bills Discounted', 'DEBTORS DEDUCTION BILLS FACTORED ASSIGNED': 'DEBITORS STATEMENT_OCR_Total Debtors_Debtors Deduction Bills Factores Assigned', 'LIMITS OS WITH KMBL BANKS 1': 'Limits OS With KMBL Banks 1', 'LIMITS OS WITH KMBL BANKS 2': 'Limits OS With KMBL Banks 2', 'LIMITS OS WITH KMBL BANKS 3': 'Limits OS With KMBL Banks 3', 'LIMITS OS WITH OTHER BANKS 1': 'Limits OS With Other Banks 1', 'LIMITS OS WITH OTHER BANKS 2': 'Limits OS With Other Banks 2', 'LIMITS OS WITH OTHER BANKS 3': 'Limits OS With Other Banks 3', 'LIMITS OS WITH OTHER BANKS 4': 'Limits OS With Other Banks 4', 'LIMITS OS WITH OTHER BANKS 5': 'Limits OS With Other Banks 5'}
            merged_data = {key: final_result.get(keys.get(key), np.nan) for key in keys}
            values_dict = {key: value[0] if isinstance(value, pd.Series) else value for key, value in merged_data.items()}
            result_df = pd.DataFrame([values_dict])
            print(result_df,'result_df12-------------')
            columns=['CASE_ID','REQUEST ID', 'CRN', 'ACCOUNT NUMBER', 'STOCK STATEMENT MONTH AS ON DATE', 'STOCKS IMPORTED RM 0 5 DAYS', 'STOCKS IMPORTED RM 6 30 DAYS', 'STOCKS IMPORTED RM 31 60 DAYS', 'STOCKS IMPORTED RM 61 90 DAYS', 'STOCKS IMPORTED RM 91 120 DAYS', 'STOCKS IMPORTED RM 121 150 DAYS', 'STOCKS IMPORTED RM 151 180 DAYS', 'STOCKS IMPORTED RM GREATER THAN180 DAYS', 'STOCKS WIP 0 5 DAYS', 'STOCKS WIP 6 30 DAYS', 'STOCKS WIP 31 60 DAYS', 'STOCKS WIP 61 90 DAYS', 'STOCKS WIP 91 120 DAYS', 'STOCKS WIP 121 150 DAYS', 'STOCKS WIP 151 180 DAYS', 'STOCKS WIP GREATER THAN180 DAYS', 'STOCKS VEHICLE STOCKS 0 5 DAYS', 'STOCKS VEHICLE STOCKS 6 30 DAYS', 'STOCKS VEHICLE STOCKS 31 60 DAYS', 'STOCKS VEHICLE STOCKS 61 90 DAYS', 'STOCKS VEHICLE STOCKS 91 120 DAYS', 'STOCKS VEHICLE STOCKS 121 150 DAYS', 'STOCKS VEHICLE STOCKS 151 180 DAYS', 'STOCKS VEHICLE STOCKS GREATER THAN180 DAYS', 'STOCKS SPARES STOCKS 0 5 DAYS', 'STOCKS SPARES STOCKS 6 30 DAYS', 'STOCKS SPARES STOCKS 31 60 DAYS', 'STOCKS SPARES STOCKS 61 90 DAYS', 'STOCKS SPARES STOCKS 91 120 DAYS', 'STOCKS SPARES STOCKS 121 150 DAYS', 'STOCKS SPARES STOCKS 151 180 DAYS', 'STOCKS SPARES STOCKS GREATER THAN180 DAYS', 'STOCKS FG TRADING 0 5 DAYS', 'STOCKS FG TRADING 6 30 DAYS', 'STOCKS FG TRADING 31 60 DAYS', 'STOCKS FG TRADING 61 90 DAYS', 'STOCKS FG TRADING 91 120 DAYS', 'STOCKS FG TRADING 121 150 DAYS', 'STOCKS FG TRADING 151 180 DAYS', 'STOCKS FG TRADING GREATER THAN180 DAYS', 'STOCKS CONSUMABLES PACKING MATERIAL', 'STOCKS STOCK IN TRANSIT', 'STOCKS OTHER STOCK 1', 'STOCKS OTHER STOCK 2', 'STOCKS RM LESS THAN 90 DAYS', 'STOCKS RM 120 180 DAYS', 'STOCKS WIP LESS THAN 60 DAYS', 'STOCKS FG LESS THAN 90 DAYS', 'STOCKS FG 120 180 DAYS', 'STOCKS VEHICLE STOCK UPTO 120 DAYS', 'STOCKS SPARES STOCK 6 60 DAYS', 'STOCKS SPARES STOCK LESS THAN 60 DAYS', 'STOCKS TOTAL SPARES STOCK', 'STOCK LESS THAN 60 DAYS', 'STOCK LESS THAN 90 DAYS', 'STOCK LESS THAN 120 DAYS', 'STOCK LESS THAN 180 DAYS', 'STOCKS WIP LESS THAN 180 DAYS', 'TOTAL STOCKS', 'DEBTORS 0 5 DAYS', 'DEBTORS 6 30 DAYS', 'DEBTORS 31 60 DAYS', 'DEBTORS 61 90 DAYS', 'DEBTORS 91 120 DAYS', 'DEBTORS 121 150 DAYS', 'DEBTORS 151 180 DAYS', 'DEBTORS GREATER THAN180 DAYS', 'DEBTORS SPARES DEBTORS 0 5 DAYS', 'DEBTORS SPARES DEBTORS 6 30 DAYS', 'DEBTORS SPARES DEBTORS 31 60 DAYS', 'DEBTORS SPARES DEBTORS 61 90 DAYS', 'DEBTORS SPARES DEBTORS 91 120 DAYS', 'DEBTORS SPARES DEBTORS 121 150 DAYS', 'DEBTORS SPARES DEBTORS 151 180 DAYS', 'DEBTORS SPARES DEBTORS GREATER THAN180 DAYS', 'DEBTORS VEHICLE DEBTORS 0 5 DAYS', 'DEBTORS VEHICLE DEBTORS 6 30 DAYS', 'DEBTORS VEHICLE DEBTORS 31 60 DAYS', 'DEBTORS VEHICLE DEBTORS 61 90 DAYS', 'DEBTORS VEHICLE DEBTORS 91 120 DAYS', 'DEBTORS VEHICLE DEBTORS 121 150 DAYS', 'DEBTORS VEHICLE DEBTORS 151 180 DAYS', 'DEBTORS VEHICLE DEBTORS GREATER THAN180 DAYS', 'DEBTORS SPARES DEBTORS 6 60 DAYS', 'DEBTORS WORKSHOP SERVICE DEBTORS LESS THAN60 DAYS', 'DEBTORS WORKSHOP SERVICE DEBTORS 6 60 DAYS', 'DEBTORS WORKSHOP SERVICE DEBTORS 60 90 DAYS', 'DEBTORS WORKSHOP SERVICE 90 120 DAYS', 'DEBTORS WORKSHOP SERVICE120 150 DAYS', 'DEBTORS VEHICLE DEBTORS 6 60 DAYS', 'DEBTORS VEHICLE DEBTORS LESS THAN 15 DAYS', 'DEBTORS SPARE DEBTORS LESS THAN 15 DAYS', 'DEBTORS VEHICLE DEBTORS LESS THAN 30 DAYS', 'DEBTORS SPARE DEBTORS LESS THAN 30 DAYS', 'DEBTORS DEBTORS EXCHANGE TRACTORS', 'DEBTORS VEHICLE DEBTORS LESS THAN 45 DAYS', 'DEBTORS VEHICLE DEBTORS LESS THAN 60 DAYS', 'DEBTORS SPARES DEBTORS LESS THAN 45 DAYS', 'DEBTORS SPARES DEBTORS LESS THAN 60 DAYS', 'DEBTORS DEDUCTION INTERGROUP RECEIVABLES 1', 'DEBTORS DEDUCTION INTERGROUP RECEIVABLES 2', 'DEBTORS LESS THAN 30 DAYS', 'DEBTORS LESS THAN 60 DAYS', 'DEBTORS LESS THAN 90 DAYS', 'DEBTORS GREATER THAN 90 DAYS', 'DEBTORS GREATER THAN 90 LESS THAN 150 DAYS', 'DEBTORS LESS THAN 120 DAYS', 'DEBTORS GREATER THAN 120 LESS THAN 150 DAYS', 'DEBTORS LESS THAN 180 DAYS', 'DEBTORS EXCHANGE TRACTORS', 'TOTAL DEBTORS', 'DEBTORS UNBILLED DEBTORS LESS THAN60 DAYS', 'DEBTORS UNBILLED DEBTORS LESS THAN90DAYS', 'DEBTORS TOTAL UNBILLED DEBTORS', 'CREDITORS SUNDRY CREDITORS FOR GOODS', 'CREDITORS SUNDRY CREDITORS FOR CAPITAL EXPENSES', 'CREDITORS SUNDRY CREDITORS FOR EXPENSES', 'CREDITORS LC BACKED CREDITORS FOR GOODS', 'CREDITORS LC BACKED CREDITORS FOR CAPITAL EXPENSES', 'CREDITORS LC BACKED CREDITORS FOR EXPENSES', 'CREDITORS OTHER CREDITORS 1', 'CREDITORS OTHER CREDITORS 2', 'CREDITORS SPARES CREDITORS', 'CREDITORS GROUP CONCERN CREDITORS', 'CREDITORS UPTO 90 DAYS', 'CREDITORS UPTO 91 150 DAYS', 'CREDITORS GREATER THAN151 DAYS', 'TOTAL CREDITORS', 'DEBTORS ADVANCE PAID TO SUPPLIER', 'SECURITY DEPOSIT', 'CVD DUTY REFUNDABLE', 'CREDITORS ADVANCES RECEIVED FROM CUSTOMERS', 'DEBTORS ADVANCE TRACTORS', 'DEBTORS DEDUCTION BILLS DISCOUNTED', 'DEBTORS DEDUCTION BILLS FACTORED ASSIGNED', 'LIMITS OS WITH KMBL BANKS 1', 'LIMITS OS WITH KMBL BANKS 2', 'LIMITS OS WITH KMBL BANKS 3', 'LIMITS OS WITH OTHER BANKS 1', 'LIMITS OS WITH OTHER BANKS 2', 'LIMITS OS WITH OTHER BANKS 3', 'LIMITS OS WITH OTHER BANKS 4', 'LIMITS OS WITH OTHER BANKS 5', 'OUTSTANDING IN SPG CF INF TA', 'ADVANCE TRACTORS', 'ADVANCE TRACTORS LESS THAN 90 DAYS', 'ADVANCE TRACTORS GREATER THAN91 DAYS TO 120 DAYS', 'ADVANCE TRACTORS GREATER THAN121 DAYS TO 150 DAYS', 'PURCHASES 1', 'PURCHASES 2', 'PURCHASES 3', 'PURCHASE AMOUNT', 'PURCHASE AMOUNT VEHICLES', 'PURCHASE AMOUNT SPARES', 'PURCHASE AMOUNT WORKSHOP SERVICE', 'TOTAL PURCHASES', 'SALES 1', 'SALES 2', 'SALES 3', 'SALES AMOUNT VEHICLES', 'SALES AMOUNT SPARES', 'SALES AMOUNT WORKSHOP SERVICE', 'TOTAL SALES', 'INVENTORY HOLDING PERIOD', 'COLLECTION PERIOD', 'CREDIT PERIOD', 'NET CYCLE', 'NET CREDIT', 'SALES REALISATION', 'NET CREDIT TO SALES REALISATION', 'OD LIMIT', 'CC LIMIT', 'ADDITIONAL FIELDS NUM1', 'ADDITIONAL FIELDS NUM2', 'ADDITIONAL FIELDS NUM3', 'ADDITIONAL FIELDS NUM4', 'ADDITIONAL FIELDS NUM5', 'ADDITIONAL FIELDS NUM6', 'ADDITIONAL FIELDS NUM7', 'ADDITIONAL FIELDS NUM8', 'ADDITIONAL FIELDS NUM9', 'ADDITIONAL FIELDS NUM10', 'ADDITIONAL FIELDS TXT1', 'ADDITIONAL FIELDS TXT2', 'ADDITIONAL FIELDS TXT3', 'ADDITIONAL FIELDS TXT4', 'ADDITIONAL FIELDS TXT5', 'ADDITIONAL FIELDS TXT6', 'ADDITIONAL FIELDS TXT7', 'ADDITIONAL FIELDS TXT8', 'ADDITIONAL FIELDS TXT9', 'ADDITIONAL FIELDS TXT10', 'DEVIATION APPLICABLE', 'DEBTORS<90 days as per SS', 'Last 3 month Sales', 'Debtors > Sales', 'Last 3 month purchases', 'Stock > Purchases', 'System DP', 'Computed DP', 'Sanction Limit', 'Final DP', 'DP Formula']

            res = pd.DataFrame({key: [] for key in columns})
            result_df = pd.concat([result_df, df], axis=1) 
            print(f"columns list {result_df.columns.tolist()}")
            res = res.append(result_df, ignore_index=True)


            dp_query=f"SELECT `dp_formula` from ocr where case_id='{case_id}'"
            dp_res = extraction_db.execute_(dp_query).to_dict(orient="records")[-1]
            logging.info(f"{dp_res}")
            if dp_res['dp_formula'] !=  None:
        
                equation_list = re.findall(r'\b\w+_\w+\b', dp_res['dp_formula'] )
                rbg_fields={'STOCKS_RM_0_5_DAYS': 'STOCKS IMPORTED RM 0 5 DAYS', 'STOCKS_RM_6_30_DAYS': 'STOCKS IMPORTED RM 6 30 DAYS', 'STOCKS_RM_31_60_DAYS': 'STOCKS IMPORTED RM 31 60 DAYS', 'STOCKS_RM_61_90_DAYS': 'STOCKS IMPORTED RM 61 90 DAYS', 'STOCKS_RM_91_120_DAYS': 'STOCKS IMPORTED RM 91 120 DAYS', 'STOCKS_RM_121_150_DAYS': 'STOCKS IMPORTED RM 121 150 DAYS', 'STOCKS_RM_151_180_DAYS': 'STOCKS IMPORTED RM 151 180 DAYS', 'STOCKS_IMPORTED_RM_GREATER_THAN180_DAYS': 'STOCKS IMPORTED RM GREATER THAN180 DAYS', 'STOCKS_WIP_0_5_DAYS': 'STOCKS WIP 0 5 DAYS', 'STOCKS_WIP_6_30_DAYS': 'STOCKS WIP 6 30 DAYS', 'STOCKS_WIP_31_60_DAYS': 'STOCKS WIP 31 60 DAYS', 'STOCKS_WIP_61_90_DAYS': 'STOCKS WIP 61 90 DAYS', 'STOCKS_WIP_91_120_DAYS': 'STOCKS WIP 91 120 DAYS', 'STOCKS_WIP_121_150_DAYS': 'STOCKS WIP 121 150 DAYS', 'STOCKS_WIP_151_180_DAYS': 'STOCKS WIP 151 180 DAYS', 'STOCKS_WIP_GREATER_THAN180_DAYS': 'STOCKS WIP GREATER THAN180 DAYS', 'STOCKS_VEHICLE_STOCKS_0_5_DAY': 'STOCKS VEHICLE STOCKS 0 5 DAYS', 'STOCKS_VEHICLE_STOCKS_6_30_DAY': 'STOCKS VEHICLE STOCKS 6 30 DAYS', 'STOCKS_VEHICLE_STOCKS_31_60DAY': 'STOCKS VEHICLE STOCKS 31 60 DAYS', 'STOCKS_VEHICLE_STOCKS_61_90DAY': 'STOCKS VEHICLE STOCKS 61 90 DAYS', 'STOCKS_VEHICLE_STOCKS_91_120_DAY': 'STOCKS VEHICLE STOCKS 91 120 DAYS', 'STOCKS_VEHICLE_STOCK121_150DAY': 'STOCKS VEHICLE STOCKS 121 150 DAYS', 'STOCKS_VEHICLE_STOCK151_180DAY': 'STOCKS VEHICLE STOCKS 151 180 DAYS', 'STOCKS_VEHICLE_STOCKS_GREATER_THAN180_DAYS': 'STOCKS VEHICLE STOCKS GREATER THAN180 DAYS', 'STOCKS_SPARES_STOCKS_0_5_DAY': 'STOCKS SPARES STOCKS 0 5 DAYS', 'STOCKS_SPARES_STOCKS_6_30_DAY': 'STOCKS SPARES STOCKS 6 30 DAYS', 'STOCKS_SPARES_STOCKS_31_60_DAY': 'STOCKS SPARES STOCKS 31 60 DAYS', 'STOCKS_SPARES_STOCKS_61_90_DAY': 'STOCKS SPARES STOCKS 61 90 DAYS', 'STOCKS_SPARES_STOCKS_91_120DAY': 'STOCKS SPARES STOCKS 91 120 DAYS', 'STOCKS_SPARES_STOCKS_121_150_DAY': 'STOCKS SPARES STOCKS 121 150 DAYS', 'STOCKS_SPARES_STOCKS151_180DAYs': 'STOCKS SPARES STOCKS 151 180 DAYS', 'STOCKS_SPARES_STOCKS_GREATER_THAN180_DAYS': 'STOCKS SPARES STOCKS GREATER THAN180 DAYS', 'STOCKS_FG_TRADING_0_5_DAYS': 'STOCKS FG TRADING 0 5 DAYS', 'STOCKS_FG_TRADING_6_30_DAYS': 'STOCKS FG TRADING 6 30 DAYS', 'STOCKS_FG_TRADING_31_60_DAYS': 'STOCKS FG TRADING 31 60 DAYS', 'STOCKS_FG_TRADING_61_90_DAYS': 'STOCKS FG TRADING 61 90 DAYS', 'STOCKS_FG_TRADING_91_120_DAYS': 'STOCKS FG TRADING 91 120 DAYS', 'STOCKS_FG_TRADING_121_150_DAYS': 'STOCKS FG TRADING 121 150 DAYS', 'STOCKS_FG_TRADING_151_180_DAYS': 'STOCKS FG TRADING 151 180 DAYS', 'STOCKS_FG_TRADING_GREATER_THAN180_DAYS': 'STOCKS FG TRADING GREATER THAN180 DAYS', 'STK_CONSUMABLES_PKG_MATERIAL': 'STOCKS CONSUMABLES PACKING MATERIAL', 'STOCKS_STOCK_IN_TRANSIT': 'STOCKS STOCK IN TRANSIT', 'STOCKS_RM_LESS_THAN_90_DAYS': 'STOCKS RM LESS THAN 90 DAYS', 'STOCKS_TOTAL_SPARES_STOCK': 'STOCKS TOTAL SPARES STOCK', 'STOCK_LESS_THAN_60_DAYS': 'STOCK LESS THAN 60 DAYS', 'STOCK_LESS_THAN_90_DAYS': 'STOCK LESS THAN 90 DAYS', 'STOCK_LESS_THAN_120_DAYS': 'STOCK LESS THAN 120 DAYS', 'STOCK_LESS_THAN_180_DAYS': 'STOCK LESS THAN 180 DAYS', 'STOCKS_WIP_LESS_THAN_180_DAYS': 'STOCKS WIP LESS THAN 180 DAYS', 'TOTAL_STOCKS': 'TOTAL STOCKS', 'DEBTORS_0_5_DAYS': 'DEBTORS 0 5 DAYS', 'DEBTORS_6_30_DAYS': 'DEBTORS 6 30 DAYS', 'DEBTORS_31_60_DAYS': 'DEBTORS 31 60 DAYS', 'DEBTORS_61_90_DAYS': 'DEBTORS 61 90 DAYS', 'DEBTORS_91_120_DAYS': 'DEBTORS 91 120 DAYS', 'DEBTORS_121_150_DAYS': 'DEBTORS 121 150 DAYS', 'DEBTORS_151_180_DAYS': 'DEBTORS 151 180 DAYS', 'DEBTORS_GREATER_THAN180_DAYS': 'TOTAL DEBTORS', 'DR_DED_INTERGROUP_RECEIVABLES1': 'DEBTORS DEDUCTION INTERGROUP RECEIVABLES 1', 'DR_DED_INTERGROUP_RECEIVABLES2': 'DEBTORS DEDUCTION INTERGROUP RECEIVABLES 2', 'DEBTORS_LESS_THAN_30_DAYS': 'DEBTORS LESS THAN 30 DAYS', 'DEBTORS_LESS_THAN_60_DAYS': 'DEBTORS LESS THAN 60 DAYS', 'DEBTORS_LESS_THAN_90_DAYS': 'DEBTORS LESS THAN 90 DAYS', 'DEBTORS_GRT_THAN_90_DAYS': 'DEBTORS GREATER THAN 90 DAYS', 'DR_GRT_THN_90_LESS_THN_150_DAY': 'DEBTORS GREATER THAN 90 LESS THAN 150 DAYS', 'DEBTORS_LESS_THAN_120_DAYS': 'DEBTORS LESS THAN 120 DAYS', 'DR_GRT_THAN_120LESSTHN_150_DAY': 'DEBTORS GREATER THAN 120 LESS THAN 150 DAYS', 'DEBTORS_LESS_THAN_180_DAYS': 'DEBTORS LESS THAN 180 DAYS', 'TOTAL_DEBTORS': 'TOTAL DEBTORS', 'DR_UNBILLED_DR_LESS_THAN60DAYS': 'DEBTORS UNBILLED DEBTORS LESS THAN60 DAYS', 'DR_UNBILLED_DR_LESS_THAN90DAYS': 'DEBTORS UNBILLED DEBTORS LESS THAN90DAYS', 'DEBTORS_TOTAL_UNBILLED_DEBTORS': 'DEBTORS TOTAL UNBILLED DEBTORS', 'CREDITORS_UPTO_90_DAYS': 'CREDITORS UPTO 90 DAYS', 'CREDITORS_UPTO_91_150_DAYS': 'CREDITORS UPTO 91 150 DAYS', 'CREDITORS_GREATER_THAN151_DAYS': 'CREDITORS GREATER THAN151 DAYS', 'TOTAL_CREDITORS': 'TOTAL CREDITORS', 'DR_ADV_PAID_TO_SUPPLIER': 'DEBTORS ADVANCE PAID TO SUPPLIER', 'SECURITY_DEPOSIT': 'SECURITY DEPOSIT', 'CREDITORS_ADV_REC_FRM_CUST': 'CREDITORS ADVANCES RECEIVED FROM CUSTOMERS', 'DR_DEDUCTION_BILLS_DISC': 'DEBTORS DEDUCTION BILLS DISCOUNTED', 'DR_DEDUCTION_BILLS_FACT_ASSI': 'DEBTORS DEDUCTION BILLS FACTORED ASSIGNED', 'LIMITS_OS_WITH_KMBL_BANKS_1': 'LIMITS OS WITH KMBL BANKS 1', 'LIMITS_OS_WITH_KMBL_BANKS_2': 'LIMITS OS WITH KMBL BANKS 2', 'LIMITS_OS_WITH_KMBL_BANKS_3': 'LIMITS OS WITH KMBL BANKS 3', 'LIMITS_OS_WITH_OTHER_BANKS_1': 'LIMITS OS WITH OTHER BANKS 1', 'LIMITS_OS_WITH_OTHER_BANKS_2': 'LIMITS OS WITH OTHER BANKS 2', 'LIMITS_OS_WITH_OTHER_BANKS_3': 'LIMITS OS WITH OTHER BANKS 3', 'LIMITS_OS_WITH_OTHER_BANKS_4': 'LIMITS OS WITH OTHER BANKS 4', 'LIMITS_OS_WITH_OTHER_BANKS_5': 'LIMITS OS WITH OTHER BANKS 5','STOCKS_FG_LESS_THAN_90_DAYS':'STOCKS FG TRADING 61 90 DAYS'}
                categories=['STOCKS_RM_0_5_DAYS', 'STOCKS_RM_6_30_DAYS', 'STOCKS_RM_31_60_DAYS', 'STOCKS_RM_61_90_DAYS', 'STOCKS_RM_91_120_DAYS', 'STOCKS_RM_121_150_DAYS', 'STOCKS_RM_151_180_DAYS', 'STOCKS_IMPORTED_RM_GREATER_THAN180_DAYS', 'STOCKS_WIP_0_5_DAYS', 'STOCKS_WIP_6_30_DAYS', 'STOCKS_WIP_31_60_DAYS', 'STOCKS_WIP_61_90_DAYS', 'STOCKS_WIP_91_120_DAYS', 'STOCKS_WIP_121_150_DAYS', 'STOCKS_WIP_151_180_DAYS', 'STOCKS_WIP_GREATER_THAN180_DAYS', 'STOCKS_VEHICLE_STOCKS_0_5_DAY', 'STOCKS_VEHICLE_STOCKS_6_30_DAY', 'STOCKS_VEHICLE_STOCKS_31_60DAY', 'STOCKS_VEHICLE_STOCKS_61_90DAY', 'STOCKS_VEHICLE_STOCKS_91_120_DAY', 'STOCKS_VEHICLE_STOCK121_150DAY', 'STOCKS_VEHICLE_STOCK151_180DAY', 'STOCKS_VEHICLE_STOCKS_GREATER_THAN180_DAYS', 'STOCKS_SPARES_STOCKS_0_5_DAY', 'STOCKS_SPARES_STOCKS_6_30_DAY', 'STOCKS_SPARES_STOCKS_31_60_DAY', 'STOCKS_SPARES_STOCKS_61_90_DAY', 'STOCKS_SPARES_STOCKS_91_120DAY', 'STOCKS_SPARES_STOCKS_121_150_DAY', 'STOCKS_SPARES_STOCKS151_180DAYs', 'STOCKS_SPARES_STOCKS_GREATER_THAN180_DAYS', 'STOCKS_FG_TRADING_0_5_DAYS', 'STOCKS_FG_TRADING_6_30_DAYS', 'STOCKS_FG_TRADING_31_60_DAYS', 'STOCKS_FG_TRADING_61_90_DAYS', 'STOCKS_FG_TRADING_91_120_DAYS', 'STOCKS_FG_TRADING_121_150_DAYS', 'STOCKS_FG_TRADING_151_180_DAYS', 'STOCKS_FG_TRADING_GREATER_THAN180_DAYS', 'STK_CONSUMABLES_PKG_MATERIAL', 'STOCKS_STOCK_IN_TRANSIT', 'STOCKS_RM_LESS_THAN_90_DAYS', 'STOCKS_TOTAL_SPARES_STOCK', 'STOCK_LESS_THAN_60_DAYS', 'STOCK_LESS_THAN_90_DAYS', 'STOCK_LESS_THAN_120_DAYS', 'STOCK_LESS_THAN_180_DAYS', 'STOCKS_WIP_LESS_THAN_180_DAYS', 'TOTAL_STOCKS', 'DEBTORS_0_5_DAYS', 'DEBTORS_6_30_DAYS', 'DEBTORS_31_60_DAYS', 'DEBTORS_61_90_DAYS', 'DEBTORS_91_120_DAYS', 'DEBTORS_121_150_DAYS', 'DEBTORS_151_180_DAYS', 'DEBTORS_GREATER_THAN180_DAYS', 'DR_DED_INTERGROUP_RECEIVABLES1', 'DR_DED_INTERGROUP_RECEIVABLES2', 'DEBTORS_LESS_THAN_30_DAYS', 'DEBTORS_LESS_THAN_60_DAYS', 'DEBTORS_LESS_THAN_90_DAYS', 'DEBTORS_GRT_THAN_90_DAYS', 'DR_GRT_THN_90_LESS_THN_150_DAY', 'DEBTORS_LESS_THAN_120_DAYS', 'DR_GRT_THAN_120LESSTHN_150_DAY', 'DEBTORS_LESS_THAN_180_DAYS', 'TOTAL_DEBTORS', 'DR_UNBILLED_DR_LESS_THAN60DAYS', 'DR_UNBILLED_DR_LESS_THAN90DAYS', 'DEBTORS_TOTAL_UNBILLED_DEBTORS', 'CREDITORS_UPTO_90_DAYS', 'CREDITORS_UPTO_91_150_DAYS', 'CREDITORS_GREATER_THAN151_DAYS', 'TOTAL_CREDITORS', 'DR_ADV_PAID_TO_SUPPLIER', 'SECURITY_DEPOSIT', 'CREDITORS_ADV_REC_FRM_CUST', 'DR_DEDUCTION_BILLS_DISC', 'DR_DEDUCTION_BILLS_FACT_ASSI', 'LIMITS_OS_WITH_KMBL_BANKS_1', 'LIMITS_OS_WITH_KMBL_BANKS_2', 'LIMITS_OS_WITH_KMBL_BANKS_3', 'LIMITS_OS_WITH_OTHER_BANKS_1', 'LIMITS_OS_WITH_OTHER_BANKS_2', 'LIMITS_OS_WITH_OTHER_BANKS_3', 'LIMITS_OS_WITH_OTHER_BANKS_4', 'LIMITS_OS_WITH_OTHER_BANKS_5','STOCKS_FG_LESS_THAN_90_DAYS']
                non_dp_formula_fields=[]
                for i in categories:
                    if i not in equation_list:                     
                        non_dp_formula_fields.append(rbg_fields[i])
                for field in non_dp_formula_fields:
                    res[field] = None
                res = res.drop(columns=['Stock Due Date','Received Date','CC_LIMIT','INSURANCE_AMOUNT','DEBTORS_MARGIN','STOCK_MARGIN','NAME OF CUSTOMER','ZMT'])

            logging.info(f"result dataframe####  {res}")
            res=res.to_dict(orient='list')
            logging.info(f"result dataframe####  {res}")
            query=f"update `ocr` set `dp_formula_json` = '{json.dumps(res)}' where `case_id` ='{case_id}'"
            extraction_db.execute_(query)
            logging.info(f"result dataframe####  {res}")
            queue_name="rbg"
    
   

        response = {
            'flag': True,
            'data': {
                    'message':'success'
                }
        }

        
        return response
    except Exception:
        logging.exception(f'Something went wrong exporting data')
        return {'flag': False, 'message': 'Unable to export data.'}
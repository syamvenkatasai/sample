"""
Author 1: Satyanarayana Y
Author 2: Gopi Teja B
Created Date: 18-08-2022
"""

from ace_logger import Logging
from db_utils import DB
from datetime import datetime,timedelta
from flask import request, jsonify
from pathlib import Path
from py_zipkin.util import generate_random_64bit_string
from app import app
from app.elasticsearch_utils import elasticsearch_search
from py_zipkin.zipkin import zipkin_span,ZipkinAttrs, create_http_headers_for_new_span
from elasticsearch import Elasticsearch
import ast
try:
    from generate_reports import reports_consumer, generate_report
except:
    from app.generate_reports import reports_consumer, generate_report
from time import time as tt
from pandas import Timestamp
import openpyxl


import base64
import json
import uuid
import pytz
import requests
import os
import re
import datetime as dt
import psutil
import pandas as pd
import numpy as np


es_dns = os.environ.get('ELASTIC_SEARCH_FULL_SEARCH_DNS','')
es_port = os.environ.get('ELASTIC_SEARCH_FULL_PORT', '')
es_scheme = os.environ.get('ELASTIC_SEARCH_FULL_SEARCH_SCHEME','')
es = Elasticsearch(
    [f'{es_dns}'],
    http_auth=('elastic','MagicWord'),
    scheme=f"{es_scheme}",
    port=es_port,
)


logging = Logging(name='reports_api')

db_config = {
    'host': os.environ['HOST_IP'],
    'password': os.environ['LOCAL_DB_PASSWORD'],
    'user': os.environ['LOCAL_DB_USER'],
    'port': os.environ['LOCAL_DB_PORT']
}


def measure_memory_usage():
    process = psutil.Process()
    memory_info = process.memory_info()
    return memory_info.rss  # Resident Set Size (RSS) in bytes


def http_transport(encoded_span):
    # The collector expects a thrift-encoded list of spans. Instead of
    # decoding and re-encoding the already thrift-encoded message, we can just
    # add header bytes that specify that what follows is a list of length 1.
    body = encoded_span
    requests.post(
        'http://servicebridge:5002/zipkin',
        data=body,
        headers={'Content-Type': 'application/x-thrift'},
    )


def get_reports_column_mapping(columns):
    return_columns_map = {}
    if columns:
        for column in columns:
            return_columns_map[column.replace('_', ' ').title()] = column
    return return_columns_map

def insert_into_audit(case_id, data):
    tenant_id = data.pop('tenant_id')
    db_config['tenant_id'] = tenant_id
    stats_db = DB('stats', **db_config)
    stats_db.insert_dict(data, 'audit')
    return True

db_column_types = {
    "mysql": {
        "int": "number",
        "tinyint": "number",
        "smallint": "number",
        "mediumint": "number",
        "bigint": "number",
        "double": "number",

        "date": "date",
        "datetime": "date",
        "timestamp": "date",
        "time": "date",

        "char": "string",
        "varchar": "string",
        "blob": "string",
        "text": "string",
        "tinytext": "string",
        "mediumtext": "string",
        "longtext": "string",
        "tinyblob": "string",
        "mediumblob": "string",
        "longblob": "string",
        "enum": "string"
    },
    "mssql": {
        "bigint": "number",
        "bit": "number",
        "decimal": "number",
        "int": "number",
        "money": "number",
        "numeric": "number",
        "smallint": "number",
        "smallmoney": "number",
        "tinyint": "number",
        "float": "number",

        "date": "date",
        "datetime2": "date",
        "datetime": "date",
        "datetimeoffset": "date",
        "smalldatetime": "date",
        "time": "date",

        "char": "string",
        "text": "string",
        "varchar": "string",
        "nchar": "string",
        "nvarchar": "string",
        "ntext": "string",


    },
    "oracle": {
        "NUMBER": "number",
        "BINARY_FLOAT": "number",
        "BINARY_DOUBLE": "number",
        "DATE": "date",
        "TIMESTAMP": "date",
        "CHAR": "string",
        "NCHAR": "string",
        "VARCHAR2": "string",
        "NVARCHAR2": "string",
        "CLOB": "string",
        "NCLOB": "string",
        "BLOB": "string"
    }
}


def create_index(tenant_ids, sources=[]):
    body = {
        "settings": {
            "analysis": {
                "analyzer": {
                    "default": {
                        "type": "custom",
                        "tokenizer": "whitespace",
                        "filter": [
                            "lowercase"
                        ]
                    }
                }
            }
        },
        "mappings": {
            "date_detection": "false",
            "numeric_detection": "false"
        }
    }
    
    body_with_date = {
            "settings": {
            "analysis": {
                "analyzer": {
                    "default": {
                        "type": "custom",
                        "tokenizer": "whitespace",
                        "filter": [
                          "lowercase"
                        ]
                    }
                }
            }
            },
        "mappings": {
            "properties": {
                "@timestamp": {
                    "type": "date"
                },
                "@version": {
                    "type": "text",
                    "fields": {
                        "keyword": {
                            "type": "keyword",
                            "ignore_above": 256
                        }
                    }
                },
                "ocr": {
                    "properties": {
                        "created_date": {
                            "type": "date"
                        },
                        "last_updated": {
                            "type": "date"
                        }
                    }
                },
                "process_queue": {
                    "properties": {
                        "created_date": {
                            "type": "date"
                        },
                        "last_updated": {
                            "type": "date"
                        },
                        "freeze": {
                            "type": "boolean"
                        },
                        "case_lock":{
                            "type": "boolean"
                        }
                    }
                }
            },
            "date_detection": "false",
            "numeric_detection": "false"
        }
        }
    indexes = []
    
    for tenant_id in tenant_ids:
        indexes.extend(get_search_indexes(sources, tenant_id))

    for ind in indexes:
        es.indices.delete(index=ind, ignore=[400, 404])
        if 'processqueue' in ind:
            es.indices.create(index=ind, body=body_with_date, ignore=400)
        else:
            es.indices.create(index=ind, body=body, ignore=400)


def get_search_indexes(sources, temp_tenant_id=''):
    """
    Author : Akshat Goyal
    :param sources:
    :return:
    """
    tenant_id = temp_tenant_id.replace('.', '')
    if not sources:
        return '_all'
    indexes = []
    if isinstance(sources, list):
        for source in sources:
            new_source = tenant_id + source if tenant_id else source
            new_source = new_source.replace('.', '').replace('_', '')
            indexes.append(new_source)
    elif isinstance(sources, str):
        new_source = tenant_id + '_' + sources if tenant_id else sources
        new_source = new_source.replace('.', '').replace('_', '')
        indexes.append(new_source)

    return indexes


def master_search(tenant_id, text, table_name, start_point, offset, columns_list, header_name):
    elastic_input = {}
    
    print(f'#######{tenant_id}')
    print(f'#######{table_name}')
    print(f'#######{start_point}')
    print(f'#######{offset}')
    print(f'#######{columns_list}')
    print(f'#######{header_name}')
    elastic_input['columns'] = columns_list
    elastic_input['start_point'] = start_point
    elastic_input['size'] = offset
    if header_name:
        header_name=header_name.upper()
        elastic_input['filter'] = [{'field': header_name, 'value': "*" + text + "*"}]
    else:
        elastic_input['text'] = text
    elastic_input['source'] = table_name
    elastic_input['tenant_id'] = tenant_id
    print(f"output of the elastic_input---------{elastic_input}")
    files, total = elasticsearch_search(elastic_input)
    print(f"output----------{files,total}")
    return files, total


def get_reports_column_data_types(columns, tenant_id):

    logging.info(f"^^^^^^^^^^^^^^^ COLUMNS ARE {columns}")

    db_config['tenant_id'] = tenant_id
    reports_db = DB('reports', **db_config)

    column_query = """SELECT *
        FROM report_requests
        FETCH FIRST 1 ROW ONLY"""
    reports_columns = list(reports_db.execute_(column_query).columns.values)

    # query = "SELECT DATA_TYPE FROM INFORMATION_SCHEMA.COLUMNS WHERE table_name = 'report_requests'"
    query = """SELECT DATA_TYPE
        FROM USER_TAB_COLUMNS
        WHERE TABLE_NAME = 'REPORT_REQUESTS'"""
    reports_column_types = list(reports_db.execute_(query)['data_type'])

    reports_dt_map = dict(zip(reports_columns, reports_column_types))

    logging.info(
        f"#########################33 reports VALUE IS {reports_dt_map} ")

    col_data_types = {}

    db_type = os.environ.get('DB_TYPE', 'mysql').lower()

    for column in columns:
        col_data_types[column] = db_column_types[db_type].get(
            reports_dt_map.get(column, ""), "unknown")

    return col_data_types



def get_group_ids(user, db):
    logging.info(f'Getting group IDs for user `{user}`')

    # query = 'SELECT organisation_attributes.attribute, user_organisation_mapping.value \
    #         FROM `user_organisation_mapping`, `active_directory`, `organisation_attributes` \
    #         WHERE active_directory.username=%s AND \
    #         active_directory.id=user_organisation_mapping.user_id AND \
    #         organisation_attributes.id=user_organisation_mapping.organisation_attribute'
    query="""SELECT organisation_attributes.attribute, user_organisation_mapping.value 
        FROM user_organisation_mapping 
        JOIN active_directory ON active_directory.id = user_organisation_mapping.user_id 
        JOIN organisation_attributes ON organisation_attributes.id = user_organisation_mapping.organisation_attribute 
        WHERE active_directory.username = %s"""

    user_group = db.execute_(query, params=[user])

    if user_group.empty:
        logging.error(f'No user organisation mapping for user `{user}`')
        return

    user_group_dict = dict(zip(user_group.attribute, user_group.value))
    user_group_dict = {key: [value] for key, value in user_group_dict.items()}
    # group_def_df = db.get_all('GROUP_DEFINITION')
    query="select * from GROUP_DEFINITION"
    group_def_df = db.execute_(query)

    if group_def_df.empty:
        logging.debug(f'Groups not defined in `group_definition`')

        return
    logging.info(f"####groupgroup_def_df {group_def_df}")

    user_group = group_def_df.to_dict()
    group_def = {}
    for i in range(len(user_group['id'])):
        id_key = str(user_group['id'][i])
        group_def[id_key] = {
            'group_name': user_group['group_name'][i],
            'group_definition': user_group['group_definition'][i],
            'group_definition_template': user_group['group_definition_template'][i]
        }

    logging.info(f"####ggroup_def {group_def}")
    group_ids = []
    for index, group in group_def.items():
        logging.debug(f'Index: {index}')
        logging.debug(f'Group: {group}')

        try:
            group_dict = json.loads(group['group_definition'])
        except:
            logging.error('Could not load group definition dict.')
            break

        # Check if all key-value from group definition is there in the user group
        if group_dict.items() == user_group_dict.items():
            group_ids.append(index)

    logging.info(f'Group IDs: {group_ids}')
    return group_ids

@app.route('/get_reports_queue', methods=['POST', 'GET'])
def get_reports_queue():
    try:
        memory_before = measure_memory_usage()
        start_time = tt()
    except:
        logging.warning("Failed to start ram and time calc")
        pass
    data = request.json

    logging.info(f'Data recieved: {data}')
    tenant_id = data.get('tenant_id', None)
    attr = ZipkinAttrs(
        trace_id=generate_random_64bit_string(),
        span_id=generate_random_64bit_string(),
        parent_span_id=None,
        flags=None,
        is_sampled=False,
        tenant_id=tenant_id
    )

    with zipkin_span(
        service_name='reports_api',
        span_name='get_reports_queue',
        transport_handler=http_transport,
        zipkin_attrs=attr,
        port=5010,
        sample_rate=0.5):
        try:
            
            user = data.get('user', None)
            total_count = 0  # pagination variable
            filtered_text = data.get('filter_text','')
            try:
                start_point = data.get('start', 1)
                end_point = data.get('end', 20)
            except Exception as e:
                logging.debug(e)
                start_point = 1
                end_point = 20

            try:
                offset = end_point - start_point
            except:
                start_point = 1
                end_point = 20
                offset = 20
            offset_ = 0 if start_point == 1 else start_point - 1

            # Sanity check
            if user is None:
                message = 'User not provided in request.'
                logging.error(message)
                return jsonify({'flag': False, 'message': message})

            # Connect to reports database
            reports_db_config = {
                                'host': os.environ['HOST_IP'],
                                'password': os.environ['LOCAL_DB_PASSWORD'],
                                'user': os.environ['LOCAL_DB_USER'],
                                'port': os.environ['LOCAL_DB_PORT'],
                                'tenant_id': tenant_id
                            }
            reports_db = DB('reports', **reports_db_config)
            group_access_db = DB('group_access', **reports_db_config)

            # query = f"select `First_Name`,`Last_Name` from `active_directory` where username ='{user}'"

            # query = f"SELECT first_name , last_name FROM ACTIVE_DIRECTORY where username = '{user}'"
            # query = f"SELECT first_name , last_name FROM ACTIVE_DIRECTORY where username = '{user}'"
            # query_data = group_access_db.execute_(query).to_dict(orient='records')[0]
            # user__=query_data['first_name']+" "+query_data['last_name']

            user__=user

            # Fetch reports
            logging.debug(f'########### Fetching reports for user `{user}`')


            if filtered_text!='':
                text = filtered_text
                table_name = 'report_requests'
                start_point = 0
                end_point = 9999
                header_name = None
                offset = end_point - start_point
                columns_list = ['REFERENCE_ID', 'REPORT_NAME', 'GENERATED_DATETIME', 'REQUESTED_BY', 'STATUS', 'REQUESTED_DATETIME','TAGS']
                columns_list_ = ["Reference Id","Report Name","Generated Datetime","Requested By","Status","Requested Datetime","Tags"]
                files, total = master_search(tenant_id = tenant_id, text = text, table_name = table_name, start_point = 0, offset = 9999, columns_list = columns_list, header_name=header_name)

                if end_point > total:
                    end_point = total
                if start_point == 1:
                    pass
                else:
                    start_point += 1
                print(f'#########Files got are: {files}')
                data = pd.DataFrame(files)
                filtered_df = data[data['REQUESTED_BY'] == user]
                files = filtered_df.to_dict(orient='records')
                print(f'Files got are: {files}')
                for i in files:
                    for key,value in i.items():
                        if key=='TAGS' or key=='tags':
                            try:
                                i[key]=ast.literal_eval(value)
                            except:
                                pass
                end_point = len(files)
                column_data_types = {"GENERATED_DATETIME":"unknown","REFERENCE_ID":"unknown","REPORT_NAME":"unknown","REQUESTED_BY":"unknown","REQUESTED_DATETIME":"unknown","STATUS":"unknown","TAGS":"unknown","generated_datetime":"string","reference_id":"number","report_name":"string","requested_by":"string","requested_datetime":"string","status":"date","tags":"string"}
                column_mapping = {"Generated Datetime":"GENERATED_DATETIME","Reference Id":"REFERENCE_ID","Report Name":"REPORT_NAME","Requested By":"REQUESTED_BY","Requested Datetime":"REQUESTED_DATETIME","Status":"STATUS","Tags":"TAGS"}
                pagination = {"start": start_point, "end": end_point, "total": end_point}
                final_data = []
                for i in files:
                    dic={}
                    for key,value in i.items():
                        key=key.lower()
                        dic[key]=value
                    res_dic = {**i,**dic}
                    final_data.append(res_dic)
                #return jsonify({"flag": True, "data": files, "pagination":pagination})
                respose_data = {
                'flag': True,
                'data': {
                    'files': final_data,
                    'buttons': [],
                    'field': [],
                    'tabs': [],
                    'column_data_types': column_data_types,
                    'column_mapping': column_mapping,
                    'excel_source_data': {},
                    'tab_type_mapping': {},
                    'pagination': pagination,
                    'column_order': columns_list_,
                    'children_dropdown': [],
                    'biz_rules': [],
                    'dropdown_values': {},
                    'cascade_object': "{}",
                    'get_report_view' : True}
                }
                return jsonify(response_data)

            # pagination_query = "SELECT reference_id,report_name,generated_datetime,requested_by,status,requested_datetime,tags FROM report_requests where requested_by=%s and parent_id is NULL ORDER BY requested_datetime desc LIMIT %s OFFSET %s"
            pagination_query=f"""
            SELECT REFERENCE_ID, REPORT_NAME, GENERATED_DATETIME, REQUESTED_BY, STATUS, REQUESTED_DATETIME,TAGS
            FROM report_requests 
            WHERE requested_by = %s AND parent_id IS NULL 
            ORDER BY requested_datetime DESC 
            OFFSET %s ROWS FETCH NEXT %s ROWS ONLY"""
            try:
                user_reports_data = reports_db.execute_(
                    pagination_query, params=[user,  offset_ , offset+1])

                user_reports_data_json = []

                logging.info(
                    f"############### user_reports_data {user_reports_data}")

                for idx, value in enumerate(user_reports_data["generated_datetime"].isnull()):
                    user_reports_data.loc[idx, "requested_datetime"] = str(
                        user_reports_data.loc[idx, "requested_datetime"])
                    if value:
                        user_reports_data.loc[idx, "generated_datetime"] = ""
                        user_reports_data.loc[idx, "generated_datetime"] = None
                    else:
                        user_reports_data.loc[idx, "generated_datetime"] = str(
                            user_reports_data.loc[idx, "generated_datetime"])

                user_reports_data_json = user_reports_data.to_dict('records')
                for i, report_data in enumerate(user_reports_data_json):
                    logging.info(f"############################ {report_data}")
                    if report_data['tags'] == "" or report_data['tags'] == '' or report_data['tags'] == None:
                        report_data['tags'] = {}
                    else:
                        report_data['tags'] = json.loads(report_data['tags'])
                    try:
                        report_data['requested_by']=user__
                        report_data['generated_datetime'] = datetime.strptime(report_data['generated_datetime'], "%Y-%m-%d %H:%M:%S")
                        report_data['generated_datetime']=report_data['generated_datetime'].strftime(f'%d-%b-%Y %H:%M:%S')
                        report_data['requested_datetime']=datetime.strptime(report_data['requested_datetime'], "%Y-%m-%d %H:%M:%S")
                        report_data['requested_datetime']=report_data['requested_datetime'].strftime(f'%d-%b-%Y %H:%M:%S')
                    except Exception as e:
                        print(e)
                        pass
                    print(f'report_data##{report_data}')
                    try:
                        report_data['TAGS']=json.loads(report_data['TAGS'])
                    except:
                        pass
                 

                    user_reports_data_json[i] = report_data

            except:
                message = 'Error fetching data from database'
                logging.exception(message)
                return jsonify({'flag': False, 'message': message})

            total_reports_query = f"SELECT COUNT(*) AS COUNT FROM report_requests where requested_by='{user}'"
            total_reports_df = reports_db.execute_(total_reports_query)

            total_count = list(total_reports_df['COUNT'])[0]
            logging.debug(total_count)


            # Get report types
            # Get group_id to know which reports are accessable
            group_db = DB("group_access", **reports_db_config)
            user_groups = get_group_ids(user, group_db)
            logging.info(f"##### user_groups: {user_groups}")

            reports_query = f"select reports_id from reports_access where group_id = {user_groups[0]}"
            reports_query_data = group_db.execute_(reports_query)['reports_id'].tolist()

            if len(reports_query_data) > 1:
                reports_query_data = tuple(reports_query_data)
                reports_query_data = "in " + str(reports_query_data)
            else:
                reports_query_data = "= " + str(reports_query_data[0])

            published_flag = 1
            reports_query = f"SELECT report_id,report_name FROM report_master where publish_flag = {published_flag} and parent_id is NULL and report_id {reports_query_data}"
            report_df = reports_db.execute_(reports_query)

            report_dict = []

            report_dict = report_df.to_dict('records')
            logging.debug(f"############ reports dict {report_dict}")

            for i, report in enumerate(report_dict):
                filters_query = f"select * from report_filter where report_id={report['report_id']}"
                filter_list = reports_db.execute_(
                    filters_query).to_dict('records')
                if len(filter_list) > 0:
                    for j, filter_data in enumerate(filter_list):
                        if filter_data['filter_options'] is not "" or filter_data['filter_options'] is not None:
                            string_list = filter_data["filter_options"].split("#$")
                            logging.info(
                                f"########### FIlter : {filter_data}and Filter Options:{string_list}")
                            if len(string_list) > 0:
                                for k, string_json in enumerate(string_list):
                                    if string_json is not '' and string_json is not "":
                                        string_list[k] = json.loads(string_json)
                            filter_list[j]["filter_options"] = string_list
                report['filters'] = filter_list
                report_dict[i] = report

            logging.debug(f"start {start_point} end {end_point} offset {offset}")

            logging.debug('***')
            logging.debug(f"start {start_point} end {end_point} offset {offset}")
            if start_point > end_point:
                start_point = end_point
            logging.debug('&&&&')
            logging.debug(f"start {start_point} end {end_point} offset {offset}")
            if end_point > total_count:
                end_point = total_count
            else:
                end_point = (int(offset) + int(start_point))
            logging.debug('####')
            logging.debug(f"start {start_point} end {end_point} offset {offset}")

            logging.debug(f"end : {end_point}")

            # define columns to display

            # column_mapping
            # column_order
            # column_data_types
            # reports
            # files
            # pagination

            ######################### Creating return Data ###############
            reports_columns = [col_name for col_name in user_reports_data.columns if col_name not in [
                'rn', 'report_id', 'request_id']]

            reports_coulmn_mapping = get_reports_column_mapping(reports_columns)
            reports_column_order = [] if not reports_coulmn_mapping else list(
                reports_coulmn_mapping.keys())

            reports_column_dt_types = get_reports_column_data_types(
                reports_columns, tenant_id)
            print(user_reports_data_json)
           
            
            
            # user_reports_data_json['TAGS'] = json.loads(user_reports_data_json['TAGS'])

            respose_data = {
                'flag': True,
                'data': {
                    'files': user_reports_data_json,
                    'buttons': [],
                    'field': [],
                    'tabs': [],
                    'excel_source_data': {},
                    'tab_type_mapping': {},
                    'pagination': {
                        'start': start_point,
                        'end': end_point,
                        'total': total_count},

                    'column_mapping': reports_coulmn_mapping,
                    'column_order': reports_column_order,
                    'children_dropdown': [],
                    'pdf_type': 'folder' if tenant_id else 'blob',
                    'biz_rules': [],
                    'dropdown_values': {},
                    'cascade_object': "{}",
                    'column_data_types': reports_column_dt_types,
                    'reports': report_dict,
                    'get_report_view' : True
                }
            }
            #logging.info(f'Response data: {respose_data}')
            
        except:
            logging.exception(
                'Something went wrong while getting reports queue. Check trace.')
            response_data = {
                'flag': False, 'message': 'System error [/get_reports_queue]! Please contact your system administrator.'}
        try:
            memory_after = measure_memory_usage()
            memory_consumed = (memory_after - memory_before) / \
                (1024 * 1024 * 1024)
            end_time = tt()
            time_consumed = str(end_time-start_time)
        except:
            logging.warning("Failed to calc end of ram and time")
            logging.exception("ram calc went wrong")
            memory_consumed = None
            time_consumed = None
            pass
        logging.info(f"## Reports API get_reports_queue Time and Ram checkpoint, Time consumed: {time_consumed}, Ram Consumed: {memory_consumed}")

        return jsonify(respose_data)


@app.route('/generate_reports', methods=['POST', 'GET'])
def generate_reports():
    try:
        memory_before = measure_memory_usage()
        start_time = tt()
    except:
        logging.warning("Failed to start ram and time calc")
        pass

    # Get data from UI
    data = request.json
    logging.info(f"## Reports info request data -- {data}")
    tenant_id = data.get('tenant_id', None)
    attr = ZipkinAttrs(
        trace_id=generate_random_64bit_string(),
        span_id=generate_random_64bit_string(),
        parent_span_id=None,
        flags=None,
        is_sampled=False,
        tenant_id=tenant_id
    )

    with zipkin_span(
        service_name='reports_api',
        span_name='generate_reports',
        transport_handler=http_transport,
        zipkin_attrs=attr,
        port=5010,
        sample_rate=0.5):
        
        requested_by = data.get('user', None)
        report_name = data.get('report_name', None)
        report_id = data.get("report_id", -1)

        filters = data.get('filters', {})
        try:
            filters["start_date"]=filters["start_date"]+' 00:00:00'
            filters["end_date"]=filters["end_date"]+' 23:59:59'
        except:
            filters["start_date"]='00-00-00 00:00:00'
            filters["end_date"]='00-00-00 00:00:00'


        text_filters = json.dumps(filters)

        # Sanity check
        if tenant_id is None:
            message = 'tenant_id not provided in request.'
            logging.error(message)
            return jsonify({'flag': False, 'message': message})

        if report_name is None:
            message = 'Report type not provided in request.'
            logging.error(message)
            return jsonify({'flag': False, 'message': message})

        logging.info(f"debugging report_name from front_end : {report_name}")
        logging.info(f"debugging report_id from front_end : {report_id}")

        # Generate file name if not given
        ist = pytz.timezone("Asia/Calcutta")
        timestamp = datetime.now(ist)
        timestamp_actual = timestamp
        timestamp__=timestamp.strftime(f'%d-%b-%Y')
        timestamp1 = timestamp.strftime(f'%d-%m-%Y %H:%M:%S')

        timestamp = str(timestamp)[:-13]
        # Add file to database
        reports_db_config = {
                                'host': os.environ['HOST_IP'],
                                'password': os.environ['LOCAL_DB_PASSWORD'],
                                'user': os.environ['LOCAL_DB_USER'],
                                'port': os.environ['LOCAL_DB_PORT'],
                                'tenant_id': tenant_id
                            }
        logging.debug(reports_db_config)


        # FETCHING DATA FROM REPORT CONFIG TABLES
        report_master_db = DB('reports', **reports_db_config)

        sheet_name = None
        report_id_idx = 0
        parent_start_time = -1
        parent_ref_id = -1
        parent_req_id = -1
        parent_id = -1
        #last_report_id = -1
        parent_query = f"SELECT report_id FROM report_master where parent_id = '{report_id}'"
        parent_id_val = report_master_db.execute_(parent_query)
        report_ids = list(parent_id_val['report_id'])
        if len(report_ids) == 0:
            report_ids = [report_id]
        else:
            parent_id = report_id
            report_ids.insert(0, report_id)

        logging.debug(f"List of report ids to process {report_ids}")
        for report_id in report_ids:
            logging.info(f"processing this {report_id}")
            # Generate a reference ID (10 digits)
            reference_id = uuid.uuid4().hex[:10].upper()
            if report_id_idx == 0:
                parent_ref_id = reference_id
            query = f'SELECT report_filename,report_out_format FROM report_template WHERE report_id={report_id}'
            report_data = report_master_db.execute_(query).to_dict('records')[0]

            filename_var = report_data["report_filename"]

            # If the file name in report Template is not mentioned it will  take report name as file name
            if filename_var == "" or filename_var == None:
                filename_var = report_name

            # Checking Fund Name
            fund_name_var = ""
            if 'fund_name' in filters:
                fund_name_var = filters['fund_name']

            file_type = report_data["report_out_format"]
            if parent_id == -1:
                file_name = f'{filename_var}-{fund_name_var}-{timestamp1}.{file_type}'
            else:
                if parent_ref_id == reference_id:
                    file_name = ""
                else:
                    file_name = f'{filename_var}-{timestamp1}#{reference_id}.{file_type}'
                    report_sheet_name_query = f"select report_name from report_master where report_id={report_id}"
                    sheet_name = report_master_db.execute_(
                        report_sheet_name_query).values.tolist()[0][0]
                    report_name = sheet_name
                    if report_id_idx == 0:
                        parent_start_time = 1

            base_query = f"select query_type,report_query,route from report_master where report_id={report_id}"

            base_query_df = report_master_db.execute_(base_query)
            base_query_list = base_query_df['report_query'].to_list()
            route_list = base_query_df['route'].tolist()
            query_type = base_query_df['query_type'].to_list()[0]

            # Updating ETA for reports
            average_query = f"SELECT AVG(process_time) AS ptime FROM report_requests WHERE report_id={report_id} AND report_name   ='{report_name}' AND status='Download' AND process_time IS NOT NULL"
            report_eta = report_master_db.execute_(average_query)
            eta = 0
            if len(report_eta['ptime']) > 0:
                if report_eta['ptime'][0] != None:
                    eta = int(report_eta['ptime'][0])
            # eta = sum(list_eta)/len(list_eta)
            # sum_query = f"SELECT SUM(process_time) AS sum FROM report_requests WHERE status='Processing' AND requested_datetime<'{timestamp}'"
            # report_eta = report_master_db.execute_(sum_query)
            formatted_requested_time = datetime.strptime(timestamp, '%Y-%m-%d %H:%M:%S')
            query = f"SELECT SUM(process_time) AS sum FROM report_requests WHERE status = 'Processing' AND requested_datetime < TO_DATE('{formatted_requested_time}', 'YYYY-MM-DD HH24:MI:SS')"
            report_eta = report_master_db.execute_(query)
            processing = 0
            if len(report_eta['sum']) > 0:
                if report_eta['sum'][0] != None:
                    processing = int(report_eta['sum'][0])
            eta += processing
            eta_datetime = dt.timedelta(seconds=int(eta))
            eta_datetime += timestamp_actual

            eta_datetime = str(eta_datetime)[:-13]

            filters_data = ""
            route = ""
            exec_query = ""
            if len(base_query_list) > 0 and query_type == 'query':
                exec_query = base_query_list[0]
                if bool(filters):


                    for column, value in filters.items():

                        if "{"+column in exec_query:
                            if len(report_ids) == 1:
                                column_query = f"select * from report_filter where report_id={report_id} and unique_name='{column}'"
                            elif len(report_ids) > 1:
                                column_query = f"select * from report_filter where report_id={parent_id} and unique_name='{column}'"

                            try:
                                column_df = report_master_db.execute_(column_query)
                                filter_type = column_df['filter_type'].to_list()[0]
                                logging.info("########## FILTER TYPE: ", filter_type)

                                if filter_type.strip().lower() == "string":
                                    exec_query = exec_query.replace(
                                        "{"+column+"}", "'"+value+"'")
                                    # exec_query=exec_query%(value)

                                elif filter_type.strip().lower() == "int" or filter_type.strip().lower() == "integer":
                                    exec_query = exec_query.replace(
                                        "{"+column+"}", value)
                                    # exec_query=exec_query%(value)
                                elif filter_type.strip().lower() == "date_picker":
                                    logging.debug(f"BEFORE {exec_query}")
                                    exec_query = exec_query.replace(
                                        "{"+column+"}", "'"+value+"'")
                                    logging.debug(f"AFTER {exec_query}")
                                    # exec_query=exec_query%(value)
                                elif filter_type.strip().lower() == "date_range":
                                    if "{"+column+"_begin"+"}" in exec_query:
                                        exec_query = exec_query.replace(
                                            "{"+column+"_begin"+"}", "'"+value['begin']+"'")
                                        # exec_query=exec_query%(value['begin'])
                                    elif "{"+column+"_start"+"}" in exec_query:
                                        exec_query = exec_query.replace(
                                            "{"+column+"_start"+"}", "'"+value['begin']+"'")
                                        # exec_query=exec_query%(value['begin'])
                                    if "{"+column+"_end"+"}" in exec_query:
                                        exec_query = exec_query.replace(
                                            "{"+column+"_end"+"}", "'"+value['end']+"'")
                                        # exec_query=exec_query%(value['end'])

                                elif filter_type.strip().lower() == "float":
                                    pass

                            except Exception as e:
                                logging.exception(e)
                                return jsonify({'flag': False, 'message': f'Failed to get the filter data {column}'})

                    # If there single quotes inside the select insert statement will fail, adding escape single quotes
                    # if "'" in exec_query:
                    #    exec_query= exec_query.replace("'","''")
            elif len(route_list) > 0 and query_type == 'route':
                if len(filters) > 0:
                    filters_data = json.dumps(filters)
                elif len(filters) == 0:
                    filters_data = json.dumps({})
           


                exec_query = route_list[0]

            report_id_idx = report_id_idx + 1

            reports_db = DB('reports', **reports_db_config)
            file_name = str(file_name).replace(' ', '_')
            file_name = str(file_name).replace(':', '_')
            logging.debug(f"################### File name after {file_name}")
            
            # insert_data = {
            #     'reference_id': reference_id,
            #     "report_name": report_name,
            #     'requested_by': requested_by,
            #     'report_output': file_name,
            #     'status': 'Processing',
            #     'requested_datetime': timestamp,
            #     'actual_report_query': exec_query,
            #     'report_id': str(report_id),
            #     'query_type': query_type,
            #     'tags': filters_data,
            #     'query_params': filters_data,
            #     'filters': text_filters,
            #     'fund_name': fund_name_var,
            #     'eta': eta_datetime,

            #     # 'scheduled_datetime' : scheduled_datetime

            # }
            formatted_requested_time = datetime.strptime(timestamp, '%Y-%m-%d %H:%M:%S')
            formatted_eta_datetime = datetime.strptime(eta_datetime, '%Y-%m-%d %H:%M:%S')
            print(formatted_requested_time,type(formatted_requested_time),'###req')
            print(formatted_eta_datetime,type(formatted_eta_datetime),"##eta")
            logging.info(type(eta_datetime),type(timestamp),eta_datetime,timestamp)
            insert_data={
                'REFERENCE_ID': reference_id,
                'REPORT_NAME': report_name,
                'REQUESTED_BY': requested_by,
                'REPORT_OUTPUT': file_name,
                'STATUS': 'Processing',
                'REQUESTED_DATETIME': formatted_requested_time,
                'ACTUAL_REPORT_QUERY': exec_query,
                'REPORT_ID': str(report_id),
                'QUERY_TYPE': query_type,
                'TAGS': filters_data,
                'QUERY_PARAMS': filters_data,
                'FILTERS': text_filters,
                'FUND_NAME': fund_name_var,
                'ETA': formatted_eta_datetime
            }

            logging.info(f"##################### INSERT DATA {insert_data}")
            report_master_db.insert_dict(insert_data, 'report_requests')
            logging.info(f"debug sheet name : {sheet_name}")
            logging.info(f"debug report name : {report_name}")
            logging.info(f"debug parent_id : {parent_id}")
            logging.info(f"debug report_id : {report_id}")
            if parent_id != -1:
                if parent_ref_id == reference_id:
                    logging.info("################Parent request id generated")
                    parent_req_id_query = "SELECT MAX(request_id) FROM report_requests"
                    parent_req_id = report_master_db.execute_(
                        parent_req_id_query).values[0][0]

                if parent_ref_id != reference_id:
                    query = f'SELECT `tags` from report_requests where reference_id = "{reference_id}"'
                    tags_df = report_master_db.execute_(query)
                    tags_query_list = tags_df['tags'].to_list()
                    tags_ = tags_query_list[0]
                    query_ex = f"UPDATE report_requests SET tags='{tags_}' WHERE reference_id = '{parent_ref_id}'"
                    report_master_db.execute_(query_ex)
                    req_id_query = "SELECT MAX(request_id) FROM report_requests"
                    req_id = report_master_db.execute_(req_id_query).values[0][0]
                    query = f"UPDATE report_requests SET parent_id='{parent_req_id}' WHERE request_id={req_id}"
                    template_query = f"UPDATE report_template SET report_sheetname='{sheet_name}' WHERE report_id={report_id}"
                    try:
                        report_master_db.execute_(query)
                        logging.info("sheet name updated in report template")
                    except:
                        logging.info(
                            "Sheet name is not updated in report template")
                    logging.info("before sheetname update")
                    report_master_db.execute_(template_query)
                    logging.info("parent id is updated in report_requests table")

            logging.info('##### sending to generate report')
            if parent_id == -1:
                parent_ref_id = -1
            # Additional data to producer for parent
            insert_data['parent_id'] = str(parent_id)
            insert_data['parent_req_id'] = str(parent_req_id)
            # insert_data['last_report_id'] = last_report_id
            insert_data['parent_ref_id'] = str(parent_ref_id)
            insert_data['parent_start_time'] = str(parent_start_time)
            insert_data['fund_name'] = fund_name_var

            logging.debug(insert_data)
            # Produce to reports consumer
            if report_id != parent_id:
                logging.info(f"########################producing {insert_data}")
                # produce('reports_consumer', {
                #         'tenant_id': tenant_id, 'report_name': report_name, **insert_data})
                
                result_message= reports_consumer({'tenant_id': tenant_id, 'report_name': report_name, **insert_data})
                logging.info(f"########################result_message {result_message}")
            if parent_id != -1:
                reference_id = parent_ref_id
        try:
            memory_after = measure_memory_usage()
            memory_consumed = (memory_after - memory_before) / \
                (1024 * 1024 * 1024)
            end_time = tt()
            time_consumed = str(end_time-start_time)
        except:
            logging.warning("Failed to calc end of ram and time")
            logging.exception("ram calc went wrong")
            memory_consumed = None
            time_consumed = None
            pass
        logging.info(f"## Reports API generate_reports Time and Ram checkpoint, Time consumed: {time_consumed}, Ram Consumed: {memory_consumed}")
        report_status_query = f"SELECT `status` FROM `report_requests` where reference_id like '%{reference_id}%'"
        report_status = reports_db.execute_(report_status_query).to_dict(orient="records")[0]['status']
        if report_status == 'Failed':
            return jsonify({'flag': True, 'message': f'The report not generated (Ref No. {reference_id})', 'reference_id': reference_id})
        else:
            return jsonify({'flag': True, 'message': f'The report will be available soon to download. (Ref No. {reference_id})', 'reference_id': reference_id})


@app.route('/get_report_view', methods=['POST', 'GET'])
def get_report_view():
    try:
        memory_before = measure_memory_usage()
        start_time = tt()
    except:
        logging.warning("Failed to start ram and time calc")
        pass
    data = request.json
    
    tenant_id = data.get('tenant_id', '')

    attr = ZipkinAttrs(
        trace_id=generate_random_64bit_string(),
        span_id=generate_random_64bit_string(),
        parent_span_id=None,
        flags=None,
        is_sampled=False,
        tenant_id=tenant_id
    )

    with zipkin_span(
        service_name='reports_api',
        span_name='get_report_view',
        transport_handler=http_transport,
        zipkin_attrs=attr,
        port=5010,
        sample_rate=0.5):

        reference_id = data.get('reference_id',None)
        db_config['tenant_id'] = tenant_id
        db = DB('reports', **db_config)
        query = 'SELECT * FROM `report_requests` WHERE `reference_id`=%s'
        try:
            report_data = db.execute_(query, params=[reference_id]).to_dict('records')[0]
            logging.info(f"############### Report Data: {report_data}")
            html_out = report_data["html_report"]
            if html_out is None:
                request_id = report_data['request_id']
                query1 = f"select * from `report_requests` WHERE parent_id = {request_id}"
                html_out = db.execute_(query1)['html_report'].to_list()
            return_data = {'flag': True, 'data': html_out}
        except Exception as e:
            logging.error(e)
            message = 'Report preview failed'
            return_data = {'flag': False, 'data': message}
        try:
            memory_after = measure_memory_usage()
            memory_consumed = (memory_after - memory_before) / \
                (1024 * 1024 * 1024)
            end_time = tt()
            time_consumed = str(end_time-start_time)
        except:
            logging.warning("Failed to calc end of ram and time")
            logging.exception("ram calc went wrong")
            memory_consumed = None
            time_consumed = None
            pass
        logging.info(f"## Reports API get_report_view Time and Ram checkpoint, Time consumed: {time_consumed}, Ram Consumed: {memory_consumed}")

        return return_data


@app.route('/download_report', methods=['POST', 'GET'])
def download_report():
    try:
        memory_before = measure_memory_usage()
        start_time = tt()
    except:
        logging.warning("Failed to start ram and time calc")
        pass
    # Get data from UI
    data = request.json
    logging.info(f'Recieved data: {data}')
    tenant_id = data.get('tenant_id', None)
    attr = ZipkinAttrs(
        trace_id=generate_random_64bit_string(),
        span_id=generate_random_64bit_string(),
        parent_span_id=None,
        flags=None,
        is_sampled=False,
        tenant_id=tenant_id
    )

    with zipkin_span(
        service_name='reports_api',
        span_name='download_report',
        transport_handler=http_transport,
        zipkin_attrs=attr,
        port=5010,
        sample_rate=0.5):

        reference_id = data.get('reference_id', None)
        report_name = data.get('report_name', None)
        generated_datetime = data.get('generated_datetime','')

        if reference_id is None:
            message = 'Reference ID is not provided.'
            logging.error(message)
            return jsonify({'flag': False, 'message': message})

        # Add file to database
        reports_db_config = {
                                'host': os.environ['HOST_IP'],
                                'password': os.environ['LOCAL_DB_PASSWORD'],
                                'user': os.environ['LOCAL_DB_USER'],
                                'port': os.environ['LOCAL_DB_PORT'],
                                'tenant_id': tenant_id
                            }
        reports_db = DB('reports', **reports_db_config)

        query = 'SELECT * FROM `report_requests` WHERE `reference_id`=%s'
        report_info = reports_db.execute_(query, params=[reference_id])

        report_file_name = list(report_info['report_output'])[0]
        logging.info(f"report_name####{report_name}")
        if 'Completed' in report_name :
            file='completed'
        if 'Rejected' in report_name :
            file='rejected'
        if 'Onhold' in report_name :
            file='onhold'

        if 'Completed' in report_name or 'process report' in report_name or 'Onhold' in report_name or  'Rejected' in report_name :
            # parsed_date = datetime.strptime(str(generated_datetime), "%d-%b-%Y %H:%M:%S")
            parsed_date = datetime.strptime(generated_datetime, "%a, %d %b %Y %H:%M:%S %Z")
            formatted_date = parsed_date.strftime("%m-%d-%Y_%H_%M")
            report_file_name=f'{file}_report--{formatted_date}.xlsx'
        if 'Process Report' in report_name:
            # parsed_date = datetime.strptime(str(generated_datetime), "%d-%b-%Y %H:%M:%S")
            parsed_date = datetime.strptime(generated_datetime, "%a, %d %b %Y %H:%M:%S %Z")
            formatted_date = parsed_date.strftime("%m-%d-%Y_%H_%M")
            report_file_name=f'Process_Report--{formatted_date}.xlsx'

        logging.info(f"report_file_name:{report_file_name}")
        report_path = Path(f'./reports/{report_file_name}')
        logging.info(f"@@@@@@@report_path: {report_path}")

        # for file in os.listdir("./reports"):
        #     if file.endswith(".docx"):
        #         print(os.path.join("./reports", file))
       
        if 'Completed' in str(report_file_name) or 'Process_Report' in str(report_file_name) or 'Onhold' in str(report_file_name) or  'Rejected' in str(report_file_name):
            try:
                with open(report_path, 'rb') as f:
                    report_blob = base64.b64encode(f.read())
            except:
                timestamp_match = re.search(r'_(\d{2})_(\d{2})\.xlsx', str(report_file_name))
                if timestamp_match:
                    hours, minutes = map(int, timestamp_match.groups())
                    timestamp = datetime.strptime(f"{hours:02d}_{minutes:02d}", '%H_%M')
                    updated_timestamp = timestamp - timedelta(minutes=1)
                    updated_timestamp_str = updated_timestamp.strftime('%H_%M')
                    updated_filename = str(report_file_name).replace(f"{hours:02d}_{minutes:02d}", updated_timestamp_str)
                    report_path = Path(f'./reports/{updated_filename}')
                    logging.info(f"@@@@@@@report_path: {report_path}")
                    with open(report_path, 'rb') as f:
                        report_blob = base64.b64encode(f.read())
        try:
            workbook = openpyxl.load_workbook(report_path)
            sheet = workbook.active
            workbook_modified = openpyxl.Workbook()
            worksheet_modified = workbook_modified.active

            # Copy values and styles from the original worksheet to the new worksheet
            for row_num, row in enumerate(sheet.iter_rows(min_row=1), start=1):
                for col_num, cell in enumerate(row, start=1):
                    new_cell = worksheet_modified.cell(row=row_num, column=col_num, value=cell.value)
                    if isinstance(cell.value, (int, float)):
                        new_cell.number_format = '0'
            workbook_modified.save(report_path)
        except Exception as e:
            logging.info(f'######{e} exception')


        else:
            with open(report_path, 'rb') as f:
                    report_blob = base64.b64encode(f.read())

        try:
            logging.debug('DOWNLOADING REPORT')
            return_data = {'flag': True, 'blob': report_blob.decode('utf-8'), 'filename': f'{report_file_name}'}
        except:
            message = 'Something went wrong while downloading report.'
            logging.exception(message)
            return_data = {'flag': False, 'message': message}
        try:
            memory_after = measure_memory_usage()
            memory_consumed = (memory_after - memory_before) / \
                (1024 * 1024 * 1024)
            end_time = tt()
            time_consumed = str(end_time-start_time)
        except:
            logging.warning("Failed to calc end of ram and time")
            logging.exception("ram calc went wrong")
            memory_consumed = None
            time_consumed = None
            pass
        logging.info(f"## Report API download_report Time and Ram checkpoint, Time consumed: {time_consumed}, Ram Consumed: {memory_consumed}")


        return return_data

### -------------------- from here , the code is responsible to genearate the accuarcy report ----------------


def get_total_mandatory_fields(queues_db):
    get_mad_filds=f"SELECT display_name,unique_name FROM `field_definition` WHERE mandatory=1"
    mad_df=queues_db.execute_(get_mad_filds)
    mandatory_fields=mad_df['unique_name'].to_list()
    total_mandatory_fields=len(mandatory_fields)
#     print(f"#### mad fields is {mandatory_fields} and its length is {len(mandatory_fields)}")
    return total_mandatory_fields,mandatory_fields

def get_edited_fields(queues_db):
    qry=f"SELECT `case_id`,`fields_changed` FROM `field_accuracy`"
    field_accuracy_df=queues_db.execute_(qry) 
    return field_accuracy_df

def get_ocr_fields(extraction_db,mandatory_fields,start_date,end_date):
    mad_flds_str=','.join(mandatory_fields)
    logging.info(f"### mad_flds str is {mad_flds_str}")
    qry=f"select `case_id`,{mad_flds_str} from ocr where created_date between '{start_date}' and '{end_date}'"
    # qry=f"select `case_id`,`entity`,`facility`,`date_of_application`,`purpose_of_application`,`applicant_customers_name`,`c_consignor_exporter_name`,`c_consignee_importer_name`,`c_consignor_exporter_address`,`c_consignee_importer_address`,`c_name_of_authorised_agent`,`beneficiary_name`,`applicant_customers_address_country`,`beneficiary_address`,`bl_reference_number`,`bl_shipper_name`,`bl_consignee_name`,`bl_consignee_importer_address`,`bl_shipper_country`,`bl_consignee_country`,`bl_notify_party_name`,`bl_notify_party_address`,`bl_place_of_receipt_location_country`,`bl_port_loading_location_country`,`bl_port_discharge_location_country`,`bl_place_of_delivery_location_country`,`bl_final_destination_location_country`,`email_of_aqad`,`product_type_f_v`,`ba_ba_draft_no`,`do_supplier_beneficiary_name`,`do_buyer_name`,`do_buyer_address`,`do_supplier_beneficiary_address`,`do_buyer_country`,`do_supplier_country`,`inv_supplier_beneficiary_name`,`inv_buyer_drawee_name`,`inv_invoice_amount_figure`,`inv_invoice_amount_currency` from ocr limit 7"
    ocr_df=extraction_db.execute_(qry)
    return ocr_df

### MERGE the dataframes to proceed furthur
def merge_df(df1,df2):
    merged_df = pd.merge(df1, df2, on='case_id', how='right')
    return merged_df

def generate_case_wise_accuracy(merged_df_dict,total_mandatory,mandatory_fields):
    case_accuracy={}
    for field_data in merged_df_dict:
        try:
            logging.info(f"#### field_data is {field_data['fields_changed']}")
            field_changed=field_data['fields_changed']
            case_id=field_data['case_id']
    #         if math.isnan(field_changed):
            if type(field_changed)==str:
                fields_changed=json.dumps(field_changed)
                field_changed=list(json.loads(json.loads(fields_changed)).keys())
                filtered_fields_changed = [field for field in field_changed if field in mandatory_fields]
                field_data_unchanged = {key: value for key, value in field_data.items() if key not in field_changed}
                condition = lambda x: x is None
                empty_vals = [value for value in field_data_unchanged.values() if condition(value)]
                not_extracted_flds=len(empty_vals)
                edited_flds=len(filtered_fields_changed)
    #             print(f"### filtered_values are {empty_vals}")
                extract_fields_cnt=total_mandatory-not_extracted_flds
                case_accuracy[case_id]={'Mandatory Fields':total_mandatory,'Extracted Fields':extract_fields_cnt,'Edited Fields':edited_flds,'Not Extracted':not_extracted_flds}
            else:
                condition = lambda x: x is None
                # Using a list comprehension
                empty_vals = [value for value in field_data.values() if condition(value)]
                not_extracted_flds=len(empty_vals)
    #             print(f"### filtered_values are {empty_vals}")
                extract_fields_cnt=total_mandatory-not_extracted_flds
                case_accuracy[case_id]={'Mandatory Fields':total_mandatory,'Extracted Fields':extract_fields_cnt,'Edited Fields':0,'Not Extracted':not_extracted_flds}
        except Exception as e:
            logging.info(f"### Exception occured {e}")
            continue

    return case_accuracy

#### Generate excel with the Final data
def generate_excel(data,file_name):
    df = pd.DataFrame(data).T
    summary_data={}
    mandatory_fields=df['Mandatory Fields'].to_list()
    extracted_fields=df['Extracted Fields'].to_list()
    edited_fields=df['Edited Fields'].to_list()
    not_extracted_fields=df['Not Extracted'].to_list()
    
    summary_data['Mandatory Fields']=round(sum(mandatory_fields),1)
    summary_data['Extracted Fields']=round(sum(extracted_fields),1)
    summary_data['Edited Fields']=round(sum(edited_fields),1)
    summary_data['Not Extracted']=round(sum(not_extracted_fields),1)
    ## Formula to calculate Accuracy (correctly extracted fields / (total no.of fields- unavailable fields ))*100
    accuracy = 100-(100*sum(edited_fields)/sum(mandatory_fields))
    summary_data['Accuracy']=accuracy


    logging.info(f"#### SUMMARY DATA is {summary_data}")
    df_summary = pd.DataFrame([summary_data])

    try:
        path=f'/var/www/reports_api/reports/{file_name}'
        logging.info(f"### writing to excel files path is {path}")
        with pd.ExcelWriter(path) as writer:
            df_summary.to_excel(writer, sheet_name='Summary',index=False)
            df.to_excel(writer, sheet_name='Case Wise', index_label="Case ID")

            # Access the worksheet and set column widths
            summary_sheet = writer.sheets['Summary']
            case_sheet = writer.sheets['Case Wise']

            # Set column width for 'Summary' sheet
            for i, column in enumerate(df_summary.columns):
                column_len = max(df_summary[column].astype(str).apply(len).max(), len(column))
                print(f"######## column len for summary sheet {column_len}")
                summary_sheet.set_column(i, i, column_len)

            # Set column width for 'Case Wise' sheet
            for i, column in enumerate(df.columns):
                column_len = max(df[column].astype(str).apply(len).max(), len(column))
                print(f"######## column len for case wise sheet{column_len}")
                case_sheet.set_column(i, i,20)
    except Exception as e:
        logging.error(f"### Error writing to Excel file: {e}")
    
    return

   

@app.route('/generate_accuracy_report', methods=['POST', 'GET'])
def generate_accuracy_report():
    try:
        memory_before = measure_memory_usage()
        start_time = tt()
    except:
        logging.warning("Failed to start ram and time calc")
        pass
    data = request.json

    logging.info(f"##### Request data in GENERATE ACCURACY REPORT IS {data}")
    
    tenant_id = data.get('tenant_id', '')
    ui_data=data['ui_data']
    file_name=ui_data.get('report_output','Accurary_report')
    start_date=data['start_date']
    end_date=data['end_date']

    if start_date == end_date:
        logging.info(f"start and end dates are same so end date incrementing for 1 day")
        logging.info(f"end_date is {end_date} and type is {type(end_date)}")
        end_date = datetime.strptime(end_date, "%Y-%m-%d")
        end_date = end_date + timedelta(days=1)
        end_date= end_date.strftime("%Y-%m-%d")
        logging.info(f"end date after increment is {end_date} and type is {type(end_date)}")
    else:
        pass

    attr = ZipkinAttrs(
        trace_id=generate_random_64bit_string(),
        span_id=generate_random_64bit_string(),
        parent_span_id=None,
        flags=None,
        is_sampled=False,
        tenant_id=tenant_id
    )

    with zipkin_span(
        service_name='reports_api',
        span_name='generate_accuracy_report',
        transport_handler=http_transport,
        zipkin_attrs=attr,
        port=5010,
        sample_rate=0.5):

        try:
            db_config['tenant_id'] = tenant_id
            queues_db=DB('queues',**db_config)
            extraction_db=DB('extraction',**db_config)
            
            total_mandatory,mandatory_fields=get_total_mandatory_fields(queues_db)
            field_accuracy_df=get_edited_fields(queues_db)
            # print(f"### field_accuracy_df is {field_accuracy_df}")
            ocr_df=get_ocr_fields(extraction_db,mandatory_fields,start_date,end_date)
            # print(f" ########## ocr df is {ocr_df}")
            merged_df=merge_df(field_accuracy_df,ocr_df)
            merged_df_dict=merged_df.to_dict(orient='records')

            case_wise_accuracy=generate_case_wise_accuracy(merged_df_dict,total_mandatory,mandatory_fields)
            logging.info(f"#### CASE WISE ACCURACY IS {case_wise_accuracy}")

            excel_status=generate_excel(case_wise_accuracy,file_name)
            logging.info(f"#### GENERATE EXCEL STATUS is {excel_status}")

            response_data={"flag":True,"report_data":{"message":"Successfully excel generated"}}
        except Exception as e:
            logging.info(f"### Issue while generating the accuracy report {e}")
            response_data={"flag":False,"report_data":{"message":"Excel not generated"}}

    return jsonify(response_data)


@app.route('/audit_report', methods=['POST', 'GET'])
def audit_report():
    try:
        memory_before = measure_memory_usage()
        start_time = tt()
    except:
        logging.warning("Failed to start ram and time calc")
        pass
    data = request.json  
    tenant_id = data['ui_data']['tenant_id']
    trace_id = generate_random_64bit_string()

    attr = ZipkinAttrs(
        trace_id=generate_random_64bit_string(),
        span_id=generate_random_64bit_string(),
        parent_span_id=None,
        flags=None,
        is_sampled=False,
        tenant_id=tenant_id if tenant_id is not None else ''
    )

    with zipkin_span(
            service_name='folder_monitor',
            zipkin_attrs=attr,
            span_name='folder_monitor',
            transport_handler=http_transport,
            sample_rate=0.5
    ) as zipkin_context:
        data = request.json
        logging.info(f"request data: {data}")

        try:
            #date_filters = data.get('filters',{})
            start_date = data['start_date']
            end_date = data['end_date']
            user=data['ui_data']['REQUESTED_BY']

            tenant_id = data['ui_data']['tenant_id']
            db_config['tenant_id'] = tenant_id

            
            query = """
                SELECT pq.case_id, ql.queue 
                FROM PROCESS_QUEUE pq 
                JOIN QUEUE_LIST ql ON ql.case_id = pq.case_id 
                WHERE CREATED_DATE >= TO_DATE(:date_param, 'YYYY-MM-DD HH24:MI:SS') 
                AND CREATED_DATE <= TO_DATE(:date_param1, 'YYYY-MM-DD HH24:MI:SS')
                AND ql.queue like '%maker%'
            """
            queue_db = DB('queues', **db_config)
            stats_db = DB('stats', **db_config)
            extraction_db = DB("extraction", **db_config)

            params = {'date_param': start_date, 'date_param1': end_date}
            df = queue_db.execute_(query, params=params)

            case_ids = df['case_id'].tolist()
            logging.info(f"case_ids{case_ids}")

            outputs = []
            #serial_number = 1
            for serial_number, case_id in enumerate(case_ids, start=1):
                audit_query = f"SELECT case_id, updated_date, ingested_queue, api_service FROM audit_ WHERE case_id LIKE '%{case_id}%'  AND api_service IN ('create_case_id', 'update_queue')"
                audit_case_data = stats_db.execute_(audit_query)
                
                logging.info(f"case_id{case_id}")
                
                grouped = audit_case_data.groupby('ingested_queue')
                
                none_group = audit_case_data[audit_case_data['ingested_queue'].isna()]
            
                query_file_recieved = f"SELECT CREATED_DATE from PROCESS_QUEUE where case_id like '%{case_id}%'"
                df_file_recieved = queue_db.execute_(query_file_recieved)
                logging.info(f"df_file_recieved{df_file_recieved}")
                query_party = f"SELECT PARTY_ID , PARTY_NAME , MOVED_BY from OCR where case_id like '%{case_id}%'"
                df_party = extraction_db.execute_(query_party)
                logging.info(f"df_party{df_party}")
                try:
                    party_id=df_party['party_id'][0]
                    party_name = df_party['party_name'][0]
                    moved_by = df_party['MOVED_BY'][0]
                except:
                    party_id=None
                    party_name = None
                    moved_by = None


                if not df_file_recieved.empty:
                    desired_updated_date = df_file_recieved['created_date'].iloc[0]
                else:
                    desired_updated_date=Timestamp('00:00:00')

                
                # Calculate relevant timestamps and time differences
                try:
                    file_received_time = desired_updated_date
                except :
                    pass
                try:
                    maker_ingestion_time = grouped.get_group('maker_queue')['updated_date'].min()
                except:
                    pass
            
                try:
                    completed_ingested_time = grouped.get_group('accepted_queue')['updated_date'].min()
                except:
                    
                    completed_ingested_time = '0000-00-00 00:00:00'
                    pass
            

                try:
                    rejected_ingested_time = grouped.get_group('rejected_queue')['updated_date'].min()
                except:
                    
                    rejected_ingested_time = '0000-00-00 00:00:00'
                    pass
            
                

                if rejected_ingested_time == '0000-00-00 00:00:00':
                    try:
                        
                        Total_handling_time= completed_ingested_time - maker_ingestion_time
                    except:
                        Total_handling_time = '0000-00-00 00:00:00'
                else:
                    try:
                        
                        Total_handling_time=rejected_ingested_time - maker_ingestion_time
                    except:
                        Total_handling_time = '0000-00-00 00:00:00'
        
                output = {
                        'serial_number': serial_number,
                    
                        'case_id': case_id,
                        'Party_id':party_id,
                        'Party_name':party_name,
                        'Business Rules Validation' : 'NA',
                        'CLIMS Create API Request time':'NA',
                        'CLIMS Create API Response Time':'NA',
                        
                        'Case creation time stamp': file_received_time,
            
                        'Maker queue in Time Stamp': maker_ingestion_time,
                        'Completed queue in time Stamp':completed_ingested_time,
                        'Rejected queue in time Stamp':rejected_ingested_time,
                        'Total Handling time':Total_handling_time,
                        'Maker Name':moved_by,
                        'User ID':user

                        
                        
                        
                        
                    }
                outputs.append(output)
                logging.info(f"output{output}")
            
            logging.info(f"outputs##{outputs}")
            for output in outputs:
                output['Case creation time stamp'] = str(output['Case creation time stamp'])
                output['Maker queue in Time Stamp'] = str(output['Maker queue in Time Stamp'])
                output['Completed queue in time Stamp'] = str(output['Completed queue in time Stamp'])
                output['Rejected queue in time Stamp'] = str(output['Rejected queue in time Stamp'])
                output['Total Handling time'] = str(output['Total Handling time'])
            return_json_data = {}
            logging.info(f'{return_json_data}###return_json_data#########return_json_data')
            return_json_data['message']='Successfully generated the report'
            return_json_data['excel_flag']= 1
            return_json_data['flag'] = True
            return_json_data['data'] = [{'row_data':outputs}]
            data['report_data'] = return_json_data

            #generate_report({'tenant_id': tenant_id, **data})
            logging.info(f'{return_json_data}###############return_json_data')
            
        except Exception as e:
            logging.info(f"error at audit_Report {e}")
            logging.debug(f"{e} ####issue")
            return_json_data = {}
            return_json_data['flag'] = False
            return_json_data['message'] = 'Failed!!'
            return jsonify(return_json_data)
    try:
        memory_after = measure_memory_usage()
        memory_consumed = (memory_after - memory_before) / \
            (1024 * 1024 * 1024)
        end_time = tt()
        memory_consumed = f"{memory_consumed:.10f}"
        logging.info(f"checkpoint memory_after - {memory_after},memory_consumed - {memory_consumed}, end_time - {end_time}")
        time_consumed = str(round(end_time-start_time,3))
    except:
        logging.warning("Failed to calc end of ram and time")
        logging.exception("ram calc went wrong")
        memory_consumed = None
        time_consumed = None
        pass

    # insert audit
    try:
        audit_data = {"tenant_id": tenant_id, "user_": "", "case_id": "",
                        "api_service": "folder_monitor", "service_container": "reportsapi",
                        "changed_data": "New file received","tables_involved": "","memory_usage_gb": str(memory_consumed), 
                        "time_consumed_secs": time_consumed, "request_payload": json.dumps(data), 
                        "response_data": json.dumps(return_json_data), "trace_id": trace_id,
                        "session_id": "","status":json.dumps(return_json_data['flag'])}
  
        insert_into_audit(audit_data)
    except:
        logging.info(f"issue in the query formation")
    return jsonify(return_json_data)



    
def extract_numeric_value(value):
    if isinstance(value, str):
        cleaned_value = re.sub(r'[^\d.]', '', value)  # Remove non-numeric characters except dot
        return pd.to_numeric(cleaned_value, errors='coerce') if cleaned_value else None
    return None

def replace_empty_with_none(value):
    return None if value == "" else value

def create_dataframe_from_json(json_data_list):
    columns_data = {}

    try:
        for idx, json_data in enumerate(json_data_list):
            if json_data is not None and isinstance(json_data, dict):  # Check if json_data is not None and is a dictionary
                for key, value in json_data.items():
                    if value is not None and value != 'null':
                        try:
                            data_dict = json.loads(value)
                            for nested_key, nested_value in data_dict.items():
                                # nested_key_cleaned = re.sub(r'\W', '_', nested_key)  # Replace non-alphanumeric characters with underscore
                                column_name = f"{nested_key}"
                                if column_name not in columns_data:
                                    columns_data[column_name] = []
                                numeric_value = extract_numeric_value(replace_empty_with_none(nested_value))
                                columns_data[column_name].append(numeric_value)
                        except json.JSONDecodeError as json_error:
                            print(f"Error decoding JSON at index {idx}: {json_error}")
                            # Handle the error if needed
                    else:
                        # Handle the case where value is None
                        for nested_key in columns_data.keys():
                            columns_data[nested_key].append(None)
            else:
                # Handle the case where json_data is None or not a dictionary
                print(f"Skipping invalid data at index {idx}: {json_data}")
                for nested_key in columns_data.keys():
                    columns_data[nested_key].append(None)

        # Fill missing keys with None
        if len(columns_data) !=0:
            max_len = max(len(v) for v in columns_data.values())
            for key, values in columns_data.items():
                values.extend([None] * (max_len - len(values)))

        df = pd.DataFrame(columns_data)
        return df
    except Exception as e:
        logging.info(f"ERROR : {e}")
        return None


def convert_to_custom_format(date_str):
    formats = [
        "%d-%m-%Y",
        "%d/%m/%Y",
        "%b %Y",
        "%d %b %Y",
        "%B %Y",
        "%m-%Y",
        "%m/%Y",
        "%d %B, %Y",
        "%B-%Y",
    ]

    for date_format in formats:
        try:
            dt = datetime.strptime(date_str, date_format)
            return dt.strftime("%d-%b-%y")
        except ValueError as e:
            logging.info(f"{e}###wrong format")
            pass

    return None



@app.route('/consolidated_report', methods=['POST', 'GET'])
def consolidated_report():
    logging.info(f'Calling Consolidated report')
    try:
        memory_before = measure_memory_usage()
        start_time = tt()
    except:
        logging.warning("Failed to start ram and time calc")
        pass
    data = request.json
    tenant_id = data['ui_data']['tenant_id']
        
    
    trace_id = generate_random_64bit_string()

    attr = ZipkinAttrs(
        trace_id=generate_random_64bit_string(),
        span_id=generate_random_64bit_string(),
        parent_span_id=None,
        flags=None,
        is_sampled=False,
        tenant_id=tenant_id if tenant_id is not None else ''
    )

    with zipkin_span(
            service_name='folder_monitor',
            zipkin_attrs=attr,
            span_name='folder_monitor',
            transport_handler=http_transport,
            sample_rate=0.5
    ) as zipkin_context:
        data=request.json
        tenant_id = data['ui_data']['tenant_id']
        db_config['tenant_id'] = tenant_id
        logging.info(f"Request Data is: {data}")
        queue_db = DB('queues', **db_config)
        queue_id=data.get("queue_id",{})
        case_id=data['ui_data']['REFERENCE_ID']
        reports_name=data['ui_data']['report_name']
        extraction_db = DB('extraction', **db_config)
        group_db = DB("group_access", **db_config)
        queues_db = DB("queues", **db_config)
        user=data['ui_data']['REQUESTED_BY']

        start_date = data['start_date']
        end_date = data['end_date']
        zonee = data.get('zone', '')
        ist = pytz.timezone("Asia/Calcutta")
        timestamp = datetime.now(ist)
        timestamp_actual = timestamp
        timestamp1 = timestamp.strftime(f'%d-%m-%Y %H:%M:%S')

        timestamp = str(timestamp)[:-13]

        res= {}
        logging.info(f"reports_name{reports_name}")
        try:
            def final_result_fun(case_ids,res_columns,queue,case_id_db_data):
                print(f'########case_id_db_data is : {case_id_db_data}')
                if len(case_id_db_data):
                    if len(case_id_db_data)==1:
                        case_id_tuple = case_id_db_data['case_id'][0]
                    else:
                        case_id_tuple = tuple(case_id_db_data['case_id'])
                else:
                    case_id_tuple="('"+'0'+"')"
                print(f'case_id_tuple is: {case_id_tuple}')
                if queue=='Maker':
                    
                    if len(case_id_db_data)==1:
                        query=f"""SELECT OCR.CASE_ID, OCR.PARTY_ID, OCR.PARTY_NAME, OCR.CUSTOMER_NAME, OCR.DRAWING_POWER ,OCR.COMMENTS
                        FROM OCR WHERE OCR.CASE_ID='{case_id_tuple}'
                        """
                    else:
                        query=f"""SELECT OCR.CASE_ID, OCR.PARTY_ID, OCR.PARTY_NAME, OCR.CUSTOMER_NAME, OCR.DRAWING_POWER ,OCR.COMMENTS
                        FROM OCR WHERE OCR.CASE_ID in {case_id_tuple}
                        """
                else:
                    if len(case_id_db_data)==1:
                        query=f"""SELECT OCR.CASE_ID, OCR.PARTY_ID, OCR.PARTY_NAME, OCR.CUSTOMER_NAME, OCR.DRAWING_POWER,OCR.REJECTED_COMMENTS 
                        FROM OCR WHERE OCR.CASE_ID='{case_id_tuple}'
                        """
                    else:
                        query=f"""SELECT OCR.CASE_ID, OCR.PARTY_ID, OCR.PARTY_NAME, OCR.CUSTOMER_NAME, OCR.DRAWING_POWER,OCR.REJECTED_COMMENTS 
                            FROM OCR WHERE OCR.CASE_ID in {case_id_tuple}
                            """
                df=extraction_db.execute_(query)
                logging.info(f"query df is {df}")
                df.rename(columns={
                    'case_id': 'ACE Case ID',
                    'party_id': 'Party ID',
                    'party_name': 'Party name',
                    'customer_name': 'User name',
                    'drawing_power': 'Drawing power Amount',
                    'hold_comments':'Hold Comments',
                    'rejected_comments':'Rejected Comments'
                }, inplace=True)
                res=[]



                for each_case in case_ids:
                    case = each_case['case_id']
                    logging.info(f'{case}#####')
                    
                    try:
                        
                        qry=f"SELECT STOCKS,CREDITORS,DEBTORS FROM OCR WHERE  CASE_ID='{case}'" 
                        json_data_list = extraction_db.execute_(qry).to_dict(orient="records")
                        values= create_dataframe_from_json(json_data_list)
                        values = values.to_dict()
                        query = f"SELECT PARTY_ID from OCR where CASE_ID = '{case}'"
                        Party = extraction_db.execute_(query)
                        

                        final_result={}
                        try:
                            for i in values:
                                final_result[i]=values[i][0]
                        except:
                            logging.info(f"exception in appending final_result")
                        if not Party.empty:
                            query = f"SELECT COMPONENT_NAME, MARGIN from AGE_MARGIN_WORKING_UAT where PARTY_ID = '{Party['party_id'][0]}'"
                            margins = extraction_db.execute_(query)
                            try:
                                rm_margin = margins.loc[margins['component_name'] == 'RAW MATERIALS INSURED', 'margin'].values[0] if not margins.loc[margins['component_name'] == 'RAW MATERIALS INSURED', 'margin'].empty else None
                                wip_margin = margins.loc[margins['component_name'] == 'WORK IN PROGRESS INSURED', 'margin'].values[0] if not margins.loc[margins['component_name'] == 'WORK IN PROGRESS INSURED', 'margin'].empty else None
                                fg_margin = margins.loc[margins['component_name'] == 'FINISHED GOODS INSURED', 'margin'].values[0] if not margins.loc[margins['component_name'] == 'FINISHED GOODS INSURED', 'margin'].empty else None
                                stores_and_spares_margin = margins.loc[margins['component_name'] == 'STOCK & STORES INSURED', 'margin'].values[0] if not margins.loc[margins['component_name'] == 'STOCK & STORES INSURED', 'margin'].empty else None
                                book_debts_margin = margins.loc[margins['component_name'] == 'BOOK DEBTS', 'margin'].values[0] if not margins.loc[margins['component_name'] == 'BOOK DEBTS', 'margin'].empty else None
                                upto_90_days_debts_margin = margins.loc[margins['component_name'] == 'BOOK DEBTS UPTO 90 DAYS', 'margin'].values[0] if not margins.loc[margins['component_name'] == 'BOOK DEBTS UPTO 90 DAYS', 'margin'].empty else None
                                upto_120_days_debts_margin = margins.loc[margins['component_name'] == 'BOOK DEBTS UPTO 120 DAYS', 'margin'].values[0] if not margins.loc[margins['component_name'] == 'BOOK DEBTS UPTO 120 DAYS', 'margin'].empty else None
                                creditors_margin = margins.loc[margins['component_name'] == 'CREDITORS', 'margin'].values[0] if not margins.loc[margins['component_name'] == 'CREDITORS', 'margin'].empty else None

                                
                                final_result['RM Margin'] = rm_margin
                                final_result['WIP Margi'] = wip_margin
                                final_result['FG Margin'] = fg_margin
                                final_result['Stores and spares margin'] = stores_and_spares_margin
                                final_result['Book debts Margin'] = book_debts_margin
                                final_result['Upto 90 days-Debts margin'] = upto_90_days_debts_margin
                                final_result['Upto 120 days-Debts margin'] = upto_120_days_debts_margin
                                final_result['Creditors margin'] = creditors_margin
                            except Exception as e:
                                logging.info(f"####issue in margins {e}")

                        res.append(final_result)
                        
                        
                    except Exception as e:
                        logging.info(f"Error at Printing  Json data {e}")
                result_df = pd.DataFrame(res)
                renamed_columns = {
                    'Drawing Power': 'Drawing power Amount',
                    'Raw Materials': 'Raw materials',
                    'Total Creditors': 'Creditors',
                    'Work in Process': 'Work in progress',
                    'Stock in Transit': 'Stock in transit'
                }

                # Check if 'stores and spares' or 'Stores & Spares' exists and rename accordingly
                if 'stores and spares' in result_df.columns:
                    renamed_columns['stores and spares'] = 'Stores and spares'
                elif 'Stores  & Spares' in result_df.columns:
                    renamed_columns['Stores  & Spares'] = 'Stores and spares'

                # Rename columns
                result_df.rename(columns=renamed_columns, inplace=True)
                
                
                logging.info(f"#####result_df{result_df}")
                logging.info(f"#####df{df}")
                final_df = pd.concat([result_df, df], axis=1)
                logging.info(f"#####final_df{final_df}")
                logging.info(f'{res_columns} ###res_columns')
                logging.info(f"{type(res_columns)} ###res_columns")
                final_df.reset_index(drop=True, inplace=True)
                res_columns.reset_index(drop=True, inplace=True)
                final_df = final_df.loc[:,~final_df.columns.duplicated()]
                res = res_columns.append(final_df, ignore_index=True)

                logging.info(f"length of the res :{len(res)}")
                res.insert(0, 'S.No', range(1, len(res) + 1))
                #logging.info(f"res for rejected comments{res["REJECTED_COMMENTS"].tolist()}")
                logging.info(f"#####res{res}")
                return res

                

            
            if 'completed' in reports_name.lower():
                file='completed'
                
                query_case_ids=f"""SELECT PROCESS_QUEUE.CASE_ID 
                        FROM hdfc_queues.PROCESS_QUEUE 
                        INNER JOIN HDFC_QUEUES.QUEUE_LIST ON PROCESS_QUEUE.CASE_ID = QUEUE_LIST.CASE_ID 
                        WHERE QUEUE_LIST.QUEUE = 'accepted_queue' 
                        AND PROCESS_QUEUE.LAST_UPDATED >= TO_DATE('{start_date}', 'YYYY-MM-DD HH24:MI:SS') 
                        AND PROCESS_QUEUE.LAST_UPDATED <= TO_DATE('{end_date}', 'YYYY-MM-DD HH24:MI:SS')"""
                case_ids = queues_db.execute_(query_case_ids).to_dict(orient="records")
                case_id_db_data = queues_db.execute_(query_case_ids)
                

                

                columns = [ "ACE Case ID", "Party ID", "Party name", "User name", "Stock", "Stock margin", "Raw materials", "RM Margin", "Work in progress", "WIP Margin", "Finished Goods", "FG Margin", "Stores and spares", "Stores and spares margin", "Book Debts & Receivables (Domestic&Export)", "Book debts Margin", "Upto 90 days", "Upto 90 days-Debts margin", "Upto 120 days", "Upto 120 days-Debts margin", "Creditors", "Creditors margin", "Advance from suppliers", "Advance to suppliers", "Stock in transit", "Group company debtors", "GST receivables (tax components)", "Unbilled debtors", "Bullion stock", "Obsolete stock", "Under deposits/Bhishi's Scheme's", "Standard Gold", "WCL sanctioned in CAM", "WCL sanctioned in Limit module", "Unsecured utilizations", "Total utilizations", "Drawing power Amount", "Remarks" ]
                queue = 'accepted_queue'
                res_columns = pd.DataFrame(columns=columns)
                res=final_result_fun(case_ids,res_columns,queue ,case_id_db_data)
                given_data = res.to_dict()
                final_data = {'row_data': []}

                # Iterating through each entry in the given data dictionary
                for i in range(len(given_data['ACE Case ID'])):
                    row_entry = {
                        'S_No': i + 1,
                        'ACE_Case_ID': given_data['ACE Case ID'][i],
                        'Party_ID': given_data['Party ID'][i],
                        'Party_Name': given_data['Party name'][i],
                        'Stock': given_data['Stock'][i],
                        'Stock_margin': given_data['Stock margin'][i],
                        'Raw_materials': given_data['Raw materials'][i],
                        'RM_Margin': given_data['RM Margin'][i],
                        'Work_in_progress': given_data['Work in progress'][i],
                        'WIP_Margin': given_data['WIP Margin'][i],
                        'Finished_Goods': given_data['Finished Goods'][i],
                        'FG_Margin': given_data['FG Margin'][i],
                        'stores_and_spares': given_data['Stores and spares'][i],
                        'Stores_and_spares_margin': given_data['Stores and spares margin'][i],
                        'Book_Debts_Receviables_Domestic_Export': given_data['Book Debts & Receivables (Domestic&Export)'][i],
                        'Book_debts_Margin': given_data['Book debts Margin'][i],
                        'Upto_90_days': given_data['Upto 90 days'][i],
                        'Upto_90_days_Debts_margin': given_data['Upto 90 days-Debts margin'][i],
                        'upto_120days': given_data['Upto 120 days'][i],
                        'upto_120days_Debts_margin': given_data['Upto 120 days-Debts margin'][i],
                        'Creditors': given_data['Creditors'][i],
                        'Credtiors_margin': given_data['Creditors margin'][i],
                        'Advance_from_suppliers': given_data['Advance from suppliers'][i],
                        'Advance_to_suppliers': given_data['Advance to suppliers'][i],
                        'Stock_in_transit': given_data['Stock in transit'][i],
                        'Group_company_debtors': given_data['Group company debtors'][i],
                        'GST_receivables_tax_components': given_data['GST receivables (tax components)'][i],
                        'Unbilled_debtors': given_data['Unbilled debtors'][i],
                        'Bullion_stock': given_data['Bullion stock'][i],
                        'Obselete_stock': given_data['Obsolete stock'][i],
                        "under_deposits_Bhishis_Schemes": given_data["Under deposits/Bhishi's Scheme's"][i],
                        'Standard_Gold': given_data['Standard Gold'][i],
                        'WCL_sanctioned_in_CAM': given_data['WCL sanctioned in CAM'][i],
                        'WCL_sanctioned_in_Limit_module': given_data['WCL sanctioned in Limit module'][i],
                        'Unsecured_utilisations': given_data['Unsecured utilizations'][i],
                        'Total_utiliisations': given_data['Total utilizations'][i],
                        'Drawing_power_Amount': given_data['Drawing power Amount'][i],
                        'Remarks': given_data['Remarks'][i],
                        
                        'User_name': given_data['User name'][i]
                    }
                    final_data['row_data'].append(row_entry)
            if 'rejected' in reports_name.lower():
                file='rejected'
                query_case_ids=f"""SELECT PROCESS_QUEUE.CASE_ID 
                        FROM hdfc_queues.PROCESS_QUEUE 
                        INNER JOIN HDFC_QUEUES.QUEUE_LIST ON PROCESS_QUEUE.CASE_ID = QUEUE_LIST.CASE_ID 
                        WHERE QUEUE_LIST.QUEUE = 'rejected_queue' 
                        AND PROCESS_QUEUE.LAST_UPDATED >= TO_DATE('{start_date}', 'YYYY-MM-DD HH24:MI:SS') 
                        AND PROCESS_QUEUE.LAST_UPDATED <= TO_DATE('{end_date}', 'YYYY-MM-DD HH24:MI:SS')"""
                
                case_ids = queues_db.execute_(query_case_ids).to_dict(orient="records")
                case_id_db_data = queues_db.execute_(query_case_ids)
                columns = [ "ACE Case ID", "Party ID", "Party name", "User name","Reject reason","Rejected Comments", "Stock", "Stock margin", "Raw materials", "RM Margin", "Work in progress", "WIP Margin", "Finished Goods", "FG Margin", "Stores and spares", "Stores and spares margin", "Book Debts & Receivables (Domestic&Export)", "Book debts Margin", "Upto 90 days", "Upto 90 days-Debts margin", "Upto 120 days", "Upto 120 days-Debts margin", "Creditors", "Creditors margin", "Advance from suppliers", "Advance to suppliers", "Stock in transit", "Group company debtors", "GST receivables (tax components)", "Unbilled debtors", "Bullion stock", "Obsolete stock", "Under deposits/Bhishi's Scheme's", "Standard Gold", "WCL sanctioned in CAM", "WCL sanctioned in Limit module", "Unsecured utilizations", "Total utilizations", "Drawing power Amount", "Remarks" ]
                queue = 'rejected_queue'
                res_columns = pd.DataFrame(columns=columns)
                res=final_result_fun(case_ids,res_columns,queue,case_id_db_data)
                given_data = res.to_dict()
                final_data = {'row_data': []}
                res = res.drop(columns=['REJECTED_COMMENTS'])

                # Iterating through each entry in the given data dictionary
                for i in range(len(given_data['ACE Case ID'])):
                    
                    row_entry = {
                        'S_No': i + 1,
                        'ACE_Case_ID': given_data['ACE Case ID'][i],
                        'Party_ID': given_data['Party ID'][i],
                        'Party_Name': given_data['Party name'][i],
                        'Reject_reason': given_data['Reject reason'][i],
                        'Rejected_Comments': given_data['Rejected Comments'][i],
                        'Stock': given_data['Stock'][i],
                        'Stock_margin': given_data['Stock margin'][i],
                        'Raw_materials': given_data['Raw materials'][i],
                        'RM_Margin': given_data['RM Margin'][i],
                        'Work_in_progress': given_data['Work in progress'][i],
                        'WIP_Margin': given_data['WIP Margin'][i],
                        'Finished_Goods': given_data['Finished Goods'][i],
                        'FG_Margin': given_data['FG Margin'][i],
                        'stores_and_spares': given_data['Stores and spares'][i],
                        'Stores_and_spares_margin': given_data['Stores and spares margin'][i],
                        'Book_Debts_Receviables_Domestic_Export': given_data['Book Debts & Receivables (Domestic&Export)'][i],
                        'Book_debts_Margin': given_data['Book debts Margin'][i],
                        'Upto_90_days': given_data['Upto 90 days'][i],
                        'Upto_90_days_Debts_margin': given_data['Upto 90 days-Debts margin'][i],
                        'upto_120days': given_data['Upto 120 days'][i],
                        'upto_120days_Debts_margin': given_data['Upto 120 days-Debts margin'][i],
                        'Creditors': given_data['Creditors'][i],
                        'Credtiors_margin': given_data['Creditors margin'][i],
                        'Advance_from_suppliers': given_data['Advance from suppliers'][i],
                        'Advance_to_suppliers': given_data['Advance to suppliers'][i],
                        'Stock_in_transit': given_data['Stock in transit'][i],
                        'Group_company_debtors': given_data['Group company debtors'][i],
                        'GST_receivables_tax_components': given_data['GST receivables (tax components)'][i],
                        'Unbilled_debtors': given_data['Unbilled debtors'][i],
                        'Bullion_stock': given_data['Bullion stock'][i],
                        'Obselete_stock': given_data['Obsolete stock'][i],
                        "under_deposits_Bhishis_Schemes": given_data["Under deposits/Bhishi's Scheme's"][i],
                        'Standard_Gold': given_data['Standard Gold'][i],
                        'WCL_sanctioned_in_CAM': given_data['WCL sanctioned in CAM'][i],
                        'WCL_sanctioned_in_Limit_module': given_data['WCL sanctioned in Limit module'][i],
                        'Unsecured_utilisations': given_data['Unsecured utilizations'][i],
                        'Total_utiliisations': given_data['Total utilizations'][i],
                        'Drawing_power_Amount': given_data['Drawing power Amount'][i],
                        'Remarks': given_data['Remarks'][i],
                        
                        'User_name': given_data['User name'][i]
                    }
                    final_data['row_data'].append(row_entry)
            
            if 'onhold' in reports_name.lower():
                logging.info(f"entered hold report block")
                file='onhold'
                # query_case_ids=f"""SELECT OCR.CASE_ID FROM hdfc_extraction.OCR INNER JOIN hdfc_queues.QUEUE_LIST ON OCR.CASE_ID = QUEUE_LIST.CASE_ID WHERE OCR.STATUS like 'hold' AND OCR.CREATED_DATE >= TO_DATE('{start_date}', 'YYYY-MM-DD HH24:MI:SS') AND OCR.CREATED_DATE <= TO_DATE('{end_date}', 'YYYY-MM-DD HH24:MI:SS')
                # """
                logging.info(f"#####start_date{start_date}")
                logging.info(f"#####end_date{end_date}")
                
                query_case_ids=f"""SELECT PROCESS_QUEUE.CASE_ID 
                        FROM hdfc_queues.PROCESS_QUEUE 
                        INNER JOIN HDFC_QUEUES.QUEUE_LIST ON PROCESS_QUEUE.CASE_ID = QUEUE_LIST.CASE_ID 
                        WHERE PROCESS_QUEUE.STATUS = 'Hold'
                        AND PROCESS_QUEUE.LAST_UPDATED >= TO_DATE('{start_date}', 'YYYY-MM-DD HH24:MI:SS') 
                        AND PROCESS_QUEUE.LAST_UPDATED <= TO_DATE('{end_date}', 'YYYY-MM-DD HH24:MI:SS')"""
                
                
             
                logging.info(f"#####query_case_ids{query_case_ids}")
                case_ids = queues_db.execute_(query_case_ids).to_dict(orient="records")
                case_id_db_data = queues_db.execute_(query_case_ids)
                logging.info(f"#####case_ids {case_ids}")
                logging.info(f"#####case_id_db_data {case_id_db_data}")
                columns = [ "ACE Case ID", "Party ID", "Party name", "User name", "Hold reason","Hold Comments","Stock", "Stock margin", "Raw materials", "RM Margin", "Work in progress", "WIP Margin", "Finished Goods", "FG Margin", "Stores and spares", "Stores and spares margin", "Book Debts & Receivables (Domestic&Export)", "Book debts Margin", "Upto 90 days", "Upto 90 days-Debts margin", "Upto 120 days", "Upto 120 days-Debts margin", "Creditors", "Creditors margin", "Advance from suppliers", "Advance to suppliers", "Stock in transit", "Group company debtors", "GST receivables (tax components)", "Unbilled debtors", "Bullion stock", "Obsolete stock", "Under deposits/Bhishi's Scheme's", "Standard Gold", "WCL sanctioned in CAM", "WCL sanctioned in Limit module", "Unsecured utilizations", "Total utilizations", "Drawing power Amount", "Remarks" ]
                queue = 'Maker'
                res_columns = pd.DataFrame(columns=columns)
                res=final_result_fun(case_ids,res_columns,queue,case_id_db_data)
                logging.debug(f"response of res{res}")
                given_data = res.to_dict()
                final_data = {'row_data': []}
                res = res.drop(columns=['COMMENTS','Hold Comments'])

                # Iterating through each entry in the given data dictionary
                for i in range(len(given_data['ACE Case ID'])):
                    row_entry = {
                        
                        'S_No': i + 1,
                        'ACE_Case_ID': given_data['ACE Case ID'][i],
                        'Party_ID': given_data['Party ID'][i],
                        'Party_Name': given_data['Party name'][i],
                        'Hold_reason': given_data['Hold reason'][i],
                        'Hold_Comments': given_data['Hold Comments'][i],
                        'Stock': given_data['Stock'][i],
                        'Stock_margin': given_data['Stock margin'][i],
                        'Raw_materials': given_data['Raw materials'][i],
                        'RM_Margin': given_data['RM Margin'][i],
                        'Work_in_progress': given_data['Work in progress'][i],
                        'WIP_Margin': given_data['WIP Margin'][i],
                        'Finished_Goods': given_data['Finished Goods'][i],
                        'FG_Margin': given_data['FG Margin'][i],
                        'stores_and_spares': given_data['Stores and spares'][i],
                        'Stores_and_spares_margin': given_data['Stores and spares margin'][i],
                        'Book_Debts_Receviables_Domestic_Export': given_data['Book Debts & Receivables (Domestic&Export)'][i],
                        'Book_debts_Margin': given_data['Book debts Margin'][i],
                        'Upto_90_days': given_data['Upto 90 days'][i],
                        'Upto_90_days_Debts_margin': given_data['Upto 90 days-Debts margin'][i],
                        'upto_120days': given_data['Upto 120 days'][i],
                        'upto_120days_Debts_margin': given_data['Upto 120 days-Debts margin'][i],
                        'Creditors': given_data['Creditors'][i],
                        'Credtiors_margin': given_data['Creditors margin'][i],
                        'Advance_from_suppliers': given_data['Advance from suppliers'][i],
                        'Advance_to_suppliers': given_data['Advance to suppliers'][i],
                        'Stock_in_transit': given_data['Stock in transit'][i],
                        'Group_company_debtors': given_data['Group company debtors'][i],
                        'GST_receivables_tax_components': given_data['GST receivables (tax components)'][i],
                        'Unbilled_debtors': given_data['Unbilled debtors'][i],
                        'Bullion_stock': given_data['Bullion stock'][i],
                        'Obselete_stock': given_data['Obsolete stock'][i],
                        "under_deposits_Bhishis_Schemes": given_data["Under deposits/Bhishi's Scheme's"][i],
                        'Standard_Gold': given_data['Standard Gold'][i],
                        'WCL_sanctioned_in_CAM': given_data['WCL sanctioned in CAM'][i],
                        'WCL_sanctioned_in_Limit_module': given_data['WCL sanctioned in Limit module'][i],
                        'Unsecured_utilisations': given_data['Unsecured utilizations'][i],
                        'Total_utiliisations': given_data['Total utilizations'][i],
                        'Drawing_power_Amount': given_data['Drawing power Amount'][i],
                        'Remarks': given_data['Remarks'][i],
                        
                        'User_name': given_data['User name'][i]
                    }
                    final_data['row_data'].append(row_entry)
                    res.rename(columns={
                    'comments': 'Hold Comments',
                }, inplace=True)
                
                
            res = res.drop(columns=['CASE_ID','PARTY_ID','PARTY_NAME','CUSTOMER_NAME','DRAWING_POWER'])
        
    

            # Assuming df is your dataframe
            columns_to_drop = [
            "Total Stock", "Obsolete Stock", "Pledge Stock", "Consumable & Spares",
            "Goods in Transit", "PCFC Stock", "Inventory Financed Stock", "Domestic Stock",
            "Export Stock", "Indigenous Stock", "Creditors Domestic", "Creditors Exports",
            "Creditors PC", "Unpaid Stocks", "GSS (Gold Saving Scheme)", "DALC", "LC Creditors",
            "Stock under LC", "Buyers Credit Outstanding", "Margin Money", "Fixed Deposit",
            "Trade Creditors", "Group Company Creditors", "Receivables", "Sales", "Total Debtors",
            "Debtors <30 days", "Debtors <60 days", "Debtors <90 days", "Debtors <120 days",
            "Debtors <150 days", "Debtors <180 days", "Debtors - Exports", "Debtors - Domestic",
            "Bill Discounted debtors", "Debtors of Group Companies", "Unbilled Debtors",
            "Domestic receivables", "Export receivables", "Gst Receivables", 
            "Customer Duty Refundable", "Duty Drawback", "Government Receivables",
            "Unbilled Revenue", "Retention Money", "Unbilled Receivables",
            "Buyers Credit Outstabding", "WIP Margi","Obselete stock",	"Pledge stock",	"goods in transist",
            "Inventory Financed stock","Indegenious Stock","GSS(gold saving scheme)","Debtiors of Group Companies",	"debtors_>180",	"Purchases"]

            # Drop the specified columns
            columns_to_drop_existing = [col for col in columns_to_drop if col in res.columns]

            # Drop the specified columns
            res = res.drop(columns=columns_to_drop_existing)
            
            logging.info(f'{timestamp},{timestamp1}---namesssss')
            parsed_date = datetime.strptime(timestamp, "%Y-%m-%d %H:%M:%S")
            formatted_date = parsed_date.strftime("%m-%d-%Y_%H_%M")
            filename = f'{file}_report--{formatted_date}.xlsx'
            path=f'/var/www/reports_api/reports/{filename}'
            try:
                res.to_excel(path, index=False)
            except Exception as e:
                logging.info(f"######in exception for excel conversion error is {e}")
        
            try:

                return_json_data = {}
                return_json_data['message']='Successfully generated the report'
                return_json_data['excel_flag']= 1
                return_json_data['flag'] = True
                return_json_data['data'] = final_data
                
                data['report_data'] = return_json_data
                logging.info(f'{return_json_data}###############return_json_data')
                
            except Exception as e:
                logging.info(f"error at process_Report {e}")
                logging.debug(f"{e} ####issue")
                return_json_data = {}
                return_json_data['flag'] = False
                return_json_data['message'] = 'Failed!!'
                #return jsonify(return_json_data)

        except Exception as e:
            logging.exception(f'Something went wrong exporting data : {e}')
            return {'flag': False, 'message': 'Unable to export data.'}
    try:
        memory_after = measure_memory_usage()
        memory_consumed = (memory_after - memory_before) / \
            (1024 * 1024 * 1024)
        end_time = tt()
        memory_consumed = f"{memory_consumed:.10f}"
        logging.info(f"checkpoint memory_after - {memory_after},memory_consumed - {memory_consumed}, end_time - {end_time}")
        time_consumed = str(round(end_time-start_time,3))
    except:
        logging.warning("Failed to calc end of ram and time")
        logging.exception("ram calc went wrong")
        memory_consumed = None
        time_consumed = None
        pass

    # insert audit
    try:
        audit_data = {"tenant_id": tenant_id, "user_": "", "case_id": "",
                        "api_service": "consolidated_report", "service_container": "reportsapi",
                        "changed_data": "New file received","tables_involved": "","memory_usage_gb": str(memory_consumed), 
                        "time_consumed_secs": time_consumed, "request_payload": json.dumps(data), 
                        "response_data": json.dumps(return_json_data), "trace_id": trace_id,
                        "session_id": "","status":json.dumps(return_json_data['flag'])}
        
        insert_into_audit(audit_data)
    except:
        logging.info(f"issue in the query formation")
    return jsonify(return_json_data)



@app.route('/process_report_agri', methods=['POST', 'GET'])
def process_report_agri():
    logging.info(f'Calling agri process report')
    try:
        memory_before = measure_memory_usage()
        start_time = tt()
    except:
        logging.warning("Failed to start ram and time calc")
        pass
        
    data = request.json
    trace_id = generate_random_64bit_string()
    tenant_id = data['ui_data']['tenant_id']

    attr = ZipkinAttrs(
        trace_id=generate_random_64bit_string(),
        span_id=generate_random_64bit_string(),
        parent_span_id=None,
        flags=None,
        is_sampled=False,
        tenant_id=tenant_id if tenant_id is not None else ''
    )

    with zipkin_span(
            service_name='folder_monitor',
            zipkin_attrs=attr,
            span_name='folder_monitor',
            transport_handler=http_transport,
            sample_rate=0.5
    ) as zipkin_context:
        data = request.json
        tenant_id = data['ui_data']['tenant_id']
        db_config['tenant_id'] = tenant_id
        logging.info(f"Request Data is: {data}")
        queue_db = DB('queues', **db_config)
        queue_id = data.get("queue_id", {})
        case_id = data.get('REFERENCE_ID', None)
        extraction_db = DB('extraction', **db_config)
        queues_db = DB('queues', **db_config)
        group_db = DB("group_access", **db_config)
        user = data['ui_data']['REQUESTED_BY']
        start_date = data['start_date']
        end_date = data['end_date']
        # segment= data['segment']
        ist = pytz.timezone("Asia/Calcutta")
        timestamp = datetime.now(ist)
        timestamp_actual = timestamp
        timestamp1 = timestamp.strftime(f'%d-%m-%Y %H:%M:%S')

        timestamp = str(timestamp)[:-13]
        try:
            # query_case_ids=f"""SELECT OCR.CASE_ID FROM hdfc_extraction.OCR INNER JOIN hdfc_queues.QUEUE_LIST ON OCR.CASE_ID = QUEUE_LIST.CASE_ID WHERE (hdfc_queues.QUEUE_LIST.queue LIKE '%accepted_queue%' or hdfc_queues.QUEUE_LIST.queue LIKE '%maker%' )  AND OCR.CREATED_DATE >= TO_DATE('{start_date}', 'YYYY-MM-DD HH24:MI:SS') AND OCR.CREATED_DATE <= TO_DATE('{end_date}', 'YYYY-MM-DD HH24:MI:SS')"""
            query_case_ids  =f"""SELECT PROCESS_QUEUE.CASE_ID 
                    FROM hdfc_queues.PROCESS_QUEUE 
                    INNER JOIN HDFC_QUEUES.QUEUE_LIST ON PROCESS_QUEUE.CASE_ID = QUEUE_LIST.CASE_ID 
                    WHERE (QUEUE_LIST.QUEUE = 'accepted_queue' OR QUEUE_LIST.QUEUE LIKE '%maker%')  
                    AND PROCESS_QUEUE.CREATED_DATE >= TO_DATE('{start_date}', 'YYYY-MM-DD HH24:MI:SS') 
                    AND PROCESS_QUEUE.CREATED_DATE <= TO_DATE('{end_date}', 'YYYY-MM-DD HH24:MI:SS')"""
            case_ids=queues_db.execute_(query_case_ids).to_dict(orient="records")

            columns = ["ACE Case ID", "Party ID", "Party name", "Rules Accuracy","Signature and Rubber stamp", "Stock statement month", "DP Share Margin -CAM", "DP Share Margin -Credit noting", "DP Share Margin -DP letter", "Stock", "stock margin", "Raw materials", "RM Margin", "Work in progress", "WIP Margin", "Finished Goods", "FG Margin", "stores and spares", "Stores and spares margin", "Book Debts & Receviables (Domestic&Export)", "Book debts Margin", "Upto 90 days", "Upto 90 days-Debts margin", "upto 120days", "upto 120days-Debts margin", "Creditors", "Creditors margin", "Advance from suppliers", "Advance to suppliers", "Stock in transit", "Group company debtors", "GST receivables (tax components)", "Unbilled debtors", "Bullion stock", "Obselete stock", "under deposits/Bhishi's Scheme's", "Standard Gold", "WCL sanctioned in CAM", "WCL sanctioned in Limit module", "Unsecured utilisations", "Total utiliisations", "Drawing power Amount", "Remarks" ]
            res = pd.DataFrame(columns=columns)


            for i in case_ids:
                case=i['case_id']
                query=f"""SELECT CASE_ID, PARTY_ID, PARTY_NAME, CUSTOMER_NAME FROM OCR where CASE_ID like '{case}'"""
                df=extraction_db.execute_(query)
                logging.info(f"dfresult {df}")
                logging.info(f"dfresult {df.columns}")

                df.rename(columns={
                    'CASE_ID': 'ACE Case ID',
                    'PARTY_ID': 'Party ID',
                    'PARTY_NAME': 'Party name'
                    
                    
                }, inplace=True)
                df=df.to_dict(orient="records")
                df = pd.DataFrame(df)
                df.drop(['case_id', 'party_id', 'party_name'], axis=1, inplace=True)
                logging.info(f"{df.columns}###renamed df columns")
                qry=f"SELECT STOCKS,CREDITORS,DEBTORS FROM OCR WHERE  CASE_ID='{case}'" 
                
                json_data_list = extraction_db.execute_(qry).to_dict(orient="records")
                logging.info(f"The json data list is :{json_data_list}")
                def replace_null_with_none(value):
                        return None if value == "null" or value == '' else value

                # Iterate through each dictionary and perform the replacement
                for data in json_data_list:
                    for key, value in data.items():
                        if isinstance(value, str):
                            data[key] = replace_null_with_none(value)
                extraction_data=create_dataframe_from_json(json_data_list)

                values = extraction_data.to_dict()
                final_result={}
                for i in values:
                    final_result[i]=values[i][0]
                query = f"SELECT PARTY_ID from OCR where CASE_ID = '{case}'"
                Party = extraction_db.execute_(query)

                if not Party.empty:
                    query = f"SELECT COMPONENT_NAME, MARGIN from AGE_MARGIN_WORKING_UAT where PARTY_ID = '{Party['party_id'][0]}'"
                    margins = extraction_db.execute_(query)
                    try:
                        rm_margin = margins.loc[margins['component_name'] == 'RAW MATERIALS INSURED', 'margin'].values[0] if not margins.loc[margins['component_name'] == 'RAW MATERIALS INSURED', 'margin'].empty else None
                        wip_margin = margins.loc[margins['component_name'] == 'WORK IN PROGRESS INSURED', 'margin'].values[0] if not margins.loc[margins['component_name'] == 'WORK IN PROGRESS INSURED', 'margin'].empty else None
                        fg_margin = margins.loc[margins['component_name'] == 'FINISHED GOODS INSURED', 'margin'].values[0] if not margins.loc[margins['component_name'] == 'FINISHED GOODS INSURED', 'margin'].empty else None
                        stores_and_spares_margin = margins.loc[margins['component_name'] == 'STOCK & STORES INSURED', 'margin'].values[0] if not margins.loc[margins['component_name'] == 'STOCK & STORES INSURED', 'margin'].empty else None
                        book_debts_margin = margins.loc[margins['component_name'] == 'BOOK DEBTS', 'margin'].values[0] if not margins.loc[margins['component_name'] == 'BOOK DEBTS', 'margin'].empty else None
                        upto_90_days_debts_margin = margins.loc[margins['component_name'] == 'BOOK DEBTS UPTO 90 DAYS', 'margin'].values[0] if not margins.loc[margins['component_name'] == 'BOOK DEBTS UPTO 90 DAYS', 'margin'].empty else None
                        upto_120_days_debts_margin = margins.loc[margins['component_name'] == 'BOOK DEBTS UPTO 120 DAYS', 'margin'].values[0] if not margins.loc[margins['component_name'] == 'BOOK DEBTS UPTO 120 DAYS', 'margin'].empty else None
                        creditors_margin = margins.loc[margins['component_name'] == 'CREDITORS', 'margin'].values[0] if not margins.loc[margins['component_name'] == 'CREDITORS', 'margin'].empty else None

                        
                        final_result['RM Margin'] = rm_margin
                        final_result['WIP Margi'] = wip_margin
                        final_result['FG Margin'] = fg_margin
                        final_result['Stores and spares margin'] = stores_and_spares_margin
                        final_result['Book debts Margin'] = book_debts_margin
                        final_result['Upto 90 days-Debts margin'] = upto_90_days_debts_margin
                        final_result['Upto 120 days-Debts margin'] = upto_120_days_debts_margin
                        final_result['Creditors margin'] = creditors_margin
                    except Exception as e:
                        logging.info(f"####exception for margins###{e}")
                final_result = pd.DataFrame([final_result])
                renamed_columns = {
                    'Drawing Power': 'Drawing power Amount',
                    'Raw Materials': 'Raw materials',
                    'Total Creditors': 'Creditors',
                    'Work in Process': 'Work in progress',
                    'Stock in Transit': 'Stock in transit',
                    'Upto 120 days-Debts margin':'upto 120days-Debts margin',
                }

                # Check if 'stores and spares' or 'Stores & Spares' exists and rename accordingly
                if 'stores and spares' in final_result.columns:
                    renamed_columns['stores and spares'] = 'stores and spares'
                elif 'Stores  & Spares' in final_result.columns:
                    renamed_columns['Stores  & Spares'] = 'stores and spares'

                # Rename columns
                final_result.rename(columns=renamed_columns, inplace=True)


                final_df = pd.concat([final_result, df], axis=1)

            # #     
                res = res.append(final_df, ignore_index=True)
            columns_to_drop = ['User name','customer_name','CUSTOMER_NAME',

            "Total Stock", "Obsolete Stock", "Pledge Stock", "Consumable & Spares",
            "Goods in Transit", "PCFC Stock", "Inventory Financed Stock", "Domestic Stock",
            "Export Stock", "Indigenous Stock", "Creditors Domestic", "Creditors Exports",
            "Creditors PC", "Unpaid Stocks", "GSS (Gold Saving Scheme)", "DALC", "LC Creditors",
            "Stock under LC", "Buyers Credit Outstanding", "Margin Money", "Fixed Deposit",
            "Trade Creditors", "Group Company Creditors", "Receivables", "Sales", "Total Debtors",
            "Debtors <30 days", "Debtors <60 days", "Debtors <90 days", "Debtors <120 days",
            "Debtors <150 days", "Debtors <180 days", "Debtors - Exports", "Debtors - Domestic",
            "Bill Discounted debtors", "Debtors of Group Companies", "Unbilled Debtors",
            "Domestic receivables", "Export receivables", "Gst Receivables", 
            "Customer Duty Refundable", "Duty Drawback", "Government Receivables",
            "Unbilled Revenue", "Retention Money", "Unbilled Receivables",
            "Buyers Credit Outstabding", "WIP Margi",	"Pledge stock",	"goods in transist",
            "Inventory Financed stock","Indegenious Stock","GSS(gold saving scheme)","Debtiors of Group Companies",	"debtors_>180",	"Purchases"]

            # Drop the specified columns
            columns_to_drop_existing = [col for col in columns_to_drop if col in res.columns]

# Drop the specified columns
            res = res.drop(columns=columns_to_drop_existing)
            # new_row = {" Signature and Rubber stamp":"Source: "," Stock statement month ":"Source: "," DP Share Margin -CAM":"Source: "," DP Share Margin -Credit noting":"Source:"," DP Share Margin -DP letter":"Source: "," Stock ":"Source: Stock Statement "," stock margin":"Source: ","Raw materials":"Source: Stock Statement " ,"RM Margin":"Source:",  "Work in progress":"Source: Stock Statement ", "WIP Margin":"Source: ", "Finished Goods":"Source: Stock Statement ", "FG Margin":"Source: ", "stores and spares":"Source: Stock Statement ", "Stores and spares margin":"Source: ", "Book Debts & Receviables (Domestic&Export)": "Source: Stock Statement ", "Book debts Margin":"Source: ", "Upto 90 days":"Source: Stock Statement ", "Upto 90 days-Debts margin":"Source: Stock Statement ", "upto 120days":"Source: Stock Statement ", "upto 120days-Debts margin":"Source: ", "Creditors":"Source: Stock Statement ", "Credtiors margin":"Source: ", "Advance from suppliers":"Source: Stock Statement ", "Advance to suppliers":"Source: Stock Statement ", "Stock in transit":"Source: Stock Statement ", "Group company debtors":"Source: Stock Statement ", "GST receivables (tax components)": "Source: Stock Statement ", "Unbilled debtors":"Source: Stock Statement ", "Bullion stock":"Source: Stock Statement ", "Obselete stock":"Source: Stock Statement ", "under deposits/Bhishi's Scheme's":"Source: Stock Statement ", "Standard Gold":"Source: Stock Statement ", "WCL sanctioned in CAM":"Source: Stock Statement ", "WCL sanctioned in Limit module":"Source: Stock Statement ", "Unsecured utilisations":"Source: Stock Statement ", "Total utiliisations":"Source: Stock Statement ", "Drawing power Amount":"Source: Stock Statement ", "Remarks":"Source: " }                #Use the loc method to insert the new row at index 0
            new_row = {" Signature and Rubber stamp":"Source: "," Stock statement month ":"Source: "," DP Share Margin -CAM":"Source:Master "," DP Share Margin -Credit noting":"Source:Master"," DP Share Margin -DP letter":"Source:Master"," Stock ":"Source: Stock Statement "," stock margin":"Source:Master","Raw materials":"Source: Stock Statement " ,"RM Margin":"Source:Master",  "Work in progress":"Source: Stock Statement ", "WIP Margin":"Source:Master", "Finished Goods":"Source: Stock Statement ", "FG Margin":"Source:Master", "stores and spares":"Source: Stock Statement ", "Stores and spares margin":"Source:Master", "Book Debts & Receviables (Domestic&Export)": "Source: Stock Statement ", "Book debts Margin":"Source:Master", "Upto 90 days":"Source: Stock Statement ", "Upto 90 days-Debts margin":"Source: Master ", "upto 120days":"Source: Stock Statement ", "upto 120days-Debts margin":"Source:Master", "Creditors":"Source: Stock Statement ", "Credtiors margin":"Source:Master", "Advance from suppliers":"Source: Stock Statement ", "Advance to suppliers":"Source: Stock Statement ", "Stock in transit":"Source: Stock Statement ", "Group company debtors":"Source: Stock Statement ", "GST receivables (tax components)": "Source: Stock Statement ", "Unbilled debtors":"Source: Stock Statement ", "Bullion stock":"Source: Stock Statement ", "Obselete stock":"Source: Stock Statement ", "under deposits/Bhishi's Scheme's":"Source: Stock Statement ", "Standard Gold":"Source: Stock Statement ", "WCL sanctioned in CAM":"Source: Stock Statement ", "WCL sanctioned in Limit module":"Source: Stock Statement ", "Unsecured utilisations":"Source: Stock Statement ", "Total utiliisations":"Source: Stock Statement ", "Drawing power Amount":"Source: Stock Statement ", "Remarks":"Source: " }
            res.loc[-1] = new_row
            res.index = res.index + 1
            res = res.sort_index()
            parsed_date = datetime.strptime(timestamp, "%Y-%m-%d %H:%M:%S")
            formatted_date = parsed_date.strftime("%m-%d-%Y_%H_%M")
            filename = f'Process_Report--{formatted_date}.xlsx'
            path = f'/var/www/reports_api/reports/{filename}'
            try:
                # logging.info(f'inside try {res}')
                writer = pd.ExcelWriter(path, engine='xlsxwriter')
                res.to_excel(writer, sheet_name='Sheet1', index=False)
                # Get the xlsxwriter workbook and worksheet objects
                workbook  = writer.book
                worksheet = writer.sheets['Sheet1']

                # #Apply your desired styles to the worksheet
                # For example, setting the number format for all cells with numeric values
                format_numeric = workbook.add_format({'num_format': '0'})
                worksheet.set_column('A:Z', None, format_numeric)
                writer.save()
                # logging.info(f'inside after try {res}')
            except Exception as e:
                # logging.info(f'inside except {res}')
                logging.info(f"######in exception for excel conversion error is {e}")
            given_data=res.to_dict()
            final_data = {'row_data': []}
    

            # Iterating through each entry in the given data dictionary
            logging.info(f"#given data{given_data}")
            if given_data:
                for i in range(len(given_data['ACE Case ID'])):
                    row_entry = {
                    'S_No': i + 1,
                    'case_id': given_data['ACE Case ID'][i],
                    'Party_Id': given_data['Party ID'][i],
                    'Party_Name': given_data['Party name'][i],
                    'Rules_Accuracy': given_data['Rules Accuracy'][i],
                    'Signature_and_Rubber_stamp': given_data['Signature and Rubber stamp'][i],
                    'Stock_statement_month': given_data['Stock statement month'][i],
                    'DP_Share_Margin_CAM': given_data['DP Share Margin -CAM'][i],
                    'DP_Share_Margin_Credit_noting': given_data['DP Share Margin -Credit noting'][i],
                    'DP_Share_Margin_DP_letter': given_data['DP Share Margin -DP letter'][i],
                    'Stock': given_data['Stock'][i],
                    'stock_margin': given_data['stock margin'][i],
                    'Raw_materials': given_data['Raw materials'][i],
                    'RM_Margin': given_data['RM Margin'][i],
                    'Work_in_progress': given_data['Work in progress'][i],
                    'WIP_Margin': given_data['WIP Margin'][i],
                    'Finished_Goods': given_data['Finished Goods'][i],
                    'FG_Margin': given_data['FG Margin'][i],
                    'stores_and_spares': given_data['stores and spares'][i],
                    'Stores_and_spares_margin': given_data['Stores and spares margin'][i],
                    'Book_Debts_Receviables_Domestic_Export': given_data['Book Debts & Receviables (Domestic&Export)'][i],
                    'Book_debts_Margin': given_data['Book debts Margin'][i],
                    'Upto_90_days': given_data['Upto 90 days'][i],
                    'Upto_90_days_Debts_margin': given_data['Upto 90 days-Debts margin'][i],
                    'upto_120days': given_data['upto 120days'][i],
                    'upto_120days_Debts_margin': given_data['upto 120days-Debts margin'][i],
                    'Creditors': given_data['Creditors'][i],
                    'Credtiors_margin': given_data['Creditors margin'][i],
                    'Advance_from_suppliers': given_data['Advance from suppliers'][i],
                    'Advance_to_suppliers': given_data['Advance to suppliers'][i],
                    'Stock_in_transit': given_data['Stock in transit'][i],
                    'Group_company_debtors': given_data['Group company debtors'][i],
                    'GST_receivables_tax_components': given_data['GST receivables (tax components)'][i],
                    'Unbilled_debtors': given_data['Unbilled debtors'][i],
                    'Bullion_stock': given_data['Bullion stock'][i],
                    'Obselete_stock': given_data['Obselete stock'][i],
                    "under_deposits_Bhishis_Schemes": given_data["under deposits/Bhishi's Scheme's"][i],
                    'Standard_Gold': given_data['Standard Gold'][i],
                    'WCL_sanctioned_in_CAM': given_data['WCL sanctioned in CAM'][i],
                    'WCL_sanctioned_in_Limit_module': given_data['WCL sanctioned in Limit module'][i],
                    'Unsecured_utilisations': given_data['Unsecured utilisations'][i],
                    'Total_utiliisations': given_data['Total utiliisations'][i],
                    'Drawing_power_Amount': given_data['Drawing power Amount'][i],
                    'Remarks': given_data['Remarks'][i],
                    # 'Total_debtors': given_data['Total debtors'][i],
                    # 'User_name': given_data['User name'][i]
                }

                    final_data['row_data'].append(row_entry)
            try:

                return_json_data = {}
                return_json_data['message']='Successfully generated the report'
                return_json_data['excel_flag']= 1
                return_json_data['flag'] = True
                return_json_data['data'] = final_data
                
                data['report_data'] = return_json_data

                #generate_report({'tenant_id': tenant_id, **data})
                logging.info(f'{return_json_data}###############return_json_data')
                return jsonify(return_json_data)
            except Exception as e:
                logging.info(f"error at process_Report {e}")
                logging.debug(f"{e} ####issue")
                return_json_data = {}
                return_json_data['flag'] = False
                return_json_data['message'] = 'Failed!!'
                
        except Exception as e:
            logging.exception(f'Something went wrong exporting data : {e}')
            return {'flag': False, 'message': 'Unable to export data.'}
    try:
        memory_after = measure_memory_usage()
        memory_consumed = (memory_after - memory_before) / \
            (1024 * 1024 * 1024)
        end_time = tt()
        memory_consumed = f"{memory_consumed:.10f}"
        logging.info(f"checkpoint memory_after - {memory_after},memory_consumed - {memory_consumed}, end_time - {end_time}")
        time_consumed = str(round(end_time-start_time,3))
    except:
        logging.warning("Failed to calc end of ram and time")
        logging.exception("ram calc went wrong")
        memory_consumed = None
        time_consumed = None
        pass

    # insert audit
    try:
        audit_data = {"tenant_id": tenant_id, "user_": "", "case_id": "",
                    "api_service": "process_report_agri", "service_container": "reportsapi",
                    "changed_data": "New file received","tables_involved": "","memory_usage_gb": str(memory_consumed), 
                    "time_consumed_secs": time_consumed, "request_payload": json.dumps(data), 
                    "response_data": json.dumps(return_json_data), "trace_id": trace_id,
                    "session_id": "","status":json.dumps(return_json_data['flag'])}
    
        insert_into_audit(audit_data)
    except:
        logging.info(f"issue in the query formation")
    return jsonify(return_json_data)









